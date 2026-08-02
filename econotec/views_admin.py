"""
Vistas del Registro Administrativo: dashboard de egresos/ingresos del taller.
Solo accesible por administradores.
"""
from datetime import date, time, timedelta
from decimal import Decimal
from io import BytesIO
import json

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.db import transaction
from django.db.models import (
    Count, DecimalField, ExpressionWrapper, F, OuterRef, Prefetch, Q,
    Subquery, Sum, Value,
)
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST

from .forms import EgresoForm
from .gamificacion import (
    SALIDA_BUENA_ESTADOS,
    SALIDA_GARANTIA_ESTADOS,
    SALIDA_MALA_ESTADOS,
    calcular_puntaje_gamificacion,
)
from .busqueda import filtrar_objetos_normalizado, texto_salida_busqueda, total_resultados
from .bitacora import construir_bitacora_usuario
from .date_filters import aplicar_rango_fecha, contexto_rango_fecha, obtener_rango_fecha
from .models import (
    IngresoEquipo, SalidaEquipo, Abono, Egreso, CategoriaEgreso, Cliente,
    AvisoPanel, BitacoraTecnico, HorarioTecnico, InventarioItem,
    SEDES_EQUIPOS, VentaInventarioItem,
)
from .pagination import paginar_resultados
from .permisos import GRUPOS_TECNICO, admin_requerido, es_admin, tecnico_requerido


MESES_ES = [
    '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
]

SALIDAS_POSITIVAS_ADMIN = SALIDA_BUENA_ESTADOS
SALIDAS_NEGATIVAS_ADMIN = SALIDA_MALA_ESTADOS


def _ingresos_dinero_mes(year, month):
    """Suma de dinero recibido en el mes (anticipos + abonos + cobros finales)."""
    anticipos = IngresoEquipo.objects.filter(
        fecha_ingreso__year=year, fecha_ingreso__month=month,
    ).aggregate(s=Sum('abono_anticipo'))['s'] or Decimal('0.00')

    diagnosticos_rapidos = IngresoEquipo.objects.filter(
        fecha_ingreso__year=year, fecha_ingreso__month=month,
        diagnostico_inmediato='si'
    ).aggregate(s=Sum('valor_diagnostico'))['s'] or Decimal('0.00')

    abonos = Abono.objects.filter(
        fecha__year=year, fecha__month=month,
    ).aggregate(s=Sum('monto'))['s'] or Decimal('0.00')

    salidas_mes = SalidaEquipo.objects.filter(
        fecha_salida__year=year, fecha_salida__month=month,
    )
    
    # Reparaciones (todo menos ventas)
    cobros_finales = salidas_mes.exclude(
        estado_reparacion='cliente_no_acepta'
    ).exclude(
        ingreso__sede='ventas'
    ).aggregate(s=Sum('valor_final_cobrado'))['s'] or Decimal('0.00')

    # Ventas de producto
    cobros_ventas = IngresoEquipo.objects.filter(
        fecha_ingreso__year=year, fecha_ingreso__month=month,
        sede='ventas'
    ).aggregate(s=Sum('valor_acordado'))['s'] or Decimal('0.00')

    cobros_diagnostico = salidas_mes.filter(
        estado_reparacion='cliente_no_acepta'
    ).aggregate(s=Sum('valor_final_cobrado'))['s'] or Decimal('0.00')

    return {
        'anticipos': anticipos,
        'diagnosticos_rapidos': diagnosticos_rapidos,
        'abonos': abonos,
        'cobros_finales': cobros_finales,
        'cobros_ventas': cobros_ventas,
        'cobros_diagnostico': cobros_diagnostico,
        'total': anticipos + diagnosticos_rapidos + abonos + cobros_finales + cobros_ventas + cobros_diagnostico,
    }


def _egresos_mes(year, month):
    return Egreso.objects.filter(
        fecha__year=year, fecha__month=month,
    ).aggregate(s=Sum('monto'))['s'] or Decimal('0.00')


def _nombre_usuario(user):
    return user.get_full_name() or user.username


def _horarios_tecnicos_dashboard():
    User = get_user_model()
    ahora = timezone.now()
    hoy = timezone.localdate(ahora)
    tecnicos = (
        User.objects
        .filter(is_active=True, groups__name__in=['Tecnicos', 'Tecnico'])
        .distinct()
        .order_by('first_name', 'username')
    )
    horarios = []
    avisos = []
    avisos_fuera = []
    for tecnico in tecnicos:
        horario, _ = HorarioTecnico.objects.get_or_create(tecnico=tecnico)
        ultimo = horario.ultima_notificacion_laboral
        aviso_hoy = bool(ultimo and timezone.localdate(ultimo) == hoy)
        ultimo_fuera = horario.ultima_notificacion_fuera_laboral
        aviso_fuera_hoy = bool(ultimo_fuera and timezone.localdate(ultimo_fuera) == hoy)
        item = {
            'user': tecnico,
            'nombre': _nombre_usuario(tecnico),
            'horario': horario,
            'dias': [
                (campo, label, getattr(horario, campo))
                for campo, label in HorarioTecnico.DIAS
            ],
            'es_dia_laboral': horario.es_dia_laboral(hoy),
            'esta_en_horario': horario.esta_en_horario(ahora),
            'aviso_hoy': aviso_hoy,
            'aviso_fuera_hoy': aviso_fuera_hoy,
        }
        horarios.append(item)
        if aviso_hoy:
            avisos.append({
                'user': tecnico,
                'nombre': item['nombre'],
                'momento': ultimo,
                'horario': horario,
            })
        if aviso_fuera_hoy:
            motivo = (
                'fuera de su día laboral'
                if horario.ultima_notificacion_fuera_motivo == 'dia'
                else 'fuera de su horario laboral'
            )
            avisos_fuera.append({
                'user': tecnico,
                'nombre': item['nombre'],
                'momento': ultimo_fuera,
                'motivo': motivo,
                'horario': horario,
            })

    avisos.sort(key=lambda aviso: aviso['momento'], reverse=True)
    avisos_fuera.sort(key=lambda aviso: aviso['momento'], reverse=True)
    return horarios, avisos, avisos_fuera


def _bitacoras_tecnicos_admin(dia):
    User = get_user_model()
    tecnicos = (
        User.objects
        .filter(is_active=True, groups__name__in=GRUPOS_TECNICO)
        .distinct()
        .order_by('first_name', 'username')
    )

    bitacoras = []
    primera_con_datos = False
    for tecnico in tecnicos:
        bitacora = construir_bitacora_usuario(tecnico, dia=dia)
        abrir = bitacora['tiene_datos'] and not primera_con_datos
        if abrir:
            primera_con_datos = True
        bitacoras.append({
            'user': tecnico,
            'nombre': _nombre_usuario(tecnico),
            'bitacora': bitacora,
            'abrir': abrir,
        })
    return bitacoras


def _equipos_mes_resumen(year, month):
    equipos = (
        IngresoEquipo.objects
        .filter(sede__in=SEDES_EQUIPOS)
        .filter(
            Q(fecha_ingreso__year=year, fecha_ingreso__month=month) |
            Q(salida__fecha_salida__year=year, salida__fecha_salida__month=month)
        )
        .select_related('cliente', 'tecnico_encargado', 'salida', 'salida__tecnico_reparo')
        .distinct()
        .order_by('-fecha_ingreso', '-numero_equipo')
    )

    resumen = []
    for ingreso in equipos:
        try:
            salida = ingreso.salida
        except SalidaEquipo.DoesNotExist:
            salida = None

        ingresado_en_mes = ingreso.fecha_ingreso.year == year and ingreso.fecha_ingreso.month == month
        entregado_en_mes = bool(
            salida and
            salida.fecha_salida.year == year and
            salida.fecha_salida.month == month
        )

        resumen.append({
            'ingreso': ingreso,
            'salida': salida,
            'ingresado_en_mes': ingresado_en_mes,
            'entregado_en_mes': entregado_en_mes,
            'tecnico_ingreso_nombre': ingreso.tecnico_encargado_nombre,
        })

    return resumen


def _filtrar_resumen_por_tecnico(queryset, campo, tecnico_filtro):
    if tecnico_filtro == 'sin_asignar':
        return queryset.filter(**{f'{campo}__isnull': True})
    if tecnico_filtro.isdigit():
        return queryset.filter(**{f'{campo}_id': int(tecnico_filtro)})
    return queryset


def _ingresos_asignados_tecnicos_mes(year, month, tecnico_filtro=''):
    """Resume ingresos por técnico asignado, sin mezclar datos de la salida."""
    ingresos_qs = (
        IngresoEquipo.objects
        .filter(
            sede__in=SEDES_EQUIPOS,
            fecha_ingreso__year=year,
            fecha_ingreso__month=month,
        )
        .select_related('cliente', 'tecnico_encargado', 'salida')
    )
    ingresos_qs = _filtrar_resumen_por_tecnico(
        ingresos_qs,
        'tecnico_encargado',
        tecnico_filtro,
    )

    total = ingresos_qs.count()
    asignados = ingresos_qs.filter(tecnico_encargado__isnull=False).count()
    sin_asignar = total - asignados
    tecnicos = (
        ingresos_qs
        .exclude(tecnico_encargado__isnull=True)
        .values('tecnico_encargado_id')
        .distinct()
        .count()
    )

    resumen = list(
        ingresos_qs
        .order_by()
        .values(
            'tecnico_encargado_id',
            'tecnico_encargado__first_name',
            'tecnico_encargado__last_name',
            'tecnico_encargado__username',
        )
        .annotate(
            total=Count('id'),
            en_taller=Count('id', filter=Q(salida__isnull=True)),
            con_salida=Count('id', filter=Q(salida__isnull=False)),
        )
        .order_by('-total', 'tecnico_encargado__first_name', 'tecnico_encargado__username')
    )
    for posicion, fila in enumerate(resumen, start=1):
        nombre = 'Sin técnico asignado'
        if fila['tecnico_encargado_id']:
            nombre = (
                f"{fila['tecnico_encargado__first_name']} "
                f"{fila['tecnico_encargado__last_name']}"
            ).strip() or fila['tecnico_encargado__username']
        fila.update({
            'posicion': posicion,
            'tecnico_id': fila['tecnico_encargado_id'],
            'tecnico_nombre': nombre,
            'participacion': round((fila['total'] / total) * 100, 1) if total else 0,
        })

    registros = list(
        ingresos_qs.order_by(
            'tecnico_encargado__first_name',
            'tecnico_encargado__username',
            '-fecha_ingreso',
            '-numero_equipo',
        )
    )

    return {
        'resumen': resumen,
        'registros': registros,
        'total': total,
        'asignados': asignados,
        'sin_asignar': sin_asignar,
        'tecnicos': tecnicos,
        'fecha_desde': date(year, month, 1),
        'fecha_hasta': date(
            year + (month == 12),
            1 if month == 12 else month + 1,
            1,
        ) - timedelta(days=1),
    }


def _salidas_reparadas_tecnicos_mes(year, month, tecnico_filtro=''):
    """Resume salidas por quien terminó la reparación (`tecnico_reparo`)."""
    salidas_qs = (
        SalidaEquipo.objects
        .filter(
            fecha_salida__year=year,
            fecha_salida__month=month,
        )
        .select_related('ingreso', 'ingreso__cliente', 'tecnico_reparo')
    )
    salidas_qs = _filtrar_resumen_por_tecnico(
        salidas_qs,
        'tecnico_reparo',
        tecnico_filtro,
    )

    metricas = salidas_qs.aggregate(
        total=Count('id'),
        positivas=Count('id', filter=Q(estado_reparacion__in=SALIDAS_POSITIVAS_ADMIN)),
        negativas=Count('id', filter=Q(estado_reparacion__in=SALIDAS_NEGATIVAS_ADMIN)),
        recaudado=Sum('valor_final_cobrado'),
    )
    total = metricas['total'] or 0
    positivas = metricas['positivas'] or 0
    negativas = metricas['negativas'] or 0

    resumen = list(
        salidas_qs
        .order_by()
        .values(
            'tecnico_reparo_id',
            'tecnico_reparo__first_name',
            'tecnico_reparo__last_name',
            'tecnico_reparo__username',
        )
        .annotate(
            total=Count('id'),
            retirados=Count('id', filter=Q(estado_reparacion='retirado')),
            pendientes=Count('id', filter=Q(estado_reparacion='pendiente_retiro')),
            positivas=Count('id', filter=Q(estado_reparacion__in=SALIDAS_POSITIVAS_ADMIN)),
            negativas=Count('id', filter=Q(estado_reparacion__in=SALIDAS_NEGATIVAS_ADMIN)),
            recaudado=Sum('valor_final_cobrado'),
        )
        .order_by('-positivas', '-total', 'tecnico_reparo__first_name', 'tecnico_reparo__username')
    )
    for posicion, fila in enumerate(resumen, start=1):
        nombre = 'Sin técnico que reparó'
        if fila['tecnico_reparo_id']:
            nombre = (
                f"{fila['tecnico_reparo__first_name']} "
                f"{fila['tecnico_reparo__last_name']}"
            ).strip() or fila['tecnico_reparo__username']
        fila.update({
            'posicion': posicion,
            'tecnico_id': fila['tecnico_reparo_id'],
            'tecnico_nombre': nombre,
            'otras': fila['total'] - fila['positivas'] - fila['negativas'],
            'efectividad': round((fila['positivas'] / fila['total']) * 100, 1) if fila['total'] else 0,
            'recaudado': fila['recaudado'] or Decimal('0.00'),
        })

    registros = list(
        salidas_qs.order_by(
            'tecnico_reparo__first_name',
            'tecnico_reparo__username',
            '-fecha_salida',
            '-id',
        )
    )
    for salida in registros:
        if salida.estado_reparacion in SALIDAS_POSITIVAS_ADMIN:
            salida.clasificacion_admin = 'positive'
            salida.clasificacion_admin_nombre = 'Positiva'
        elif salida.estado_reparacion in SALIDAS_NEGATIVAS_ADMIN:
            salida.clasificacion_admin = 'negative'
            salida.clasificacion_admin_nombre = 'Negativa'
        else:
            salida.clasificacion_admin = 'neutral'
            salida.clasificacion_admin_nombre = 'Otra'

    return {
        'resumen': resumen,
        'registros': registros,
        'total': total,
        'positivas': positivas,
        'negativas': negativas,
        'otras': total - positivas - negativas,
        'efectividad': round((positivas / total) * 100, 1) if total else 0,
        'recaudado': metricas['recaudado'] or Decimal('0.00'),
    }


def _periodo_admin_request(request, metodo='GET'):
    datos = request.POST if metodo == 'POST' else request.GET
    hoy = date.today()
    try:
        year = int(datos.get('ano') or hoy.year)
        month = int(datos.get('mes') or hoy.month)
    except (TypeError, ValueError):
        year, month = hoy.year, hoy.month
    month = min(max(month, 1), 12)
    return year, month


@admin_requerido
def admin_dashboard(request):
    """Dashboard financiero mensual."""
    hoy = date.today()
    year = int(request.GET.get('ano') or hoy.year)
    month = int(request.GET.get('mes') or hoy.month)
    tecnico_resumen_filtro = (request.GET.get('tecnico_resumen') or '').strip()
    if tecnico_resumen_filtro != 'sin_asignar' and not tecnico_resumen_filtro.isdigit():
        tecnico_resumen_filtro = ''

    User = get_user_model()
    tecnico_ids_periodo = set(
        IngresoEquipo.objects.filter(
            sede__in=SEDES_EQUIPOS,
            fecha_ingreso__year=year,
            fecha_ingreso__month=month,
            tecnico_encargado__isnull=False,
        ).values_list('tecnico_encargado_id', flat=True)
    )
    tecnico_ids_periodo.update(
        SalidaEquipo.objects.filter(
            fecha_salida__year=year,
            fecha_salida__month=month,
            tecnico_reparo__isnull=False,
        ).values_list('tecnico_reparo_id', flat=True)
    )
    tecnicos_resumen = (
        User.objects
        .filter(
            Q(is_active=True, groups__name__in=GRUPOS_TECNICO)
            | Q(pk__in=tecnico_ids_periodo)
        )
        .distinct()
        .order_by('first_name', 'last_name', 'username')
    )
    tecnico_resumen_nombre = 'Todos los técnicos'
    if tecnico_resumen_filtro == 'sin_asignar':
        tecnico_resumen_nombre = 'Sin técnico asignado'
    elif tecnico_resumen_filtro.isdigit():
        tecnico_seleccionado = User.objects.filter(pk=int(tecnico_resumen_filtro)).first()
        if tecnico_seleccionado:
            tecnico_resumen_nombre = _nombre_usuario(tecnico_seleccionado)
        else:
            tecnico_resumen_filtro = ''

    dinero_in = _ingresos_dinero_mes(year, month)
    egresos_total = _egresos_mes(year, month)
    utilidad = dinero_in['total'] - egresos_total

    # Equipos del mes
    equipos_ingresados = IngresoEquipo.objects.filter(
        sede__in=SEDES_EQUIPOS,
        fecha_ingreso__year=year, fecha_ingreso__month=month,
    ).count()
    equipos_entregados = SalidaEquipo.objects.filter(
        fecha_salida__year=year, fecha_salida__month=month,
    ).count()
    equipos_mes_resumen = _equipos_mes_resumen(year, month)
    ingresos_tecnicos = _ingresos_asignados_tecnicos_mes(
        year,
        month,
        tecnico_resumen_filtro,
    )
    salidas_tecnicos = _salidas_reparadas_tecnicos_mes(
        year,
        month,
        tecnico_resumen_filtro,
    )

    # Desglose por tipo de salida
    salidas_por_estado = (
        SalidaEquipo.objects
        .filter(fecha_salida__year=year, fecha_salida__month=month)
        .values('estado_reparacion')
        .annotate(total=Sum('valor_final_cobrado'), count=Sum('id'))
    )
    salidas_por_estado = (
        SalidaEquipo.objects
        .filter(fecha_salida__year=year, fecha_salida__month=month)
        .values('estado_reparacion')
        .annotate(total=Sum('valor_final_cobrado'), count=Count('id'))
    )
    map_estados = dict(SalidaEquipo.ESTADO_REPARACION)
    salidas_resumen = [
        {
            'estado': map_estados.get(s['estado_reparacion'], s['estado_reparacion']),
            'count': s['count'],
            'total': s['total'] or Decimal('0.00'),
        }
        for s in salidas_por_estado
    ]

    # Bodegaje: totals for this month (based on actual retiro fecha)
    bodegaje_cobrado = SalidaEquipo.objects.filter(
        fecha_retiro_real__year=year, fecha_retiro_real__month=month,
        bodegaje_monto_congelado__gt=0, bodegaje_aplicado_al_pago=True,
    ).aggregate(s=Sum('bodegaje_monto_congelado'))['s'] or Decimal('0.00')
    bodegaje_perdonado = SalidaEquipo.objects.filter(
        fecha_retiro_real__year=year, fecha_retiro_real__month=month,
        bodegaje_monto_congelado__gt=0, bodegaje_aplicado_al_pago=False,
    ).aggregate(s=Sum('bodegaje_monto_congelado'))['s'] or Decimal('0.00')

    bodegaje_cobrado_count = SalidaEquipo.objects.filter(
        fecha_retiro_real__year=year, fecha_retiro_real__month=month,
        bodegaje_monto_congelado__gt=0, bodegaje_aplicado_al_pago=True,
    ).count()
    bodegaje_perdonado_count = SalidaEquipo.objects.filter(
        fecha_retiro_real__year=year, fecha_retiro_real__month=month,
        bodegaje_monto_congelado__gt=0, bodegaje_aplicado_al_pago=False,
    ).count()

    # Facturas en salidas del mes
    facturas_salidas_mes = SalidaEquipo.objects.filter(
        fecha_salida__year=year,
        fecha_salida__month=month,
    )
    facturas_si_count = facturas_salidas_mes.filter(factura_realizada='si').count()
    facturas_no_count = facturas_salidas_mes.filter(factura_realizada='no').count()

    # Egresos por categoría
    egresos_por_cat = (
        Egreso.objects
        .filter(fecha__year=year, fecha__month=month)
        .values('categoria__nombre', 'categoria__color', 'categoria__icono')
        .annotate(total=Sum('monto'), count=Count('id'))
        .order_by('-total')
    )

    # Lista de años con datos
    anos_disp = sorted(set(
        list(IngresoEquipo.objects.dates('fecha_ingreso', 'year').values_list('fecha_ingreso__year', flat=True)) +
        list(Egreso.objects.dates('fecha', 'year').values_list('fecha__year', flat=True))
    ), reverse=True)
    if not anos_disp:
        anos_disp = [hoy.year]

    # Datos para gráficos
    # 1. Ganancias por Técnico (Gráfico circular)
    ganancias_qs = (
        SalidaEquipo.objects
        .filter(fecha_salida__year=year, fecha_salida__month=month, ingreso__tecnico_encargado__isnull=False)
        .values('ingreso__tecnico_encargado__first_name', 'ingreso__tecnico_encargado__username')
        .annotate(total=Sum('valor_final_cobrado'))
        .filter(total__gt=0)
        .order_by('-total')
    )
    ganancias_tecnicos_labels = [g['ingreso__tecnico_encargado__first_name'] or g['ingreso__tecnico_encargado__username'] for g in ganancias_qs]
    ganancias_tecnicos_data = [float(g['total']) for g in ganancias_qs]

    # 2. Tendencia Anual (Gráfico de líneas y columnas)
    tendencia_ingresos = []
    tendencia_egresos = []
    for m in range(1, 13):
        tendencia_ingresos.append(float(_ingresos_dinero_mes(year, m)['total']))
        tendencia_egresos.append(float(_egresos_mes(year, m)))

    # 3. Egresos por categoría (Gráfico de columnas)
    column_egresos_labels = [e['categoria__nombre'] for e in egresos_por_cat]
    column_egresos_data = [float(e['total']) for e in egresos_por_cat]

    # Usuarios en línea
    from django.contrib.auth.models import User
    from django.utils import timezone
    from datetime import timedelta
    
    limite_online = timezone.now() - timedelta(minutes=5)
    usuarios_activos = User.objects.select_related('actividad').order_by('-actividad__ultima_conexion')
    lista_usuarios = []
    for u in usuarios_activos:
        en_linea = False
        ultima = getattr(u, 'actividad', None)
        if ultima and ultima.ultima_conexion >= limite_online:
            en_linea = True
        lista_usuarios.append({
            'user': u,
            'en_linea': en_linea,
            'ultima_conexion': ultima.ultima_conexion if ultima else None
        })

    horarios_tecnicos, avisos_laborales_hoy, avisos_fuera_laboral_hoy = _horarios_tecnicos_dashboard()

    return render(request, 'admin_panel/dashboard.html', {
        'year': year,
        'month': month,
        'mes_nombre': MESES_ES[month],
        'hoy': hoy,
        'anos_disp': anos_disp,
        'lista_usuarios': lista_usuarios,
        'horarios_tecnicos': horarios_tecnicos,
        'avisos_laborales_hoy': avisos_laborales_hoy,
        'avisos_fuera_laboral_hoy': avisos_fuera_laboral_hoy,
        'dinero_in': dinero_in,
        'egresos_total': egresos_total,
        'utilidad': utilidad,
        'equipos_ingresados': equipos_ingresados,
        'equipos_entregados': equipos_entregados,
        'equipos_mes_resumen': equipos_mes_resumen,
        'ingresos_tecnicos_resumen': ingresos_tecnicos['resumen'],
        'ingresos_tecnicos_registros': ingresos_tecnicos['registros'],
        'ingresos_tecnicos_total': ingresos_tecnicos['total'],
        'ingresos_tecnicos_asignados': ingresos_tecnicos['asignados'],
        'ingresos_tecnicos_sin_asignar': ingresos_tecnicos['sin_asignar'],
        'ingresos_tecnicos_count': ingresos_tecnicos['tecnicos'],
        'ingresos_tecnicos_desde': ingresos_tecnicos['fecha_desde'],
        'ingresos_tecnicos_hasta': ingresos_tecnicos['fecha_hasta'],
        'salidas_tecnicos_resumen': salidas_tecnicos['resumen'],
        'salidas_tecnicos_registros': salidas_tecnicos['registros'],
        'salidas_tecnicos_total': salidas_tecnicos['total'],
        'salidas_tecnicos_positivas': salidas_tecnicos['positivas'],
        'salidas_tecnicos_negativas': salidas_tecnicos['negativas'],
        'salidas_tecnicos_otras': salidas_tecnicos['otras'],
        'salidas_tecnicos_efectividad': salidas_tecnicos['efectividad'],
        'salidas_tecnicos_recaudado': salidas_tecnicos['recaudado'],
        'tecnicos_resumen': tecnicos_resumen,
        'tecnico_resumen_filtro': tecnico_resumen_filtro,
        'tecnico_resumen_nombre': tecnico_resumen_nombre,
        'salidas_resumen': salidas_resumen,
        'egresos_por_cat': egresos_por_cat,
        'bodegaje_cobrado': bodegaje_cobrado,
        'bodegaje_perdonado': bodegaje_perdonado,
        'bodegaje_cobrado_count': bodegaje_cobrado_count,
        'bodegaje_perdonado_count': bodegaje_perdonado_count,
        'facturas_si_count': facturas_si_count,
        'facturas_no_count': facturas_no_count,
        'anos_disp': anos_disp,
        'meses_es': MESES_ES,
        'chart_ganancias_labels': json.dumps(ganancias_tecnicos_labels),
        'chart_ganancias_data': json.dumps(ganancias_tecnicos_data),
        'chart_tendencia_labels': json.dumps(MESES_ES[1:]),
        'chart_tendencia_ingresos': json.dumps(tendencia_ingresos),
        'chart_tendencia_egresos': json.dumps(tendencia_egresos),
        'chart_egresos_cat_labels': json.dumps(column_egresos_labels),
        'chart_egresos_cat_data': json.dumps(column_egresos_data),
    })


@admin_requerido
def admin_bitacoras_tecnicos(request):
    hoy = timezone.localdate()
    fecha_param = (request.GET.get('fecha') or '').strip()
    try:
        dia = date.fromisoformat(fecha_param) if fecha_param else hoy
    except ValueError:
        dia = hoy

    bitacoras_tecnicos = _bitacoras_tecnicos_admin(dia)

    return render(request, 'admin_panel/bitacoras_tecnicos.html', {
        'dia': dia,
        'hoy': hoy,
        'fecha_iso': dia.isoformat(),
        'fecha_txt': dia.strftime('%d/%m/%Y'),
        'es_hoy': dia == hoy,
        'dashboard_year': dia.year,
        'dashboard_month': dia.month,
        'bitacoras_tecnicos': bitacoras_tecnicos,
        'bitacoras_tecnicos_total': sum(item['bitacora']['total'] for item in bitacoras_tecnicos),
    })


def _periodo_ventas_inventario(request):
    hoy = timezone.localdate()
    try:
        year = int(request.GET.get('ano') or hoy.year)
    except (TypeError, ValueError):
        year = hoy.year
    try:
        month = int(request.GET.get('mes') or hoy.month)
    except (TypeError, ValueError):
        month = hoy.month
    return year, min(max(month, 1), 12)


def _ventas_admin_con_totales(queryset):
    decimal_field = DecimalField(max_digits=14, decimal_places=2)
    abonos_total = (
        Abono.objects
        .filter(ingreso_id=OuterRef('pk'))
        .values('ingreso_id')
        .annotate(total=Sum('monto'))
        .values('total')[:1]
    )
    return (
        queryset
        .annotate(
            abonos_posteriores_admin=Coalesce(
                Subquery(abonos_total, output_field=decimal_field),
                Value(Decimal('0.00')),
                output_field=decimal_field,
            ),
        )
        .annotate(
            total_pagado_admin_db=ExpressionWrapper(
                Coalesce(
                    F('abono_anticipo'),
                    Value(Decimal('0.00')),
                    output_field=decimal_field,
                ) + F('abonos_posteriores_admin'),
                output_field=decimal_field,
            ),
        )
    )


def _preparar_venta_admin(venta):
    total = venta.valor_acordado or Decimal('0.00')
    pagado = (
        (venta.abono_anticipo or Decimal('0.00'))
        + (venta.abonos_posteriores_admin or Decimal('0.00'))
    )
    saldo = max(total - pagado, Decimal('0.00'))
    if total <= pagado:
        estado = 'Pagado'
        estado_clase = 'pagado'
    elif pagado > 0:
        estado = 'Parcial'
        estado_clase = 'parcial'
    else:
        estado = 'Pendiente'
        estado_clase = 'pendiente'

    productos = list(venta.productos_inventario.all())
    venta.total_admin = total
    venta.pagado_admin = pagado
    venta.saldo_admin = saldo
    venta.estado_pago_admin = estado
    venta.estado_pago_clase_admin = estado_clase
    venta.unidades_admin = sum(producto.cantidad for producto in productos)
    venta.productos_admin = productos
    venta.productos_resumen_admin = ', '.join(
        f'{producto.cantidad} x {producto.inventario_item.producto}'
        for producto in productos
    ) or venta.problema_reportado or 'Venta de producto'
    venta.metodos_pago_admin = venta.resumen_metodos_pago or 'Sin pago registrado'
    return venta


@admin_requerido
def admin_ventas_inventario(request):
    """Centro administrativo de ventas de producto e inventario."""
    year, month = _periodo_ventas_inventario(request)
    tab = (request.GET.get('tab') or 'ventas').strip().lower()
    if tab not in {'ventas', 'actividad', 'movimientos', 'inventario'}:
        tab = 'ventas'

    q = (request.GET.get('q') or '').strip()
    estado_pago = (request.GET.get('estado_pago') or '').strip().lower()
    tecnico_id = (request.GET.get('tecnico') or '').strip()
    registrador_id = (request.GET.get('registrador') or '').strip()
    actividad_usuario_id = (request.GET.get('actividad_usuario') or '').strip()
    inventario_sede = (request.GET.get('inventario_sede') or '').strip()
    inventario_ubicacion = (request.GET.get('inventario_ubicacion') or '').strip()
    inventario_estado = (request.GET.get('inventario_estado') or '').strip()
    inventario_categoria = (request.GET.get('inventario_categoria') or '').strip()
    inventario_stock = (request.GET.get('inventario_stock') or '').strip()

    ventas_periodo = IngresoEquipo.objects.filter(
        sede='ventas',
        fecha_ingreso__year=year,
        fecha_ingreso__month=month,
    )
    ventas_metricas = list(
        _ventas_admin_con_totales(ventas_periodo)
        .values('valor_acordado', 'abono_anticipo', 'abonos_posteriores_admin')
    )
    total_facturado = sum(
        (venta['valor_acordado'] or Decimal('0.00'))
        for venta in ventas_metricas
    )
    total_cobrado = sum(
        (venta['abono_anticipo'] or Decimal('0.00'))
        + (venta['abonos_posteriores_admin'] or Decimal('0.00'))
        for venta in ventas_metricas
    )
    total_saldo = Decimal('0.00')
    ventas_pagadas = 0
    ventas_parciales = 0
    ventas_pendientes = 0
    for venta in ventas_metricas:
        total = venta['valor_acordado'] or Decimal('0.00')
        pagado = (
            (venta['abono_anticipo'] or Decimal('0.00'))
            + (venta['abonos_posteriores_admin'] or Decimal('0.00'))
        )
        total_saldo += max(total - pagado, Decimal('0.00'))
        if total <= pagado:
            ventas_pagadas += 1
        elif pagado > 0:
            ventas_parciales += 1
        else:
            ventas_pendientes += 1

    unidades_vendidas = (
        VentaInventarioItem.objects
        .filter(
            venta__sede='ventas',
            venta__fecha_ingreso__year=year,
            venta__fecha_ingreso__month=month,
        )
        .aggregate(total=Sum('cantidad'))['total']
        or 0
    )

    inventario_valores = list(
        InventarioItem.objects.values('cantidad', 'costo', 'estado')
    )
    inventario_unidades = sum(item['cantidad'] for item in inventario_valores)
    inventario_valor = sum(
        Decimal(item['cantidad']) * (item['costo'] or Decimal('0.00'))
        for item in inventario_valores
    )
    inventario_agotado = sum(
        1 for item in inventario_valores if item['cantidad'] == 0
    )
    inventario_bajo = sum(
        1 for item in inventario_valores if 0 < item['cantidad'] <= 5
    )
    inventario_no_disponible = sum(
        1 for item in inventario_valores if item['estado'] == 'no_disponible'
    )

    anos_disp = list(
        IngresoEquipo.objects
        .filter(sede='ventas')
        .dates('fecha_ingreso', 'year')
        .values_list('fecha_ingreso__year', flat=True)
    )
    anos_disp = sorted(set(anos_disp + [year]), reverse=True)

    User = get_user_model()
    personal_ids = set(
        ventas_periodo.exclude(tecnico_encargado_id=None)
        .values_list('tecnico_encargado_id', flat=True)
    )
    registradores_ids = set(
        ventas_periodo.exclude(registrado_por_id=None)
        .values_list('registrado_por_id', flat=True)
    )
    personal_ventas = User.objects.filter(
        pk__in=personal_ids | registradores_ids
    ).order_by('first_name', 'last_name', 'username')

    categorias_inventario = list(
        InventarioItem.objects
        .exclude(categoria='')
        .order_by('categoria')
        .values_list('categoria', flat=True)
        .distinct()
    )

    page_obj = None
    querystring = ''
    if tab == 'ventas':
        ventas_qs = ventas_periodo
        if q:
            filtro_q = (
                Q(cliente__nombres__icontains=q)
                | Q(cliente__cedula__icontains=q)
                | Q(cliente__whatsapp__icontains=q)
                | Q(problema_reportado__icontains=q)
                | Q(productos_inventario__inventario_item__producto__icontains=q)
                | Q(productos_inventario__inventario_item__codigo__icontains=q)
                | Q(productos_inventario__inventario_item__marca__icontains=q)
                | Q(productos_inventario__inventario_item__modelo__icontains=q)
            )
            codigo_q = q.upper()
            if codigo_q.startswith('P'):
                codigo_q = codigo_q[1:]
            if codigo_q.isdigit():
                filtro_q |= Q(numero_equipo=int(codigo_q))
            ventas_qs = ventas_qs.filter(filtro_q).distinct()
        if tecnico_id.isdigit():
            ventas_qs = ventas_qs.filter(tecnico_encargado_id=int(tecnico_id))
        if registrador_id.isdigit():
            ventas_qs = ventas_qs.filter(registrado_por_id=int(registrador_id))

        ventas_qs = _ventas_admin_con_totales(ventas_qs)
        if estado_pago == 'pagado':
            ventas_qs = ventas_qs.filter(
                valor_acordado__isnull=False,
                total_pagado_admin_db__gte=F('valor_acordado'),
            )
        elif estado_pago == 'parcial':
            ventas_qs = ventas_qs.filter(
                valor_acordado__isnull=False,
                total_pagado_admin_db__gt=0,
                total_pagado_admin_db__lt=F('valor_acordado'),
            )
        elif estado_pago == 'pendiente':
            ventas_qs = ventas_qs.filter(
                Q(valor_acordado__isnull=True)
                | Q(total_pagado_admin_db__lte=0)
            )

        ventas_qs = (
            ventas_qs
            .select_related('cliente', 'tecnico_encargado', 'registrado_por')
            .prefetch_related(
                Prefetch(
                    'productos_inventario',
                    queryset=VentaInventarioItem.objects.select_related('inventario_item'),
                ),
                Prefetch('abonos', queryset=Abono.objects.order_by('fecha', 'creado')),
            )
            .order_by('-fecha_ingreso', '-numero_equipo')
        )
        page_obj, querystring = paginar_resultados(request, ventas_qs)
        for venta in page_obj.object_list:
            _preparar_venta_admin(venta)

    elif tab == 'actividad':
        actividad_qs = (
            BitacoraTecnico.objects
            .select_related('user', 'ingreso', 'ingreso__cliente', 'abono')
            .filter(
                momento__year=year,
                momento__month=month,
            )
            .filter(
                Q(tipo__in=['venta_producto', 'venta_editada'])
                | Q(ingreso__sede='ventas')
                | Q(abono__ingreso__sede='ventas')
            )
        )
        if q:
            actividad_qs = actividad_qs.filter(
                Q(texto__icontains=q)
                | Q(codigo__icontains=q)
                | Q(usuario_nombre__icontains=q)
                | Q(ingreso__cliente__nombres__icontains=q)
                | Q(ingreso__cliente__cedula__icontains=q)
            )
        if actividad_usuario_id.isdigit():
            actividad_qs = actividad_qs.filter(user_id=int(actividad_usuario_id))
        page_obj, querystring = paginar_resultados(
            request,
            actividad_qs.order_by('-momento', '-pk'),
        )

    elif tab == 'movimientos':
        movimientos_qs = (
            VentaInventarioItem.objects
            .select_related(
                'venta',
                'venta__cliente',
                'venta__tecnico_encargado',
                'venta__registrado_por',
                'inventario_item',
            )
            .filter(
                venta__sede='ventas',
                venta__fecha_ingreso__year=year,
                venta__fecha_ingreso__month=month,
            )
        )
        if q:
            movimientos_qs = movimientos_qs.filter(
                Q(venta__cliente__nombres__icontains=q)
                | Q(venta__cliente__cedula__icontains=q)
                | Q(inventario_item__producto__icontains=q)
                | Q(inventario_item__codigo__icontains=q)
                | Q(inventario_item__marca__icontains=q)
                | Q(inventario_item__modelo__icontains=q)
            )
        if inventario_sede in {'guayaquil', 'quito'}:
            movimientos_qs = movimientos_qs.filter(
                inventario_item__sede=inventario_sede
            )
        if inventario_ubicacion in dict(InventarioItem.UBICACIONES):
            if (
                inventario_sede == 'guayaquil'
                and inventario_ubicacion not in {'guayaquil_norte', 'guayaquil_centro'}
            ):
                inventario_ubicacion = ''
            elif inventario_sede == 'quito' and inventario_ubicacion != 'quito':
                inventario_ubicacion = ''
            else:
                movimientos_qs = movimientos_qs.filter(
                    inventario_item__ubicacion=inventario_ubicacion
                )
        page_obj, querystring = paginar_resultados(
            request,
            movimientos_qs.order_by('-venta__fecha_ingreso', '-creado', '-pk'),
        )

    else:
        inventario_qs = InventarioItem.objects.select_related('registrado_por')
        if q:
            inventario_qs = inventario_qs.filter(
                Q(producto__icontains=q)
                | Q(codigo__icontains=q)
                | Q(marca__icontains=q)
                | Q(modelo__icontains=q)
                | Q(serie__icontains=q)
                | Q(causa_no_disponible__icontains=q)
            )
        if inventario_sede in {'guayaquil', 'quito'}:
            inventario_qs = inventario_qs.filter(sede=inventario_sede)
        ubicaciones_validas = dict(InventarioItem.UBICACIONES)
        if inventario_ubicacion in ubicaciones_validas:
            if (
                inventario_sede == 'guayaquil'
                and inventario_ubicacion not in {'guayaquil_norte', 'guayaquil_centro'}
            ):
                inventario_ubicacion = ''
            elif inventario_sede == 'quito' and inventario_ubicacion != 'quito':
                inventario_ubicacion = ''
            else:
                inventario_qs = inventario_qs.filter(
                    ubicacion=inventario_ubicacion
                )
        if inventario_estado in dict(InventarioItem.ESTADOS):
            inventario_qs = inventario_qs.filter(estado=inventario_estado)
        if inventario_categoria in categorias_inventario:
            inventario_qs = inventario_qs.filter(categoria=inventario_categoria)
        if inventario_stock == 'agotado':
            inventario_qs = inventario_qs.filter(cantidad=0)
        elif inventario_stock == 'bajo':
            inventario_qs = inventario_qs.filter(cantidad__gt=0, cantidad__lte=5)
        elif inventario_stock == 'disponible':
            inventario_qs = inventario_qs.filter(
                cantidad__gt=5,
                estado='disponible',
            )

        page_obj, querystring = paginar_resultados(
            request,
            inventario_qs.order_by('producto', 'codigo'),
        )
        for item in page_obj.object_list:
            item.valor_stock_admin = (
                Decimal(item.cantidad) * (item.costo or Decimal('0.00'))
            )

    return render(request, 'admin_panel/ventas_inventario.html', {
        'tab': tab,
        'year': year,
        'month': month,
        'mes_nombre': MESES_ES[month],
        'meses_es': MESES_ES,
        'anos_disp': anos_disp,
        'q': q,
        'estado_pago': estado_pago,
        'tecnico_id': tecnico_id,
        'registrador_id': registrador_id,
        'actividad_usuario_id': actividad_usuario_id,
        'inventario_sede': inventario_sede,
        'inventario_ubicacion': inventario_ubicacion,
        'inventario_estado': inventario_estado,
        'inventario_categoria': inventario_categoria,
        'inventario_stock': inventario_stock,
        'personal_ventas': personal_ventas,
        'categorias_inventario': categorias_inventario,
        'ubicaciones_inventario': InventarioItem.UBICACIONES,
        'estados_inventario': InventarioItem.ESTADOS,
        'page_obj': page_obj,
        'querystring': querystring,
        'total_ventas': len(ventas_metricas),
        'total_facturado': total_facturado,
        'total_cobrado': total_cobrado,
        'total_saldo': total_saldo,
        'ventas_pagadas': ventas_pagadas,
        'ventas_parciales': ventas_parciales,
        'ventas_pendientes': ventas_pendientes,
        'unidades_vendidas': unidades_vendidas,
        'inventario_productos': len(inventario_valores),
        'inventario_unidades': inventario_unidades,
        'inventario_valor': inventario_valor,
        'inventario_agotado': inventario_agotado,
        'inventario_bajo': inventario_bajo,
        'inventario_no_disponible': inventario_no_disponible,
    })


@admin_requerido
def admin_equipos_mes_exportar(request):
    year, month = _periodo_admin_request(request)
    resumen = _equipos_mes_resumen(year, month)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        messages.error(request, 'No se pudo generar el Excel (falta openpyxl).')
        return redirect(f"{reverse('econotec:admin_dashboard')}?ano={year}&mes={month}#equipos-mes-resumen")

    wb = Workbook()
    ws = wb.active
    ws.title = f'Equipos {MESES_ES[month][:3]} {year}'
    ws.append([f'Resumen de equipos del mes - {MESES_ES[month]} {year}'])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color='F97618')

    headers = [
        'Codigo', 'Movimiento', 'Fecha ingreso', 'Fecha entrega', 'Cliente',
        'Cedula', 'WhatsApp', 'Tipo', 'Marca', 'Modelo / Serie', 'Sede',
        'Tecnico asignado al ingreso', 'Estado actual', 'Estado salida', 'Valor acordado',
        'Valor salida',
    ]
    ws.append(headers)
    fill = PatternFill('solid', fgColor='F97618')
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill

    for item in resumen:
        ingreso = item['ingreso']
        salida = item['salida']
        movimientos = []
        if item['ingresado_en_mes']:
            movimientos.append('Ingresado')
        if item['entregado_en_mes']:
            movimientos.append('Entregado')
        ws.append([
            ingreso.codigo_equipo,
            ' / '.join(movimientos) or 'Relacionado',
            ingreso.fecha_ingreso.strftime('%d/%m/%Y'),
            salida.fecha_salida.strftime('%d/%m/%Y') if salida else '',
            ingreso.cliente.nombres,
            ingreso.cliente.cedula,
            ingreso.cliente.whatsapp,
            ingreso.tipo_equipo_display,
            ingreso.marca,
            ingreso.modelo_serie_detalle,
            ingreso.get_sede_display(),
            item['tecnico_ingreso_nombre'],
            ingreso.get_estado_display(),
            salida.get_estado_reparacion_display() if salida else 'Sin salida',
            float(ingreso.valor_acordado or 0),
            float(salida.valor_final_cobrado or 0) if salida else 0,
        ])

    widths = [14, 18, 14, 14, 26, 16, 16, 16, 18, 34, 16, 18, 26, 30, 16, 16]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=2, column=idx).column_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"equipos_mes_{year}_{month:02d}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@admin_requerido
@require_POST
def admin_equipos_mes_borrar(request):
    year, month = _periodo_admin_request(request, metodo='POST')
    mes_nombre = MESES_ES[month]
    admin_password = request.POST.get('admin_password', '')
    dashboard_url = f"{reverse('econotec:admin_dashboard')}?ano={year}&mes={month}#equipos-mes-resumen"

    if not request.user.check_password(admin_password):
        messages.error(request, f'Contraseña de administrador incorrecta. No se eliminó {mes_nombre} {year}.')
        return redirect(dashboard_url)

    ingresos_mes_ids = list(
        IngresoEquipo.objects.filter(
            sede__in=SEDES_EQUIPOS,
            fecha_ingreso__year=year,
            fecha_ingreso__month=month,
        ).values_list('id', flat=True)
    )
    salidas_qs = SalidaEquipo.objects.filter(
        Q(fecha_salida__year=year, fecha_salida__month=month, ingreso__sede__in=SEDES_EQUIPOS) |
        Q(ingreso_id__in=ingresos_mes_ids)
    )
    abonos_qs = Abono.objects.filter(
        Q(fecha__year=year, fecha__month=month, ingreso__sede__in=SEDES_EQUIPOS) |
        Q(ingreso_id__in=ingresos_mes_ids)
    )
    salidas_ids = list(salidas_qs.values_list('id', flat=True))

    ingresos_count = len(ingresos_mes_ids)
    salidas_count = len(salidas_ids)
    abonos_count = abonos_qs.count()

    with transaction.atomic():
        abonos_qs.delete()
        SalidaEquipo.objects.filter(id__in=salidas_ids).delete()
        IngresoEquipo.objects.filter(id__in=ingresos_mes_ids).delete()

    messages.success(
        request,
        f'{mes_nombre} {year} eliminado: {ingresos_count} ingresos, '
        f'{salidas_count} salidas y {abonos_count} abonos. Los clientes no fueron borrados.'
    )
    return redirect(dashboard_url)


@admin_requerido
@require_POST
def admin_horario_tecnico_guardar(request, user_id):
    User = get_user_model()
    tecnico_qs = (
        User.objects
        .filter(pk=user_id, is_active=True, groups__name__in=['Tecnicos', 'Tecnico'])
        .distinct()
    )
    tecnico = get_object_or_404(tecnico_qs)
    horario, _ = HorarioTecnico.objects.get_or_create(tecnico=tecnico)

    try:
        hora_inicio = time.fromisoformat(request.POST.get('hora_inicio') or '09:00')
        hora_fin = time.fromisoformat(request.POST.get('hora_fin') or '18:00')
    except ValueError:
        messages.error(request, 'Horario inválido. Usa formato HH:MM.')
        return redirect(reverse('econotec:admin_dashboard') + '#horarios-tecnicos')

    if hora_inicio >= hora_fin:
        messages.error(request, 'La hora de entrada debe ser menor que la hora de salida.')
        return redirect(reverse('econotec:admin_dashboard') + '#horarios-tecnicos')

    horario.activo = request.POST.get('activo') == 'on'
    for campo, _label in HorarioTecnico.DIAS:
        setattr(horario, campo, request.POST.get(campo) == 'on')
    horario.hora_inicio = hora_inicio
    horario.hora_fin = hora_fin
    horario.save()

    messages.success(request, f'Horario laboral de {_nombre_usuario(tecnico)} actualizado.')
    return redirect(reverse('econotec:admin_dashboard') + '#horarios-tecnicos')


@tecnico_requerido
def salida_facturas_lista(request):
    """Listado de salidas que sí tienen factura realizada."""
    hoy = date.today()
    year = int(request.GET.get('ano') or hoy.year)
    mes_param = (request.GET.get('mes') or str(hoy.month)).strip().lower()
    month = None if mes_param == 'todos' else int(mes_param)
    q = (request.GET.get('q') or '').strip()
    fecha_desde, fecha_hasta, fecha_preset = obtener_rango_fecha(request)
    rango_activo = bool(fecha_desde or fecha_hasta)

    base_qs = (
        SalidaEquipo.objects
        .select_related('ingreso', 'ingreso__cliente', 'registrado_por', 'tecnico_reparo')
        .filter(factura_realizada='si')
        .order_by('-fecha_salida', '-creado')
    )

    if rango_activo:
        base_qs = aplicar_rango_fecha(base_qs, 'fecha_salida', fecha_desde, fecha_hasta)
        periodo_label = contexto_rango_fecha(
            fecha_desde,
            fecha_hasta,
            fecha_preset,
            etiqueta='Fecha factura',
        )['fecha_resumen'].replace('Fecha factura: ', '')
    else:
        base_qs = base_qs.filter(fecha_salida__year=year)
        if month is not None:
            base_qs = base_qs.filter(fecha_salida__month=month)
        periodo_label = f'{MESES_ES[month] if month else "Todos los meses"} {year}'

    total_periodo = base_qs.count()
    qs = base_qs
    qs = filtrar_objetos_normalizado(qs, q, texto_salida_busqueda)
    total = total_resultados(qs)
    page_obj, querystring = paginar_resultados(request, qs)

    anos_disp = sorted(set(
        list(SalidaEquipo.objects.dates('fecha_salida', 'year').values_list('fecha_salida__year', flat=True))
    ), reverse=True)
    if not anos_disp:
        anos_disp = [hoy.year]

    context = {
        'year': year,
        'month': month,
        'mes_param': mes_param,
        'mes_nombre': MESES_ES[month] if month else 'Todos los meses',
        'periodo_label': periodo_label,
        'meses_es': MESES_ES,
        'anos_disp': anos_disp,
        'q': q,
        'salidas': page_obj.object_list,
        'page_obj': page_obj,
        'querystring': querystring,
        'total': total,
        'total_periodo': total_periodo,
    }
    context.update(contexto_rango_fecha(
        fecha_desde,
        fecha_hasta,
        fecha_preset,
        etiqueta='Fecha factura',
    ))
    return render(request, 'admin_panel/facturas_salidas.html', context)


@admin_requerido
def admin_bodegajes(request):
    """Lista administrativa de bodegajes (cobrados y no cobrados) por mes."""
    hoy = date.today()
    year = int(request.GET.get('ano') or hoy.year)
    month = int(request.GET.get('mes') or hoy.month)

    base_qs = SalidaEquipo.objects.select_related('ingreso', 'ingreso__cliente')
    cobrados = base_qs.filter(
        fecha_retiro_real__year=year, fecha_retiro_real__month=month,
        bodegaje_monto_congelado__gt=0, bodegaje_aplicado_al_pago=True,
    ).order_by('-fecha_retiro_real')
    no_cobrados = base_qs.filter(
        fecha_retiro_real__year=year, fecha_retiro_real__month=month,
        bodegaje_monto_congelado__gt=0, bodegaje_aplicado_al_pago=False,
    ).order_by('-fecha_retiro_real')

    bodegaje_cobrado = cobrados.aggregate(s=Sum('bodegaje_monto_congelado'))['s'] or Decimal('0.00')
    bodegaje_perdonado = no_cobrados.aggregate(s=Sum('bodegaje_monto_congelado'))['s'] or Decimal('0.00')

    bodegaje_cobrado_count = cobrados.count()
    bodegaje_perdonado_count = no_cobrados.count()

    # Lista de años disponibles (reusar lógica)
    anos_disp = sorted(set(
        list(IngresoEquipo.objects.dates('fecha_ingreso', 'year').values_list('fecha_ingreso__year', flat=True)) +
        list(Egreso.objects.dates('fecha', 'year').values_list('fecha__year', flat=True))
    ), reverse=True)
    if not anos_disp:
        anos_disp = [hoy.year]

    return render(request, 'admin_panel/bodegajes.html', {
        'year': year,
        'month': month,
        'mes_nombre': MESES_ES[month],
        'cobrados': cobrados,
        'no_cobrados': no_cobrados,
        'bodegaje_cobrado': bodegaje_cobrado,
        'bodegaje_perdonado': bodegaje_perdonado,
        'bodegaje_cobrado_count': bodegaje_cobrado_count,
        'bodegaje_perdonado_count': bodegaje_perdonado_count,
        'anos_disp': anos_disp,
        'meses_es': MESES_ES,
    })


@admin_requerido
def admin_activos_bodegaje(request):
    """Manejo de equipos en bodegaje activo y chatarrerización."""
    from .alertas import salidas_bodegaje_qs
    import json
    from django.http import JsonResponse
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            accion = data.get('accion')
            salida_id = data.get('salida_id')
            
            salida = SalidaEquipo.objects.get(pk=salida_id)
            if accion == 'retirado':
                salida.estado_reparacion = 'retirado'
                salida.fecha_retiro_real = date.today()
            elif accion == 'chatarrerizacion':
                salida.estado_reparacion = 'chatarrerizacion'
                salida.fecha_retiro_real = date.today()  # Detiene el bodegaje
                
            salida.save()
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'msg': str(e)})

    activos_qs = salidas_bodegaje_qs(incluir_silenciados=True).select_related('ingreso', 'ingreso__cliente')
    from .alertas import whatsapp_link_bodegaje
    
    # Procesar bodegaje para la vista
    activos = []
    hoy = date.today()
    for s in activos_qs:
        bod = s.calcular_bodegaje()
        if bod['dias'] > 0:
            dias_totales = (hoy - s.fecha_salida).days if s.fecha_salida else 0
            activos.append({
                'salida': s,
                'dias_bodegaje': bod['dias'],
                'monto_bodegaje': bod['monto'],
                'dias_totales': dias_totales,
                'wa_link': whatsapp_link_bodegaje(s),
            })
            
    chatarrerizacion = SalidaEquipo.objects.filter(
        estado_reparacion='chatarrerizacion'
    ).select_related('ingreso', 'ingreso__cliente').order_by('-fecha_retiro_real')
    
    return render(request, 'admin_panel/activos_bodegaje.html', {
        'activos': activos,
        'chatarrerizacion': chatarrerizacion,
        'count_chatarrerizacion': chatarrerizacion.count(),
    })

def _equipos_administrativos_filtrados(request, incluir_egreso=False):
    """Consulta común sin calcular totales financieros para vistas generales."""
    estado_filtro = (request.GET.get('estado') or '').strip()
    q = (request.GET.get('q') or '').strip()
    relaciones = ['cliente', 'registrado_por', 'tecnico_encargado']
    if incluir_egreso:
        relaciones.extend([
            'egreso_compra',
            'egreso_compra__categoria',
            'egreso_compra__registrado_por',
        ])
    qs = (
        IngresoEquipo.objects
        .filter(estado__in=('donado', 'equipo_a_comprar'))
        .select_related(*relaciones)
        .order_by('-fecha_ingreso', '-numero_equipo')
    )
    if estado_filtro in ('donado', 'equipo_a_comprar'):
        qs = qs.filter(estado=estado_filtro)
    if q:
        filtros_busqueda = (
            Q(marca__icontains=q)
            | Q(modelo_serie__icontains=q)
            | Q(serie__icontains=q)
            | Q(cliente__nombres__icontains=q)
            | Q(cliente__cedula__icontains=q)
        )
        if q.isdigit():
            filtros_busqueda |= Q(numero_equipo=int(q))
        elif len(q) > 1 and q[0].upper() in {'G', 'U', 'P'} and q[1:].isdigit():
            sede_por_prefijo = {'G': 'guayaquil', 'U': 'quito', 'P': 'ventas'}
            filtros_busqueda |= Q(
                sede=sede_por_prefijo[q[0].upper()],
                numero_equipo=int(q[1:]),
            )
        qs = qs.filter(filtros_busqueda)
    return qs, estado_filtro, q


@tecnico_requerido
def equipos_administrativos_general(request):
    """Consulta separada para técnicos y asesores, sin resumen ni egresos."""
    if es_admin(request.user):
        return redirect('econotec:admin_equipos_administrativos')

    qs, estado_filtro, q = _equipos_administrativos_filtrados(
        request,
        incluir_egreso=False,
    )
    return render(request, 'admin_panel/equipos_administrativos.html', {
        'ingresos': qs,
        'estado_filtro': estado_filtro,
        'q': q,
        'vista_admin': False,
    })


@admin_requerido
def admin_equipos_administrativos(request):
    """Bandeja administrativa completa para Donados y Equipos a comprar."""
    qs, estado_filtro, q = _equipos_administrativos_filtrados(
        request,
        incluir_egreso=True,
    )

    resumen = {
        'total': qs.count(),
        'donados': qs.filter(estado='donado').count(),
        'compras': qs.filter(estado='equipo_a_comprar').count(),
        'valor_compras': qs.filter(estado='equipo_a_comprar').aggregate(
            total=Sum('valor_acordado')
        )['total'] or Decimal('0.00'),
    }
    return render(request, 'admin_panel/equipos_administrativos.html', {
        'ingresos': qs,
        'resumen': resumen,
        'estado_filtro': estado_filtro,
        'q': q,
        'vista_admin': True,
    })


@admin_requerido
def admin_equipos_cortesia(request):
    """Bandeja mensual de ingresos y salidas registrados como cortesía."""
    hoy = date.today()
    try:
        year = int(request.GET.get('ano') or hoy.year)
        month = int(request.GET.get('mes') or hoy.month)
    except (TypeError, ValueError):
        year, month = hoy.year, hoy.month
    month = min(max(month, 1), 12)
    q = (request.GET.get('q') or '').strip()

    cortesias = (
        IngresoEquipo.objects
        .filter(estado='cortesia')
        .filter(
            Q(fecha_ingreso__year=year, fecha_ingreso__month=month)
            | Q(salida__fecha_salida__year=year, salida__fecha_salida__month=month)
        )
        .select_related(
            'cliente',
            'tecnico_encargado',
            'registrado_por',
            'salida',
            'salida__tecnico_reparo',
            'salida__registrado_por',
        )
        .distinct()
        .order_by('-fecha_ingreso', '-numero_equipo')
    )
    if q:
        filtros = (
            Q(marca__icontains=q)
            | Q(modelo_serie__icontains=q)
            | Q(serie__icontains=q)
            | Q(cliente__nombres__icontains=q)
            | Q(cliente__cedula__icontains=q)
        )
        if q.isdigit():
            filtros |= Q(numero_equipo=int(q))
        cortesias = cortesias.filter(filtros)

    items = []
    for ingreso in cortesias:
        try:
            salida = ingreso.salida
        except SalidaEquipo.DoesNotExist:
            salida = None
        items.append({
            'ingreso': ingreso,
            'salida': salida,
            'ingresado_en_mes': (
                ingreso.fecha_ingreso.year == year
                and ingreso.fecha_ingreso.month == month
            ),
            'salio_en_mes': bool(
                salida
                and salida.fecha_salida.year == year
                and salida.fecha_salida.month == month
            ),
        })

    page_obj, querystring = paginar_resultados(request, items)
    anos_disp = sorted(set(
        IngresoEquipo.objects
        .filter(estado='cortesia')
        .dates('fecha_ingreso', 'year')
        .values_list('fecha_ingreso__year', flat=True)
    ), reverse=True)
    if year not in anos_disp:
        anos_disp.append(year)
        anos_disp.sort(reverse=True)

    return render(request, 'admin_panel/equipos_cortesia.html', {
        'items': page_obj.object_list,
        'page_obj': page_obj,
        'querystring': querystring,
        'q': q,
        'year': year,
        'month': month,
        'mes_nombre': MESES_ES[month],
        'meses_es': MESES_ES,
        'anos_disp': anos_disp,
        'ingresos_mes': IngresoEquipo.objects.filter(
            estado='cortesia',
            fecha_ingreso__year=year,
            fecha_ingreso__month=month,
        ).count(),
        'salidas_mes': SalidaEquipo.objects.filter(
            estado_reparacion='cortesia',
            fecha_salida__year=year,
            fecha_salida__month=month,
        ).count(),
        'pendientes_salida': IngresoEquipo.objects.filter(
            estado='cortesia',
            salida__isnull=True,
        ).count(),
    })


@admin_requerido
def egresos_lista(request):
    cat_filtro = (request.GET.get('cat') or '').strip()
    ano_filtro = (request.GET.get('ano') or '').strip()
    mes_filtro = (request.GET.get('mes') or '').strip()

    qs = Egreso.objects.select_related('categoria', 'registrado_por').order_by('-fecha', '-creado')
    if cat_filtro and cat_filtro.isdigit():
        qs = qs.filter(categoria_id=int(cat_filtro))
    if ano_filtro and ano_filtro.isdigit():
        qs = qs.filter(fecha__year=int(ano_filtro))
    if mes_filtro and mes_filtro.isdigit():
        qs = qs.filter(fecha__month=int(mes_filtro))

    total = qs.aggregate(s=Sum('monto'))['s'] or Decimal('0.00')

    return render(request, 'admin_panel/egresos_lista.html', {
        'egresos': qs,
        'total': total,
        'categorias': CategoriaEgreso.objects.filter(activo=True),
        'cat_filtro': cat_filtro,
        'ano_filtro': ano_filtro,
        'mes_filtro': mes_filtro,
        'meses_es': MESES_ES,
    })


@admin_requerido
def egreso_crear(request):
    if request.method == 'POST':
        form = EgresoForm(request.POST)
        if form.is_valid():
            egreso = form.save(commit=False)
            egreso.registrado_por = request.user
            egreso.save()
            messages.success(request, f'Egreso registrado: ${egreso.monto}.')
            return redirect('econotec:admin_egresos_lista')
    else:
        form = EgresoForm(initial={'fecha': date.today()})
    return render(request, 'admin_panel/egreso_form.html', {
        'form': form,
        'modo': 'crear',
        'titulo': 'Nuevo Egreso',
    })


@admin_requerido
def egreso_editar(request, pk):
    egreso = get_object_or_404(Egreso, pk=pk)
    if request.method == 'POST':
        form = EgresoForm(request.POST, instance=egreso)
        if form.is_valid():
            form.save()
            messages.success(request, 'Egreso actualizado.')
            return redirect('econotec:admin_egresos_lista')
    else:
        form = EgresoForm(instance=egreso)
    return render(request, 'admin_panel/egreso_form.html', {
        'form': form,
        'egreso': egreso,
        'modo': 'editar',
        'titulo': f'Editar egreso: {egreso.concepto}',
    })


@admin_requerido
@require_POST
def egreso_eliminar(request, pk):
    egreso = get_object_or_404(Egreso, pk=pk)
    concepto = egreso.concepto
    egreso.delete()
    messages.success(request, f'Egreso "{concepto}" eliminado.')
    return redirect('econotec:admin_egresos_lista')


# ─────────────────────────────────────────────────────────
# Exportación
# ─────────────────────────────────────────────────────────

@admin_requerido
def export_reporte_mes(request):
    """Exporta el reporte mensual (ingresos vs egresos) a Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    hoy = date.today()
    year = int(request.GET.get('ano') or hoy.year)
    month = int(request.GET.get('mes') or hoy.month)

    dinero_in = _ingresos_dinero_mes(year, month)
    egresos_total = _egresos_mes(year, month)
    utilidad = dinero_in['total'] - egresos_total

    bodegaje_cobrado = SalidaEquipo.objects.filter(
        fecha_retiro_real__year=year, fecha_retiro_real__month=month,
        bodegaje_monto_congelado__gt=0, bodegaje_aplicado_al_pago=True,
    ).aggregate(s=Sum('bodegaje_monto_congelado'))['s'] or Decimal('0.00')
    
    bodegaje_perdonado = SalidaEquipo.objects.filter(
        fecha_retiro_real__year=year, fecha_retiro_real__month=month,
        bodegaje_monto_congelado__gt=0, bodegaje_aplicado_al_pago=False,
    ).aggregate(s=Sum('bodegaje_monto_congelado'))['s'] or Decimal('0.00')

    wb = Workbook()
    ws = wb.active
    ws.title = f'Reporte {MESES_ES[month][:3]}-{year}'

    ws.merge_cells('A1:D1')
    title = ws.cell(row=1, column=1,
                    value=f'Reporte financiero — {MESES_ES[month]} {year} — Econotec')
    title.font = Font(bold=True, size=14, color='F97618')
    title.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 26

    rows = [
        ('Anticipos recibidos', dinero_in['anticipos']),
        ('Diagnósticos rápidos (adicionales)', dinero_in['diagnosticos_rapidos']),
        ('Abonos posteriores', dinero_in['abonos']),
        ('Cobros en salida (reparaciones)', dinero_in['cobros_finales']),
        ('Ventas de productos', dinero_in['cobros_ventas']),
        ('Cobros por diagnóstico (no reparado)', dinero_in['cobros_diagnostico']),
        ('Bodegaje cobrado (en retiros)', bodegaje_cobrado),
        ('N° casos bodegaje cobrados', SalidaEquipo.objects.filter(
            fecha_retiro_real__year=year, fecha_retiro_real__month=month,
            bodegaje_monto_congelado__gt=0, bodegaje_aplicado_al_pago=True,
        ).count()),
        ('Bodegaje no cobrado / perdonado', bodegaje_perdonado),
        ('N° casos bodegaje no cobrados', SalidaEquipo.objects.filter(
            fecha_retiro_real__year=year, fecha_retiro_real__month=month,
            bodegaje_monto_congelado__gt=0, bodegaje_aplicado_al_pago=False,
        ).count()),
        ('TOTAL DINERO RECIBIDO', dinero_in['total']),
        ('', None),
        ('TOTAL EGRESOS', egresos_total),
        ('', None),
        ('UTILIDAD DEL MES', utilidad),
    ]
    for r, (label, value) in enumerate(rows, start=3):
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        if value is not None:
            ws.cell(row=r, column=2, value=float(value))

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="reporte_{year}_{month:02d}_econotec.xlsx"'
    )
    return response


@admin_requerido
def export_egresos(request):
    """Exporta egresos a Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Egresos Econotec'

    headers = ['Fecha', 'Categoría', 'Concepto', 'Monto', 'Notas', 'Registrado por']
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='F97618')
        c.alignment = Alignment(horizontal='center')

    qs = Egreso.objects.select_related('categoria', 'registrado_por').order_by('-fecha')
    for row, e in enumerate(qs, start=2):
        ws.cell(row=row, column=1, value=e.fecha.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=e.categoria.nombre)
        ws.cell(row=row, column=3, value=e.concepto)
        ws.cell(row=row, column=4, value=float(e.monto))
        ws.cell(row=row, column=5, value=e.notas)
        ws.cell(row=row, column=6,
                value=e.registrado_por.get_full_name() if e.registrado_por else '')

    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 22

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="egresos_econotec.xlsx"'
    return response


@admin_requerido
def admin_mantenimiento_reset(request):
    """
    Vista de Mantenimiento para respaldar y borrar todos los datos transaccionales.
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'download_backup':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font
            except ImportError:
                messages.error(request, 'No se pudo generar el Excel (falta openpyxl).')
                return redirect('econotec:admin_mantenimiento_reset')
                
            wb = Workbook()
            
            # Pestaña 1: Clientes
            ws_clientes = wb.active
            ws_clientes.title = 'Clientes'
            ws_clientes.append(['ID', 'Nombres', 'Cedula', 'WhatsApp', 'Correo', 'Sector', 'Fecha Registro'])
            for col in range(1, 8): ws_clientes.cell(row=1, column=col).font = Font(bold=True)
            for c in Cliente.objects.all():
                ws_clientes.append([c.id, c.nombres, c.cedula, c.whatsapp, c.correo, c.sector, str(c.creado.date())])

            # Pestaña 2: Equipos
            ws_equipos = wb.create_sheet(title='Equipos')
            ws_equipos.append(['Codigo', 'Cliente', 'Tipo', 'Marca', 'Modelo', 'Serie', 'Sede', 'Estado', 'Valor Acordado', 'Fecha Ingreso'])
            for col in range(1, 11): ws_equipos.cell(row=1, column=col).font = Font(bold=True)
            for eq in IngresoEquipo.objects.select_related('cliente').all():
                ws_equipos.append([
                    eq.codigo_equipo, eq.cliente.nombres, eq.tipo_equipo, eq.marca, eq.modelo_serie, eq.serie,
                    eq.get_sede_display(), eq.get_estado_display(), 
                    float(eq.valor_acordado or 0), str(eq.fecha_ingreso)
                ])
                
            # Pestaña 3: Pagos
            ws_pagos = wb.create_sheet(title='Pagos (Abonos)')
            ws_pagos.append(['Recibo', 'Equipo', 'Cliente', 'Monto', 'Metodo', 'Fecha'])
            for col in range(1, 7): ws_pagos.cell(row=1, column=col).font = Font(bold=True)
            for p in Abono.objects.select_related('ingreso__cliente').all():
                ws_pagos.append([
                    p.numero_recibo, p.ingreso.codigo_equipo, p.ingreso.cliente.nombres, 
                    float(p.monto), p.get_metodo_display(), str(p.fecha)
                ])

            # Pestaña 4: Egresos
            ws_egresos = wb.create_sheet(title='Egresos')
            ws_egresos.append(['Concepto', 'Categoria', 'Monto', 'Fecha'])
            for col in range(1, 5): ws_egresos.cell(row=1, column=col).font = Font(bold=True)
            for eg in Egreso.objects.select_related('categoria').all():
                ws_egresos.append([
                    eg.concepto, eg.categoria.nombre if eg.categoria else '', 
                    float(eg.monto), str(eg.fecha)
                ])

            # Pestaña 5: Bodegaje
            from .alertas import salidas_bodegaje_qs
            ws_bodegaje = wb.create_sheet(title='Bodegaje')
            ws_bodegaje.append(['Equipo', 'Cliente', 'Sede', 'Dias en Bodega', 'Monto Acumulado', 'Fecha Salida'])
            for col in range(1, 7): ws_bodegaje.cell(row=1, column=col).font = Font(bold=True)
            for s in salidas_bodegaje_qs(incluir_silenciados=True).select_related('ingreso__cliente'):
                bod = s.calcular_bodegaje()
                if bod['dias'] > 0:
                    ws_bodegaje.append([
                        s.ingreso.codigo_equipo, s.ingreso.cliente.nombres, 
                        s.ingreso.get_sede_display(), bod['dias'], 
                        float(bod['monto']), str(s.fecha_salida)
                    ])

            # Pestaña 6: Chatarrizacion
            ws_chatarra = wb.create_sheet(title='Chatarrizados')
            ws_chatarra.append(['Equipo', 'Cliente', 'Sede', 'Fecha Salida', 'Fecha Chatarr.'])
            for col in range(1, 6): ws_chatarra.cell(row=1, column=col).font = Font(bold=True)
            for s in SalidaEquipo.objects.filter(estado_reparacion='chatarrerizacion').select_related('ingreso__cliente'):
                ws_chatarra.append([
                    s.ingreso.codigo_equipo, s.ingreso.cliente.nombres, 
                    s.ingreso.get_sede_display(), str(s.fecha_salida), str(s.fecha_retiro_real)
                ])

            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            
            response = HttpResponse(
                buf.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            filename = f"econotec_full_backup_{date.today().strftime('%Y%m%d')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        elif action == 'reset_all':
            admin_password = request.POST.get('admin_password', '')
            if not request.user.check_password(admin_password):
                messages.error(request, 'Contraseña de administrador incorrecta. Se ha cancelado el reinicio por seguridad.')
                return redirect('econotec:admin_mantenimiento_reset')
                
            # Borrar en orden inverso de dependencias para no violar foreign keys
            Egreso.objects.all().delete()
            SalidaEquipo.objects.all().delete()
            Abono.objects.all().delete()
            IngresoEquipo.objects.all().delete()
            Cliente.objects.all().delete()
            
            messages.success(request, '¡La base de datos transaccional ha sido reseteada por completo! Los correlativos de G y U empezarán desde 1.')
            return redirect('econotec:admin_dashboard')
            
    return render(request, 'admin/mantenimiento_confirmar.html')


# ═════════════════════════════════════════════════════════════════
# Gamificación - Reseteo y Exportación
# ═════════════════════════════════════════════════════════════════
from django.contrib.auth.models import User
from django.utils import timezone
from .models import UsuarioActividad

def _obtener_estadisticas_gamificacion():
    stats = []
    usuarios = User.objects.filter(is_active=True, groups__name__in=['Tecnicos', 'Asesores']).distinct()
    
    for u in usuarios:
        fecha_reinicio = None
        if hasattr(u, 'actividad') and u.actividad.fecha_reinicio_perfil:
            fecha_reinicio = u.actividad.fecha_reinicio_perfil

        ingresos_qs = IngresoEquipo.objects.filter(
            sede__in=SEDES_EQUIPOS,
            tecnico_encargado=u,
        )
        # El nivel se calcula por las salidas que el técnico REPARÓ, no por
        # las que registró ni por los ingresos (misma regla que api_perfil).
        salidas_qs = SalidaEquipo.objects.filter(tecnico_reparo=u)
        ventas_producto_qs = IngresoEquipo.objects.filter(
            sede='ventas',
            tecnico_encargado=u,
        )

        if fecha_reinicio:
            ingresos_qs = ingresos_qs.filter(creado__gte=fecha_reinicio)
            salidas_qs = salidas_qs.filter(creado__gte=fecha_reinicio)
            ventas_producto_qs = ventas_producto_qs.filter(creado__gte=fecha_reinicio)

        ingresos = ingresos_qs.count()
        ventas_producto = ventas_producto_qs.count()
        salidas_buenas = salidas_qs.filter(estado_reparacion__in=SALIDA_BUENA_ESTADOS).count()
        salidas_malas = salidas_qs.filter(estado_reparacion__in=SALIDA_MALA_ESTADOS).count()
        salidas_garantia = salidas_qs.filter(estado_reparacion__in=SALIDA_GARANTIA_ESTADOS).count()
        
        total = calcular_puntaje_gamificacion(
            salidas_buenas,
            ventas_producto,
            salidas_malas,
            salidas_garantia,
        )
        
        if total <= 49:
            nivel = 'Novato'
        elif total <= 99:
            nivel = 'Intermedio'
        elif total <= 499:
            nivel = 'Avanzado'
        elif total <= 999:
            nivel = 'Experto'
        elif total <= 3999:
            nivel = 'Maestro'
        else:
            nivel = 'God Tec Econotec'
            
        stats.append({
            'usuario': f"{u.first_name} {u.last_name}".strip() or u.username,
            'ingresos': ingresos,
            'buenas': salidas_buenas,
            'producto': ventas_producto,
            'malas': salidas_malas,
            'total': total,
            'nivel': nivel
        })
    
    # Ordenar por puntaje total descendente
    stats.sort(key=lambda x: x['total'], reverse=True)
    return stats

@admin_requerido
@require_POST
def admin_perfiles_reiniciar(request):
    password = request.POST.get('password', '')
    if not request.user.check_password(password):
        messages.error(request, 'Contraseña incorrecta. No se han reiniciado los perfiles.')
        return redirect('econotec:admin_dashboard')
        
    ahora = timezone.now()
    usuarios = User.objects.filter(groups__name__in=['Tecnicos', 'Asesores']).distinct()
    for u in usuarios:
        act, created = UsuarioActividad.objects.get_or_create(user=u)
        act.fecha_reinicio_perfil = ahora
        act.save()
        
    messages.success(request, '¡Perfiles gamificados reiniciados con éxito! Ahora todos empiezan desde cero.')
    return redirect('econotec:admin_dashboard')

@admin_requerido
def admin_perfiles_exportar(request, formato):
    stats = _obtener_estadisticas_gamificacion()
    fecha_str = timezone.now().strftime("%Y%m%d_%H%M")
    
    if formato == 'excel':
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ranking Perfiles"
        
        # Headers
        headers = ['Posición', 'Técnico / Asesor', 'Ingresos Asignados', 'Salidas Buenas', 'Salida de Producto', 'Salidas Malas', 'Puntaje Total', 'Nivel Alcanzado']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            
        # Data
        for row_num, stat in enumerate(stats, 2):
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=stat['usuario'])
            ws.cell(row=row_num, column=3, value=stat['ingresos'])
            ws.cell(row=row_num, column=4, value=stat['buenas'])
            ws.cell(row=row_num, column=5, value=stat['producto'])
            ws.cell(row=row_num, column=6, value=stat['malas'])
            ws.cell(row=row_num, column=7, value=stat['total'])
            ws.cell(row=row_num, column=8, value=stat['nivel'])
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Ranking_Perfiles_{fecha_str}.xlsx'
        wb.save(response)
        return response
        
    elif formato == 'pdf':
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=Ranking_Perfiles_{fecha_str}.pdf'
        
        c = canvas.Canvas(response, pagesize=A4)
        y = 800
        
        c.setFont("Helvetica-Bold", 16)
        c.drawString(50, y, "Econotec - Ranking de Productividad (Gamificacion)")
        y -= 30
        
        c.setFont("Helvetica", 10)
        c.drawString(50, y, f"Fecha de reporte: {timezone.now().strftime('%d/%m/%Y %H:%M')}")
        y -= 40
        
        # Table Header
        c.setFont("Helvetica-Bold", 10)
        c.drawString(50, y, "Pos")
        c.drawString(90, y, "Usuario")
        c.drawString(220, y, "Ing. asign.")
        c.drawString(280, y, "S. Buenas")
        c.drawString(340, y, "S. Prod.")
        c.drawString(400, y, "S. Malas")
        c.drawString(460, y, "Puntos")
        c.drawString(510, y, "Nivel")
        y -= 20
        
        c.line(50, y+10, 550, y+10)
        
        c.setFont("Helvetica", 10)
        for i, stat in enumerate(stats, 1):
            if y < 50:
                c.showPage()
                y = 800
            c.drawString(50, y, str(i))
            c.drawString(90, y, str(stat['usuario'])[:25])
            c.drawString(220, y, str(stat['ingresos']))
            c.drawString(280, y, str(stat['buenas']))
            c.drawString(340, y, str(stat['producto']))
            c.drawString(400, y, str(stat['malas']))
            c.drawString(460, y, str(stat['total']))
            c.drawString(510, y, str(stat['nivel']))
            y -= 20
            
        c.save()
        return response
        
    return redirect('econotec:admin_dashboard')


# ═════════════════════════════════════════════════════════════════
# Avisos del panel principal (solo administrador)
# ═════════════════════════════════════════════════════════════════

@admin_requerido
def avisos_lista(request):
    """Lista de avisos del panel. Solo el administrador puede gestionarlos."""
    avisos = AvisoPanel.objects.select_related('creado_por').all()
    vigentes = sum(1 for a in avisos if a.vigente)
    return render(request, 'admin_panel/avisos_lista.html', {
        'avisos': avisos,
        'total': avisos.count(),
        'vigentes': vigentes,
    })


@admin_requerido
def aviso_crear(request):
    from .forms import AvisoPanelForm
    if request.method == 'POST':
        form = AvisoPanelForm(request.POST)
        if form.is_valid():
            aviso = form.save(commit=False)
            aviso.creado_por = request.user
            aviso.save()
            messages.success(request, 'Aviso creado. Se mostrará en el inicio según sus fechas.')
            return redirect('econotec:avisos_lista')
    else:
        from datetime import date as _date
        form = AvisoPanelForm(initial={
            'fecha_inicio': _date.today(),
            'fecha_fin': _date.today(),
            'activo': True,
            'tipo': 'info',
        })
    return render(request, 'admin_panel/aviso_form.html', {
        'form': form, 'modo': 'crear', 'titulo': 'Nuevo aviso',
    })


@admin_requerido
def aviso_editar(request, pk):
    from .forms import AvisoPanelForm
    aviso = get_object_or_404(AvisoPanel, pk=pk)
    if request.method == 'POST':
        form = AvisoPanelForm(request.POST, instance=aviso)
        if form.is_valid():
            form.save()
            messages.success(request, 'Aviso actualizado.')
            return redirect('econotec:avisos_lista')
    else:
        form = AvisoPanelForm(instance=aviso)
    return render(request, 'admin_panel/aviso_form.html', {
        'form': form, 'modo': 'editar', 'aviso': aviso,
        'titulo': f'Editar aviso — {aviso.titulo}',
    })


@admin_requerido
@require_POST
def aviso_eliminar(request, pk):
    aviso = get_object_or_404(AvisoPanel, pk=pk)
    aviso.delete()
    messages.success(request, 'Aviso eliminado.')
    return redirect('econotec:avisos_lista')


# ═════════════════════════════════════════════════════════════════
# Control de Registro (Auditoría) — solo administrador
# ═════════════════════════════════════════════════════════════════

@admin_requerido
def control_registro(request):
    """
    Auditoría: últimos equipos registrados y últimos pagos registrados en el
    sistema, con fecha/hora, quién los registró (asesor) y el cliente.
    Solo visible para administradores.
    """
    equipos_qs = (
        IngresoEquipo.objects
        .select_related('cliente', 'registrado_por', 'tecnico_encargado')
        .order_by('-creado')
    )

    abonos_qs = (
        Abono.objects
        .select_related('ingreso', 'ingreso__cliente', 'registrado_por')
        .order_by('-creado')
    )
    equipos_page_obj, equipos_querystring = paginar_resultados(
        request,
        equipos_qs,
        page_param='pagina_equipos',
    )
    abonos_page_obj, abonos_querystring = paginar_resultados(
        request,
        abonos_qs,
        page_param='pagina_pagos',
    )

    return render(request, 'admin_panel/control_registro.html', {
        'equipos': equipos_page_obj.object_list,
        'abonos': abonos_page_obj.object_list,
        'equipos_page_obj': equipos_page_obj,
        'abonos_page_obj': abonos_page_obj,
        'equipos_querystring': equipos_querystring,
        'abonos_querystring': abonos_querystring,
        'limite': 10,
    })
