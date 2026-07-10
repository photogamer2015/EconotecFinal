"""
Vistas de Pagos / Abonos / Historial.
"""
from collections import defaultdict
from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib import messages
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST

from .forms import AbonoForm
from .models import Abono, IngresoEquipo
from .permisos import tecnico_requerido, asesor_requerido


MESES_ES = [
    '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
]

@asesor_requerido
def pagos_lista(request):
    """
    Vista principal de Pagos (REPARACIONES): lista los ingresos de taller con
    su estado de pago. Las ventas de producto (sede='ventas') se muestran en su
    propio apartado (`pagos_ventas_lista`) para no mezclarlas.
    """
    base_qs = IngresoEquipo.objects.exclude(sede='ventas')
    context = _construir_contexto_pagos(request, base_qs)
    context['modo'] = 'reparaciones'
    return render(request, 'pagos/lista.html', context)


@asesor_requerido
def pagos_ventas_lista(request):
    """
    Control de Pagos exclusivo de VENTAS DE PRODUCTO (sede='ventas').
    Mismo formato que Pagos, pero solo con las ventas de producto.
    """
    base_qs = IngresoEquipo.objects.filter(sede='ventas')
    context = _construir_contexto_pagos(request, base_qs)
    context['modo'] = 'ventas'
    return render(request, 'pagos/lista.html', context)


def _construir_contexto_pagos(request, base_qs):
    """
    Construye el contexto de la lista de pagos a partir de un queryset base
    (reparaciones o ventas). Aplica los filtros de búsqueda / estado / sede.
    """
    q = (request.GET.get('q') or '').strip()
    estado_pago = (request.GET.get('estado_pago') or '').strip()
    sede_pago = (request.GET.get('sede_pago') or '').strip()

    qs = (base_qs
          .select_related('cliente')
          .prefetch_related('abonos')
          .order_by('-fecha_ingreso'))

    if q:
        import re
        q_filter = (
            Q(cliente__cedula__icontains=q) |
            Q(cliente__nombres__icontains=q) |
            Q(marca__icontains=q) |
            Q(numero_equipo__icontains=q) |
            Q(sede__icontains=q) |
            Q(tipo_equipo__icontains=q) |
            Q(tipo_equipo_otro__icontains=q)
        )
        
        # Si el usuario busca "G1000", extraemos "1000" para buscar en numero_equipo
        digitos = re.sub(r'\D', '', q)
        if digitos:
            q_filter |= Q(numero_equipo__icontains=digitos)
            
        qs = qs.filter(q_filter)

    if sede_pago:
        qs = qs.filter(sede=sede_pago)

    ingresos = list(qs)

    # Filtrar por estado de pago en Python (es propiedad)
    if estado_pago:
        ingresos = [i for i in ingresos if i.estado_pago.lower() == estado_pago.lower()]

    total_acordado = sum((i.valor_efectivo_a_cobrar for i in ingresos), Decimal('0.00'))
    total_pagado = sum((i.total_abonado for i in ingresos), Decimal('0.00'))
    total_pendiente = total_acordado - total_pagado
    total_bodegaje = sum((i.bodegaje_pendiente for i in ingresos), Decimal('0.00'))
    total_diagnostico = sum(((i.valor_diagnostico or Decimal('0.00')) for i in ingresos if i.diagnostico_inmediato == 'si'), Decimal('0.00'))

    return {
        'q': q,
        'estado_pago': estado_pago,
        'sede_pago': sede_pago,
        'ingresos': ingresos,
        'total_acordado': total_acordado,
        'total_pagado': total_pagado,
        'total_pendiente': total_pendiente,
        'total_bodegaje': total_bodegaje,
        'total_diagnostico': total_diagnostico,
        'total_count': len(ingresos),
    }


@asesor_requerido
def pagos_export(request):
    """Exportar pagos a Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Pagos Econotec'

    headers = [
        'N° Equipo', 'Fecha Ingreso', 'Cliente', 'Cédula',
        'Equipo', 'Marca', 'Valor Acordado', 'Adicional (Diag.)',
        'Total Abonado', 'Diferencia', 'Estado', 'Métodos de Pago'
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='F97618')
        c.alignment = Alignment(horizontal='center')

    qs = (IngresoEquipo.objects
          .select_related('cliente')
          .prefetch_related('abonos')
          .order_by('-fecha_ingreso'))

    for row, ing in enumerate(qs, start=2):
        ws.cell(row=row, column=1, value=ing.codigo_equipo)
        ws.cell(row=row, column=2, value=ing.fecha_ingreso.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=3, value=ing.cliente.nombres)
        ws.cell(row=row, column=4, value=ing.cliente.cedula)
        ws.cell(row=row, column=5, value=ing.tipo_equipo_display)
        ws.cell(row=row, column=6, value=ing.marca)
        ws.cell(row=row, column=7, value=float(ing.valor_acordado or 0))
        
        diag_val = float(ing.valor_diagnostico or 0) if ing.diagnostico_inmediato == 'si' else 0.0
        ws.cell(row=row, column=8, value=diag_val)
        
        ws.cell(row=row, column=9, value=float(ing.total_abonado))
        ws.cell(row=row, column=10, value=float(ing.diferencia))
        ws.cell(row=row, column=11, value=ing.estado_pago)
        ws.cell(row=row, column=12, value=ing.resumen_metodos_pago)

    for col in range(1, 13):
        ws.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="pagos_econotec.xlsx"'
    return response


# ═════════════════════════════════════════════════════════════════
# Abonos por ingreso
# ═════════════════════════════════════════════════════════════════

@asesor_requerido
def ingreso_abonos(request, pk):
    """Lista los abonos de un ingreso específico."""
    ingreso = get_object_or_404(
        IngresoEquipo.objects.select_related('cliente'),
        pk=pk,
    )
    abonos = ingreso.abonos.select_related('registrado_por').order_by('-fecha', '-creado')
    salida = getattr(ingreso, 'salida', None)
    return render(request, 'pagos/ingreso_abonos.html', {
        'ingreso': ingreso,
        'abonos': abonos,
        'salida': salida,
    })


@asesor_requerido
def abono_crear(request, ingreso_pk):
    """Registrar un nuevo abono para un ingreso."""
    ingreso = get_object_or_404(IngresoEquipo, pk=ingreso_pk)

    if request.method == 'POST':
        form = AbonoForm(request.POST, ingreso=ingreso)
        if form.is_valid():
            abono = form.save(commit=False)
            
            if abono.metodo == 'mixto':
                monto_1 = request.POST.get('abono_monto_1')
                metodo_1 = request.POST.get('abono_metodo_1')
                banco_1 = request.POST.get('abono_banco_1')
                
                monto_2 = request.POST.get('abono_monto_2')
                metodo_2 = request.POST.get('abono_metodo_2')
                banco_2 = request.POST.get('abono_banco_2')
                
                from decimal import Decimal
                obs_base = abono.observaciones or ''
                
                if monto_1 and Decimal(monto_1) > 0:
                    a1 = Abono(
                        ingreso=ingreso, fecha=abono.fecha, monto=Decimal(monto_1),
                        metodo=metodo_1, banco=banco_1 if metodo_1 == 'transferencia' else '',
                        bodegaje_decision=abono.bodegaje_decision,
                        bodegaje_monto_aplicado=abono.bodegaje_monto_aplicado,
                        observaciones=('Pago Mixto (Parte 1). ' + obs_base).strip(),
                        registrado_por=request.user
                    )
                    a1.save()
                
                if monto_2 and Decimal(monto_2) > 0:
                    a2 = Abono(
                        ingreso=ingreso, fecha=abono.fecha, monto=Decimal(monto_2),
                        metodo=metodo_2, banco=banco_2 if metodo_2 == 'transferencia' else '',
                        bodegaje_decision='na',
                        bodegaje_monto_aplicado=0,
                        observaciones=('Pago Mixto (Parte 2). ' + obs_base).strip(),
                        registrado_por=request.user
                    )
                    a2.save()
                    
                messages.success(request, f'Abono mixto registrado exitosamente.')
                return redirect('econotec:ingreso_abonos', pk=ingreso.pk)
            else:
                abono.ingreso = ingreso
                abono.registrado_por = request.user
                abono.save()
                messages.success(
                    request,
                    f'Abono {abono.numero_recibo} registrado por ${abono.monto}.'
                )
                return redirect('econotec:ingreso_abonos', pk=ingreso.pk)
    else:
        form = AbonoForm(initial={'fecha': date.today(), 'monto': ingreso.diferencia}, ingreso=ingreso)

    return render(request, 'pagos/abono_form.html', {
        'form': form,
        'ingreso': ingreso,
        'modo': 'crear',
        'titulo': f'Nuevo Abono — Equipo {ingreso.codigo_equipo}',
    })


@asesor_requerido
def abono_editar(request, ingreso_pk, abono_pk):
    ingreso = get_object_or_404(IngresoEquipo, pk=ingreso_pk)
    abono = get_object_or_404(Abono, pk=abono_pk, ingreso=ingreso)

    if request.method == 'POST':
        form = AbonoForm(request.POST, instance=abono, ingreso=ingreso)
        if form.is_valid():
            form.save()
            messages.success(request, 'Abono actualizado.')
            return redirect('econotec:ingreso_abonos', pk=ingreso.pk)
    else:
        form = AbonoForm(instance=abono, ingreso=ingreso)

    return render(request, 'pagos/abono_form.html', {
        'form': form,
        'ingreso': ingreso,
        'abono': abono,
        'modo': 'editar',
        'titulo': f'Editar Abono {abono.numero_recibo}',
    })


@asesor_requerido
@require_POST
def abono_eliminar(request, ingreso_pk, abono_pk):
    ingreso = get_object_or_404(IngresoEquipo, pk=ingreso_pk)
    abono = get_object_or_404(Abono, pk=abono_pk, ingreso=ingreso)
    numero = abono.numero_recibo
    abono.delete()
    messages.success(request, f'Abono {numero} eliminado.')
    return redirect('econotec:ingreso_abonos', pk=ingreso.pk)


@asesor_requerido
def abono_recibo(request, abono_pk):
    """Vista de recibo imprimible de un abono."""
    abono = get_object_or_404(
        Abono.objects.select_related('ingreso', 'ingreso__cliente', 'registrado_por'),
        pk=abono_pk,
    )
    return render(request, 'pagos/recibo.html', {
        'abono': abono,
    })


# ═════════════════════════════════════════════════════════════════
# Historial (organizado por año/mes)
# ═════════════════════════════════════════════════════════════════

@tecnico_requerido
def historial_lista(request):
    """Historial de ingresos agrupados por año y mes."""
    ano_filtro = (request.GET.get('ano') or '').strip()
    mes_filtro = (request.GET.get('mes') or '').strip()

    qs = (IngresoEquipo.objects
          .select_related('cliente')
          .prefetch_related('abonos', 'salida')
          .order_by('-fecha_ingreso'))

    if ano_filtro and ano_filtro.isdigit():
        qs = qs.filter(fecha_ingreso__year=int(ano_filtro))
    if mes_filtro and mes_filtro.isdigit():
        qs = qs.filter(fecha_ingreso__month=int(mes_filtro))

    # Agrupar por año-mes
    grupos = defaultdict(list)
    for ing in qs:
        key = (ing.fecha_ingreso.year, ing.fecha_ingreso.month)
        grupos[key].append(ing)

    # Ordenar por año-mes desc
    grupos_ordenados = []
    for (ano, mes), lista in sorted(grupos.items(), reverse=True):
        total_acordado = sum((i.valor_acordado or Decimal('0.00') for i in lista), Decimal('0.00'))
        total_pagado = sum((i.total_abonado for i in lista), Decimal('0.00'))
        
        # Agrupar internamente por tipo_equipo_display
        tipos_dict = defaultdict(list)
        for i in lista:
            tipos_dict[i.tipo_equipo_display].append(i)
            
        tipos_ordenados = []
        for tipo, items in sorted(tipos_dict.items()):
            tipos_ordenados.append({
                'tipo': tipo,
                'total': len(items),
                'ingresos': items,
            })

        grupos_ordenados.append({
            'ano': ano,
            'mes': mes,
            'mes_nombre': MESES_ES[mes],
            'tipos': tipos_ordenados,
            'total_equipos': len(lista),
            'total_acordado': total_acordado,
            'total_pagado': total_pagado,
        })

    # Años disponibles para filtrar
    anos_disponibles = sorted(
        IngresoEquipo.objects.dates('fecha_ingreso', 'year', order='DESC'),
        reverse=True,
    )
    anos_disponibles = [d.year for d in anos_disponibles]

    return render(request, 'historial/lista.html', {
        'grupos': grupos_ordenados,
        'ano_filtro': ano_filtro,
        'mes_filtro': mes_filtro,
        'anos_disponibles': anos_disponibles,
        'meses_es': MESES_ES,
    })


@tecnico_requerido
def historial_export(request):
    """Exportar historial completo a Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Historial Econotec'

    headers = [
        'N° Equipo', 'Fecha Ingreso', 'Cliente', 'Cédula',
        'Tipo', 'Marca', 'Modelo', 'Serie', 'Problema',
        'Valor Acordado', 'Total Abonado', 'Diferencia',
        'Estado Equipo', 'Estado Pago',
        'Fecha Salida', 'Estado Reparación',
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='F97618')
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[1].height = 30

    qs = (IngresoEquipo.objects
          .select_related('cliente', 'salida')
          .prefetch_related('abonos')
          .order_by('-fecha_ingreso'))

    ano_filtro = request.GET.get('ano')
    mes_filtro = request.GET.get('mes')
    if ano_filtro and ano_filtro.isdigit():
        qs = qs.filter(fecha_ingreso__year=int(ano_filtro))
    if mes_filtro and mes_filtro.isdigit():
        qs = qs.filter(fecha_ingreso__month=int(mes_filtro))

    for row, ing in enumerate(qs, start=2):
        ws.cell(row=row, column=1, value=ing.codigo_equipo)
        ws.cell(row=row, column=2, value=ing.fecha_ingreso.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=3, value=ing.cliente.nombres)
        ws.cell(row=row, column=4, value=ing.cliente.cedula)
        ws.cell(row=row, column=5, value=ing.tipo_equipo_display)
        ws.cell(row=row, column=6, value=ing.marca)
        ws.cell(row=row, column=7, value=ing.modelo_serie)
        ws.cell(row=row, column=8, value=ing.serie)
        ws.cell(row=row, column=9, value=ing.problema_reportado)
        ws.cell(row=row, column=10, value=float(ing.valor_acordado or 0))
        ws.cell(row=row, column=11, value=float(ing.total_abonado))
        ws.cell(row=row, column=12, value=float(ing.diferencia))
        ws.cell(row=row, column=13, value=ing.get_estado_display())
        ws.cell(row=row, column=14, value=ing.estado_pago)
        salida = getattr(ing, 'salida', None)
        if salida:
            ws.cell(row=row, column=15, value=salida.fecha_salida.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=16, value=salida.get_estado_reparacion_display())
        else:
            ws.cell(row=row, column=15, value='—')
            ws.cell(row=row, column=16, value='—')

    for col in range(1, 17):
        ws.column_dimensions[chr(64 + col)].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    nombre_archivo = "historial_econotec.xlsx"
    if ano_filtro and mes_filtro:
        nombre_archivo = f"historial_econotec_{ano_filtro}_{mes_filtro}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


@tecnico_requerido
def historial_imprimir(request):
    """Vista HTML para imprimir el historial de un mes específico."""
    ano_filtro = request.GET.get('ano')
    mes_filtro = request.GET.get('mes')
    
    if not (ano_filtro and mes_filtro and ano_filtro.isdigit() and mes_filtro.isdigit()):
        return redirect('econotec:historial_lista')
        
    ano, mes = int(ano_filtro), int(mes_filtro)

    qs = (IngresoEquipo.objects
          .select_related('cliente')
          .prefetch_related('abonos', 'salida')
          .filter(fecha_ingreso__year=ano, fecha_ingreso__month=mes)
          .order_by('-fecha_ingreso'))

    lista = list(qs)
    total_acordado = sum((i.valor_acordado or Decimal('0.00') for i in lista), Decimal('0.00'))
    total_pagado = sum((i.total_abonado for i in lista), Decimal('0.00'))

    tipos_dict = defaultdict(list)
    for i in lista:
        tipos_dict[i.tipo_equipo_display].append(i)

    tipos_ordenados = []
    for tipo, items in sorted(tipos_dict.items()):
        tipos_ordenados.append({
            'tipo': tipo,
            'total': len(items),
            'ingresos': items,
        })

    grupo = {
        'ano': ano,
        'mes': mes,
        'mes_nombre': MESES_ES[mes],
        'tipos': tipos_ordenados,
        'total_equipos': len(lista),
        'total_acordado': total_acordado,
        'total_pagado': total_pagado,
    }

    return render(request, 'historial/imprimir.html', {
        'grupo': grupo,
    })
