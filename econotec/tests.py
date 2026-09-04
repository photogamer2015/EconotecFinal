import json
import re
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from io import BytesIO
from unittest.mock import ANY, patch
from urllib.parse import parse_qs, urlparse
from zoneinfo import ZoneInfo

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core import mail
from django.db import connection
from django.test import TestCase, override_settings
from django.test.utils import CaptureQueriesContext
from django.urls import reverse
from django.utils import timezone

from .forms import IngresoEquipoForm
from . import views_print
from .alertas import (
    equipos_demorados_qs,
    salidas_bodegaje_qs,
    whatsapp_link_bodegaje,
    whatsapp_link_equipo_listo,
    whatsapp_link_venta_producto,
)
from .horarios import registrar_entrada_laboral
from .models import (
    Abono, BitacoraTecnico, Cliente, Egreso, HorarioTecnico, IngresoEquipo, InventarioItem,
    NotificacionAsesora, NotificacionInventarioAdmin, SalidaEquipo, UsuarioActividad,
    VentaInventarioItem,
)
from .qr_utils import token_para_ingreso
from .views_auth import CAPTCHA_SESSION_KEY, LOGIN_2FA_SESSION_KEY, LOGIN_EMAIL_SETUP_SESSION_KEY


@override_settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend')
class LoginCaptchaTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.usuario = User.objects.create_user(
            username='captcha_user',
            password='testpass123',
            email='captcha_user@example.com',
        )
        self.usuario_sin_correo = User.objects.create_user(
            username='sin_correo',
            password='testpass123',
        )

    def _captcha_answer(self):
        response = self.client.get(reverse('login'))
        self.assertEqual(response.status_code, 200)
        return str(self.client.session[CAPTCHA_SESSION_KEY])

    def test_login_modal_sede_usa_imagenes_de_sede(self):
        response = self.client.get(reverse('login'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="sede-confirm-img"')
        self.assertContains(response, 'inventario/guayaquil.jpg')
        self.assertContains(response, 'inventario/quito.jpg')

    def test_login_rechaza_captcha_incorrecto(self):
        respuesta = self._captcha_answer()

        response = self.client.post(reverse('login'), {
            'username': self.usuario.username,
            'password': 'testpass123',
            'sede': 'guayaquil',
            'captcha_respuesta': str(int(respuesta) + 1),
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Resuelve correctamente la suma de seguridad.')
        self.assertNotIn('_auth_user_id', self.client.session)

    def test_login_acepta_captcha_correcto_y_envia_codigo(self):
        respuesta = self._captcha_answer()

        response = self.client.post(reverse('login'), {
            'username': self.usuario.username,
            'password': 'testpass123',
            'sede': 'quito',
            'captcha_respuesta': respuesta,
        })

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse('login_2fa'))
        self.assertNotIn('_auth_user_id', self.client.session)
        self.assertEqual(self.client.session[LOGIN_2FA_SESSION_KEY]['sede'], 'quito')
        self.assertEqual(len(mail.outbox), 1)
        self.assertIn('Código de acceso Econotec', mail.outbox[0].subject)
        html = next(
            alternativa[0]
            for alternativa in mail.outbox[0].alternatives
            if alternativa[1] == 'text/html'
        )
        self.assertIn('Econotec', html)
        self.assertIn('Verificación segura', html)
        self.assertIn('#f97618', html)

    def test_doble_factor_acepta_codigo_correcto_y_guarda_sede(self):
        respuesta = self._captcha_answer()
        self.client.post(reverse('login'), {
            'username': self.usuario.username,
            'password': 'testpass123',
            'sede': 'quito',
            'captcha_respuesta': respuesta,
        })
        codigo = re.search(r'\b(\d{6})\b', mail.outbox[0].body).group(1)

        response = self.client.post(reverse('login_2fa'), {
            'codigo': codigo,
        })

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse('econotec:bienvenida'))
        self.assertEqual(self.client.session['sede_actual'], 'quito')
        self.assertEqual(int(self.client.session['_auth_user_id']), self.usuario.pk)
        self.assertNotIn(LOGIN_2FA_SESSION_KEY, self.client.session)

    def test_doble_factor_rechaza_codigo_incorrecto(self):
        respuesta = self._captcha_answer()
        self.client.post(reverse('login'), {
            'username': self.usuario.username,
            'password': 'testpass123',
            'sede': 'guayaquil',
            'captcha_respuesta': respuesta,
        })
        codigo = re.search(r'\b(\d{6})\b', mail.outbox[0].body).group(1)
        codigo_incorrecto = '000000' if codigo != '000000' else '111111'

        response = self.client.post(reverse('login_2fa'), {
            'codigo': codigo_incorrecto,
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Código incorrecto.')
        self.assertContains(response, 'Te quedan 9 intentos.')
        self.assertNotIn('_auth_user_id', self.client.session)

    def test_login_sin_correo_pide_registro_de_correo(self):
        respuesta = self._captcha_answer()

        response = self.client.post(reverse('login'), {
            'username': self.usuario_sin_correo.username,
            'password': 'testpass123',
            'sede': 'guayaquil',
            'captcha_respuesta': respuesta,
        })

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse('login_registrar_correo'))
        self.assertNotIn('_auth_user_id', self.client.session)
        self.assertEqual(
            self.client.session[LOGIN_EMAIL_SETUP_SESSION_KEY]['user_id'],
            self.usuario_sin_correo.pk,
        )

    def test_qr_inventario_exige_login_y_regresa_al_producto(self):
        item = InventarioItem.objects.create(
            sede='guayaquil',
            categoria='impresora',
            tipo='impresora-laser',
            producto='Tinta de prueba QR',
            marca='Epson',
            modelo='544',
            estado='disponible',
            cantidad=1,
            ubicacion='guayaquil_norte',
            registrado_por=self.usuario,
        )
        detalle_url = reverse(
            'econotec:inventario_detalle_item',
            kwargs={'codigo': item.codigo},
        )
        imprimir_url = reverse(
            'econotec:inventario_qr_imprimir',
            kwargs={'codigo': item.codigo},
        )

        response = self.client.get(detalle_url)
        imprimir_response = self.client.get(imprimir_url)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, f'{reverse("login")}?next={detalle_url}')
        self.assertEqual(imprimir_response.status_code, 302)
        self.assertEqual(
            imprimir_response.url,
            f'{reverse("login")}?next={imprimir_url}',
        )

        login_page = self.client.get(response.url)
        self.assertContains(login_page, f'name="next" value="{detalle_url}"')
        respuesta_captcha = str(self.client.session[CAPTCHA_SESSION_KEY])

        response = self.client.post(reverse('login'), {
            'username': self.usuario.username,
            'password': 'testpass123',
            'sede': 'guayaquil',
            'captcha_respuesta': respuesta_captcha,
            'next': detalle_url,
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse('login_2fa'))

        codigo = re.search(r'\b(\d{6})\b', mail.outbox[0].body).group(1)
        response = self.client.post(reverse('login_2fa'), {'codigo': codigo})

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, detalle_url)
        detalle = self.client.get(response.url)
        self.assertEqual(detalle.status_code, 200)
        self.assertContains(detalle, item.codigo)
        self.assertContains(detalle, 'Tinta de prueba QR')

    def test_registro_correo_verificado_guarda_email_y_entra(self):
        respuesta = self._captcha_answer()
        self.client.post(reverse('login'), {
            'username': self.usuario_sin_correo.username,
            'password': 'testpass123',
            'sede': 'quito',
            'captcha_respuesta': respuesta,
        })

        response = self.client.post(reverse('login_registrar_correo'), {
            'accion': 'enviar_codigo',
            'email': 'nuevo@example.com',
            'email_confirmacion': 'nuevo@example.com',
        })

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse('login_registrar_correo'))
        self.assertEqual(len(mail.outbox), 1)
        self.assertIn('Verifica tu correo Econotec', mail.outbox[0].subject)
        html = next(
            alternativa[0]
            for alternativa in mail.outbox[0].alternatives
            if alternativa[1] == 'text/html'
        )
        self.assertIn('Verifica tu correo', html)
        self.assertIn('Registrar correo', html)
        self.assertIn('#f97618', html)
        self.usuario_sin_correo.refresh_from_db()
        self.assertEqual(self.usuario_sin_correo.email, '')

        codigo = re.search(r'\b(\d{6})\b', mail.outbox[0].body).group(1)
        response = self.client.post(reverse('login_registrar_correo'), {
            'accion': 'verificar',
            'codigo': codigo,
        })

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse('econotec:bienvenida'))
        self.usuario_sin_correo.refresh_from_db()
        self.assertEqual(self.usuario_sin_correo.email, 'nuevo@example.com')
        self.assertEqual(self.client.session['sede_actual'], 'quito')
        self.assertEqual(int(self.client.session['_auth_user_id']), self.usuario_sin_correo.pk)
        self.assertNotIn(LOGIN_EMAIL_SETUP_SESSION_KEY, self.client.session)

    def test_registro_correo_rechaza_email_duplicado(self):
        respuesta = self._captcha_answer()
        self.client.post(reverse('login'), {
            'username': self.usuario_sin_correo.username,
            'password': 'testpass123',
            'sede': 'guayaquil',
            'captcha_respuesta': respuesta,
        })

        response = self.client.post(reverse('login_registrar_correo'), {
            'accion': 'enviar_codigo',
            'email': self.usuario.email,
            'email_confirmacion': self.usuario.email,
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Ese correo ya está registrado en otro usuario.')
        self.assertEqual(len(mail.outbox), 0)
        self.usuario_sin_correo.refresh_from_db()
        self.assertEqual(self.usuario_sin_correo.email, '')


class VentasTests(TestCase):
    FIRMA_PNG_DATA_URI = (
        'data:image/png;base64,'
        'iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNk+M9QDwADhgF/TV5CiwAAAABJRU5ErkJggg=='
    )

    def setUp(self):
        User = get_user_model()
        asesores = Group.objects.create(name='Asesores')
        tecnicos = Group.objects.create(name='Tecnicos')

        self.vendedor = User.objects.create_user(username='Kimberly', email='kimberly@example.com')
        self.vendedor.groups.add(asesores)

        self.usuario = User.objects.create_user(username='Yandri', email='yandri@example.com')
        self.usuario.groups.add(tecnicos)
        self.client.force_login(self.usuario)
        self.admin = User.objects.create_superuser(
            username='RootAdmin',
            email='admin@example.com',
            password='adminpass123',
        )

        self.cliente_existente = Cliente.objects.create(
            cedula='1207342716',
            nombres='Yandri Guevara',
            whatsapp='0939746169',
            correo='yandridavid@hotmail.com',
            sector='norte',
        )

    def crear_producto_venta(self):
        if hasattr(self, 'producto_venta'):
            return self.producto_venta
        self.producto_venta = InventarioItem.objects.create(
            sede='guayaquil',
            categoria='impresora',
            tipo='impresora-laser',
            producto='Tarjeta',
            marca='Epson',
            modelo='Laser-t32',
            estado='disponible',
            cantidad=10,
            ubicacion='guayaquil_norte',
            registrado_por=self.usuario,
        )
        return self.producto_venta

    def venta_post_data(self, **overrides):
        producto_venta = self.crear_producto_venta()
        data = {
            'cli-cedula': self.cliente_existente.cedula,
            'cli-nombres': self.cliente_existente.nombres,
            'cli-whatsapp': self.cliente_existente.whatsapp,
            'cli-correo': self.cliente_existente.correo,
            'cli-sector': self.cliente_existente.sector,
            'cli-sector_otro': '',
            'ing-asesor_comercial': 'Kimberly',
            'ing-fecha_ingreso': '2026-07-09',
            'ing-numero_factura': '',
            'ing-tecnico_encargado': str(self.usuario.pk),
            'ing-problema_reportado': '',
            'ing-valor_acordado': '25',
            'ing-firma_cliente_opcion': 'no',
            'ing-firma_cliente_imagen': '',
            'inventario_seleccionado': json.dumps([
                {'item_id': producto_venta.pk, 'cantidad': 2},
            ]),
            # Valores que el navegador puede enviar desde campos ocultos.
            # Se omite ing-diagnostico_metodo para cubrir el bug corregido.
            'ing-tipo_equipo': 'otro',
            'ing-tipo_equipo_otro': '',
            'ing-marca': 'N/A',
            'ing-modelo_serie': 'N/A',
            'ing-serie': '',
            'ing-accesorios_entregados': 'Ninguno',
            'ing-diagnostico_inmediato': 'no',
            'ing-valor_diagnostico': '0.00',
            'ing-abono_anticipo': '0.00',
            'ing-anticipo_metodo': 'efectivo',
            'ing-estado': 'ingresado',
            'ing-subestado_reparacion': '',
            'ing-subestado_entregado': 'con_solucion',
            'ing-equipo_garantia': '',
            'ing-equipo_garantia_manual': '',
            'ing-motivo_garantia': '',
        }
        data.update(overrides)
        return data

    def crear_ingreso_reparacion(self, **overrides):
        data = {
            'sede': 'guayaquil',
            'asesor_comercial': 'Kimberly',
            'fecha_ingreso': date(2026, 7, 9),
            'cliente': self.cliente_existente,
            'tipo_equipo': 'laptop',
            'marca': 'HP',
            'modelo_serie': 'Elitebook',
            'accesorios_entregados': '',
            'problema_reportado': 'No enciende',
            'valor_acordado': Decimal('25.00'),
            'tecnico_encargado': self.usuario,
            'estado': 'en_reparacion',
            'subestado_reparacion': 'en_reparacion',
            'registrado_por': self.usuario,
        }
        data.update(overrides)
        return IngresoEquipo.objects.create(**data)

    def salida_post_data(self, **overrides):
        data = {
            'fecha_salida': '2026-07-17',
            'estado_reparacion': 'pendiente_retiro',
            'tecnico_reparo': str(self.usuario.pk),
            'reporte_tecnico': 'Equipo revisado.',
            'observaciones': '',
            'valor_final_cobrado': '0.00',
            'valor_acordado_revision': '',
            'aplica_valor_acordado_adicional': 'no',
            'valor_acordado_adicional': '0.00',
            'motivo_valor_acordado_adicional': '',
            'metodo_pago_final': 'efectivo',
            'numero_recibo': '',
            'banco': '',
            'banco_otro': '',
            'tarjeta_app': '',
            'comprobante_url': '',
            'monto_1': '',
            'metodo_1': '',
            'banco_1': '',
            'monto_2': '',
            'metodo_2': '',
            'banco_2': '',
            'factura_realizada': 'no',
            'factura_nombres': '',
            'factura_cedula': '',
            'factura_correo': '',
            'asesora_notificacion': '',
            'mensaje_notificacion': '',
        }
        data.update(overrides)
        return data

    def salida_rapida_post_data(self, **overrides):
        data = {
            f'salida-{campo}': valor
            for campo, valor in self.salida_post_data().items()
        }
        data.update({f'salida-{campo}': valor for campo, valor in overrides.items()})
        return data

    def registrar_ingreso_negativo(self, estado_ingreso, **salida_overrides):
        self.activar_sede_guayaquil()
        ultimo_pk = IngresoEquipo.objects.order_by('-pk').values_list('pk', flat=True).first() or 0
        data = self.ingreso_registro_post_data(**{
            'ing-estado': estado_ingreso,
            'ing-subestado_reparacion': '',
            'ing-subestado_entregado': '',
            'ing-valor_acordado_estado': 'no',
            'ing-valor_acordado': '',
            'confirmar_mismo_equipo_cliente': '1',
        })
        data.update(self.salida_rapida_post_data(**salida_overrides))

        response = self.client.post(reverse('econotec:ingreso_registrar'), data)
        ingreso = IngresoEquipo.objects.get(pk__gt=ultimo_pk)
        return response, ingreso, SalidaEquipo.objects.get(ingreso=ingreso)

    def crear_notificacion_asesora(self, asesora=None, mensaje='Pendiente por cobrar.', **overrides):
        asesora = asesora or self.vendedor
        ingreso = self.crear_ingreso_reparacion(
            estado=overrides.pop('estado_ingreso', 'garantia'),
            valor_acordado=overrides.pop('valor_acordado_ingreso', Decimal('60.00')),
            marca=overrides.pop('marca', 'HP'),
            motivo_garantia='Garantía por retorno',
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 17),
            estado_reparacion=overrides.pop('estado_reparacion', 'garantia_fallos_adicionales'),
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        notificacion = NotificacionAsesora.objects.create(
            salida=salida,
            ingreso=ingreso,
            asesora=asesora,
            creado_por=self.usuario,
            valor_acordado=overrides.pop('valor_acordado', Decimal('60.00')),
            mensaje=mensaje,
            **overrides,
        )
        return notificacion

    def crear_venta_producto(self, **overrides):
        data = {
            'sede': 'ventas',
            'asesor_comercial': 'Kimberly',
            'fecha_ingreso': date.today(),
            'cliente': self.cliente_existente,
            'tipo_equipo': 'otro',
            'marca': 'N/A',
            'modelo_serie': 'Producto',
            'accesorios_entregados': 'Ninguno',
            'problema_reportado': 'Venta de producto',
            'valor_acordado': Decimal('10.00'),
            'tecnico_encargado': self.usuario,
            'estado': 'entregado',
            'subestado_entregado': 'con_solucion',
            'registrado_por': self.usuario,
        }
        data.update(overrides)
        return IngresoEquipo.objects.create(**data)

    def ingreso_edit_post_data(self, ingreso, **overrides):
        data = {
            'cli-cedula': ingreso.cliente.cedula,
            'cli-nombres': ingreso.cliente.nombres,
            'cli-whatsapp': ingreso.cliente.whatsapp,
            'cli-correo': ingreso.cliente.correo,
            'cli-sector': ingreso.cliente.sector,
            'cli-sector_otro': ingreso.cliente.sector_otro,
            'ing-numero_factura': ingreso.numero_factura,
            'ing-asesor_comercial': ingreso.asesor_comercial,
            'ing-tecnico_encargado': str(ingreso.tecnico_encargado_id or ''),
            'ing-fecha_ingreso': ingreso.fecha_ingreso.isoformat(),
            'ing-tipo_equipo': ingreso.tipo_equipo,
            'ing-tipo_equipo_otro': ingreso.tipo_equipo_otro,
            'ing-marca': ingreso.marca,
            'ing-modelo_serie': ingreso.modelo_serie,
            'ing-serie': ingreso.serie,
            'ing-accesorios_entregados': ingreso.accesorios_entregados,
            'ing-problema_reportado': ingreso.problema_reportado,
            'ing-firma_cliente_opcion': 'si' if ingreso.firma_cliente and ingreso.firma_cliente_imagen else 'no',
            'ing-firma_cliente_imagen': ingreso.firma_cliente_imagen,
            'ing-diagnostico_inmediato': ingreso.diagnostico_inmediato,
            'ing-valor_diagnostico': str(ingreso.valor_diagnostico),
            'ing-valor_acordado': str(ingreso.valor_acordado or ''),
            'ing-abono_anticipo': str(ingreso.abono_anticipo),
            'ing-diagnostico_metodo': ingreso.diagnostico_metodo,
            'ing-diagnostico_banco': ingreso.diagnostico_banco,
            'ing-diagnostico_banco_otro': ingreso.diagnostico_banco_otro,
            'ing-diagnostico_tarjeta_app': ingreso.diagnostico_tarjeta_app,
            'ing-diagnostico_comprobante_url': ingreso.diagnostico_comprobante_url,
            'ing-diagnostico_monto_1': '',
            'ing-diagnostico_metodo_1': ingreso.diagnostico_metodo_1,
            'ing-diagnostico_banco_1': ingreso.diagnostico_banco_1,
            'ing-diagnostico_monto_2': '',
            'ing-diagnostico_metodo_2': ingreso.diagnostico_metodo_2,
            'ing-diagnostico_banco_2': ingreso.diagnostico_banco_2,
            'ing-anticipo_metodo': ingreso.anticipo_metodo,
            'ing-anticipo_banco': ingreso.anticipo_banco,
            'ing-anticipo_banco_otro': ingreso.anticipo_banco_otro,
            'ing-anticipo_tarjeta_app': ingreso.anticipo_tarjeta_app,
            'ing-anticipo_comprobante_url': ingreso.anticipo_comprobante_url,
            'ing-anticipo_monto_1': '',
            'ing-anticipo_metodo_1': ingreso.anticipo_metodo_1,
            'ing-anticipo_banco_1': ingreso.anticipo_banco_1,
            'ing-anticipo_monto_2': '',
            'ing-anticipo_metodo_2': ingreso.anticipo_metodo_2,
            'ing-anticipo_banco_2': ingreso.anticipo_banco_2,
            'ing-estado': ingreso.estado,
            'ing-subestado_reparacion': ingreso.subestado_reparacion,
            'ing-subestado_entregado': ingreso.subestado_entregado,
            'ing-equipo_garantia': '',
            'ing-equipo_garantia_manual': ingreso.equipo_garantia_manual or '',
            'ing-motivo_garantia': ingreso.motivo_garantia,
        }
        data.update(overrides)
        return data

    def ingreso_registro_post_data(self, **overrides):
        data = {
            'cli-cedula': self.cliente_existente.cedula,
            'cli-nombres': self.cliente_existente.nombres,
            'cli-whatsapp': self.cliente_existente.whatsapp,
            'cli-correo': self.cliente_existente.correo,
            'cli-sector': self.cliente_existente.sector,
            'cli-sector_otro': self.cliente_existente.sector_otro,
            'ing-numero_factura': '',
            'ing-asesor_comercial': 'Kimberly',
            'ing-tecnico_encargado': str(self.usuario.pk),
            'ing-fecha_ingreso': '2026-07-09',
            'ing-tipo_equipo': 'laptop',
            'ing-tipo_equipo_otro': '',
            'ing-marca': 'MacBook M4 S',
            'ing-modelo_serie': 'MacBook M4 S',
            'ing-serie': '',
            'ing-accesorios_entregados': 'Cargador',
            'ing-problema_reportado': 'No enciende',
            'ing-firma_cliente_opcion': 'no',
            'ing-firma_cliente_imagen': '',
            'ing-diagnostico_inmediato': 'no',
            'ing-valor_diagnostico': '0.00',
            'ing-valor_acordado': '25',
            'ing-abono_anticipo': '0.00',
            'ing-diagnostico_metodo': 'efectivo',
            'ing-diagnostico_banco': '',
            'ing-diagnostico_banco_otro': '',
            'ing-diagnostico_tarjeta_app': '',
            'ing-diagnostico_comprobante_url': '',
            'ing-diagnostico_monto_1': '',
            'ing-diagnostico_metodo_1': '',
            'ing-diagnostico_banco_1': '',
            'ing-diagnostico_monto_2': '',
            'ing-diagnostico_metodo_2': '',
            'ing-diagnostico_banco_2': '',
            'ing-anticipo_metodo': 'efectivo',
            'ing-anticipo_banco': '',
            'ing-anticipo_banco_otro': '',
            'ing-anticipo_tarjeta_app': '',
            'ing-anticipo_comprobante_url': '',
            'ing-anticipo_monto_1': '',
            'ing-anticipo_metodo_1': '',
            'ing-anticipo_banco_1': '',
            'ing-anticipo_monto_2': '',
            'ing-anticipo_metodo_2': '',
            'ing-anticipo_banco_2': '',
            'ing-estado': 'en_reparacion',
            'ing-subestado_reparacion': 'en_reparacion',
            'ing-subestado_entregado': '',
            'ing-equipo_garantia': '',
            'ing-equipo_garantia_manual': '',
            'ing-motivo_garantia': '',
        }
        data.update(overrides)
        return data

    def ingreso_form_data(self, **overrides):
        data = {
            key.replace('ing-', '', 1): value
            for key, value in self.ingreso_registro_post_data().items()
            if key.startswith('ing-')
        }
        data.update(overrides)
        return data

    def activar_sede_guayaquil(self):
        session = self.client.session
        session['sede_actual'] = 'guayaquil'
        session.save()

    def test_formulario_venta_usa_selector_de_inventario(self):
        response = self.client.get(reverse('econotec:venta_registrar'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '+ Ver inventario')
        self.assertContains(response, 'Sede de inventario')
        self.assertContains(response, 'Guayaquil - Norte')
        self.assertContains(response, 'Guayaquil - Centro')
        self.assertContains(response, 'Quito')
        self.assertContains(response, 'Selecciona una sede para empezar')
        self.assertContains(response, 'Selecciona una categoria')
        self.assertContains(response, 'Selecciona el tipo')
        self.assertContains(response, 'Cantidad a llevar')
        self.assertContains(response, 'Valor / valores del producto')
        self.assertContains(response, 'Método de pago')
        self.assertContains(response, '¿Factura con datos?')
        self.assertContains(response, 'Si los datos del cliente ya están capturados')
        self.assertContains(response, '__syncFacturaVentaCliente')
        self.assertContains(response, 'Firma del Cliente')
        self.assertContains(response, 'Firmar ahora')
        self.assertContains(response, 'id_firma_cliente_imagen')
        self.assertContains(response, 'Valor desde inventario')
        self.assertContains(response, 'id_venta_valor_desde_inventario')
        self.assertContains(response, 'Ninguna')
        self.assertContains(response, 'Usar observación de inventario')
        self.assertContains(response, 'Sin observación de inventario')
        self.assertContains(response, 'Incluir observación')
        self.assertContains(response, 'data-observation-inventory')
        self.assertNotContains(response, 'Detalle adicional de la venta')

    def test_registrar_venta_no_requiere_campos_diagnostico_ocultos(self):
        response = self.client.post(
            reverse('econotec:venta_registrar'),
            self.venta_post_data(),
        )

        self.assertRedirects(response, reverse('econotec:venta_lista'))
        venta = IngresoEquipo.objects.get(sede='ventas')
        self.assertEqual(venta.cliente, self.cliente_existente)
        self.assertEqual(venta.estado, 'entregado')
        self.assertEqual(venta.subestado_entregado, 'con_solucion')
        self.assertEqual(venta.diagnostico_metodo, 'efectivo')
        self.assertEqual(venta.tecnico_encargado, self.usuario)
        self.assertEqual(venta.valor_acordado, Decimal('25.00'))
        self.assertEqual(venta.abono_anticipo, Decimal('25.00'))
        self.assertEqual(venta.diferencia, Decimal('0.00'))
        self.assertEqual(venta.estado_pago, 'Pagado')
        self.assertEqual(venta.problema_reportado, '2 x Tarjeta')
        self.producto_venta.refresh_from_db()
        self.assertEqual(self.producto_venta.cantidad, 8)
        self.assertTrue(VentaInventarioItem.objects.filter(
            venta=venta,
            inventario_item=self.producto_venta,
            cantidad=2,
        ).exists())

    def test_registrar_venta_guarda_observacion_de_producto_inventario(self):
        producto = self.crear_producto_venta()

        response = self.client.post(
            reverse('econotec:venta_registrar'),
            self.venta_post_data(
                inventario_seleccionado=json.dumps([
                    {
                        'item_id': producto.pk,
                        'cantidad': 2,
                        'observacion': 'Entregar con cable USB probado.',
                    },
                ]),
            ),
        )

        self.assertRedirects(response, reverse('econotec:venta_lista'))
        venta = IngresoEquipo.objects.get(sede='ventas')
        relacion = VentaInventarioItem.objects.get(venta=venta, inventario_item=producto)
        self.assertEqual(relacion.observacion, 'Entregar con cable USB probado.')
        productos_pdf = views_print._venta_productos_pdf_items(venta)
        detalles_pdf = ' · '.join(productos_pdf[0]['detalles'])
        self.assertIn('Observación: Entregar con cable USB probado.', detalles_pdf)

    def test_registrar_venta_puede_usar_observacion_guardada_en_inventario(self):
        producto = self.crear_producto_venta()
        producto.observacion = 'Producto probado en inventario.'
        producto.save(update_fields=['observacion'])

        response = self.client.post(
            reverse('econotec:venta_registrar'),
            self.venta_post_data(
                inventario_seleccionado=json.dumps([
                    {
                        'item_id': producto.pk,
                        'cantidad': 1,
                        'observacion': producto.observacion,
                    },
                ]),
            ),
        )

        self.assertRedirects(response, reverse('econotec:venta_lista'))
        venta = IngresoEquipo.objects.get(sede='ventas')
        relacion = VentaInventarioItem.objects.get(venta=venta, inventario_item=producto)
        self.assertEqual(relacion.observacion, 'Producto probado en inventario.')

    def test_registrar_venta_guarda_firma_cliente_y_pdf_usa_firma_tecnico(self):
        response = self.client.post(
            reverse('econotec:venta_registrar'),
            self.venta_post_data(
                **{
                    'ing-firma_cliente_opcion': 'si',
                    'ing-firma_cliente_imagen': self.FIRMA_PNG_DATA_URI,
                }
            ),
        )

        self.assertRedirects(response, reverse('econotec:venta_lista'))
        venta = IngresoEquipo.objects.get(sede='ventas')
        self.assertTrue(venta.firma_cliente)
        self.assertEqual(venta.firma_cliente_imagen, self.FIRMA_PNG_DATA_URI)
        productos_pdf = views_print._venta_productos_pdf_items(venta)
        self.assertEqual(productos_pdf[0]['titulo'], '2 x Tarjeta')
        detalles_pdf = ' · '.join(productos_pdf[0]['detalles'])
        self.assertIn('Categoría: Impresora', detalles_pdf)
        self.assertIn('Tipo: Impresora Laser', detalles_pdf)
        self.assertIn('Marca: Epson', detalles_pdf)
        self.assertIn('Modelo: Laser-t32', detalles_pdf)
        self.assertNotIn('Sede', detalles_pdf)
        self.assertNotIn('stock', detalles_pdf.lower())
        self.assertEqual(views_print._venta_metodo_pago_pdf(venta), 'Efectivo')

        with (
            patch('econotec.views_print._draw_label_value', wraps=views_print._draw_label_value) as draw_label,
            patch('econotec.views_print._draw_signature_image') as draw_signature,
            patch('econotec.views_print._draw_static_image') as draw_static,
        ):
            pdf_response = self.client.get(reverse('econotec:ingreso_pdf', kwargs={'pk': venta.pk}))

        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(pdf_response['Content-Type'], 'application/pdf')
        draw_signature.assert_any_call(ANY, self.FIRMA_PNG_DATA_URI, ANY, ANY, ANY, ANY)
        self.assertTrue(
            any(call.args[1] == 'firma_tecnico_recibe.png' for call in draw_static.call_args_list)
        )
        etiquetas_pdf = [call.args[3] for call in draw_label.call_args_list if len(call.args) > 3]
        self.assertIn('Técnico vendió:', etiquetas_pdf)
        self.assertNotIn('Categoría del producto:', etiquetas_pdf)

    def test_registrar_venta_exige_producto_del_inventario(self):
        response = self.client.post(
            reverse('econotec:venta_registrar'),
            self.venta_post_data(inventario_seleccionado='[]'),
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Selecciona al menos un producto del inventario')
        self.assertFalse(IngresoEquipo.objects.filter(sede='ventas').exists())
        self.producto_venta.refresh_from_db()
        self.assertEqual(self.producto_venta.cantidad, 10)

    def test_registrar_venta_guarda_pago_y_factura(self):
        response = self.client.post(
            reverse('econotec:venta_registrar'),
            self.venta_post_data(
                **{
                    'ing-anticipo_metodo': 'transferencia',
                    'ing-anticipo_banco': 'guayaquil',
                    'venta_factura_realizada': 'si',
                    'venta_factura_nombres': 'Yandri Guevara',
                    'venta_factura_cedula': '1207342716',
                    'venta_factura_correo': 'factura@example.com',
                }
            ),
        )

        self.assertRedirects(response, reverse('econotec:venta_lista'))
        venta = IngresoEquipo.objects.get(sede='ventas')
        self.assertEqual(venta.valor_acordado, Decimal('25.00'))
        self.assertEqual(venta.abono_anticipo, Decimal('25.00'))
        self.assertEqual(venta.anticipo_metodo, 'transferencia')
        self.assertEqual(venta.anticipo_banco, 'guayaquil')
        self.assertEqual(venta.factura_realizada, 'si')
        self.assertEqual(venta.factura_nombres, 'Yandri Guevara')
        self.assertEqual(venta.factura_cedula, '1207342716')
        self.assertEqual(venta.factura_correo, 'factura@example.com')
        self.assertEqual(venta.diferencia, Decimal('0.00'))

    def test_registrar_venta_puede_usar_valor_desde_inventario(self):
        producto = self.crear_producto_venta()
        producto.costo = Decimal('12.50')
        producto.save(update_fields=['costo'])

        response = self.client.post(
            reverse('econotec:venta_registrar'),
            self.venta_post_data(
                **{
                    'ing-valor_acordado': '1.00',
                    'venta_valor_desde_inventario': 'si',
                }
            ),
        )

        self.assertRedirects(response, reverse('econotec:venta_lista'))
        venta = IngresoEquipo.objects.get(sede='ventas')
        self.assertEqual(venta.valor_acordado, Decimal('25.00'))
        self.assertEqual(venta.abono_anticipo, Decimal('25.00'))
        self.assertEqual(venta.diferencia, Decimal('0.00'))

    def test_registrar_venta_valida_campos_de_pago(self):
        response = self.client.post(
            reverse('econotec:venta_registrar'),
            self.venta_post_data(
                **{
                    'ing-anticipo_metodo': 'transferencia',
                    'ing-anticipo_banco': '',
                }
            ),
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Indica el banco usado para la transferencia')
        self.assertFalse(IngresoEquipo.objects.filter(sede='ventas').exists())
        self.producto_venta.refresh_from_db()
        self.assertEqual(self.producto_venta.cantidad, 10)

    def test_registrar_venta_exige_tecnico_que_vendio(self):
        response = self.client.post(
            reverse('econotec:venta_registrar'),
            self.venta_post_data(**{'ing-tecnico_encargado': ''}),
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(IngresoEquipo.objects.filter(sede='ventas').exists())
        self.assertIn('tecnico_encargado', response.context['ing_form'].errors)
        self.producto_venta.refresh_from_db()
        self.assertEqual(self.producto_venta.cantidad, 10)

    def test_venta_inventario_catalogo_muestra_productos_disponibles(self):
        producto = self.crear_producto_venta()
        producto.observacion = 'Producto revisado en bodega.'
        producto.save(update_fields=['observacion'])
        InventarioItem.objects.create(
            sede='guayaquil',
            categoria='impresora',
            tipo='impresora-laser',
            producto='Toner',
            marca='Epson',
            modelo='Centro',
            estado='disponible',
            cantidad=4,
            ubicacion='guayaquil_centro',
            registrado_por=self.usuario,
        )
        no_disponible = InventarioItem.objects.create(
            sede='guayaquil',
            categoria='impresora',
            tipo='impresora-laser',
            producto='Adaptador',
            marca='Epson',
            modelo='Bloqueado',
            estado='no_disponible',
            causa_no_disponible='defectuoso',
            cantidad=2,
            ubicacion='guayaquil_norte',
            registrado_por=self.usuario,
        )

        response = self.client.get(reverse('econotec:venta_inventario_catalogo'), {
            'ubicacion': 'guayaquil_norte',
            'categoria': 'impresora',
            'tipo': 'impresora-laser',
        })

        self.assertEqual(response.status_code, 200)
        data = response.json()
        items = {item['item_id']: item for item in data['items']}
        self.assertEqual(set(items), {producto.pk, no_disponible.pk})
        self.assertEqual(items[producto.pk]['cantidad'], 10)
        self.assertTrue(items[producto.pk]['seleccionable'])
        self.assertEqual(items[producto.pk]['ubicacion'], 'Guayaquil - Norte')
        self.assertEqual(items[producto.pk]['costo'], '0.00')
        self.assertEqual(items[producto.pk]['observacion'], '')
        self.assertEqual(items[producto.pk]['observacion_inventario'], 'Producto revisado en bodega.')
        self.assertNotIn('precio_venta', items[producto.pk])
        self.assertEqual(items[no_disponible.pk]['estado'], 'no_disponible')
        self.assertEqual(items[no_disponible.pk]['estado_label'], 'No disponible')
        self.assertEqual(items[no_disponible.pk]['causa_no_disponible'], 'defectuoso')
        self.assertEqual(items[no_disponible.pk]['causa_no_disponible_label'], 'Defectuoso')
        self.assertFalse(items[no_disponible.pk]['seleccionable'])

    def test_venta_inventario_agregar_y_quitar_actualiza_stock(self):
        producto = self.crear_producto_venta()
        venta = IngresoEquipo.objects.create(
            sede='ventas',
            asesor_comercial='Kimberly',
            fecha_ingreso=date(2026, 7, 8),
            cliente=self.cliente_existente,
            tipo_equipo='otro',
            marca='N/A',
            modelo_serie='N/A',
            accesorios_entregados='Ninguno',
            problema_reportado='Venta de producto',
            valor_acordado=Decimal('10.00'),
            tecnico_encargado=self.usuario,
            estado='entregado',
            subestado_entregado='con_solucion',
        )

        response = self.client.post(
            reverse('econotec:venta_inventario_agregar', kwargs={'pk': venta.pk}),
            {'item_id': producto.pk, 'cantidad': '3', 'observacion': 'Primera nota de venta.'},
        )

        self.assertEqual(response.status_code, 200)
        relacion = VentaInventarioItem.objects.get(venta=venta, inventario_item=producto)
        producto.refresh_from_db()
        self.assertEqual(relacion.cantidad, 3)
        self.assertEqual(relacion.observacion, 'Primera nota de venta.')
        self.assertEqual(producto.cantidad, 7)

        response = self.client.post(
            reverse('econotec:venta_inventario_agregar', kwargs={'pk': venta.pk}),
            {'item_id': producto.pk, 'cantidad': '2'},
        )

        self.assertEqual(response.status_code, 200)
        relacion.refresh_from_db()
        producto.refresh_from_db()
        self.assertEqual(relacion.cantidad, 5)
        self.assertEqual(producto.cantidad, 5)

        response = self.client.get(reverse('econotec:venta_editar', kwargs={'pk': venta.pk}))
        self.assertContains(response, 'Editar cantidad')
        self.assertContains(response, 'Revertir y eliminar')

        response = self.client.post(
            reverse('econotec:venta_inventario_actualizar_cantidad', kwargs={'pk': venta.pk, 'relacion_pk': relacion.pk}),
            {'cantidad': '3'},
        )

        self.assertEqual(response.status_code, 200)
        relacion.refresh_from_db()
        producto.refresh_from_db()
        self.assertEqual(relacion.cantidad, 3)
        self.assertEqual(producto.cantidad, 7)

        response = self.client.post(
            reverse('econotec:venta_inventario_actualizar_cantidad', kwargs={'pk': venta.pk, 'relacion_pk': relacion.pk}),
            {'cantidad': '11'},
        )

        self.assertEqual(response.status_code, 409)
        relacion.refresh_from_db()
        producto.refresh_from_db()
        self.assertEqual(relacion.cantidad, 3)
        self.assertEqual(producto.cantidad, 7)

        response = self.client.post(
            reverse('econotec:venta_inventario_quitar', kwargs={'pk': venta.pk, 'relacion_pk': relacion.pk}),
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(VentaInventarioItem.objects.filter(pk=relacion.pk).exists())
        producto.refresh_from_db()
        self.assertEqual(producto.cantidad, 10)

    def test_venta_inventario_actualiza_observacion_sin_mover_stock(self):
        producto = self.crear_producto_venta()
        venta = IngresoEquipo.objects.create(
            sede='ventas',
            asesor_comercial='Kimberly',
            fecha_ingreso=date(2026, 7, 8),
            cliente=self.cliente_existente,
            tipo_equipo='otro',
            marca='N/A',
            modelo_serie='N/A',
            accesorios_entregados='Ninguno',
            problema_reportado='Venta de producto',
            valor_acordado=Decimal('10.00'),
            tecnico_encargado=self.usuario,
            estado='entregado',
            subestado_entregado='con_solucion',
        )
        relacion = VentaInventarioItem.objects.create(
            venta=venta,
            inventario_item=producto,
            cantidad=1,
        )

        response = self.client.post(
            reverse('econotec:venta_inventario_actualizar_observacion', kwargs={'pk': venta.pk, 'relacion_pk': relacion.pk}),
            {'observacion': 'Cliente pidió conservar caja.'},
        )

        self.assertEqual(response.status_code, 200)
        relacion.refresh_from_db()
        producto.refresh_from_db()
        self.assertEqual(relacion.observacion, 'Cliente pidió conservar caja.')
        self.assertEqual(producto.cantidad, 10)
        self.assertEqual(response.json()['producto']['observacion'], 'Cliente pidió conservar caja.')

    def test_editar_venta_conserva_estado_entregado(self):
        venta = IngresoEquipo.objects.create(
            sede='ventas',
            asesor_comercial='Kimberly',
            fecha_ingreso=date(2026, 7, 8),
            cliente=self.cliente_existente,
            tipo_equipo='otro',
            marca='N/A',
            modelo_serie='N/A',
            accesorios_entregados='Ninguno',
            problema_reportado='tinta anterior',
            valor_acordado=Decimal('10.00'),
            tecnico_encargado=self.usuario,
            estado='entregado',
            subestado_entregado='con_solucion',
        )

        response = self.client.post(
            reverse('econotec:venta_editar', kwargs={'pk': venta.pk}),
            self.venta_post_data(
                **{
                    'ing-problema_reportado': 'tinta negra',
                    'ing-valor_acordado': '30',
                    # Este valor llegaba desde el select oculto de la pantalla.
                    'ing-estado': 'ingresado',
                }
            ),
        )

        self.assertRedirects(response, reverse('econotec:venta_lista'))
        venta.refresh_from_db()
        self.assertEqual(venta.problema_reportado, 'tinta negra')
        self.assertEqual(venta.valor_acordado, Decimal('30.00'))
        self.assertEqual(venta.estado, 'entregado')
        self.assertEqual(venta.subestado_entregado, 'con_solucion')
        self.assertEqual(venta.tecnico_encargado, self.usuario)

    def test_perfil_suma_un_punto_por_salida_de_producto(self):
        response = self.client.post(
            reverse('econotec:venta_registrar'),
            self.venta_post_data(
                **{'ing-problema_reportado': 'tinta para perfil'}
            ),
        )
        self.assertRedirects(response, reverse('econotec:venta_lista'))

        response = self.client.get(reverse('econotec:api_perfil'))

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['email'], 'yandri@example.com')
        self.assertEqual(data['salidas_producto'], 1)
        self.assertEqual(data['total'], 1)
        self.assertGreaterEqual(data['bitacora_total'], 1)

    def test_api_bitacora_hoy_sin_datos_devuelve_vacia(self):
        response = self.client.get(reverse('econotec:api_bitacora_hoy'))

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertFalse(data['tiene_datos'])
        self.assertEqual(data['total'], 0)
        self.assertIn('Reporte del día', data['texto'])
        self.assertIn('Técnico: Yandri', data['texto'])

    def test_api_bitacora_hoy_genera_reporte_de_salida_del_tecnico(self):
        ingreso = self.crear_ingreso_reparacion(
            fecha_ingreso=timezone.localdate(),
            reporte_tecnico='Instalación de cartuchos nuevos y mantenimiento a Canon PIXMA G3110',
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=timezone.localdate(),
            estado_reparacion='pendiente_retiro',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            tecnico_reparo=self.usuario,
            registrado_por=self.usuario,
        )

        response = self.client.get(reverse('econotec:api_bitacora_hoy'))

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['tiene_datos'])
        self.assertEqual(data['total'], 1)
        self.assertIn('Reporte del día', data['texto'])
        self.assertIn('Técnico: Yandri', data['texto'])
        self.assertIn('Instalación de cartuchos nuevos y mantenimiento a Canon PIXMA G3110', data['texto'])
        self.assertIn(f'#{ingreso.codigo_equipo} lista, cliente notificado.', data['texto'])
        self.assertRegex(data['texto'], r'\*\d{1,2}:\d{2} (AM|PM)\* - Instalación')
        self.assertNotRegex(data['texto'], r'\d{1,2}:\d{2} - \d{1,2}:\d{2}')

    def test_bitacora_se_reinicia_en_medianoche_local(self):
        from .views import _construir_bitacora_usuario

        zona_local = ZoneInfo('America/Guayaquil')
        dia_anterior = date(2026, 7, 22)
        dia_nuevo = date(2026, 7, 23)

        BitacoraTecnico.objects.create(
            user=self.usuario,
            usuario_nombre='Yandri',
            momento=datetime(2026, 7, 22, 23, 59, tzinfo=zona_local),
            tipo='reporte',
            texto='Acción antes de medianoche.',
        )
        BitacoraTecnico.objects.create(
            user=self.usuario,
            usuario_nombre='Yandri',
            momento=datetime(2026, 7, 23, 0, 0, tzinfo=zona_local),
            tipo='reporte',
            texto='Acción justo a medianoche.',
        )

        reporte_anterior = _construir_bitacora_usuario(self.usuario, dia=dia_anterior)
        reporte_nuevo = _construir_bitacora_usuario(self.usuario, dia=dia_nuevo)

        self.assertEqual(reporte_anterior['fecha'], '22/07/2026')
        self.assertEqual(reporte_anterior['total'], 1)
        self.assertIn('Acción antes de medianoche.', reporte_anterior['texto'])
        self.assertNotIn('Acción justo a medianoche.', reporte_anterior['texto'])

        self.assertEqual(reporte_nuevo['fecha'], '23/07/2026')
        self.assertEqual(reporte_nuevo['total'], 1)
        self.assertIn('*12:00 AM* - Acción justo a medianoche.', reporte_nuevo['texto'])
        self.assertNotIn('Acción antes de medianoche.', reporte_nuevo['texto'])

    def test_bitacora_formatea_hora_y_separa_movimientos_para_copiar(self):
        from .bitacora import construir_bitacora_usuario

        zona_local = ZoneInfo('America/Guayaquil')
        dia = date(2026, 7, 24)
        BitacoraTecnico.objects.create(
            user=self.usuario,
            usuario_nombre='Yandri',
            momento=datetime(2026, 7, 24, 13, 32, tzinfo=zona_local),
            tipo='reporte',
            texto='Primera acción registrada.',
        )
        BitacoraTecnico.objects.create(
            user=self.usuario,
            usuario_nombre='Yandri',
            momento=datetime(2026, 7, 24, 13, 33, tzinfo=zona_local),
            tipo='estado',
            texto=(
                'Datos actualizados en Laptop HP Elitebook #G1000 para Yandri Guevara. '
                'Detalles: Marca: Epson -> HP; Valor acordado: $50.00 -> $0.00.'
            ),
        )

        reporte = construir_bitacora_usuario(self.usuario, dia=dia)

        self.assertIn('*1:32 PM* - Primera acción registrada.\n\n*1:33 PM* - Datos actualizados', reporte['detalle'])
        self.assertIn('Detalles:\n  - Marca: Epson -> HP.\n  - Valor acordado: $50.00 -> $0.00.', reporte['detalle'])

    def test_perfil_asesor_muestra_datos_basicos(self):
        self.client.force_login(self.vendedor)

        response = self.client.get(reverse('econotec:api_perfil'))

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['tipo_perfil'], 'asesor')
        self.assertEqual(data['nombre'], 'Kimberly')
        self.assertEqual(data['email'], 'kimberly@example.com')
        self.assertEqual(data['nivel'], 'Asesor registrado')
        self.assertEqual(data['total'], 0)
        self.assertEqual(data['color'], '#0d47a1')
        self.assertIn('#ec4899', data['colores_disponibles'])

    def test_perfil_asesor_guarda_color_preferido(self):
        self.client.force_login(self.vendedor)

        response = self.client.post(
            reverse('econotec:api_perfil_color'),
            data='{"color":"#ec4899"}',
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['color'], '#ec4899')
        actividad = UsuarioActividad.objects.get(user=self.vendedor)
        self.assertEqual(actividad.perfil_color_asesor, '#ec4899')

        response = self.client.get(reverse('econotec:api_perfil'))
        self.assertEqual(response.json()['color'], '#ec4899')

    def test_perfil_asesor_rechaza_color_no_permitido(self):
        self.client.force_login(self.vendedor)

        response = self.client.post(
            reverse('econotec:api_perfil_color'),
            data='{"color":"#000000"}',
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)

    def test_bienvenida_muestra_boton_perfil_para_asesor(self):
        self.client.force_login(self.vendedor)
        self.activar_sede_guayaquil()

        response = self.client.get(reverse('econotec:bienvenida'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '📦 Inventario')
        self.assertContains(response, f'href="{reverse("econotec:inventario_menu")}"')
        self.assertContains(response, 'class="top-inventory-link" title="Abrir inventario" target="_blank" rel="noopener"')
        self.assertContains(response, 'id="btn-perfil" data-perfil-trigger')
        self.assertContains(response, 'id="btn-perfil-mobile" class="mobile-profile-trigger" data-perfil-trigger')
        self.assertContains(response, 'Asesor')
        self.assertContains(response, 'Ver equipos que registré')
        self.assertContains(response, f'?registrador={self.vendedor.pk}&sede=todas')
        self.assertContains(response, 'Cambiar color del perfil')
        self.assertContains(response, 'data-color="#ec4899"')

    def test_inventario_menu_muestra_sedes_iniciales(self):
        response = self.client.get(reverse('econotec:inventario_menu'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Inventario')
        self.assertContains(response, 'Guayaquil')
        self.assertContains(response, 'Norte - Centro')
        self.assertContains(response, 'Quito')
        self.assertContains(response, 'inventario/guayaquil.jpg')
        self.assertContains(response, 'inventario/quito.jpg')
        self.assertContains(response, 'data-sede="guayaquil"')
        self.assertContains(response, 'data-sede="quito"')
        self.assertContains(response, 'Categorías de inventario')
        self.assertContains(response, 'Impresora')
        self.assertContains(response, 'Impresora Laser')
        self.assertContains(response, 'Impresora Inyección')
        self.assertContains(response, 'Computadora')
        self.assertContains(response, 'PC')
        self.assertContains(response, 'Laptops')
        self.assertContains(response, 'Consola')
        self.assertIn('consola de mesa', response.content.decode().lower())
        self.assertContains(response, 'Portatil')
        self.assertContains(response, 'Celular')
        self.assertContains(response, 'Tablet')
        self.assertContains(response, 'Mando')
        self.assertContains(response, 'Otros equipos/materiales')
        self.assertContains(response, 'inventario/otros-equipos-materiales.png')
        self.assertContains(
            response,
            reverse('econotec:inventario_categoria', kwargs={
                'sede': '__sede__',
                'categoria': 'otros-equipos-materiales',
            }),
        )
        self.assertNotContains(response, 'Venta de Producto')

    def test_inventario_otros_equipos_materiales_usa_flujo_completo(self):
        categoria_url = reverse('econotec:inventario_categoria', kwargs={
            'sede': 'guayaquil',
            'categoria': 'otros-equipos-materiales',
        })
        tabla_url = reverse('econotec:inventario_tabla', kwargs={
            'sede': 'guayaquil',
            'categoria': 'otros-equipos-materiales',
            'tipo': 'otros-equipos-materiales',
        })
        registrar_url = reverse('econotec:inventario_registrar', kwargs={
            'sede': 'guayaquil',
            'categoria': 'otros-equipos-materiales',
            'tipo': 'otros-equipos-materiales',
        })

        categoria = self.client.get(categoria_url)
        self.assertRedirects(categoria, tabla_url)

        formulario = self.client.get(registrar_url)
        self.assertEqual(formulario.status_code, 200)
        self.assertContains(formulario, 'Ingresar')
        self.assertContains(formulario, 'Otros equipos/materiales')
        self.assertContains(formulario, 'Producto')
        self.assertContains(formulario, 'Ubicación')

        response = self.client.post(registrar_url, {
            'producto': 'Kit de herramientas',
            'marca': 'Genérica',
            'modelo': 'Técnico',
            'serie': '',
            'estado': 'disponible',
            'causa_no_disponible': '',
            'cantidad': '4',
            'costo': '12.50',
            'ubicacion': 'guayaquil_norte',
            'observacion': 'Material para el taller.',
        })

        item = InventarioItem.objects.get(producto='Kit de herramientas')
        self.assertEqual(item.categoria, 'otros-equipos-materiales')
        self.assertEqual(item.tipo, 'otros-equipos-materiales')
        self.assertTrue(item.codigo.startswith('INV-GYE-OTRO-'))
        self.assertRedirects(response, tabla_url)

        tabla = self.client.get(tabla_url)
        self.assertContains(tabla, 'Kit de herramientas')
        self.assertContains(tabla, 'Material para el taller.')
        self.assertContains(tabla, 'data:image/png;base64,')
        self.assertContains(tabla, reverse(
            'econotec:inventario_detalle_item',
            kwargs={'codigo': item.codigo},
        ))

    def test_inventario_categoria_con_subtipos_muestra_opciones(self):
        response = self.client.get(reverse('econotec:inventario_categoria', kwargs={
            'sede': 'guayaquil',
            'categoria': 'impresora',
        }))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Guayaquil')
        self.assertContains(response, 'Impresora')
        self.assertContains(response, 'Impresora Laser')
        self.assertContains(response, 'Impresora Inyección')
        self.assertContains(response, reverse('econotec:inventario_tabla', kwargs={
            'sede': 'guayaquil',
            'categoria': 'impresora',
            'tipo': 'impresora-laser',
        }))

    def test_inventario_categoria_sin_subtipos_redirige_a_tabla(self):
        response = self.client.get(reverse('econotec:inventario_categoria', kwargs={
            'sede': 'quito',
            'categoria': 'celular',
        }))

        self.assertRedirects(response, reverse('econotec:inventario_tabla', kwargs={
            'sede': 'quito',
            'categoria': 'celular',
            'tipo': 'celular',
        }))

    def test_inventario_tabla_muestra_contexto_y_columnas(self):
        response = self.client.get(reverse('econotec:inventario_tabla', kwargs={
            'sede': 'guayaquil',
            'categoria': 'computadora',
            'tipo': 'pc',
        }))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Inventario')
        self.assertContains(response, 'Guayaquil')
        self.assertContains(response, 'Computadora')
        self.assertContains(response, 'PC')
        self.assertContains(response, 'Código')
        self.assertContains(response, 'Producto')
        self.assertContains(response, 'Marca')
        self.assertContains(response, 'Modelo')
        self.assertContains(response, 'Serie')
        self.assertContains(response, '(opcional)')
        self.assertContains(response, 'Estado')
        self.assertContains(response, 'Cantidad')
        self.assertContains(response, 'Costo')
        self.assertContains(response, 'Ubicación')
        self.assertContains(response, 'Acción')
        self.assertNotContains(response, 'Marca / Modelo')
        self.assertContains(response, 'Observación')
        self.assertContains(response, 'Sin registros todavía para PC en Guayaquil.')

    def test_inventario_tabla_filtra_por_texto_estado_y_ubicacion(self):
        item_norte = InventarioItem.objects.create(
            sede='guayaquil',
            categoria='impresora',
            tipo='impresora-laser',
            producto='Tarjeta Epson',
            marca='Epson',
            modelo='Laser-t32',
            estado='disponible',
            cantidad=35,
            costo=Decimal('12.50'),
            ubicacion='guayaquil_norte',
            registrado_por=self.usuario,
        )
        item_centro = InventarioItem.objects.create(
            sede='guayaquil',
            categoria='impresora',
            tipo='impresora-laser',
            producto='Cabezal Centro',
            marca='HP',
            modelo='Ink 900',
            estado='no_disponible',
            causa_no_disponible='agotado',
            cantidad=0,
            ubicacion='guayaquil_centro',
            registrado_por=self.usuario,
        )
        InventarioItem.objects.create(
            sede='quito',
            categoria='impresora',
            tipo='impresora-laser',
            producto='Producto Quito',
            marca='Canon',
            modelo='Q1',
            estado='disponible',
            cantidad=3,
            ubicacion='quito',
            registrado_por=self.usuario,
        )
        url = reverse('econotec:inventario_tabla', kwargs={
            'sede': 'guayaquil',
            'categoria': 'impresora',
            'tipo': 'impresora-laser',
        })

        response = self.client.get(url)
        self.assertContains(response, 'Filtro de búsqueda')
        self.assertContains(response, 'Guayaquil - Norte')
        self.assertContains(response, 'Guayaquil - Centro')
        self.assertNotContains(response, 'value="quito"')
        self.assertContains(response, 'Tarjeta Epson')
        self.assertContains(response, '$12,50')
        self.assertContains(response, 'Cabezal Centro')
        self.assertContains(response, 'Causa: Agotado')
        self.assertNotContains(response, 'Producto Quito')

        por_codigo = self.client.get(url, {'q': item_norte.codigo.lower()})
        self.assertContains(por_codigo, 'Tarjeta Epson')
        self.assertNotContains(por_codigo, 'Cabezal Centro')
        self.assertContains(por_codigo, 'Mostrando 1 de 2')

        por_estado = self.client.get(url, {'estado': 'no_disponible'})
        self.assertContains(por_estado, 'Cabezal Centro')
        self.assertNotContains(por_estado, 'Tarjeta Epson')
        self.assertContains(por_estado, 'selected>No disponible</option>')

        por_ubicacion = self.client.get(url, {'ubicacion': item_centro.ubicacion})
        self.assertContains(por_ubicacion, 'Cabezal Centro')
        self.assertNotContains(por_ubicacion, 'Tarjeta Epson')
        self.assertContains(por_ubicacion, 'selected>Guayaquil - Centro</option>')

        por_texto_ubicacion = self.client.get(url, {'q': 'guayaquil centro'})
        self.assertContains(por_texto_ubicacion, 'Cabezal Centro')
        self.assertNotContains(por_texto_ubicacion, 'Tarjeta Epson')

    def test_inventario_tabla_paginala_a_cinco_y_muestra_estado_siempre(self):
        url = reverse('econotec:inventario_tabla', kwargs={
            'sede': 'guayaquil',
            'categoria': 'impresora',
            'tipo': 'impresora-laser',
        })
        for numero in range(12):
            InventarioItem.objects.create(
                sede='guayaquil',
                categoria='impresora',
                tipo='impresora-laser',
                producto=f'Producto paginado {numero}',
                marca='Epson',
                modelo=f'M{numero}',
                estado='disponible',
                cantidad=2,
                ubicacion='guayaquil_norte',
                registrado_por=self.usuario,
            )

        response = self.client.get(url, {'q': 'Producto paginado'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['page_obj'].paginator.count, 12)
        self.assertEqual(len(response.context['items']), 5)
        self.assertContains(response, 'Página 1 de 3')
        self.assertContains(response, 'Mostrando máximo 5 por página')
        html = response.content.decode()
        self.assertRegex(html, r'q=Producto(?:\+|%20)paginado(?:&amp;|&)pagina=2')

        response_sin_resultados = self.client.get(url, {'q': 'sin coincidencias'})
        self.assertEqual(response_sin_resultados.status_code, 200)
        self.assertEqual(response_sin_resultados.context['page_obj'].paginator.count, 0)
        self.assertContains(response_sin_resultados, 'Página 1 de 1')
        self.assertContains(response_sin_resultados, 'Mostrando máximo 5 por página')
        self.assertContains(response_sin_resultados, 'list-pagination-disabled')

    def test_inventario_exporta_excel_con_filtros_actuales(self):
        item_filtrado = InventarioItem.objects.create(
            sede='guayaquil',
            categoria='impresora',
            tipo='impresora-laser',
            producto='Tarjeta Epson',
            marca='Epson',
            modelo='Laser-t32',
            estado='disponible',
            cantidad=1,
            costo=Decimal('10.50'),
            ubicacion='guayaquil_norte',
            observacion='Solo vender con empaque.',
            registrado_por=self.usuario,
        )
        InventarioItem.objects.create(
            sede='guayaquil',
            categoria='impresora',
            tipo='impresora-laser',
            producto='Cabezal Centro',
            marca='HP',
            modelo='Ink 900',
            estado='disponible',
            cantidad=4,
            ubicacion='guayaquil_centro',
            registrado_por=self.usuario,
        )

        response = self.client.get(reverse('econotec:inventario_export', kwargs={
            'sede': 'guayaquil',
            'categoria': 'impresora',
            'tipo': 'impresora-laser',
        }), {
            'q': item_filtrado.codigo.lower(),
            'ubicacion': 'guayaquil_norte',
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('inventario_guayaquil_impresora_impresora-laser.xlsx', response['Content-Disposition'])
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws['A1'].value, 'Código')
        self.assertEqual(ws['A2'].value, item_filtrado.codigo)
        self.assertEqual(ws['B2'].value, 'Tarjeta Epson')
        self.assertEqual(ws['H2'].value, 1)
        self.assertEqual(ws['K1'].value, 'Observación')
        self.assertEqual(ws['K2'].value, 'Solo vender con empaque.')
        self.assertEqual(ws.max_row, 2)

    def test_inventario_stock_uno_muestra_alerta_y_notifica_admin(self):
        item = InventarioItem.objects.create(
            sede='guayaquil',
            categoria='impresora',
            tipo='impresora-laser',
            producto='KSKL',
            marca='Ep-Hp',
            modelo='dd',
            estado='disponible',
            cantidad=1,
            ubicacion='guayaquil_norte',
            registrado_por=self.usuario,
        )
        url = reverse('econotec:inventario_tabla', kwargs={
            'sede': 'guayaquil',
            'categoria': 'impresora',
            'tipo': 'impresora-laser',
        })

        response = self.client.get(url)
        self.assertContains(response, 'solo queda 1 unidad de KSKL')
        self.assertContains(response, 'Solo queda 1')
        self.assertContains(response, 'Notificar al admin')

        response_post = self.client.post(
            reverse('econotec:inventario_notificar_admin', kwargs={'codigo': item.codigo}),
            {'next': url},
        )
        self.assertRedirects(response_post, url)
        notificacion = NotificacionInventarioAdmin.objects.get(inventario_item=item)
        self.assertEqual(notificacion.creado_por, self.usuario)
        self.assertFalse(notificacion.leida)

        response_pendiente = self.client.get(url)
        self.assertContains(response_pendiente, 'Admin notificado')

        self.client.force_login(self.admin)
        admin_response = self.client.get(reverse('econotec:bienvenida'))
        self.assertContains(admin_response, 'Inventario bajo')
        self.assertContains(admin_response, item.codigo)
        self.assertContains(admin_response, reverse('econotec:notificacion_inventario_admin_ver', kwargs={'pk': notificacion.pk}))

        ver_response = self.client.get(reverse('econotec:notificacion_inventario_admin_ver', kwargs={'pk': notificacion.pk}))
        self.assertRedirects(ver_response, reverse('econotec:inventario_detalle_item', kwargs={'codigo': item.codigo}))
        notificacion.refresh_from_db()
        self.assertTrue(notificacion.leida)

    def test_asesora_puede_notificar_stock_uno_al_admin(self):
        item = InventarioItem.objects.create(
            sede='quito',
            categoria='computadora',
            tipo='pc',
            producto='Tarjeta gráfica',
            marca='Nvidia',
            modelo='LGT30292',
            estado='disponible',
            cantidad=1,
            ubicacion='quito',
            registrado_por=self.usuario,
        )
        self.client.force_login(self.vendedor)

        response = self.client.post(
            reverse('econotec:inventario_notificar_admin', kwargs={'codigo': item.codigo}),
            {'next': reverse('econotec:inventario_tabla', kwargs={
                'sede': 'quito',
                'categoria': 'computadora',
                'tipo': 'pc',
            })},
        )

        self.assertEqual(response.status_code, 302)
        notificacion = NotificacionInventarioAdmin.objects.get(inventario_item=item)
        self.assertEqual(notificacion.creado_por, self.vendedor)

    def test_inventario_filtro_ubicacion_respeta_sede_actual(self):
        InventarioItem.objects.create(
            sede='quito',
            categoria='tablet',
            tipo='tablet',
            producto='Tablet Quito',
            marca='Samsung',
            modelo='A9',
            estado='disponible',
            cantidad=3,
            ubicacion='quito',
            registrado_por=self.usuario,
        )

        response = self.client.get(reverse('econotec:inventario_tabla', kwargs={
            'sede': 'quito',
            'categoria': 'tablet',
            'tipo': 'tablet',
        }))

        self.assertContains(response, 'value="quito"')
        self.assertContains(response, 'Tablet Quito')
        self.assertNotContains(response, 'Guayaquil - Norte')
        self.assertNotContains(response, 'Guayaquil - Centro')

    def test_inventario_formulario_muestra_campos_y_ubicaciones_de_sede(self):
        response = self.client.get(reverse('econotec:inventario_registrar', kwargs={
            'sede': 'guayaquil',
            'categoria': 'computadora',
            'tipo': 'pc',
        }))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Guardar equipo')
        self.assertContains(response, 'Producto')
        self.assertContains(response, 'Marca')
        self.assertContains(response, 'Modelo')
        self.assertContains(response, 'Serie')
        self.assertContains(response, 'Estado')
        self.assertContains(response, 'Disponible')
        self.assertContains(response, 'No disponible')
        self.assertNotContains(response, 'En uso')
        self.assertNotContains(response, 'Reservado')
        self.assertNotContains(response, 'Dañado')
        self.assertNotContains(response, 'Vendido')
        self.assertContains(response, 'Cantidad')
        self.assertContains(response, 'Costo (USD)')
        self.assertNotContains(response, 'Precio de venta (USD)')
        self.assertContains(response, 'Causa de no disponibilidad')
        self.assertContains(response, 'Ubicación')
        self.assertContains(response, 'Observación')
        self.assertContains(response, 'Guayaquil Norte')
        self.assertContains(response, 'Guayaquil Centro')

    def test_inventario_registrar_crea_item_con_codigo_y_qr_en_tabla(self):
        url = reverse('econotec:inventario_registrar', kwargs={
            'sede': 'guayaquil',
            'categoria': 'computadora',
            'tipo': 'pc',
        })

        response = self.client.post(url, {
            'producto': 'CPU completo',
            'marca': 'HP',
            'modelo': 'EliteDesk 800',
            'serie': '',
            'estado': 'disponible',
            'causa_no_disponible': '',
            'cantidad': '2',
            'costo': '8.75',
            'ubicacion': 'guayaquil_norte',
            'observacion': 'Guardar para cliente frecuente.',
        })

        item = InventarioItem.objects.get()
        self.assertTrue(item.codigo.startswith('INV-GYE-PC-'))
        self.assertEqual(item.producto, 'CPU completo')
        self.assertEqual(item.marca, 'HP')
        self.assertEqual(item.modelo, 'EliteDesk 800')
        self.assertEqual(item.serie, '')
        self.assertEqual(item.causa_no_disponible, '')
        self.assertEqual(item.costo, Decimal('8.75'))
        self.assertEqual(item.ubicacion, 'guayaquil_norte')
        self.assertEqual(item.observacion, 'Guardar para cliente frecuente.')
        self.assertRedirects(response, reverse('econotec:inventario_tabla', kwargs={
            'sede': 'guayaquil',
            'categoria': 'computadora',
            'tipo': 'pc',
        }))

        tabla = self.client.get(response.url)
        self.assertContains(tabla, item.codigo)
        self.assertContains(tabla, 'CPU completo')
        self.assertContains(tabla, 'HP')
        self.assertContains(tabla, 'EliteDesk 800')
        self.assertContains(tabla, 'Guayaquil Norte')
        self.assertContains(tabla, 'Guardar para cliente frecuente.')
        self.assertContains(tabla, 'Editar')
        self.assertContains(tabla, 'Eliminar')
        self.assertContains(tabla, 'data:image/png;base64,')
        self.assertContains(tabla, reverse('econotec:inventario_detalle_item', kwargs={'codigo': item.codigo}))
        self.assertContains(tabla, reverse('econotec:inventario_qr_imprimir', kwargs={'codigo': item.codigo}))
        self.assertContains(tabla, reverse('econotec:inventario_editar', kwargs={'codigo': item.codigo}))
        self.assertContains(tabla, reverse('econotec:inventario_eliminar', kwargs={'codigo': item.codigo}))

    def test_inventario_rechaza_producto_y_modelo_repetidos_sin_importar_formato(self):
        existente = InventarioItem.objects.create(
            sede='guayaquil',
            categoria='impresora',
            tipo='impresora-laser',
            producto='Tarjeta lógica',
            marca='Epson',
            modelo='Láser-t32',
            estado='disponible',
            cantidad=34,
            ubicacion='guayaquil_norte',
            registrado_por=self.usuario,
        )
        registrar_url = reverse('econotec:inventario_registrar', kwargs={
            'sede': 'guayaquil',
            'categoria': 'impresora',
            'tipo': 'impresora-laser',
        })

        response = self.client.post(registrar_url, {
            'producto': '  TARJETA   LOGICA  ',
            'marca': 'Epson',
            'modelo': '  LASER-T32  ',
            'serie': '',
            'estado': 'disponible',
            'causa_no_disponible': '',
            'cantidad': '1',
            'costo': '10.00',
            'ubicacion': 'guayaquil_norte',
            'observacion': '',
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(InventarioItem.objects.count(), 1)
        self.assertIn('producto', response.context['form'].errors)
        self.assertIn('modelo', response.context['form'].errors)
        self.assertContains(response, existente.codigo)
        self.assertContains(response, 'Ya existe un producto con este mismo nombre')
        self.assertContains(response, 'Ya existe un producto con este mismo modelo')

    def test_inventario_rechaza_producto_repetido_aunque_cambie_modelo(self):
        existente = InventarioItem.objects.create(
            sede='guayaquil',
            categoria='impresora',
            tipo='impresora-inyeccion',
            producto='Tinta cyan',
            marca='Genérica',
            modelo='1 litro',
            estado='disponible',
            cantidad=1,
            ubicacion='guayaquil_norte',
            registrado_por=self.usuario,
        )
        registrar_url = reverse('econotec:inventario_registrar', kwargs={
            'sede': 'guayaquil',
            'categoria': 'impresora',
            'tipo': 'impresora-inyeccion',
        })

        response = self.client.post(registrar_url, {
            'producto': 'TINTA CYAN',
            'marca': 'Genérica',
            'modelo': '500 ml',
            'serie': '',
            'estado': 'disponible',
            'causa_no_disponible': '',
            'cantidad': '1',
            'costo': '15.00',
            'ubicacion': 'guayaquil_norte',
            'observacion': '',
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(InventarioItem.objects.count(), 1)
        self.assertIn('producto', response.context['form'].errors)
        self.assertContains(response, existente.codigo)

    def test_inventario_rechaza_modelo_repetido_aunque_cambie_producto(self):
        existente = InventarioItem.objects.create(
            sede='guayaquil',
            categoria='impresora',
            tipo='impresora-inyeccion',
            producto='Tinta cyan',
            marca='Genérica',
            modelo='1 litro',
            estado='disponible',
            cantidad=1,
            ubicacion='guayaquil_norte',
            registrado_por=self.usuario,
        )
        registrar_url = reverse('econotec:inventario_registrar', kwargs={
            'sede': 'guayaquil',
            'categoria': 'impresora',
            'tipo': 'impresora-inyeccion',
        })

        response = self.client.post(registrar_url, {
            'producto': 'Tinta magenta',
            'marca': 'Genérica',
            'modelo': '1 LITRO',
            'serie': '',
            'estado': 'disponible',
            'causa_no_disponible': '',
            'cantidad': '1',
            'costo': '15.00',
            'ubicacion': 'guayaquil_norte',
            'observacion': '',
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(InventarioItem.objects.count(), 1)
        self.assertIn('modelo', response.context['form'].errors)
        self.assertContains(response, existente.codigo)

    def test_inventario_permite_producto_y_modelo_parecidos(self):
        InventarioItem.objects.create(
            sede='guayaquil',
            categoria='impresora',
            tipo='impresora-inyeccion',
            producto='Tinta cyan',
            marca='Genérica',
            modelo='1 litro',
            estado='disponible',
            cantidad=1,
            ubicacion='guayaquil_norte',
            registrado_por=self.usuario,
        )
        registrar_url = reverse('econotec:inventario_registrar', kwargs={
            'sede': 'guayaquil',
            'categoria': 'impresora',
            'tipo': 'impresora-inyeccion',
        })

        response = self.client.post(registrar_url, {
            'producto': 'Tinta cyan premium',
            'marca': 'Genérica',
            'modelo': '1 litro plus',
            'serie': '',
            'estado': 'disponible',
            'causa_no_disponible': '',
            'cantidad': '1',
            'costo': '18.00',
            'ubicacion': 'guayaquil_norte',
            'observacion': '',
        })

        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            InventarioItem.objects.filter(producto='Tinta cyan premium').exists()
        )

    def test_inventario_editar_el_mismo_registro_no_es_duplicado(self):
        item = InventarioItem.objects.create(
            sede='guayaquil',
            categoria='impresora',
            tipo='impresora-laser',
            producto='Tarjeta lógica',
            marca='Epson',
            modelo='Laser-t32',
            estado='disponible',
            cantidad=34,
            costo=Decimal('10.00'),
            ubicacion='guayaquil_norte',
            registrado_por=self.usuario,
        )
        editar_url = reverse(
            'econotec:inventario_editar',
            kwargs={'codigo': item.codigo},
        )

        response = self.client.post(editar_url, {
            'producto': 'TARJETA LÓGICA',
            'marca': 'Epson',
            'modelo': 'LASER-T32',
            'serie': '',
            'estado': 'disponible',
            'causa_no_disponible': '',
            'cantidad': '35',
            'costo': '10.00',
            'ubicacion': 'guayaquil_norte',
            'observacion': '',
        })

        self.assertRedirects(response, reverse(
            'econotec:inventario_detalle_item',
            kwargs={'codigo': item.codigo},
        ))
        item.refresh_from_db()
        self.assertEqual(item.cantidad, 35)

    def test_inventario_detalle_y_qr_imprimible_muestran_datos(self):
        item = InventarioItem.objects.create(
            sede='guayaquil',
            categoria='computadora',
            tipo='pc',
            producto='CPU completo',
            marca='HP',
            modelo='EliteDesk 800',
            serie='',
            estado='disponible',
            cantidad=2,
            ubicacion='guayaquil_centro',
            observacion='Equipo listo para exhibición.',
            registrado_por=self.usuario,
        )

        detalle = self.client.get(reverse('econotec:inventario_detalle_item', kwargs={'codigo': item.codigo}))
        self.assertEqual(detalle.status_code, 200)
        self.assertContains(detalle, item.codigo)
        self.assertContains(detalle, 'CPU completo')
        self.assertContains(detalle, 'Computadora')
        self.assertContains(detalle, 'HP')
        self.assertContains(detalle, 'EliteDesk 800')
        self.assertContains(detalle, 'Guayaquil Centro')
        self.assertContains(detalle, 'Equipo listo para exhibición.')
        self.assertContains(detalle, 'data:image/png;base64,')
        self.assertContains(detalle, 'Editar producto')
        self.assertContains(detalle, 'Guardar cambios')
        self.assertContains(detalle, 'target="_blank"')
        self.assertContains(detalle, 'rel="noopener"')

        imprimir = self.client.get(reverse('econotec:inventario_qr_imprimir', kwargs={'codigo': item.codigo}))
        self.assertEqual(imprimir.status_code, 200)
        self.assertContains(imprimir, 'CPU completo')
        self.assertContains(imprimir, 'HP')
        self.assertContains(imprimir, 'EliteDesk 800')
        self.assertContains(imprimir, 'Computadora')
        self.assertContains(imprimir, item.codigo)
        self.assertNotContains(imprimir, 'Producto:')
        self.assertNotContains(imprimir, 'Marca:')
        self.assertNotContains(imprimir, 'Modelo:')
        self.assertNotContains(imprimir, 'Categoría:')

    def test_inventario_editar_actualiza_producto_sin_cambiar_codigo(self):
        item = InventarioItem.objects.create(
            sede='guayaquil',
            categoria='computadora',
            tipo='pc',
            producto='CPU completo',
            marca='HP',
            modelo='EliteDesk 800',
            estado='disponible',
            cantidad=0,
            ubicacion='guayaquil_norte',
            registrado_por=self.usuario,
        )
        editar_url = reverse('econotec:inventario_editar', kwargs={'codigo': item.codigo})
        codigo_original = item.codigo

        response = self.client.post(editar_url, {
            'producto': 'CPU oficina',
            'marca': 'Dell',
            'modelo': 'OptiPlex 7050',
            'serie': '',
            'estado': 'no_disponible',
            'causa_no_disponible': 'bajo_pedido',
            'cantidad': '0',
            'costo': '0.00',
            'ubicacion': 'guayaquil_centro',
            'observacion': 'Esperando repuesto del proveedor.',
        })

        self.assertRedirects(response, reverse('econotec:inventario_detalle_item', kwargs={'codigo': item.codigo}))
        item.refresh_from_db()
        self.assertEqual(item.codigo, codigo_original)
        self.assertEqual(item.producto, 'CPU oficina')
        self.assertEqual(item.marca, 'Dell')
        self.assertEqual(item.modelo, 'OptiPlex 7050')
        self.assertEqual(item.estado, 'no_disponible')
        self.assertEqual(item.causa_no_disponible, 'bajo_pedido')
        self.assertEqual(item.cantidad, 0)
        self.assertEqual(item.ubicacion, 'guayaquil_centro')
        self.assertEqual(item.observacion, 'Esperando repuesto del proveedor.')

    def test_inventario_formulario_exige_causa_si_no_disponible(self):
        url = reverse('econotec:inventario_registrar', kwargs={
            'sede': 'guayaquil',
            'categoria': 'computadora',
            'tipo': 'pc',
        })

        response = self.client.post(url, {
            'producto': 'CPU completo',
            'marca': 'HP',
            'modelo': 'EliteDesk 800',
            'serie': '',
            'estado': 'no_disponible',
            'causa_no_disponible': '',
            'cantidad': '0',
            'costo': '8.75',
            'ubicacion': 'guayaquil_norte',
        })

        self.assertEqual(response.status_code, 200)
        self.assertIn('causa_no_disponible', response.context['form'].errors)
        self.assertFalse(InventarioItem.objects.exists())

        response = self.client.post(url, {
            'producto': 'CPU completo',
            'marca': 'HP',
            'modelo': 'EliteDesk 800',
            'serie': '',
            'estado': 'no_disponible',
            'causa_no_disponible': 'obsoleto',
            'cantidad': '0',
            'costo': '8.75',
            'ubicacion': 'guayaquil_norte',
        })

        item = InventarioItem.objects.get()
        self.assertRedirects(response, reverse('econotec:inventario_tabla', kwargs={
            'sede': 'guayaquil',
            'categoria': 'computadora',
            'tipo': 'pc',
        }))
        self.assertEqual(item.estado, 'no_disponible')
        self.assertEqual(item.causa_no_disponible, 'obsoleto')

    def test_inventario_eliminar_remueve_item_para_admin_y_tecnico(self):
        item = InventarioItem.objects.create(
            sede='guayaquil',
            categoria='computadora',
            tipo='pc',
            producto='CPU completo',
            marca='HP',
            modelo='EliteDesk 800',
            estado='disponible',
            cantidad=2,
            ubicacion='guayaquil_norte',
            registrado_por=self.usuario,
        )
        eliminar_url = reverse('econotec:inventario_eliminar', kwargs={'codigo': item.codigo})

        response = self.client.post(eliminar_url)

        self.assertRedirects(response, reverse('econotec:inventario_tabla', kwargs={
            'sede': 'guayaquil',
            'categoria': 'computadora',
            'tipo': 'pc',
        }))
        self.assertFalse(InventarioItem.objects.filter(pk=item.pk).exists())

    def test_inventario_acciones_bloquean_asesor_comercial(self):
        item = InventarioItem.objects.create(
            sede='guayaquil',
            categoria='computadora',
            tipo='pc',
            producto='CPU completo',
            marca='HP',
            modelo='EliteDesk 800',
            estado='disponible',
            cantidad=2,
            ubicacion='guayaquil_norte',
            registrado_por=self.usuario,
        )
        self.client.force_login(self.vendedor)

        tabla = self.client.get(reverse('econotec:inventario_tabla', kwargs={
            'sede': 'guayaquil',
            'categoria': 'computadora',
            'tipo': 'pc',
        }))
        self.assertNotContains(tabla, 'Acción')
        self.assertNotContains(tabla, reverse('econotec:inventario_editar', kwargs={'codigo': item.codigo}))

        edit_response = self.client.get(reverse('econotec:inventario_editar', kwargs={'codigo': item.codigo}))
        self.assertRedirects(edit_response, reverse('econotec:inventario_detalle_item', kwargs={'codigo': item.codigo}))

        delete_response = self.client.post(reverse('econotec:inventario_eliminar', kwargs={'codigo': item.codigo}))
        self.assertRedirects(delete_response, reverse('econotec:inventario_detalle_item', kwargs={'codigo': item.codigo}))
        self.assertTrue(InventarioItem.objects.filter(pk=item.pk).exists())

    def test_inventario_detalle_permite_guardar_cantidad(self):
        item = InventarioItem.objects.create(
            sede='guayaquil',
            categoria='computadora',
            tipo='pc',
            producto='CPU completo',
            marca='HP',
            modelo='EliteDesk 800',
            estado='disponible',
            cantidad=1,
            ubicacion='guayaquil_norte',
            registrado_por=self.usuario,
        )
        detalle_url = reverse(
            'econotec:inventario_detalle_item',
            kwargs={'codigo': item.codigo},
        )
        cantidad_url = reverse(
            'econotec:inventario_actualizar_cantidad',
            kwargs={'codigo': item.codigo},
        )

        detalle = self.client.get(detalle_url)
        self.assertContains(detalle, 'aria-label="Restar una unidad"')
        self.assertContains(detalle, 'aria-label="Sumar una unidad"')
        self.assertContains(detalle, 'Guardar cambios')

        response = self.client.post(cantidad_url, {'cantidad': '5'})
        self.assertRedirects(response, detalle_url)
        item.refresh_from_db()
        self.assertEqual(item.cantidad, 5)

        self.client.post(cantidad_url, {'cantidad': '-8'})
        item.refresh_from_db()
        self.assertEqual(item.cantidad, 0)

    def test_inventario_detalle_y_cantidad_requieren_sesion_y_no_modifican(self):
        item = InventarioItem.objects.create(
            sede='quito',
            categoria='tablet',
            tipo='tablet',
            producto='Tablet',
            marca='Samsung',
            modelo='A9',
            estado='disponible',
            cantidad=3,
            ubicacion='quito',
            registrado_por=self.usuario,
        )
        detalle_url = reverse(
            'econotec:inventario_detalle_item',
            kwargs={'codigo': item.codigo},
        )
        cantidad_url = reverse(
            'econotec:inventario_actualizar_cantidad',
            kwargs={'codigo': item.codigo},
        )
        self.client.logout()

        detalle = self.client.get(detalle_url)
        self.assertEqual(detalle.status_code, 302)
        self.assertEqual(detalle.url, f'{reverse("login")}?next={detalle_url}')

        response = self.client.post(cantidad_url, {'cantidad': '10'})
        self.assertEqual(response.status_code, 302)
        item.refresh_from_db()
        self.assertEqual(item.cantidad, 3)

    def test_detalle_muestra_asesor_que_registro_el_equipo(self):
        ingreso = self.crear_ingreso_reparacion(registrado_por=self.vendedor)

        response = self.client.get(reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk}))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Asesor que registró')
        self.assertContains(response, 'Kimberly')
        self.assertContains(response, 'kimberly@example.com')

    def test_detalle_muestra_check_reparacion_en_reparacion_y_garantia(self):
        ingreso = self.crear_ingreso_reparacion(subestado_reparacion='en_reparacion')
        ingreso_cliente = self.crear_ingreso_reparacion(subestado_reparacion='espera_cliente')
        ingreso_garantia = self.crear_ingreso_reparacion(
            estado='garantia',
            subestado_reparacion='',
            motivo_garantia='Garantía por retorno',
        )

        response = self.client.get(reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk}))
        response_cliente = self.client.get(
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso_cliente.pk})
        )
        response_garantia = self.client.get(
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso_garantia.pk})
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="reparacion-check-btn"')
        self.assertContains(response, 'Aún sigue reparando este equipo')
        self.assertEqual(response_cliente.status_code, 200)
        self.assertNotContains(response_cliente, 'id="reparacion-check-btn"')
        self.assertEqual(response_garantia.status_code, 200)
        self.assertContains(response_garantia, 'Garantía')
        self.assertContains(response_garantia, 'id="reparacion-check-btn"')

    def test_check_reparacion_registra_bitacora_una_vez_por_dia(self):
        ingreso = self.crear_ingreso_reparacion(
            marca='Sony',
            modelo_serie='Playstation 2',
            serie='23311',
            problema_reportado='No da video',
            accesorios_entregados='Cable de poder',
        )
        url = reverse('econotec:ingreso_reparacion_check', kwargs={'pk': ingreso.pk})
        zona_local = ZoneInfo('America/Guayaquil')

        with patch('econotec.views.timezone.localdate', return_value=date(2026, 7, 23)), \
             patch('econotec.bitacora.timezone.now', return_value=datetime(2026, 7, 23, 22, 15, tzinfo=zona_local)):
            response = self.client.post(url, HTTP_X_REQUESTED_WITH='XMLHttpRequest')
            response_repetido = self.client.post(url, HTTP_X_REQUESTED_WITH='XMLHttpRequest')

        eventos = BitacoraTecnico.objects.filter(
            user=self.usuario,
            ingreso=ingreso,
            metadata__accion='reparacion_check',
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['ok'])
        self.assertFalse(response.json()['already'])
        self.assertEqual(response_repetido.status_code, 200)
        self.assertTrue(response_repetido.json()['already'])
        self.assertEqual(eventos.count(), 1)
        self.assertIn('aún sigue reparando este equipo', eventos.first().texto)
        self.assertIn('Sony Playstation 2', eventos.first().texto)
        self.assertEqual(eventos.first().codigo, ingreso.codigo_equipo)

        from .bitacora import construir_bitacora_usuario
        reporte = construir_bitacora_usuario(self.usuario, dia=date(2026, 7, 23))
        self.assertIn('*10:15 PM* - El técnico Yandri aún sigue reparando este equipo', reporte['texto'])

        with patch('econotec.views.timezone.localdate', return_value=date(2026, 7, 24)), \
             patch('econotec.bitacora.timezone.now', return_value=datetime(2026, 7, 24, 0, 1, tzinfo=zona_local)):
            response_dia_siguiente = self.client.post(url, HTTP_X_REQUESTED_WITH='XMLHttpRequest')

        self.assertEqual(response_dia_siguiente.status_code, 200)
        self.assertFalse(response_dia_siguiente.json()['already'])
        self.assertEqual(eventos.count(), 2)

    def test_check_reparacion_omite_problema_reportado_no(self):
        ingreso = self.crear_ingreso_reparacion(
            problema_reportado='No.',
            accesorios_entregados='Ninguno',
        )

        response = self.client.post(
            reverse('econotec:ingreso_reparacion_check', kwargs={'pk': ingreso.pk}),
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        self.assertEqual(response.status_code, 200)
        evento = BitacoraTecnico.objects.get(
            user=self.usuario,
            ingreso=ingreso,
            metadata__accion='reparacion_check',
        )
        self.assertNotIn('problema reportado: No', evento.texto)
        self.assertIn('accesorios: Ninguno', evento.texto)

    def test_check_reparacion_permite_garantia(self):
        ingreso = self.crear_ingreso_reparacion(
            estado='garantia',
            subestado_reparacion='',
            motivo_garantia='Retorno por falla cubierta',
            equipo_garantia_manual='G1000',
            problema_reportado='No.',
            accesorios_entregados='Cargador',
        )

        response = self.client.post(
            reverse('econotec:ingreso_reparacion_check', kwargs={'pk': ingreso.pk}),
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['ok'])
        evento = BitacoraTecnico.objects.get(
            user=self.usuario,
            ingreso=ingreso,
            metadata__accion='reparacion_check',
        )
        self.assertNotIn('problema reportado: No', evento.texto)
        self.assertIn('garantía de G1000', evento.texto)
        self.assertIn('motivo de garantía: Retorno por falla cubierta', evento.texto)
        self.assertIn('estado confirmado: Garantía', evento.texto)

    def test_check_reparacion_admin_lo_ve_y_registra_en_bitacora_del_tecnico(self):
        ingreso = self.crear_ingreso_reparacion()
        self.client.force_login(self.admin)

        response = self.client.get(reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk}))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="reparacion-check-btn"')
        self.assertContains(response, 'bitácora de Yandri')

        url = reverse('econotec:ingreso_reparacion_check', kwargs={'pk': ingreso.pk})
        with patch('econotec.views.timezone.localdate', return_value=date(2026, 7, 23)):
            response_check = self.client.post(url, HTTP_X_REQUESTED_WITH='XMLHttpRequest')

        self.assertEqual(response_check.status_code, 200)
        self.assertTrue(response_check.json()['ok'])
        self.assertEqual(
            BitacoraTecnico.objects.filter(
                user=self.usuario,
                ingreso=ingreso,
                metadata__accion='reparacion_check',
            ).count(),
            1,
        )
        self.assertFalse(
            BitacoraTecnico.objects.filter(
                user=self.admin,
                ingreso=ingreso,
                metadata__accion='reparacion_check',
            ).exists()
        )

    def test_check_reparacion_rechaza_subestado_cliente(self):
        ingreso = self.crear_ingreso_reparacion(subestado_reparacion='espera_cliente')

        response = self.client.post(
            reverse('econotec:ingreso_reparacion_check', kwargs={'pk': ingreso.pk}),
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(BitacoraTecnico.objects.filter(ingreso=ingreso).count(), 0)

    def test_perfil_suma_cuatro_puntos_por_salida_buena_positiva(self):
        ingreso = self.crear_ingreso_reparacion()
        SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='pendiente_retiro',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            tecnico_reparo=self.usuario,
            registrado_por=self.usuario,
        )

        response = self.client.get(reverse('econotec:api_perfil'))

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['salidas_buenas'], 1)
        self.assertEqual(data['total'], 4)

    def test_perfil_no_suma_puntos_de_salida_buena_si_no_es_positiva(self):
        ingreso = self.crear_ingreso_reparacion()
        SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='no_reparable',
            cliente_recibe_conforme='no',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            tecnico_reparo=self.usuario,
            registrado_por=self.usuario,
        )

        response = self.client.get(reverse('econotec:api_perfil'))

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['salidas_buenas'], 0)
        self.assertEqual(data['salidas_malas'], 1)
        self.assertEqual(data['total'], 0)

    def test_formulario_salida_acredita_buena_y_mala_al_tecnico_seleccionado(self):
        User = get_user_model()
        tecnico_seleccionado = User.objects.create_user(
            username='TecnicoResultadoSalida',
            first_name='Tecnico',
            last_name='Resultado',
        )
        tecnico_seleccionado.groups.add(Group.objects.get(name='Tecnicos'))
        ingreso_positivo = self.crear_ingreso_reparacion(
            tecnico_encargado=self.usuario,
            fecha_ingreso=date(2026, 7, 15),
            valor_acordado=Decimal('0.00'),
        )
        response_positiva = self.client.post(
            reverse('econotec:salida_registrar', kwargs={'ingreso_pk': ingreso_positivo.pk}),
            self.salida_post_data(
                tecnico_reparo=str(tecnico_seleccionado.pk),
                estado_reparacion='pendiente_retiro',
                metodo_pago_final='sin_pago',
            ),
        )
        response_negativa, ingreso_negativo, salida_negativa = self.registrar_ingreso_negativo(
            IngresoEquipoForm.ESTADO_NO_SE_PUDO_REPARAR,
            tecnico_reparo=str(tecnico_seleccionado.pk),
            estado_reparacion='no_reparable',
            metodo_pago_final='sin_pago',
        )

        salida_positiva = SalidaEquipo.objects.get(ingreso=ingreso_positivo)
        self.assertEqual(response_positiva.status_code, 302)
        self.assertEqual(response_negativa.status_code, 302)
        self.assertEqual(salida_positiva.tecnico_reparo, tecnico_seleccionado)
        self.assertEqual(salida_negativa.tecnico_reparo, tecnico_seleccionado)
        self.assertEqual(salida_positiva.registrado_por, self.usuario)
        self.assertEqual(salida_negativa.registrado_por, self.usuario)

        self.client.force_login(tecnico_seleccionado)
        perfil_tecnico = self.client.get(reverse('econotec:api_perfil')).json()
        self.assertEqual(perfil_tecnico['salidas_buenas'], 1)
        self.assertEqual(perfil_tecnico['salidas_malas'], 1)
        self.assertEqual(perfil_tecnico['total'], 3)

        self.client.force_login(self.usuario)
        perfil_registrador = self.client.get(reverse('econotec:api_perfil')).json()
        self.assertEqual(perfil_registrador['salidas_buenas'], 0)
        self.assertEqual(perfil_registrador['salidas_malas'], 0)

        from .gamificacion import SALIDA_BUENA_ESTADOS, SALIDA_MALA_ESTADOS
        self.assertEqual(set(SALIDA_BUENA_ESTADOS), {
            'pendiente_retiro',
            'garantia',
            'garantia_fallos_adicionales',
            'retirado',
        })
        self.assertEqual(set(SALIDA_MALA_ESTADOS), {
            'no_reparable',
            'cliente_no_acepta',
            'chatarrerizacion',
        })

    def test_menu_ventas_muestra_control_de_pago_de_ventas(self):
        response = self.client.get(reverse('econotec:venta_menu'))

        self.assertContains(response, 'Control de Pago de Ventas')
        self.assertContains(response, 'Lista de Ventas con Pago Parcial')
        self.assertContains(response, reverse('econotec:venta_lista_parciales'))
        self.assertContains(response, reverse('econotec:pagos_ventas_menu'))

    def test_hoja_qr_muestra_categoria_marca_modelo_serie_y_problema(self):
        ingreso = self.crear_ingreso_reparacion(
            tipo_equipo='otro',
            tipo_equipo_otro='Consola',
            marca='Sony',
            modelo_serie='Playstation 5',
            serie='PS5-001',
            problema_reportado='No da grafica',
        )

        response = self.client.get(
            reverse('econotec:ingreso_imprimir_qr', kwargs={'pk': ingreso.pk})
        )

        self.assertContains(response, 'Consola', count=2)
        self.assertContains(response, 'Marca:', count=2)
        self.assertContains(response, 'Sony', count=2)
        self.assertContains(response, 'Modelo:', count=2)
        self.assertContains(response, 'Playstation 5', count=2)
        self.assertContains(response, 'Serie:', count=2)
        self.assertContains(response, 'PS5-001', count=2)
        self.assertContains(response, 'Problema:')
        self.assertContains(response, 'No da grafica', count=2)
        html = response.content.decode()
        self.assertLess(html.index('Consola'), html.index('Marca:'))
        self.assertLess(html.index('Marca:'), html.index('Modelo:'))
        self.assertLess(html.index('Modelo:'), html.index('Serie:'))
        self.assertLess(html.index('Serie:'), html.index('Problema:'))

    def test_hoja_qr_oculta_serie_si_no_se_registra(self):
        ingreso = self.crear_ingreso_reparacion(
            tipo_equipo='laptop',
            modelo_serie='Elitebook',
            serie='',
        )

        response = self.client.get(
            reverse('econotec:ingreso_imprimir_qr', kwargs={'pk': ingreso.pk})
        )

        self.assertContains(response, 'Modelo:', count=2)
        self.assertContains(response, 'Elitebook', count=2)
        self.assertNotContains(response, 'Serie:')

    def test_firma_cliente_imagen_no_es_obligatoria_si_cliente_no_firma(self):
        form = IngresoEquipoForm(data=self.ingreso_form_data(
            firma_cliente_opcion='no',
            firma_cliente_imagen='',
        ))

        self.assertTrue(form.is_valid(), form.errors.as_data())
        self.assertFalse(form.cleaned_data['firma_cliente'])
        self.assertEqual(form.cleaned_data['firma_cliente_imagen'], '')

    def test_firma_cliente_exige_seleccionar_si_o_no(self):
        data = self.ingreso_form_data()
        data.pop('firma_cliente_opcion')

        form = IngresoEquipoForm(data=data)

        self.assertFalse(form.is_valid())
        self.assertIn('firma_cliente_opcion', form.errors)

    def test_firma_cliente_si_exige_imagen_capturada(self):
        form = IngresoEquipoForm(data=self.ingreso_form_data(
            firma_cliente_opcion='si',
            firma_cliente_imagen='',
        ))

        self.assertFalse(form.is_valid())
        self.assertIn('firma_cliente_opcion', form.errors)

    def test_registrar_ingreso_guarda_firma_cliente_opcional(self):
        self.activar_sede_guayaquil()

        response = self.client.post(
            reverse('econotec:ingreso_registrar'),
            self.ingreso_registro_post_data(
                **{
                    'ing-firma_cliente_opcion': 'si',
                    'ing-firma_cliente_imagen': self.FIRMA_PNG_DATA_URI,
                }
            ),
        )

        ingreso = IngresoEquipo.objects.get(cliente=self.cliente_existente)
        self.assertRedirects(
            response,
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk}),
        )
        self.assertTrue(ingreso.firma_cliente)
        self.assertEqual(ingreso.firma_cliente_imagen, self.FIRMA_PNG_DATA_URI)

    @override_settings(
        EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
        INGRESO_EMAIL_AUTOMATICO=True,
        INGRESO_EMAIL_ADJUNTAR_PDF=True,
        DEFAULT_FROM_EMAIL='Econotec <no-reply@econotec.test>',
    )
    def test_registrar_ingreso_envia_correo_profesional_con_pdf(self):
        self.activar_sede_guayaquil()
        mail.outbox.clear()

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                reverse('econotec:ingreso_registrar'),
                self.ingreso_registro_post_data(),
            )

        ingreso = IngresoEquipo.objects.get(cliente=self.cliente_existente)
        self.assertRedirects(
            response,
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk}),
        )
        self.assertEqual(len(mail.outbox), 1)
        correo = mail.outbox[0]
        self.assertEqual(correo.to, [self.cliente_existente.correo])
        self.assertIn(ingreso.codigo_equipo, correo.subject)
        self.assertIn('No enciende', correo.body)
        html = next(
            alternativa[0]
            for alternativa in correo.alternatives
            if alternativa[1] == 'text/html'
        )
        self.assertIn('SOLICITUD RECIBIDA', html.upper())
        self.assertIn(ingreso.codigo_equipo, html)
        self.assertIn(self.cliente_existente.nombres, html)
        self.assertIn(self.cliente_existente.cedula, html)
        self.assertIn(self.cliente_existente.correo, html)
        self.assertIn('MacBook M4 S', html)
        self.assertNotIn('{{', html)
        self.assertNotIn('{%', html)
        self.assertEqual(len(correo.attachments), 1)
        adjunto = correo.attachments[0]
        self.assertEqual(
            adjunto[0],
            f'Solicitud_de_ingreso_{ingreso.codigo_equipo}.pdf',
        )
        self.assertTrue(adjunto[1].startswith(b'%PDF-'))
        self.assertEqual(adjunto[2], 'application/pdf')

    @override_settings(
        EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
        INGRESO_EMAIL_AUTOMATICO=True,
    )
    def test_registrar_ingreso_sin_correo_no_impide_guardar(self):
        self.activar_sede_guayaquil()
        mail.outbox.clear()

        response = self.client.post(
            reverse('econotec:ingreso_registrar'),
            self.ingreso_registro_post_data(**{'cli-correo': ''}),
        )

        ingreso = IngresoEquipo.objects.get(cliente=self.cliente_existente)
        self.assertRedirects(
            response,
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk}),
        )
        self.assertEqual(len(mail.outbox), 0)
        self.assertEqual(ingreso.cliente.correo, '')

    @override_settings(
        EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
        INGRESO_EMAIL_AUTOMATICO=True,
        INGRESO_EMAIL_ADJUNTAR_PDF=True,
    )
    def test_nuevo_cliente_tambien_recibe_correo_automatico(self):
        self.activar_sede_guayaquil()
        mail.outbox.clear()
        datos = self.ingreso_registro_post_data(**{
            'cli-cedula': '0923456789',
            'cli-nombres': 'Ana Prueba',
            'cli-whatsapp': '0987654321',
            'cli-correo': 'ana@example.com',
            'cli-sector': 'centro',
        })

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                reverse('econotec:ingreso_registrar'),
                datos,
            )

        ingreso = IngresoEquipo.objects.get(cliente__cedula='0923456789')
        self.assertRedirects(
            response,
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk}),
        )
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].to, ['ana@example.com'])

    @override_settings(
        EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
        INGRESO_EMAIL_AUTOMATICO=True,
    )
    @patch('econotec.emails.EmailMultiAlternatives.send', side_effect=OSError('SMTP no disponible'))
    def test_fallo_correo_automatico_no_revierte_ingreso(self, _send):
        self.activar_sede_guayaquil()

        with self.assertLogs('econotec.emails', level='ERROR'):
            with self.captureOnCommitCallbacks(execute=True):
                response = self.client.post(
                    reverse('econotec:ingreso_registrar'),
                    self.ingreso_registro_post_data(),
                )

        ingreso = IngresoEquipo.objects.get(cliente=self.cliente_existente)
        self.assertRedirects(
            response,
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk}),
        )
        self.assertEqual(ingreso.problema_reportado, 'No enciende')

    def test_detalle_no_muestra_boton_de_correo_manual(self):
        ingreso = self.crear_ingreso_reparacion()

        response = self.client.get(
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk}),
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'Enviar por correo')

    @override_settings(
        EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
        SALIDA_EMAIL_AUTOMATICO=True,
        EQUIPO_EMAIL_ADJUNTAR_PDF=True,
        DEFAULT_FROM_EMAIL='Econotec <no-reply@econotec.test>',
    )
    def test_confirmar_aviso_finalizacion_envia_acta_saldo_y_bodegaje(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('30.00'),
            abono_anticipo=Decimal('10.00'),
        )
        Abono.objects.create(
            ingreso=ingreso,
            fecha=date(2026, 8, 20),
            monto=Decimal('5.00'),
            metodo='efectivo',
            registrado_por=self.usuario,
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 8, 23),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        mail.outbox.clear()

        response = self.client.post(
            reverse('econotec:salida_enviar_correo_finalizacion', kwargs={'pk': salida.pk}),
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['ok'])
        self.assertEqual(len(mail.outbox), 1)
        correo = mail.outbox[0]
        self.assertEqual(correo.to, [self.cliente_existente.correo])
        self.assertIn(ingreso.codigo_equipo, correo.subject)
        self.assertIn('finalizado', correo.subject.lower())
        html = next(
            alternativa[0]
            for alternativa in correo.alternatives
            if alternativa[1] == 'text/html'
        )
        self.assertIn('TU EQUIPO ESTÁ LISTO', html.upper())
        self.assertIn('$15,00', html)
        self.assertIn('Historial completo de pagos', html)
        self.assertIn('Regla de bodegaje', html)
        self.assertIn('5 días de gracia', html)
        self.assertIn('$1,00', html)
        self.assertEqual(len(correo.attachments), 1)
        self.assertEqual(
            correo.attachments[0][0],
            f'Acta_equipo_finalizado_{ingreso.codigo_equipo}.pdf',
        )
        self.assertTrue(correo.attachments[0][1].startswith(b'%PDF-'))
        self.assertEqual(correo.attachments[0][2], 'application/pdf')

    @override_settings(
        EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
        SALIDA_EMAIL_AUTOMATICO=True,
    )
    def test_aviso_finalizacion_sin_correo_informa_y_no_falla(self):
        self.cliente_existente.correo = ''
        self.cliente_existente.save(update_fields=['correo'])
        ingreso = self.crear_ingreso_reparacion(valor_acordado=Decimal('20.00'))
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 8, 23),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        mail.outbox.clear()

        response = self.client.post(
            reverse('econotec:salida_enviar_correo_finalizacion', kwargs={'pk': salida.pk}),
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.json()['ok'])
        self.assertEqual(response.json()['codigo'], 'sin_correo')
        self.assertEqual(len(mail.outbox), 0)

    @override_settings(
        EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
        PAGO_EMAIL_AUTOMATICO=True,
        EQUIPO_EMAIL_ADJUNTAR_PDF=True,
    )
    def test_registrar_abono_envia_valor_saldo_historial_y_acta(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('30.00'),
            abono_anticipo=Decimal('5.00'),
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date.today(),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        mail.outbox.clear()

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                reverse('econotec:abono_crear', kwargs={'ingreso_pk': ingreso.pk}),
                {
                    'fecha': date.today().isoformat(),
                    'monto': '10.00',
                    'metodo': 'efectivo',
                    'banco': '',
                    'banco_otro': '',
                    'tarjeta_app': '',
                    'comprobante_url': '',
                    'numero_recibo': '',
                    'observaciones': 'Abono recibido en oficina.',
                    'factura_realizada': 'no',
                    'factura_nombres': '',
                    'factura_cedula': '',
                    'factura_correo': '',
                    'bodegaje_decision': 'na',
                    'bodegaje_monto_aplicado': '0.00',
                    'accion_abono': 'registrar',
                },
            )

        self.assertRedirects(
            response,
            reverse('econotec:ingreso_abonos', kwargs={'pk': ingreso.pk}),
        )
        ingreso.refresh_from_db()
        self.assertEqual(ingreso.diferencia, Decimal('15.00'))
        self.assertEqual(len(mail.outbox), 1)
        correo = mail.outbox[0]
        self.assertIn('Abono registrado', correo.subject)
        html = next(
            alternativa[0]
            for alternativa in correo.alternatives
            if alternativa[1] == 'text/html'
        )
        self.assertIn('Recibimos tu abono', html)
        self.assertIn('$10,00', html)
        self.assertIn('$15,00', html)
        self.assertIn('Abono recibido en oficina.', html)
        self.assertIn('Historial completo de pagos', html)
        self.assertIn('Regla de bodegaje', html)
        self.assertEqual(len(correo.attachments), 1)
        self.assertTrue(correo.attachments[0][1].startswith(b'%PDF-'))

    @override_settings(
        EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
        PAGO_EMAIL_AUTOMATICO=True,
        EQUIPO_EMAIL_ADJUNTAR_PDF=True,
    )
    def test_abono_completo_y_salida_envia_correo_de_cierre(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('30.00'),
            abono_anticipo=Decimal('0.00'),
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date.today(),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        mail.outbox.clear()

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                reverse('econotec:abono_crear', kwargs={'ingreso_pk': ingreso.pk}),
                {
                    'fecha': date.today().isoformat(),
                    'monto': '30.00',
                    'metodo': 'efectivo',
                    'banco': '',
                    'banco_otro': '',
                    'tarjeta_app': '',
                    'comprobante_url': '',
                    'numero_recibo': '',
                    'observaciones': 'Pago final.',
                    'factura_realizada': 'no',
                    'factura_nombres': '',
                    'factura_cedula': '',
                    'factura_correo': '',
                    'bodegaje_decision': 'na',
                    'bodegaje_monto_aplicado': '0.00',
                    'accion_abono': 'registrar_y_salida',
                },
            )

        self.assertRedirects(response, reverse('econotec:salida_retiros_lista'))
        salida.refresh_from_db()
        self.assertEqual(salida.estado_reparacion, 'retirado')
        self.assertEqual(len(mail.outbox), 1)
        correo = mail.outbox[0]
        self.assertIn('Salida de la oficina confirmada', correo.subject)
        html = next(
            alternativa[0]
            for alternativa in correo.alternatives
            if alternativa[1] == 'text/html'
        )
        self.assertIn('Tu equipo salió de la oficina', html)
        self.assertIn('Salida física confirmada', html)
        self.assertIn('$0,00', html)
        self.assertIn('Regla de bodegaje', html)
        self.assertEqual(len(correo.attachments), 1)

    @override_settings(PAGO_EMAIL_AUTOMATICO=True)
    @patch('econotec.emails.EmailMultiAlternatives.send', side_effect=OSError('SMTP no disponible'))
    def test_fallo_correo_abono_no_revierte_el_pago(self, _send):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('30.00'),
            abono_anticipo=Decimal('0.00'),
        )

        with self.assertLogs('econotec.emails', level='ERROR'):
            with self.captureOnCommitCallbacks(execute=True):
                response = self.client.post(
                    reverse('econotec:abono_crear', kwargs={'ingreso_pk': ingreso.pk}),
                    {
                        'fecha': '2026-08-23',
                        'monto': '10.00',
                        'metodo': 'efectivo',
                        'banco': '',
                        'banco_otro': '',
                        'tarjeta_app': '',
                        'comprobante_url': '',
                        'numero_recibo': '',
                        'observaciones': '',
                        'factura_realizada': 'no',
                        'factura_nombres': '',
                        'factura_cedula': '',
                        'factura_correo': '',
                        'bodegaje_decision': 'na',
                        'bodegaje_monto_aplicado': '0.00',
                        'accion_abono': 'registrar',
                    },
                )

        self.assertRedirects(
            response,
            reverse('econotec:ingreso_abonos', kwargs={'pk': ingreso.pk}),
        )
        ingreso.refresh_from_db()
        self.assertEqual(ingreso.abonos.count(), 1)
        self.assertEqual(ingreso.diferencia, Decimal('20.00'))

    def test_ingreso_imprimir_muestra_firma_cliente_si_existe(self):
        ingreso = self.crear_ingreso_reparacion(
            firma_cliente=True,
            firma_cliente_imagen=self.FIRMA_PNG_DATA_URI,
        )

        response = self.client.get(reverse('econotec:ingreso_imprimir', kwargs={'pk': ingreso.pk}))

        self.assertContains(response, 'alt="Firma del cliente"')
        self.assertContains(response, self.FIRMA_PNG_DATA_URI)

        pdf_response = self.client.get(reverse('econotec:ingreso_pdf', kwargs={'pk': ingreso.pk}))
        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(pdf_response['Content-Type'], 'application/pdf')

    def test_hoja_ingreso_regenerada_muestra_abonos_posteriores(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('100.00'),
            abono_anticipo=Decimal('10.00'),
        )
        Abono.objects.create(
            ingreso=ingreso,
            fecha=date(2026, 7, 10),
            monto=Decimal('25.00'),
            registrado_por=self.usuario,
        )

        response = self.client.get(
            reverse('econotec:ingreso_imprimir', kwargs={'pk': ingreso.pk})
        )

        self.assertContains(response, 'Total abonado:')
        self.assertContains(response, '$35,00')
        self.assertContains(response, '$65,00')

        from reportlab.pdfgen.canvas import Canvas
        with patch.object(Canvas, 'drawString', autospec=True) as draw_string:
            pdf_response = self.client.get(
                reverse('econotec:ingreso_pdf', kwargs={'pk': ingreso.pk})
            )

        textos_pdf = [call.args[3] for call in draw_string.call_args_list]
        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(pdf_response['Content-Type'], 'application/pdf')
        self.assertIn('$ 35.00', textos_pdf)
        self.assertIn('$ 65.00', textos_pdf)

    def test_tipo_equipo_mando_se_acepta_y_se_imprime(self):
        form = IngresoEquipoForm(data=self.ingreso_form_data(
            tipo_equipo='mando',
            tipo_equipo_otro='',
            marca='Sony',
            modelo_serie='DualSense',
        ))

        self.assertTrue(form.is_valid(), form.errors.as_json())

        ingreso = self.crear_ingreso_reparacion(
            tipo_equipo='mando',
            marca='Sony',
            modelo_serie='DualSense',
        )
        response = self.client.get(reverse('econotec:ingreso_imprimir', kwargs={'pk': ingreso.pk}))

        self.assertContains(response, 'MANDO <span class="check-box">X</span>', html=False)
        self.assertContains(response, 'Mando')

    def test_tipo_equipo_maquina_coser_se_acepta_y_se_imprime(self):
        form = IngresoEquipoForm(data=self.ingreso_form_data(
            tipo_equipo='maquina_coser',
            tipo_equipo_otro='',
            marca='Singer',
            modelo_serie='Tradition 2250',
        ))

        self.assertTrue(form.is_valid(), form.errors.as_json())

        ingreso = self.crear_ingreso_reparacion(
            tipo_equipo='maquina_coser',
            marca='Singer',
            modelo_serie='Tradition 2250',
        )
        response = self.client.get(reverse('econotec:ingreso_imprimir', kwargs={'pk': ingreso.pk}))

        self.assertContains(response, 'MAQUINA DE COSER <span class="check-box">X</span>', html=False)
        self.assertContains(response, 'Máquina de Coser')

    def test_historial_muestra_maquina_coser_como_apartado_propio(self):
        self.crear_ingreso_reparacion(
            tipo_equipo='maquina_coser',
            marca='Singer',
            modelo_serie='Tradition 2250',
        )

        response = self.client.get(reverse('econotec:historial_lista'), {
            'ano': '2026',
            'mes': '7',
        })

        self.assertContains(response, 'Máquina de Coser')
        self.assertContains(response, 'Singer')

    def test_formulario_ingreso_incluye_responsive_movil_y_firma_tactil(self):
        self.activar_sede_guayaquil()

        response = self.client.get(reverse('econotec:ingreso_registrar'))

        self.assertContains(response, '<meta name="viewport" content="width=device-width, initial-scale=1.0">')
        self.assertContains(response, '@media (max-width: 640px)')
        self.assertContains(response, 'id="firma-modal"')
        self.assertContains(response, 'id="firma-canvas"')
        self.assertContains(response, 'touch-action: none')
        self.assertContains(response, 'firma-modal-open')
        self.assertContains(response, 'accesorios-opciones')

    def test_alerta_bodegaje_usa_tecnico_de_salida_y_tiene_desplegable(self):
        User = get_user_model()
        tecnicos = Group.objects.get(name='Tecnicos')
        tecnico_entrada = User.objects.create_user(
            username='EntradaTec',
            first_name='Entrada',
            last_name='Tec',
        )
        tecnico_salida = User.objects.create_user(
            username='SalidaTec',
            first_name='Salida',
            last_name='Tec',
        )
        tecnico_entrada.groups.add(tecnicos)
        tecnico_salida.groups.add(tecnicos)
        ingreso = self.crear_ingreso_reparacion(
            tecnico_encargado=tecnico_entrada,
            estado='entregado',
            subestado_entregado='con_solucion',
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date.today() - timedelta(days=5),
            estado_reparacion='pendiente_retiro',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            tecnico_reparo=tecnico_salida,
            registrado_por=self.usuario,
        )

        self.assertEqual(list(salidas_bodegaje_qs(usuario=tecnico_salida)), [salida])
        self.assertEqual(list(salidas_bodegaje_qs(usuario=tecnico_entrada)), [])

        response = self.client.get(reverse('econotec:bienvenida'))

        self.assertContains(response, 'Téc.:')
        self.assertContains(response, 'Salida Tec')
        self.assertNotContains(response, 'Entrada Tec')
        self.assertContains(response, 'id="btn-toggle-bodegaje"')
        self.assertContains(response, 'aria-controls="alerta-bodegaje-body"')
        self.assertContains(response, 'function toggleDashboardBodegaje')
        self.assertContains(response, 'salida-oficina-btn is-bloqueado')
        self.assertContains(response, '💳')
        self.assertContains(response, 'Saldo pendiente')
        self.assertContains(response, 'Debe pagar $25,00')

        response = self.client.get(reverse('econotec:alertas_bodegaje'))

        self.assertContains(response, 'Técnico de salida', count=1)
        self.assertContains(response, 'Salida Tec')
        self.assertNotContains(response, 'Entrada Tec')

    def test_menu_bodegaje_esta_disponible_para_roles_operativos(self):
        ruta = reverse('econotec:admin_activos_bodegaje')

        for usuario in (self.usuario, self.vendedor, self.admin):
            with self.subTest(usuario=usuario.username):
                self.client.force_login(usuario)
                response = self.client.get(reverse('econotec:bienvenida'))

                self.assertEqual(response.status_code, 200)
                self.assertContains(response, 'Bodegaje / Chatarrerización')
                self.assertContains(response, f'href="{ruta}"')

    def test_panel_bodegaje_permite_avisos_a_todos_y_reserva_chatarrerizacion_al_admin(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('30.00'),
            abono_anticipo=Decimal('10.00'),
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date.today() - timedelta(days=10),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        ruta = reverse('econotec:admin_activos_bodegaje')
        accion_chatarrerizacion = (
            f"onclick=\"cambiarEstado({salida.pk}, 'chatarrerizacion')\""
        )

        for usuario in (self.usuario, self.vendedor):
            with self.subTest(usuario=usuario.username):
                self.client.force_login(usuario)
                response = self.client.get(ruta)

                self.assertEqual(response.status_code, 200)
                self.assertContains(response, '💬 Enviar WhatsApp')
                self.assertContains(response, '✉️ Enviar correo')
                self.assertContains(response, 'Servicio:')
                self.assertContains(response, 'Bodegaje:')
                self.assertContains(response, 'Total hoy:')
                self.assertNotContains(response, accion_chatarrerizacion)

        self.client.force_login(self.admin)
        response = self.client.get(ruta)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, accion_chatarrerizacion)

    def test_whatsapp_bodegaje_incluye_dias_saldo_total_y_advertencia(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('30.00'),
            abono_anticipo=Decimal('10.00'),
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date.today() - timedelta(days=10),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )

        enlace = whatsapp_link_bodegaje(salida)
        mensaje = parse_qs(urlparse(enlace).query)['text'][0]

        self.assertIn('AVISO IMPORTANTE DE RETIRO Y BODEGAJE', mensaje)
        self.assertIn(ingreso.codigo_equipo, mensaje)
        self.assertIn('10 día(s)', mensaje)
        self.assertIn('Saldo pendiente del servicio: *$20,00*', mensaje)
        self.assertIn('Bodegaje acumulado (6 día(s)): *$6,00*', mensaje)
        self.assertIn('Total a regularizar al día de hoy: *$26,00*', mensaje)
        self.assertIn('5 días de gracia', mensaje)
        self.assertIn('chatarrerización', mensaje)

    @override_settings(
        EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
        EQUIPO_EMAIL_ADJUNTAR_PDF=True,
        DEFAULT_FROM_EMAIL='Econotec <no-reply@econotec.test>',
    )
    def test_correo_bodegaje_incluye_valores_historial_regla_y_acta_pdf(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('30.00'),
            abono_anticipo=Decimal('5.00'),
        )
        Abono.objects.create(
            ingreso=ingreso,
            fecha=date.today() - timedelta(days=3),
            monto=Decimal('5.00'),
            metodo='efectivo',
            observaciones='Pago anterior en efectivo.',
            registrado_por=self.usuario,
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date.today() - timedelta(days=10),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        mail.outbox.clear()

        response = self.client.post(
            reverse('econotec:bodegaje_enviar_correo', kwargs={'pk': salida.pk}),
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['ok'])
        self.assertEqual(len(mail.outbox), 1)
        correo = mail.outbox[0]
        self.assertEqual(correo.to, [self.cliente_existente.correo])
        self.assertIn(ingreso.codigo_equipo, correo.subject)
        self.assertIn('bodegaje', correo.subject.lower())
        html = next(
            alternativa[0]
            for alternativa in correo.alternatives
            if alternativa[1] == 'text/html'
        )
        self.assertIn('Tu equipo continúa en bodegaje', html)
        self.assertIn('$20,00', html)
        self.assertIn('$6,00', html)
        self.assertIn('$26,00', html)
        self.assertIn('5 días de gracia', html)
        self.assertIn('Pago anterior en efectivo.', html)
        self.assertIn('Advertencia de chatarrerización', html)
        self.assertNotIn('{{', html)
        self.assertEqual(len(correo.attachments), 1)
        self.assertEqual(
            correo.attachments[0][0],
            f'Acta_actualizada_{ingreso.codigo_equipo}.pdf',
        )
        self.assertTrue(correo.attachments[0][1].startswith(b'%PDF-'))
        self.assertEqual(correo.attachments[0][2], 'application/pdf')

    @override_settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend')
    def test_correo_bodegaje_sin_direccion_no_envia_y_explica_el_motivo(self):
        self.cliente_existente.correo = ''
        self.cliente_existente.save(update_fields=['correo'])
        ingreso = self.crear_ingreso_reparacion()
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date.today() - timedelta(days=10),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        mail.outbox.clear()

        response = self.client.post(
            reverse('econotec:bodegaje_enviar_correo', kwargs={'pk': salida.pk}),
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.json()['ok'])
        self.assertEqual(response.json()['codigo'], 'sin_correo')
        self.assertEqual(len(mail.outbox), 0)

    def test_solo_admin_puede_confirmar_chatarrerizacion_activa(self):
        ingreso = self.crear_ingreso_reparacion()
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date.today() - timedelta(days=10),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        ruta = reverse('econotec:admin_activos_bodegaje')
        payload = json.dumps({
            'accion': 'chatarrerizacion',
            'salida_id': salida.pk,
        })

        response = self.client.post(
            ruta,
            data=payload,
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 403)
        salida.refresh_from_db()
        self.assertEqual(salida.estado_reparacion, 'pendiente_retiro')
        self.assertIsNone(salida.fecha_retiro_real)

        self.client.force_login(self.admin)
        response = self.client.post(
            ruta,
            data=payload,
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok')
        salida.refresh_from_db()
        self.assertEqual(salida.estado_reparacion, 'chatarrerizacion')
        self.assertEqual(salida.fecha_retiro_real, date.today())

    def test_salida_totales_acredita_ranking_al_tecnico_que_reparo(self):
        User = get_user_model()
        tecnicos = Group.objects.get(name='Tecnicos')
        tecnico_entrada = User.objects.create_user(
            username='EntradaRanking',
            first_name='Entrada',
            last_name='Ranking',
        )
        tecnico_salida = User.objects.create_user(
            username='SalidaRanking',
            first_name='Salida',
            last_name='Ranking',
        )
        tecnico_entrada.groups.add(tecnicos)
        tecnico_salida.groups.add(tecnicos)
        ingreso = self.crear_ingreso_reparacion(
            tecnico_encargado=tecnico_entrada,
            fecha_ingreso=date(2026, 7, 1),
            valor_acordado=Decimal('100.00'),
            abono_anticipo=Decimal('10.00'),
        )
        self.crear_venta_producto(
            tecnico_encargado=tecnico_entrada,
            fecha_ingreso=date(2026, 7, 2),
            valor_acordado=Decimal('500.00'),
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 20),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=tecnico_salida,
            valor_final_cobrado=Decimal('90.00'),
            metodo_pago_final='efectivo',
            registrado_por=self.usuario,
        )
        self.client.force_login(self.admin)

        response = self.client.get(reverse('econotec:salida_totales'), {
            'desde': '2026-07-01',
            'hasta': '2026-07-31',
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['total_equipos'], 1)
        self.assertEqual(len(response.context['ranking_ingresos']), 1)
        ranking_ingreso = response.context['ranking_ingresos'][0]
        self.assertEqual(ranking_ingreso['tecnico_id'], tecnico_entrada.pk)
        self.assertEqual(ranking_ingreso['nombre'], 'Entrada Ranking')
        self.assertEqual(ranking_ingreso['num_ingresos'], 1)
        self.assertEqual(ranking_ingreso['sin_salida'], 0)
        self.assertEqual(ranking_ingreso['con_salida'], 1)
        self.assertEqual(ranking_ingreso['total_acordado'], Decimal('100.00'))
        self.assertEqual(ranking_ingreso['total_anticipo'], Decimal('10.00'))
        self.assertEqual(len(response.context['ranking']), 1)
        ranking = response.context['ranking'][0]
        self.assertEqual(ranking['tecnico_id'], tecnico_salida.pk)
        self.assertEqual(ranking['nombre'], 'Salida Ranking')
        self.assertEqual(ranking['num_equipos'], 1)
        self.assertEqual(ranking['salidas_positivas'], 1)
        self.assertEqual(ranking['salidas_negativas'], 0)
        self.assertEqual(ranking['total_recaudado'], Decimal('90.00'))
        self.assertContains(response, 'Salida Ranking')
        self.assertContains(response, 'Entrada Ranking')
        self.assertContains(response, 'Ranking de Ingresos por Técnico Asignado')
        self.assertContains(response, 'Ranking de Técnicos por Salidas Reparadas')

    def test_tecnico_y_asesor_ven_ranking_filtrado_sin_valores_monetarios(self):
        ingreso_periodo = self.crear_ingreso_reparacion(
            fecha_ingreso=date(2026, 7, 5),
            valor_acordado=Decimal('987.65'),
            abono_anticipo=Decimal('123.45'),
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso_periodo,
            fecha_salida=date(2026, 7, 20),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('864.20'),
            metodo_pago_final='efectivo',
            registrado_por=self.usuario,
        )
        self.crear_ingreso_reparacion(
            fecha_ingreso=date(2026, 8, 5),
            valor_acordado=Decimal('555.55'),
        )

        for usuario in (self.usuario, self.vendedor):
            with self.subTest(rol=usuario.username):
                self.client.force_login(usuario)
                response = self.client.get(reverse('econotec:salida_totales'), {
                    'desde': '2026-07-01',
                    'hasta': '2026-07-31',
                })

                self.assertEqual(response.status_code, 200)
                self.assertFalse(response.context['mostrar_valores_monetarios'])
                self.assertEqual(response.context['filtros'], {
                    'desde': '2026-07-01',
                    'hasta': '2026-07-31',
                })
                self.assertEqual(response.context['total_equipos'], 1)
                self.assertEqual(response.context['total_salidas_global'], 1)
                self.assertFalse(any(
                    'total_acordado_global' in contexto
                    for contexto in response.context
                ))
                self.assertNotIn('total_acordado', response.context['ranking_ingresos'][0])
                self.assertNotIn('total_recaudado', response.context['ranking'][0])
                self.assertContains(response, 'Equipos ingresados')
                self.assertContains(response, 'Equipos finalizados positivos')
                self.assertContains(response, 'value="2026-07-01"')
                self.assertContains(response, 'value="2026-07-31"')
                self.assertContains(response, 'fecha_desde=2026-07-01')
                self.assertContains(response, 'fecha_hasta=2026-07-31')
                self.assertNotContains(response, 'Total acordado')
                self.assertNotContains(response, 'Total recaudado')
                self.assertNotContains(response, 'Diagnóstico (No reparados)')
                self.assertNotContains(response, '>Anticipos<')
                self.assertNotContains(response, '>Acordado<')
                self.assertNotContains(response, '>Recaudado<')
                self.assertNotContains(response, '$987,65')
                self.assertNotContains(response, '$123,45')
                self.assertNotContains(response, '$864,20')

    def test_superusuario_asignado_como_tecnico_tampoco_ve_dinero_en_ranking(self):
        User = get_user_model()
        superusuario_tecnico = User.objects.create_superuser(
            username='TecnicoConPermisosElevados',
            email='tecnico-elevado@example.com',
            password='testpass123',
        )
        superusuario_tecnico.groups.add(Group.objects.get(name='Tecnicos'))
        ingreso = self.crear_ingreso_reparacion(
            fecha_ingreso=date(2026, 7, 5),
            valor_acordado=Decimal('987.65'),
            abono_anticipo=Decimal('123.45'),
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 20),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('864.20'),
            metodo_pago_final='efectivo',
            registrado_por=self.usuario,
        )
        self.client.force_login(superusuario_tecnico)

        response = self.client.get(reverse('econotec:salida_totales'))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context['mostrar_valores_monetarios'])
        self.assertFalse(response.context['puede_ver_valores_ranking'])
        self.assertFalse(any(
            'total_acordado_global' in contexto
            for contexto in response.context
        ))
        self.assertNotIn('total_acordado', response.context['ranking_ingresos'][0])
        self.assertNotIn('total_recaudado', response.context['ranking'][0])
        self.assertNotContains(response, 'Total acordado')
        self.assertNotContains(response, 'Total recaudado')
        self.assertNotContains(response, '>Anticipos<')
        self.assertNotContains(response, '>Acordado<')
        self.assertNotContains(response, '>Recaudado<')
        self.assertNotContains(response, '$987,65')
        self.assertNotContains(response, '$123,45')
        self.assertNotContains(response, '$864,20')

        inicio = self.client.get(reverse('econotec:bienvenida'))
        self.assertContains(
            inicio,
            'Consulta trabajos asignados, equipos finalizados, resultados y efectividad de cada técnico.',
        )

    def test_salida_totales_separa_fuera_de_oficina_y_finalizados(self):
        ingreso_fuera = self.crear_ingreso_reparacion(
            fecha_ingreso=date(2026, 8, 1),
        )
        ingreso_en_oficina = self.crear_ingreso_reparacion(
            fecha_ingreso=date(2026, 8, 2),
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso_fuera,
            fecha_salida=date(2026, 8, 10),
            fecha_retiro_real=date(2026, 8, 12),
            # La ubicación depende de fecha_retiro_real, aunque el resultado
            # histórico no tenga el código "retirado".
            estado_reparacion='no_reparable',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso_en_oficina,
            fecha_salida=date(2026, 8, 11),
            estado_reparacion='revision',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        self.client.force_login(self.admin)

        response = self.client.get(reverse('econotec:salida_totales'), {
            'desde': '2026-08-01',
            'hasta': '2026-08-31',
        })

        self.assertEqual(response.status_code, 200)
        ranking = next(
            fila for fila in response.context['ranking']
            if fila['tecnico_id'] == self.usuario.pk
        )
        self.assertEqual(ranking['num_equipos'], 2)
        self.assertEqual(ranking['entregados'], 1)
        self.assertEqual(ranking['pendientes'], 1)
        self.assertContains(response, 'Fuera de oficina')
        self.assertContains(response, 'En oficina')
        self.assertContains(response, 'Equipos finalizados')
        self.assertContains(
            response,
            f'{reverse("econotec:salida_retiros_lista")}?tecnico_salida={self.usuario.pk}',
        )
        self.assertContains(
            response,
            f'{reverse("econotec:salida_lista")}?tecnico_salida={self.usuario.pk}',
        )

    def test_top_clientes_cuenta_equipos_reales_por_sede_sin_multiplicar(self):
        biomedics = Cliente.objects.create(
            cedula='0993018740001',
            nombres='BIOMEDICIS',
            whatsapp='0967792636',
            correo='eromero@grupobiomedics.com',
            sector='norte',
        )
        for _ in range(5):
            self.crear_ingreso_reparacion(cliente=biomedics, sede='guayaquil')
        for _ in range(2):
            self.crear_ingreso_reparacion(cliente=biomedics, sede='quito')
        for _ in range(3):
            self.crear_venta_producto(cliente=biomedics)

        response = self.client.get(reverse('econotec:cliente_top_recurrentes'))

        guayaquil = {
            cliente.pk: cliente.total_ingresos
            for cliente in response.context['clientes_guayaquil']
        }
        quito = {
            cliente.pk: cliente.total_ingresos
            for cliente in response.context['clientes_quito']
        }
        self.assertEqual(response.status_code, 200)
        self.assertEqual(guayaquil[biomedics.pk], 5)
        self.assertEqual(quito[biomedics.pk], 2)

    def test_busqueda_clientes_ignora_tildes_y_mayusculas(self):
        self.cliente_existente.nombres = 'Yandri Guevará'
        self.cliente_existente.save(update_fields=['nombres'])
        Cliente.objects.create(
            cedula='0927827281919',
            nombres='Randy Rodriguez',
            whatsapp='90939202',
            correo='photogamer2016pg@gmail.com',
            sector='norte',
        )

        response = self.client.get(reverse('econotec:cliente_lista'), {'q': 'guevara'})

        self.assertEqual(response.context['total'], 1)
        self.assertContains(response, 'Yandri Guevará')
        self.assertNotContains(response, 'Randy Rodriguez')

    def test_busqueda_lista_equipos_ignora_tildes_y_mayusculas(self):
        self.cliente_existente.nombres = 'Yandri Guevará'
        self.cliente_existente.save(update_fields=['nombres'])
        ingreso = self.crear_ingreso_reparacion()

        response = self.client.get(
            reverse('econotec:ingreso_lista'),
            {'q': 'guevara', 'sede': 'todas'},
        )

        self.assertEqual(response.context['total'], 1)
        self.assertContains(response, ingreso.codigo_equipo)
        self.assertContains(response, 'Yandri Guevará')

    def test_lista_equipos_paginala_a_cinco_y_conserva_filtros(self):
        self.activar_sede_guayaquil()
        for n in range(12):
            self.crear_ingreso_reparacion(
                marca=f'Marca {n}',
                modelo_serie=f'Modelo paginado {n}',
            )

        response = self.client.get(reverse('econotec:ingreso_lista'), {'sede': 'todas'})
        ingresos = list(response.context['ingresos'])

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['total'], 12)
        self.assertEqual(len(ingresos), 5)
        self.assertEqual(response.context['page_obj'].paginator.per_page, 5)
        self.assertContains(response, 'Página 1 de 3')
        self.assertContains(response, 'Mostrando máximo 5 por página')
        self.assertContains(response, 'sede=todas&pagina=2')

        response = self.client.get(
            reverse('econotec:ingreso_lista'),
            {'sede': 'todas', 'pagina': 2},
        )

        self.assertEqual(response.context['total'], 12)
        self.assertEqual(len(response.context['ingresos']), 5)
        self.assertContains(response, 'Página 2 de 3')

    def test_listados_ventas_y_pagos_paginalan_a_cinco(self):
        for n in range(12):
            self.crear_venta_producto(
                problema_reportado=f'Venta completa {n}',
                valor_acordado=Decimal('20.00'),
                abono_anticipo=Decimal('20.00'),
                anticipo_metodo='efectivo',
            )
            self.crear_venta_producto(
                problema_reportado=f'Venta parcial {n}',
                valor_acordado=Decimal('50.00'),
                abono_anticipo=Decimal('10.00'),
                anticipo_metodo='efectivo',
            )
            self.crear_ingreso_reparacion(
                modelo_serie=f'Reparación pago {n}',
            )

        escenarios = [
            (reverse('econotec:venta_lista'), 'total', 12),
            (reverse('econotec:venta_lista_parciales'), 'total', 12),
            (reverse('econotec:pagos_lista'), 'total_count', 12),
            (reverse('econotec:pagos_ventas_completos'), 'total_count', 12),
            (reverse('econotec:pagos_ventas_parciales'), 'total_count', 12),
        ]
        for url, total_key, total_esperado in escenarios:
            with self.subTest(url=url):
                response = self.client.get(url)

                self.assertEqual(response.status_code, 200)
                self.assertEqual(response.context[total_key], total_esperado)
                self.assertEqual(len(response.context['ingresos']), 5)
                self.assertEqual(response.context['page_obj'].paginator.per_page, 5)
                self.assertContains(response, 'Página 1 de 3')
                self.assertContains(response, 'Mostrando máximo 5 por página')

    def test_listados_salidas_clientes_facturas_y_auditoria_paginalan_a_cinco(self):
        for n in range(12):
            Cliente.objects.create(
                cedula=f'09{n:08d}',
                nombres=f'Cliente paginado {n}',
                whatsapp=f'099000{n:04d}',
                correo=f'cliente{n}@example.com',
                sector='norte',
            )
            ingreso_salida = self.crear_ingreso_reparacion(
                estado='entregado',
                modelo_serie=f'Salida paginada {n}',
            )
            SalidaEquipo.objects.create(
                ingreso=ingreso_salida,
                fecha_salida=date(2026, 7, 9),
                estado_reparacion='retirado',
                cliente_recibe_conforme='si',
                valor_final_cobrado=Decimal('25.00'),
                metodo_pago_final='efectivo',
                factura_realizada='no',
                registrado_por=self.usuario,
            )
            ingreso_factura = self.crear_ingreso_reparacion(
                estado='entregado',
                modelo_serie=f'Factura paginada {n}',
            )
            SalidaEquipo.objects.create(
                ingreso=ingreso_factura,
                fecha_salida=date(2026, 7, 10),
                estado_reparacion='retirado',
                cliente_recibe_conforme='si',
                valor_final_cobrado=Decimal('30.00'),
                metodo_pago_final='efectivo',
                factura_realizada='si',
                factura_nombres=f'Cliente facturado {n}',
                factura_cedula=f'10{n:08d}',
                factura_correo=f'factura{n}@example.com',
                registrado_por=self.usuario,
            )
            ingreso_pago = self.crear_ingreso_reparacion(
                modelo_serie=f'Pago auditoría {n}',
            )
            Abono.objects.create(
                ingreso=ingreso_pago,
                fecha=date(2026, 7, 11),
                monto=Decimal('5.00'),
                metodo='efectivo',
                registrado_por=self.usuario,
            )

        response_salidas = self.client.get(reverse('econotec:salida_lista'))
        self.assertEqual(response_salidas.status_code, 200)
        self.assertEqual(response_salidas.context['total'], 24)
        self.assertEqual(len(response_salidas.context['salidas']), 5)
        self.assertEqual(response_salidas.context['page_obj'].paginator.per_page, 5)
        self.assertContains(response_salidas, 'Página 1 de 5')

        response_clientes = self.client.get(
            reverse('econotec:cliente_lista'),
            {'q': 'Cliente paginado'},
        )
        self.assertEqual(response_clientes.status_code, 200)
        self.assertEqual(response_clientes.context['total'], 12)
        self.assertEqual(len(response_clientes.context['clientes']), 5)
        self.assertContains(response_clientes, 'q=Cliente+paginado&pagina=2')

        self.client.force_login(self.admin)
        response_facturas = self.client.get(
            reverse('econotec:salida_facturas_lista'),
            {'ano': '2026', 'mes': '7'},
        )
        self.assertEqual(response_facturas.status_code, 200)
        self.assertEqual(response_facturas.context['total'], 12)
        self.assertEqual(len(response_facturas.context['salidas']), 5)
        self.assertContains(response_facturas, 'Página 1 de 3')

        response_auditoria = self.client.get(reverse('econotec:control_registro'))
        self.assertEqual(response_auditoria.status_code, 200)
        self.assertEqual(len(response_auditoria.context['equipos']), 5)
        self.assertEqual(len(response_auditoria.context['abonos']), 5)
        self.assertEqual(response_auditoria.context['equipos_page_obj'].paginator.per_page, 5)
        self.assertEqual(response_auditoria.context['abonos_page_obj'].paginator.per_page, 5)
        self.assertContains(response_auditoria, 'pagina_equipos=2#panel-equipos')
        self.assertContains(response_auditoria, 'pagina_pagos=2#panel-pagos')

    def test_paginacion_se_muestra_con_menos_de_cinco_y_sin_resultados(self):
        ingreso = self.crear_ingreso_reparacion(estado='entregado')
        SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='retirado',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('25.00'),
            metodo_pago_final='efectivo',
            factura_realizada='no',
            registrado_por=self.usuario,
        )

        response_salidas = self.client.get(reverse('econotec:salida_lista'))

        self.assertEqual(response_salidas.status_code, 200)
        self.assertEqual(len(response_salidas.context['salidas']), 1)
        self.assertContains(response_salidas, 'Página 1 de 1')
        self.assertContains(response_salidas, 'Mostrando máximo 5 por página')
        self.assertContains(response_salidas, 'list-pagination-disabled')
        self.assertNotContains(response_salidas, 'pagina=2')

        self.client.force_login(self.admin)
        response_facturas = self.client.get(
            reverse('econotec:salida_facturas_lista'),
            {'ano': '2026', 'mes': '7', 'q': 'sin coincidencias'},
        )

        self.assertEqual(response_facturas.status_code, 200)
        self.assertEqual(response_facturas.context['total'], 0)
        self.assertContains(response_facturas, 'No hay facturas realizadas para este filtro')
        self.assertContains(response_facturas, 'Página 1 de 1')
        self.assertContains(response_facturas, 'Mostrando máximo 5 por página')
        self.assertContains(response_facturas, 'list-pagination-disabled')

    def test_admin_ventas_inventario_tabs_paginalan_a_cinco(self):
        self.client.force_login(self.admin)
        producto_base = InventarioItem.objects.create(
            sede='guayaquil',
            categoria='impresora',
            tipo='impresora-laser',
            producto='Producto base admin',
            marca='Epson',
            modelo='Base',
            estado='disponible',
            cantidad=20,
            costo=Decimal('5.00'),
            ubicacion='guayaquil_norte',
            registrado_por=self.usuario,
        )
        for n in range(12):
            venta = self.crear_venta_producto(
                fecha_ingreso=date(2026, 7, 25),
                problema_reportado=f'Venta admin paginada {n}',
                valor_acordado=Decimal('20.00'),
                abono_anticipo=Decimal('20.00'),
                anticipo_metodo='efectivo',
            )
            VentaInventarioItem.objects.create(
                venta=venta,
                inventario_item=producto_base,
                cantidad=1,
            )
            BitacoraTecnico.objects.create(
                user=self.usuario,
                usuario_nombre='Yandri',
                momento=datetime(
                    2026, 7, 25, 10, n,
                    tzinfo=ZoneInfo('America/Guayaquil'),
                ),
                tipo='venta_producto',
                texto=f'Venta admin paginada {n}',
                codigo=venta.codigo_equipo,
                ingreso=venta,
            )
            InventarioItem.objects.create(
                sede='guayaquil',
                categoria='computadora',
                tipo='pc',
                producto=f'Inventario admin paginado {n}',
                marca='Dell',
                modelo=f'M{n}',
                estado='disponible',
                cantidad=3,
                costo=Decimal('10.00'),
                ubicacion='guayaquil_norte',
                registrado_por=self.usuario,
            )

        for tab in ('ventas', 'actividad', 'movimientos', 'inventario'):
            with self.subTest(tab=tab):
                response = self.client.get(
                    reverse('econotec:admin_ventas_inventario'),
                    {'ano': '2026', 'mes': '7', 'tab': tab},
                )

                self.assertEqual(response.status_code, 200)
                self.assertEqual(len(response.context['page_obj'].object_list), 5)
                self.assertEqual(response.context['page_obj'].paginator.per_page, 5)
                self.assertContains(response, 'Mostrando máximo 5 por página')

    def test_lista_equipos_filtra_por_firma_cliente(self):
        self.activar_sede_guayaquil()
        ingreso_con_firma = self.crear_ingreso_reparacion(
            marca='Epson',
            modelo_serie='L3250 con firma',
            firma_cliente=True,
            firma_cliente_imagen=self.FIRMA_PNG_DATA_URI,
        )
        ingreso_sin_firma = self.crear_ingreso_reparacion(
            marca='HP',
            modelo_serie='Elitebook sin firma',
            firma_cliente=False,
            firma_cliente_imagen='',
        )

        response_con_firma = self.client.get(
            reverse('econotec:ingreso_lista'),
            {'firma': 'con_firma'},
        )
        response_sin_firma = self.client.get(
            reverse('econotec:ingreso_lista'),
            {'firma': 'sin_firma'},
        )

        self.assertEqual(response_con_firma.context['total'], 1)
        self.assertContains(response_con_firma, ingreso_con_firma.codigo_equipo)
        self.assertNotContains(response_con_firma, ingreso_sin_firma.codigo_equipo)
        self.assertContains(response_con_firma, 'value="con_firma" selected')

        self.assertEqual(response_sin_firma.context['total'], 1)
        self.assertContains(response_sin_firma, ingreso_sin_firma.codigo_equipo)
        self.assertNotContains(response_sin_firma, ingreso_con_firma.codigo_equipo)
        self.assertContains(response_sin_firma, 'value="sin_firma" selected')

    def test_lista_equipos_filtra_por_rango_fecha_ingreso(self):
        ingreso_julio = self.crear_ingreso_reparacion(
            fecha_ingreso=date(2026, 7, 20),
            marca='Epson',
            modelo_serie='L3250 rango',
        )
        ingreso_junio = self.crear_ingreso_reparacion(
            fecha_ingreso=date(2026, 6, 25),
            marca='HP',
            modelo_serie='Elitebook fuera rango',
        )

        response = self.client.get(reverse('econotec:ingreso_lista'), {
            'fecha_desde': '2026-07-01',
            'fecha_hasta': '2026-07-31',
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['total'], 1)
        self.assertContains(response, ingreso_julio.codigo_equipo)
        self.assertNotContains(response, ingreso_junio.codigo_equipo)
        self.assertContains(response, 'Fecha ingreso: 01/07/2026 - 31/07/2026')

    def test_busqueda_lista_salidas_ignora_tildes_y_mayusculas(self):
        self.cliente_existente.nombres = 'Yandri Guevará'
        self.cliente_existente.save(update_fields=['nombres'])
        ingreso = self.crear_ingreso_reparacion()
        SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='pendiente_retiro',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            tecnico_reparo=self.usuario,
            registrado_por=self.usuario,
        )

        response = self.client.get(reverse('econotec:salida_lista'), {'q': 'GUEVARA'})

        self.assertEqual(response.context['total'], 1)
        self.assertContains(response, ingreso.codigo_equipo)
        self.assertContains(response, 'Yandri Guevará')

    def test_lista_salidas_filtra_por_rango_fecha_salida(self):
        ingreso_julio = self.crear_ingreso_reparacion(estado='entregado')
        ingreso_junio = self.crear_ingreso_reparacion(estado='entregado')
        SalidaEquipo.objects.create(
            ingreso=ingreso_julio,
            fecha_salida=date(2026, 7, 20),
            estado_reparacion='pendiente_retiro',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso_junio,
            fecha_salida=date(2026, 6, 25),
            estado_reparacion='pendiente_retiro',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )

        response = self.client.get(reverse('econotec:salida_lista'), {
            'fecha_desde': '2026-07-01',
            'fecha_hasta': '2026-07-31',
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['total'], 1)
        self.assertContains(response, ingreso_julio.codigo_equipo)
        self.assertNotContains(response, ingreso_junio.codigo_equipo)
        self.assertContains(response, 'Fecha de finalización: 01/07/2026 - 31/07/2026')

    def test_estado_visual_muestra_pendiente_retiro_si_salida_esta_pendiente(self):
        ingreso = self.crear_ingreso_reparacion(
            estado='entregado',
            subestado_entregado='con_solucion',
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='pendiente_retiro',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        ingreso.refresh_from_db()

        self.assertEqual(ingreso.estado_visual_key, 'pendiente_retiro')
        self.assertEqual(ingreso.estado_visual_display, 'Pendiente de retiro')
        self.assertEqual(ingreso.subestado_visual_display, 'Reparado - pendiente de retiro')

        response = self.client.get(reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk}))

        self.assertContains(response, 'Pendiente de retiro')
        self.assertContains(response, 'Reparado - pendiente de retiro')
        self.assertNotContains(response, 'Listo para entrega')

    def test_alerta_diagnostico_excluye_ingreso_con_salida_registrada(self):
        ingreso = self.crear_ingreso_reparacion(
            estado='ingresado',
            subestado_reparacion='',
            fecha_ingreso=date.today() - timedelta(days=10),
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date.today() - timedelta(days=6),
            estado_reparacion='pendiente_retiro',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        IngresoEquipo.objects.filter(pk=ingreso.pk).update(
            estado='ingresado',
            subestado_entregado='',
        )
        ingreso.refresh_from_db()

        self.assertEqual(ingreso.estado_visual_display, 'Pendiente de retiro')
        self.assertNotIn(ingreso, list(equipos_demorados_qs(usuario=None)))

    def test_editar_ingreso_con_salida_muestra_estado_bloqueado(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('100.00'),
            diagnostico_inmediato='si',
            valor_diagnostico=Decimal('10.00'),
            diagnostico_metodo='mixto',
            diagnostico_monto_1=Decimal('5.00'),
            diagnostico_metodo_1='transferencia',
            diagnostico_banco_1='pichincha',
            diagnostico_monto_2=Decimal('5.00'),
            diagnostico_metodo_2='efectivo',
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='pendiente_retiro',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )

        response = self.client.get(
            reverse('econotec:ingreso_editar', kwargs={'pk': ingreso.pk})
        )

        self.assertContains(response, 'Pendiente de retiro')
        self.assertContains(response, 'Equipo finalizado')
        self.assertContains(response, 'El diagnóstico inmediato y su método de pago quedan bloqueados')
        self.assertContains(response, 'El valor acordado de este ingreso queda bloqueado')
        self.assertContains(response, '100.00')
        self.assertContains(response, '10.00')
        self.assertContains(response, 'disabled')
        self.assertContains(response, 'value="entregado"')
        self.assertNotContains(response, 'Ingresado / En diagnóstico')

    def test_editar_ingreso_con_salida_ignora_estado_posteado(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('100.00'),
            diagnostico_inmediato='si',
            valor_diagnostico=Decimal('10.00'),
            diagnostico_metodo='mixto',
            diagnostico_monto_1=Decimal('5.00'),
            diagnostico_metodo_1='transferencia',
            diagnostico_banco_1='pichincha',
            diagnostico_monto_2=Decimal('5.00'),
            diagnostico_metodo_2='efectivo',
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='pendiente_retiro',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        IngresoEquipo.objects.filter(pk=ingreso.pk).update(
            estado='ingresado',
            subestado_entregado='',
        )
        ingreso.refresh_from_db()

        response = self.client.post(
            reverse('econotec:ingreso_editar', kwargs={'pk': ingreso.pk}),
            self.ingreso_edit_post_data(
                ingreso,
                **{
                    'ing-estado': 'ingresado',
                    'ing-subestado_entregado': '',
                    'ing-valor_acordado': '999.00',
                    'ing-valor_acordado_estado': 'si',
                    'ing-diagnostico_inmediato': 'no',
                    'ing-valor_diagnostico': '99.00',
                    'ing-diagnostico_metodo': 'efectivo',
                    'ing-diagnostico_monto_1': '',
                    'ing-diagnostico_metodo_1': '',
                    'ing-diagnostico_banco_1': '',
                    'ing-diagnostico_monto_2': '',
                    'ing-diagnostico_metodo_2': '',
                    'ing-diagnostico_banco_2': '',
                    'ing-modelo_serie': 'Elitebook actualizado',
                },
            ),
        )

        self.assertRedirects(
            response,
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk})
        )
        ingreso.refresh_from_db()
        self.assertEqual(ingreso.estado, 'entregado')
        self.assertEqual(ingreso.subestado_entregado, 'pendiente_retiro')
        self.assertEqual(ingreso.valor_acordado, Decimal('100.00'))
        self.assertEqual(ingreso.diagnostico_inmediato, 'si')
        self.assertEqual(ingreso.valor_diagnostico, Decimal('10.00'))
        self.assertEqual(ingreso.diagnostico_metodo, 'mixto')
        self.assertEqual(ingreso.diagnostico_monto_1, Decimal('5.00'))
        self.assertEqual(ingreso.diagnostico_metodo_1, 'transferencia')
        self.assertEqual(ingreso.diagnostico_banco_1, 'pichincha')
        self.assertEqual(ingreso.diagnostico_monto_2, Decimal('5.00'))
        self.assertEqual(ingreso.diagnostico_metodo_2, 'efectivo')
        self.assertEqual(ingreso.modelo_serie, 'Elitebook actualizado')

    def test_editar_ingreso_muestra_resultados_negativos_en_estado_abierto(self):
        escenarios = (
            ('ingresado', ''),
            ('en_reparacion', 'en_reparacion'),
            ('en_reparacion', 'espera_repuesto'),
            ('en_reparacion', 'espera_cliente'),
        )
        for estado, subestado in escenarios:
            with self.subTest(estado=estado, subestado=subestado):
                ingreso = self.crear_ingreso_reparacion(
                    estado=estado,
                    subestado_reparacion=subestado,
                    subestado_entregado='',
                )

                response = self.client.get(
                    reverse('econotec:ingreso_editar', kwargs={'pk': ingreso.pk})
                )

                opciones = dict(response.context['ing_form'].fields['estado'].choices)
                self.assertNotIn('entregado', opciones)
                self.assertEqual(
                    opciones[IngresoEquipoForm.ESTADO_NO_QUISO_REPARAR],
                    'Cliente no quiso reparar',
                )
                self.assertEqual(
                    opciones[IngresoEquipoForm.ESTADO_NO_SE_PUDO_REPARAR],
                    'No se pudo reparar',
                )
                self.assertContains(response, 'id="finalizacion-rapida"')

    def test_editar_ingreso_advierte_pagos_antes_de_resultado_negativo(self):
        escenarios = (
            {
                'nombre': 'anticipo',
                'abono_anticipo': Decimal('20.00'),
                'diagnostico_inmediato': 'no',
                'valor_diagnostico': Decimal('0.00'),
                'texto': 'Abonos / anticipos',
                'total': '20.00',
            },
            {
                'nombre': 'diagnostico',
                'abono_anticipo': Decimal('0.00'),
                'diagnostico_inmediato': 'si',
                'valor_diagnostico': Decimal('15.00'),
                'texto': 'Diagnóstico inmediato',
                'total': '15.00',
            },
        )
        for escenario in escenarios:
            with self.subTest(nombre=escenario['nombre']):
                ingreso = self.crear_ingreso_reparacion(
                    abono_anticipo=escenario['abono_anticipo'],
                    diagnostico_inmediato=escenario['diagnostico_inmediato'],
                    valor_diagnostico=escenario['valor_diagnostico'],
                )

                response = self.client.get(
                    reverse('econotec:ingreso_editar', kwargs={'pk': ingreso.pk})
                )

                self.assertContains(response, 'id="confirmacion-estado-modal"')
                self.assertContains(response, escenario['texto'])
                self.assertContains(response, f'${escenario["total"]}')
                self.assertContains(response, 'No estoy de acuerdo')
                self.assertContains(response, 'Sí, estoy de acuerdo')

    def test_editar_ingreso_solo_valor_acordado_no_muestra_advertencia(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('80.00'),
            abono_anticipo=Decimal('0.00'),
            diagnostico_inmediato='no',
            valor_diagnostico=Decimal('0.00'),
        )

        response = self.client.get(
            reverse('econotec:ingreso_editar', kwargs={'pk': ingreso.pk})
        )

        self.assertNotContains(response, 'id="confirmacion-estado-modal"')

    def test_editar_ingreso_bloquea_resultado_negativo_pagado_sin_confirmacion(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('40.00'),
            abono_anticipo=Decimal('20.00'),
            estado='en_reparacion',
            subestado_reparacion='espera_cliente',
            subestado_entregado='',
        )
        data = self.ingreso_edit_post_data(
            ingreso,
            **{
                'ing-estado': IngresoEquipoForm.ESTADO_NO_SE_PUDO_REPARAR,
                'ing-subestado_reparacion': '',
                'ing-subestado_entregado': '',
                'ing-valor_acordado_estado': 'no',
                'ing-valor_acordado': '',
                'ing-abono_anticipo': '0.00',
            },
        )
        data.update(self.salida_rapida_post_data(
            estado_reparacion='no_reparable',
        ))

        response = self.client.post(
            reverse('econotec:ingreso_editar', kwargs={'pk': ingreso.pk}),
            data,
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('estado', response.context['ing_form'].errors)
        self.assertFalse(SalidaEquipo.objects.filter(ingreso=ingreso).exists())
        ingreso.refresh_from_db()
        self.assertEqual(ingreso.estado, 'en_reparacion')
        self.assertEqual(ingreso.subestado_reparacion, 'espera_cliente')
        self.assertEqual(ingreso.abono_anticipo, Decimal('20.00'))

    def test_editar_ingreso_acepta_resultado_negativo_pagado_confirmado(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('40.00'),
            abono_anticipo=Decimal('20.00'),
            estado='en_reparacion',
            subestado_reparacion='espera_cliente',
            subestado_entregado='',
        )
        data = self.ingreso_edit_post_data(
            ingreso,
            **{
                'confirmar_finalizacion_con_pago': '1',
                'ing-estado': IngresoEquipoForm.ESTADO_NO_QUISO_REPARAR,
                'ing-subestado_reparacion': '',
                'ing-subestado_entregado': '',
                'ing-valor_acordado_estado': 'no',
                'ing-valor_acordado': '',
                'ing-abono_anticipo': '0.00',
            },
        )
        data.update(self.salida_rapida_post_data(
            estado_reparacion='cliente_no_acepta',
        ))

        response = self.client.post(
            reverse('econotec:ingreso_editar', kwargs={'pk': ingreso.pk}),
            data,
        )

        salida = SalidaEquipo.objects.get(ingreso=ingreso)
        self.assertRedirects(
            response,
            reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}),
        )
        ingreso.refresh_from_db()
        self.assertEqual(ingreso.estado, 'entregado')
        self.assertEqual(ingreso.subestado_entregado, 'no_quiso_reparar')

    def test_editar_ingreso_finaliza_negativo_con_cobro_adicional(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('40.00'),
            abono_anticipo=Decimal('0.00'),
            estado='en_reparacion',
            subestado_reparacion='espera_cliente',
            subestado_entregado='',
        )
        data = self.ingreso_edit_post_data(
            ingreso,
            **{
                'ing-estado': IngresoEquipoForm.ESTADO_NO_QUISO_REPARAR,
                'ing-subestado_reparacion': '',
                'ing-subestado_entregado': '',
                'ing-valor_acordado_estado': 'no',
                'ing-valor_acordado': '',
                'ing-abono_anticipo': '0.00',
            },
        )
        data.update(self.salida_rapida_post_data(
            estado_reparacion='cliente_no_acepta',
            aplica_valor_acordado_adicional='si',
            valor_acordado_adicional='15.00',
            motivo_valor_acordado_adicional='Revisión avanzada autorizada.',
            valor_final_cobrado='5.00',
            metodo_pago_final='efectivo',
            asesora_notificacion=str(self.vendedor.pk),
        ))

        response = self.client.post(
            reverse('econotec:ingreso_editar', kwargs={'pk': ingreso.pk}),
            data,
        )

        salida = SalidaEquipo.objects.get(ingreso=ingreso)
        self.assertRedirects(
            response,
            reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}),
        )
        ingreso.refresh_from_db()
        self.assertEqual(ingreso.estado, 'entregado')
        self.assertEqual(ingreso.subestado_entregado, 'no_quiso_reparar')
        self.assertEqual(ingreso.valor_acordado, Decimal('0.00'))
        self.assertEqual(salida.estado_reparacion, 'cliente_no_acepta')
        self.assertEqual(salida.valor_acordado_adicional, Decimal('15.00'))
        self.assertEqual(salida.valor_final_cobrado, Decimal('5.00'))
        self.assertEqual(ingreso.diferencia, Decimal('10.00'))
        notificacion = NotificacionAsesora.objects.get(salida=salida)
        self.assertEqual(notificacion.tipo, NotificacionAsesora.TIPO_COBRO_ADICIONAL)
        self.assertEqual(notificacion.valor_acordado, Decimal('10.00'))

    def test_bitacora_edicion_ingreso_muestra_detalle_de_cambios(self):
        ingreso = self.crear_ingreso_reparacion(
            marca='Epson',
            modelo_serie='WF-2750',
            serie='23311',
            valor_acordado=Decimal('50.00'),
            estado='en_reparacion',
            subestado_reparacion='en_reparacion',
            subestado_entregado='',
        )

        response = self.client.post(
            reverse('econotec:ingreso_editar', kwargs={'pk': ingreso.pk}),
            self.ingreso_edit_post_data(
                ingreso,
                **{
                    'ing-marca': 'HP',
                    'ing-valor_acordado_estado': 'si',
                    'ing-estado': 'garantia',
                    'ing-subestado_reparacion': '',
                    'ing-subestado_entregado': '',
                    'ing-motivo_garantia': 'Retorno por falla adicional',
                },
            ),
        )

        self.assertRedirects(
            response,
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk}),
        )
        evento = BitacoraTecnico.objects.filter(
            user=self.usuario,
            ingreso=ingreso,
            tipo='ingreso_editado',
        ).latest('pk')

        self.assertIn('Datos actualizados en Laptop HP WF-2750', evento.texto)
        self.assertIn('Detalles:', evento.texto)
        self.assertIn('Estado del equipo: En reparación -> Garantía', evento.texto)
        self.assertIn('Detalle de reparación: En reparación -> —', evento.texto)
        self.assertIn('Motivo de garantía: — -> Retorno por falla adicional', evento.texto)
        self.assertIn('Marca: Epson -> HP', evento.texto)
        self.assertIn('Valor acordado: $50.00 -> $0.00', evento.texto)
        self.assertNotIn('Estado del valor acordado', evento.texto)
        self.assertNotIn('Firma del cliente: — -> No firma', evento.texto)

    def test_dashboard_modal_total_equipos_muestra_pendiente_retiro_visual(self):
        ingreso = self.crear_ingreso_reparacion(
            estado='entregado',
            subestado_entregado='con_solucion',
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='pendiente_retiro',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )

        response = self.client.get(
            reverse('econotec:dashboard_details', kwargs={'tipo': 'equipos_total'})
        )

        self.assertContains(response, ingreso.codigo_equipo)
        self.assertContains(response, 'Pendiente de retiro')
        self.assertNotContains(response, 'Entregado al cliente')

    def test_dashboard_total_equipos_no_mezcla_ventas_producto(self):
        ingreso = self.crear_ingreso_reparacion(fecha_ingreso=date.today())
        venta = self.crear_venta_producto()

        response = self.client.get(reverse('econotec:bienvenida'))

        self.assertEqual(response.context['stats']['total_ingresos'], 1)
        self.assertEqual(response.context['stats']['ingresos_mes'], 1)
        self.assertEqual(response.context['equipos_top'][0]['nombre'], 'Laptop')
        self.assertNotEqual(ingreso.codigo_equipo[0], venta.codigo_equipo[0])

    def test_bienvenida_muestra_resumen_ingresos_y_salidas_por_sede_para_asesor(self):
        ingreso_g_1 = self.crear_ingreso_reparacion(sede='guayaquil')
        ingreso_g_2 = self.crear_ingreso_reparacion(sede='guayaquil')
        ingreso_u_1 = self.crear_ingreso_reparacion(sede='quito')
        venta = self.crear_venta_producto()
        SalidaEquipo.objects.create(
            ingreso=ingreso_g_1,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='pendiente_retiro',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso_u_1,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='pendiente_retiro',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        self.client.force_login(self.vendedor)

        response = self.client.get(reverse('econotec:bienvenida'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Total de ingresos y equipos finalizados')
        self.assertContains(response, 'Sedes G / U')
        self.assertContains(response, '<details class="movement-summary"', html=False)
        self.assertContains(response, "abrirModalDashboard('ingresos_sede_guayaquil')")
        self.assertContains(response, "abrirModalDashboard('ingresos_sede_quito')")
        self.assertContains(response, "abrirModalDashboard('salidas_sede_guayaquil')")
        self.assertContains(response, "abrirModalDashboard('salidas_sede_quito')")
        self.assertEqual(response.context['resumen_movimientos']['ingresos']['guayaquil'], 2)
        self.assertEqual(response.context['resumen_movimientos']['ingresos']['quito'], 1)
        self.assertEqual(response.context['resumen_movimientos']['ingresos']['total'], 3)
        self.assertEqual(response.context['resumen_movimientos']['salidas']['guayaquil'], 1)
        self.assertEqual(response.context['resumen_movimientos']['salidas']['quito'], 1)
        self.assertEqual(response.context['resumen_movimientos']['salidas']['total'], 2)
        self.assertEqual(response.context['resumen_movimientos']['total_general'], 5)

        response_ingresos_g = self.client.get(
            reverse('econotec:dashboard_details', kwargs={'tipo': 'ingresos_sede_guayaquil'})
        )
        response_ingresos_u = self.client.get(
            reverse('econotec:dashboard_details', kwargs={'tipo': 'ingresos_sede_quito'})
        )
        response_salidas_g = self.client.get(
            reverse('econotec:dashboard_details', kwargs={'tipo': 'salidas_sede_guayaquil'})
        )
        response_salidas_u = self.client.get(
            reverse('econotec:dashboard_details', kwargs={'tipo': 'salidas_sede_quito'})
        )

        self.assertContains(response_ingresos_g, ingreso_g_1.codigo_equipo)
        self.assertContains(response_ingresos_g, ingreso_g_2.codigo_equipo)
        self.assertNotContains(response_ingresos_g, ingreso_u_1.codigo_equipo)
        self.assertNotContains(response_ingresos_g, venta.codigo_equipo)
        self.assertContains(response_ingresos_u, ingreso_u_1.codigo_equipo)
        self.assertNotContains(response_ingresos_u, ingreso_g_1.codigo_equipo)
        self.assertContains(response_salidas_g, ingreso_g_1.codigo_equipo)
        self.assertNotContains(response_salidas_g, ingreso_u_1.codigo_equipo)
        self.assertContains(response_salidas_u, ingreso_u_1.codigo_equipo)
        self.assertNotContains(response_salidas_u, ingreso_g_1.codigo_equipo)

    def test_dashboard_modal_total_equipos_excluye_ventas_producto(self):
        ingreso = self.crear_ingreso_reparacion(fecha_ingreso=date.today())
        venta = self.crear_venta_producto()

        response = self.client.get(
            reverse('econotec:dashboard_details', kwargs={'tipo': 'equipos_total'})
        )

        self.assertContains(response, ingreso.codigo_equipo)
        self.assertNotContains(response, venta.codigo_equipo)

    def test_modales_dashboard_incluyen_filtro_sede_excepto_clientes(self):
        ingreso_g = self.crear_ingreso_reparacion(
            sede='guayaquil',
            fecha_ingreso=date.today(),
            tipo_equipo='laptop',
        )
        ingreso_u = self.crear_ingreso_reparacion(
            sede='quito',
            fecha_ingreso=date.today(),
            tipo_equipo='impresora',
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso_u,
            fecha_salida=date.today(),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )

        for tipo in ('equipos_total', 'ingresos_mes', 'pendientes', 'salidas_mes'):
            with self.subTest(tipo=tipo):
                response = self.client.get(
                    reverse('econotec:dashboard_details', kwargs={'tipo': tipo})
                )

                self.assertEqual(response.status_code, 200)
                self.assertNotContains(response, 'id="dashModalSearch"')
                self.assertNotContains(response, 'id="dashModalEquipo"')
                self.assertNotContains(response, 'id="dashModalOrden"')
                self.assertNotContains(response, 'id="dashModalClear"')
                self.assertContains(response, 'name="dashModalSede"', count=3)
                self.assertContains(response, 'value="guayaquil"')
                self.assertContains(response, 'value="quito"')
                self.assertContains(response, 'data-dashboard-row')
                self.assertContains(response, 'data-label=')

        total = self.client.get(
            reverse('econotec:dashboard_details', kwargs={'tipo': 'equipos_total'})
        )
        self.assertContains(total, f'data-codigo="{ingreso_g.codigo_equipo}"')
        self.assertContains(total, 'data-sedes="guayaquil"')

        clientes = self.client.get(
            reverse('econotec:dashboard_details', kwargs={'tipo': 'clientes'})
        )
        self.assertContains(clientes, ingreso_g.codigo_equipo)
        self.assertContains(clientes, ingreso_u.codigo_equipo)
        self.assertContains(clientes, 'data-sedes="guayaquil|||quito"')
        self.assertNotContains(clientes, 'name="dashModalSede"')
        self.assertNotContains(clientes, 'value="guayaquil"')
        self.assertNotContains(clientes, 'value="quito"')

        inicio = self.client.get(reverse('econotec:bienvenida'))
        self.assertContains(inicio, 'function inicializarFiltrosModalDashboard()')
        self.assertContains(inicio, '@media (max-width: 900px)')
        self.assertContains(inicio, '@media (max-width: 640px)')
        self.assertContains(inicio, '@media (max-width: 420px)')

    def test_dashboard_modal_clientes_usa_consultas_acotadas(self):
        cliente_extra = Cliente.objects.create(
            cedula='0912345678',
            nombres='Cliente Extra',
            whatsapp='099000111',
            correo='extra@example.com',
        )
        self.crear_ingreso_reparacion(cliente=cliente_extra, sede='guayaquil')
        self.crear_ingreso_reparacion(cliente=cliente_extra, sede='quito')

        with CaptureQueriesContext(connection) as consultas:
            response = self.client.get(
                reverse('econotec:dashboard_details', kwargs={'tipo': 'clientes'})
            )

        self.assertEqual(response.status_code, 200)
        self.assertLessEqual(len(consultas.captured_queries), 10)
        self.assertContains(response, 'Cliente Extra')
        self.assertNotContains(response, 'name="dashModalSede"')

    def test_admin_dashboard_equipos_mes_excluye_ventas_producto(self):
        User = get_user_model()
        admin = User.objects.create_superuser(username='Admin', password='x')
        self.client.force_login(admin)
        hoy = date.today()
        self.crear_ingreso_reparacion(fecha_ingreso=hoy)
        self.crear_venta_producto(fecha_ingreso=hoy)

        response = self.client.get(
            reverse('econotec:admin_dashboard'),
            {'ano': str(hoy.year), 'mes': str(hoy.month)},
        )

        self.assertEqual(response.context['equipos_ingresados'], 1)

    def test_admin_dashboard_separa_tecnico_de_ingreso_y_tecnico_de_salida(self):
        User = get_user_model()
        tecnico_salida = User.objects.create_user(
            username='TecnicoSalidaAdmin',
            first_name='Tecnico',
            last_name='Salida',
        )
        tecnico_salida.groups.add(Group.objects.get(name='Tecnicos'))
        ingreso_asignado = self.crear_ingreso_reparacion(
            fecha_ingreso=date(2026, 7, 8),
            tecnico_encargado=self.usuario,
            modelo_serie='Ingreso asignado a Yandri',
            registrado_por=tecnico_salida,
        )
        ingreso_sin_asignar = self.crear_ingreso_reparacion(
            fecha_ingreso=date(2026, 7, 9),
            tecnico_encargado=None,
            modelo_serie='Ingreso pendiente de asignacion',
        )
        ingreso_salida_negativa = self.crear_ingreso_reparacion(
            fecha_ingreso=date(2026, 7, 10),
            tecnico_encargado=self.usuario,
            modelo_serie='Ingreso con salida negativa',
            registrado_por=tecnico_salida,
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso_asignado,
            fecha_salida=date(2026, 7, 20),
            fecha_retiro_real=date(2026, 7, 22),
            estado_reparacion='retirado',
            tecnico_reparo=tecnico_salida,
            valor_final_cobrado=Decimal('25.00'),
            metodo_pago_final='efectivo',
            bodegaje_dias_congelado=5,
            bodegaje_monto_congelado=Decimal('5.00'),
            bodegaje_aplicado_al_pago=True,
            registrado_por=self.usuario,
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso_salida_negativa,
            fecha_salida=date(2026, 7, 21),
            estado_reparacion='no_reparable',
            tecnico_reparo=tecnico_salida,
            valor_final_cobrado=Decimal('5.00'),
            metodo_pago_final='efectivo',
            registrado_por=self.usuario,
        )
        self.client.force_login(self.admin)

        response = self.client.get(
            reverse('econotec:admin_dashboard'),
            {'ano': '2026', 'mes': '7'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['ingresos_tecnicos_total'], 3)
        self.assertEqual(response.context['ingresos_tecnicos_asignados'], 2)
        self.assertEqual(response.context['ingresos_tecnicos_sin_asignar'], 1)
        self.assertEqual(response.context['ingresos_tecnicos_count'], 1)
        resumen_por_tecnico = {
            fila['tecnico_id']: fila
            for fila in response.context['ingresos_tecnicos_resumen']
        }
        self.assertEqual(resumen_por_tecnico[self.usuario.pk]['total'], 2)
        self.assertEqual(resumen_por_tecnico[self.usuario.pk]['con_salida'], 2)
        self.assertEqual(resumen_por_tecnico[None]['total'], 1)
        self.assertNotIn(tecnico_salida.pk, resumen_por_tecnico)
        registros = response.context['ingresos_tecnicos_registros']
        self.assertEqual({registro.pk for registro in registros}, {
            ingreso_asignado.pk,
            ingreso_sin_asignar.pk,
            ingreso_salida_negativa.pk,
        })
        salidas_por_tecnico = {
            fila['tecnico_id']: fila
            for fila in response.context['salidas_tecnicos_resumen']
        }
        self.assertNotIn(self.usuario.pk, salidas_por_tecnico)
        self.assertEqual(salidas_por_tecnico[tecnico_salida.pk]['total'], 2)
        self.assertEqual(salidas_por_tecnico[tecnico_salida.pk]['positivas'], 1)
        self.assertEqual(salidas_por_tecnico[tecnico_salida.pk]['negativas'], 1)
        self.assertEqual(salidas_por_tecnico[tecnico_salida.pk]['retirados'], 1)
        self.assertEqual(salidas_por_tecnico[tecnico_salida.pk]['pendientes'], 1)
        self.assertEqual(salidas_por_tecnico[tecnico_salida.pk]['efectividad'], 50.0)
        self.assertEqual(response.context['salidas_tecnicos_recaudado'], Decimal('30.00'))
        self.assertEqual(response.context['salidas_fisicas_admin_total'], 1)
        self.assertEqual(response.context['salidas_fisicas_admin_con_bodegaje'], 1)
        self.assertEqual(response.context['salidas_fisicas_admin_bodegaje_cobrado'], 1)
        self.assertEqual(response.context['salidas_fisicas_admin_bodegaje_total'], Decimal('5.00'))
        self.assertEqual(
            [salida.ingreso_id for salida in response.context['salidas_fisicas_admin_registros']],
            [ingreso_asignado.pk],
        )
        resumen_general = next(
            item for item in response.context['equipos_mes_resumen']
            if item['ingreso'].pk == ingreso_asignado.pk
        )
        self.assertEqual(resumen_general['tecnico_ingreso_nombre'], self.usuario.username)
        self.assertContains(response, 'Ingresos de equipos asignados a técnicos')
        self.assertContains(response, 'No representa quién hizo la reparación')
        self.assertContains(response, 'Técnico asignado al ingresar')
        self.assertContains(response, 'Técnico que deseas consultar')
        self.assertContains(response, 'Equipos finalizados por técnico que terminó la reparación')
        self.assertContains(response, 'Equipos finalizados del mes')
        self.assertContains(response, 'Resultados positivos')
        self.assertContains(response, 'Resultados negativos')
        self.assertContains(response, 'Salidas físicas confirmadas')
        self.assertContains(response, 'Ver lista completa →')
        self.assertContains(response, reverse('econotec:salida_retiros_lista'))
        self.assertContains(response, 'Fecha de salida física')
        self.assertContains(response, '✅ Salió de la oficina')
        self.assertNotContains(response, 'Salidas por técnico que terminó la reparación')

        response_tecnico_salida = self.client.get(
            reverse('econotec:admin_dashboard'),
            {
                'ano': '2026',
                'mes': '7',
                'tecnico_resumen': str(tecnico_salida.pk),
            },
        )
        self.assertEqual(response_tecnico_salida.context['tecnico_resumen_nombre'], 'Tecnico Salida')
        self.assertEqual(response_tecnico_salida.context['ingresos_tecnicos_total'], 0)
        self.assertEqual(response_tecnico_salida.context['salidas_tecnicos_total'], 2)
        self.assertEqual(response_tecnico_salida.context['salidas_tecnicos_positivas'], 1)
        self.assertEqual(response_tecnico_salida.context['salidas_tecnicos_negativas'], 1)

        response_tecnico_ingreso = self.client.get(
            reverse('econotec:admin_dashboard'),
            {
                'ano': '2026',
                'mes': '7',
                'tecnico_resumen': str(self.usuario.pk),
            },
        )
        self.assertEqual(response_tecnico_ingreso.context['ingresos_tecnicos_total'], 2)
        self.assertEqual(response_tecnico_ingreso.context['salidas_tecnicos_total'], 0)

        from .views_admin import _obtener_estadisticas_gamificacion
        perfiles = {
            fila['usuario']: fila
            for fila in _obtener_estadisticas_gamificacion()
        }
        self.assertEqual(perfiles[self.usuario.username]['ingresos'], 2)
        self.assertEqual(perfiles['Tecnico Salida']['ingresos'], 0)

        from openpyxl import load_workbook
        export_response = self.client.get(
            reverse('econotec:admin_equipos_mes_exportar'),
            {'ano': '2026', 'mes': '7'},
        )
        worksheet = load_workbook(BytesIO(export_response.content)).active
        self.assertEqual(worksheet.cell(row=2, column=12).value, 'Tecnico asignado al ingreso')
        fila_ingreso = next(
            row for row in range(3, worksheet.max_row + 1)
            if worksheet.cell(row=row, column=1).value == ingreso_asignado.codigo_equipo
        )
        self.assertEqual(worksheet.cell(row=fila_ingreso, column=12).value, self.usuario.username)

    def test_admin_dashboard_resumen_equipos_mes_separa_periodos(self):
        self.client.force_login(self.admin)
        ingreso_julio = self.crear_ingreso_reparacion(
            fecha_ingreso=date(2026, 7, 5),
            modelo_serie='Equipo ingresado en julio',
        )
        ingreso_junio_entregado_julio = self.crear_ingreso_reparacion(
            fecha_ingreso=date(2026, 6, 28),
            modelo_serie='Equipo junio entregado julio',
        )
        ingreso_agosto = self.crear_ingreso_reparacion(
            fecha_ingreso=date(2026, 8, 2),
            modelo_serie='Equipo ingresado en agosto',
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso_julio,
            fecha_salida=date(2026, 7, 20),
            estado_reparacion='retirado',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('15.00'),
            metodo_pago_final='efectivo',
            registrado_por=self.usuario,
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso_junio_entregado_julio,
            fecha_salida=date(2026, 7, 21),
            estado_reparacion='retirado',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('20.00'),
            metodo_pago_final='efectivo',
            registrado_por=self.usuario,
        )

        response_julio = self.client.get(
            reverse('econotec:admin_dashboard'),
            {'ano': '2026', 'mes': '7'},
        )
        resumen_julio = response_julio.context['equipos_mes_resumen']

        self.assertEqual(response_julio.context['equipos_ingresados'], 1)
        self.assertEqual(response_julio.context['equipos_entregados'], 2)
        self.assertEqual(len(resumen_julio), 2)
        self.assertContains(response_julio, 'Resumen de Equipos del Mes')
        self.assertContains(response_julio, '<details class="equipos-mes-details">', html=False)
        self.assertContains(response_julio, 'Guardar Excel')
        self.assertContains(response_julio, 'Borrar mes Julio')
        self.assertContains(response_julio, ingreso_julio.codigo_equipo)
        self.assertContains(response_julio, ingreso_junio_entregado_julio.codigo_equipo)
        self.assertNotContains(response_julio, ingreso_agosto.codigo_equipo)

        response_agosto = self.client.get(
            reverse('econotec:admin_dashboard'),
            {'ano': '2026', 'mes': '8'},
        )
        resumen_agosto = response_agosto.context['equipos_mes_resumen']

        self.assertEqual(response_agosto.context['equipos_ingresados'], 1)
        self.assertEqual(response_agosto.context['equipos_entregados'], 0)
        self.assertEqual(len(resumen_agosto), 1)
        self.assertContains(response_agosto, ingreso_agosto.codigo_equipo)
        self.assertNotContains(response_agosto, ingreso_julio.codigo_equipo)

    def test_admin_bitacoras_tecnicos_muestra_movimientos_por_tecnico(self):
        self.client.force_login(self.admin)
        zona_local = ZoneInfo('America/Guayaquil')
        BitacoraTecnico.objects.create(
            user=self.usuario,
            usuario_nombre='Yandri',
            momento=datetime(2026, 7, 10, 10, 15, tzinfo=zona_local),
            tipo='reporte',
            texto='Reporte técnico visible para admin.',
            codigo='G777',
        )
        BitacoraTecnico.objects.create(
            user=self.usuario,
            usuario_nombre='Yandri',
            momento=datetime(2026, 7, 11, 0, 0, tzinfo=zona_local),
            tipo='reporte',
            texto='Acción después de medianoche.',
            codigo='G778',
        )

        response = self.client.get(
            reverse('econotec:admin_bitacoras_tecnicos'),
            {'fecha': '2026-07-10'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Bitácoras de Técnicos')
        self.assertContains(response, self.usuario.username)
        self.assertContains(response, 'Reporte técnico visible para admin.')
        self.assertContains(response, 'G777')
        self.assertNotContains(response, 'Acción después de medianoche.')
        self.assertNotContains(response, 'G778')

    def test_admin_ventas_inventario_es_exclusivo_y_resume_operacion(self):
        producto = InventarioItem.objects.create(
            sede='guayaquil',
            categoria='computadora',
            tipo='pc',
            producto='Monitor administrativo',
            marca='Dell',
            modelo='P2422H',
            estado='disponible',
            cantidad=8,
            costo=Decimal('75.00'),
            ubicacion='guayaquil_norte',
            registrado_por=self.usuario,
        )
        venta = self.crear_venta_producto(
            fecha_ingreso=date(2026, 7, 25),
            problema_reportado='2 x Monitor administrativo',
            valor_acordado=Decimal('200.00'),
            abono_anticipo=Decimal('100.00'),
            anticipo_metodo='efectivo',
        )
        VentaInventarioItem.objects.create(
            venta=venta,
            inventario_item=producto,
            cantidad=2,
        )
        Abono.objects.create(
            ingreso=venta,
            fecha=date(2026, 7, 25),
            monto=Decimal('50.00'),
            metodo='transferencia',
            banco='pichincha',
            registrado_por=self.usuario,
        )
        BitacoraTecnico.objects.create(
            user=self.usuario,
            usuario_nombre='Yandri',
            momento=datetime(
                2026, 7, 25, 14, 30,
                tzinfo=ZoneInfo('America/Guayaquil'),
            ),
            tipo='venta_producto',
            texto='Venta administrativa registrada para auditoría.',
            codigo=venta.codigo_equipo,
            ingreso=venta,
        )

        response = self.client.get(
            reverse('econotec:admin_ventas_inventario'),
            {'ano': '2026', 'mes': '7'},
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse('econotec:bienvenida'))

        self.client.force_login(self.admin)
        response = self.client.get(
            reverse('econotec:admin_ventas_inventario'),
            {'ano': '2026', 'mes': '7'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['total_ventas'], 1)
        self.assertEqual(response.context['total_facturado'], Decimal('200.00'))
        self.assertEqual(response.context['total_cobrado'], Decimal('150.00'))
        self.assertEqual(response.context['total_saldo'], Decimal('50.00'))
        self.assertEqual(response.context['ventas_parciales'], 1)
        self.assertEqual(response.context['unidades_vendidas'], 2)
        self.assertEqual(response.context['inventario_valor'], Decimal('600.00'))
        self.assertContains(response, 'Administración de')
        self.assertContains(response, 'Ventas e Inventario')
        self.assertContains(response, venta.codigo_equipo)
        self.assertContains(response, 'Monitor administrativo')

        response_actividad = self.client.get(
            reverse('econotec:admin_ventas_inventario'),
            {'ano': '2026', 'mes': '7', 'tab': 'actividad'},
        )
        self.assertEqual(response_actividad.status_code, 200)
        self.assertContains(response_actividad, 'Venta administrativa registrada para auditoría.')

        response_movimientos = self.client.get(
            reverse('econotec:admin_ventas_inventario'),
            {'ano': '2026', 'mes': '7', 'tab': 'movimientos'},
        )
        self.assertEqual(response_movimientos.status_code, 200)
        self.assertContains(response_movimientos, 'Salidas de inventario')
        self.assertContains(response_movimientos, 'Monitor administrativo')

    def test_admin_ventas_inventario_filtra_ubicacion_segun_sede(self):
        norte = InventarioItem.objects.create(
            sede='guayaquil',
            categoria='computadora',
            tipo='pc',
            producto='Monitor Norte',
            marca='Dell',
            modelo='N24',
            estado='disponible',
            cantidad=4,
            costo=Decimal('80.00'),
            ubicacion='guayaquil_norte',
            registrado_por=self.usuario,
        )
        quito = InventarioItem.objects.create(
            sede='quito',
            categoria='computadora',
            tipo='pc',
            producto='Monitor Quito',
            marca='Dell',
            modelo='Q24',
            estado='disponible',
            cantidad=7,
            costo=Decimal('85.00'),
            ubicacion='quito',
            registrado_por=self.usuario,
        )
        self.client.force_login(self.admin)

        response = self.client.get(
            reverse('econotec:admin_ventas_inventario'),
            {
                'tab': 'inventario',
                'ano': '2026',
                'mes': '7',
                'inventario_sede': 'guayaquil',
                'inventario_ubicacion': 'guayaquil_norte',
            },
        )

        self.assertEqual(response.status_code, 200)
        ids = [item.pk for item in response.context['page_obj'].object_list]
        self.assertIn(norte.pk, ids)
        self.assertNotIn(quito.pk, ids)
        self.assertContains(response, 'Monitor Norte')
        self.assertNotContains(response, 'Monitor Quito')

    def test_admin_exporta_y_borra_resumen_equipos_mes_con_password(self):
        from openpyxl import load_workbook

        self.client.force_login(self.admin)
        ingreso_julio = self.crear_ingreso_reparacion(
            fecha_ingreso=date(2026, 7, 5),
            modelo_serie='Equipo julio para borrar',
        )
        ingreso_junio = self.crear_ingreso_reparacion(
            fecha_ingreso=date(2026, 6, 28),
            modelo_serie='Equipo junio con salida julio',
        )
        ingreso_agosto = self.crear_ingreso_reparacion(
            fecha_ingreso=date(2026, 8, 2),
            modelo_serie='Equipo agosto protegido',
        )
        salida_julio = SalidaEquipo.objects.create(
            ingreso=ingreso_julio,
            fecha_salida=date(2026, 7, 20),
            estado_reparacion='retirado',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('15.00'),
            metodo_pago_final='efectivo',
            registrado_por=self.usuario,
        )
        salida_junio_julio = SalidaEquipo.objects.create(
            ingreso=ingreso_junio,
            fecha_salida=date(2026, 7, 21),
            estado_reparacion='retirado',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('20.00'),
            metodo_pago_final='efectivo',
            registrado_por=self.usuario,
        )
        abono_julio = Abono.objects.create(
            ingreso=ingreso_junio,
            fecha=date(2026, 7, 10),
            monto=Decimal('5.00'),
            metodo='efectivo',
            registrado_por=self.usuario,
        )
        abono_agosto = Abono.objects.create(
            ingreso=ingreso_agosto,
            fecha=date(2026, 8, 3),
            monto=Decimal('8.00'),
            metodo='efectivo',
            registrado_por=self.usuario,
        )

        export_response = self.client.get(
            reverse('econotec:admin_equipos_mes_exportar'),
            {'ano': '2026', 'mes': '7'},
        )

        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(
            export_response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        workbook = load_workbook(BytesIO(export_response.content))
        worksheet = workbook.active
        self.assertEqual(worksheet['A1'].value, 'Resumen de equipos del mes - Julio 2026')
        self.assertEqual(worksheet['A2'].value, 'Codigo')
        codigos = [worksheet.cell(row=row, column=1).value for row in range(3, worksheet.max_row + 1)]
        self.assertIn(ingreso_julio.codigo_equipo, codigos)
        self.assertIn(ingreso_junio.codigo_equipo, codigos)

        wrong_response = self.client.post(
            reverse('econotec:admin_equipos_mes_borrar'),
            {'ano': '2026', 'mes': '7', 'admin_password': 'mal'},
        )

        self.assertEqual(wrong_response.status_code, 302)
        self.assertTrue(IngresoEquipo.objects.filter(pk=ingreso_julio.pk).exists())
        self.assertTrue(SalidaEquipo.objects.filter(pk=salida_julio.pk).exists())

        ok_response = self.client.post(
            reverse('econotec:admin_equipos_mes_borrar'),
            {'ano': '2026', 'mes': '7', 'admin_password': 'adminpass123'},
        )

        self.assertEqual(ok_response.status_code, 302)
        self.assertFalse(IngresoEquipo.objects.filter(pk=ingreso_julio.pk).exists())
        self.assertFalse(SalidaEquipo.objects.filter(pk=salida_julio.pk).exists())
        self.assertFalse(SalidaEquipo.objects.filter(pk=salida_junio_julio.pk).exists())
        self.assertFalse(Abono.objects.filter(pk=abono_julio.pk).exists())
        self.assertTrue(IngresoEquipo.objects.filter(pk=ingreso_junio.pk).exists())
        self.assertTrue(IngresoEquipo.objects.filter(pk=ingreso_agosto.pk).exists())
        self.assertTrue(Abono.objects.filter(pk=abono_agosto.pk).exists())

    def test_admin_dashboard_muestra_horarios_y_avisos_laborales(self):
        self.client.force_login(self.admin)
        HorarioTecnico.objects.create(
            tecnico=self.usuario,
            hora_inicio=time(8, 30),
            hora_fin=time(17, 30),
            ultima_notificacion_laboral=timezone.now(),
        )

        response = self.client.get(reverse('econotec:admin_dashboard'))

        self.assertContains(response, 'Horarios laborales de técnicos')
        self.assertContains(response, 'Yandri')
        self.assertContains(response, '08:30 - 17:30')
        self.assertContains(response, 'entró a su día laboral')
        self.assertEqual(len(response.context['avisos_laborales_hoy']), 1)

    def test_admin_puede_actualizar_horario_laboral_de_tecnico(self):
        self.client.force_login(self.admin)

        response = self.client.post(
            reverse('econotec:admin_horario_tecnico_guardar', kwargs={'user_id': self.usuario.pk}),
            {
                'activo': 'on',
                'lunes': 'on',
                'miercoles': 'on',
                'viernes': 'on',
                'hora_inicio': '08:15',
                'hora_fin': '17:45',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn('#horarios-tecnicos', response.url)
        horario = HorarioTecnico.objects.get(tecnico=self.usuario)
        self.assertTrue(horario.activo)
        self.assertTrue(horario.lunes)
        self.assertFalse(horario.martes)
        self.assertTrue(horario.miercoles)
        self.assertFalse(horario.jueves)
        self.assertTrue(horario.viernes)
        self.assertEqual(horario.hora_inicio.strftime('%H:%M'), '08:15')
        self.assertEqual(horario.hora_fin.strftime('%H:%M'), '17:45')

    def test_registrar_entrada_laboral_avisa_una_vez_por_dia(self):
        dia_lunes = date(2026, 7, 20)
        momento_lunes = datetime(2026, 7, 20, 10, 0, tzinfo=ZoneInfo('America/Guayaquil'))

        with patch('econotec.horarios.timezone.localdate', return_value=dia_lunes), \
             patch('econotec.horarios.timezone.now', return_value=momento_lunes):
            primer_aviso = registrar_entrada_laboral(self.usuario)
            segundo_aviso = registrar_entrada_laboral(self.usuario)

        self.assertIsNotNone(primer_aviso)
        self.assertIsNone(segundo_aviso)
        horario = HorarioTecnico.objects.get(tecnico=self.usuario)
        self.assertIsNotNone(horario.ultima_notificacion_laboral)

    def test_registrar_entrada_fuera_de_dia_laboral_avisa_en_admin(self):
        self.client.force_login(self.admin)
        dia_miercoles = date(2026, 7, 22)
        momento_miercoles = datetime(2026, 7, 22, 10, 0, tzinfo=ZoneInfo('America/Guayaquil'))
        HorarioTecnico.objects.create(
            tecnico=self.usuario,
            lunes=True,
            martes=False,
            miercoles=False,
            jueves=False,
            viernes=True,
        )

        with patch('econotec.horarios.timezone.localdate', return_value=dia_miercoles), \
             patch('econotec.horarios.timezone.now', return_value=momento_miercoles):
            primer_aviso = registrar_entrada_laboral(self.usuario)
            segundo_aviso = registrar_entrada_laboral(self.usuario)

        horario = HorarioTecnico.objects.get(tecnico=self.usuario)
        self.assertIsNotNone(primer_aviso)
        self.assertIsNone(segundo_aviso)
        self.assertIsNotNone(horario.ultima_notificacion_fuera_laboral)
        self.assertEqual(horario.ultima_notificacion_fuera_motivo, 'dia')

        with patch('econotec.views_admin.timezone.localdate', return_value=dia_miercoles), \
             patch('econotec.views_admin.timezone.now', return_value=momento_miercoles):
            response = self.client.get(reverse('econotec:admin_dashboard'))

        self.assertContains(response, 'entró fuera de su día laboral')
        self.assertEqual(len(response.context['avisos_fuera_laboral_hoy']), 1)

    def test_estado_visual_conserva_entregado_con_solucion_si_cliente_retiro(self):
        ingreso = self.crear_ingreso_reparacion(estado='entregado')
        SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='retirado',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        ingreso.refresh_from_db()

        self.assertEqual(ingreso.estado_visual_key, 'entregado')
        self.assertEqual(ingreso.estado_visual_display, 'Entregado al cliente')
        self.assertEqual(ingreso.subestado_visual_display, 'Con solución')

        response = self.client.get(reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk}))

        self.assertContains(response, 'Entregado al cliente')
        self.assertContains(response, 'Con solución')

    def test_estado_visual_sin_solucion_usa_color_rojo(self):
        ingreso = self.crear_ingreso_reparacion(
            estado='entregado',
            subestado_entregado='sin_solucion',
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='no_reparable',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        ingreso.refresh_from_db()

        self.assertEqual(ingreso.estado_visual_key, 'no_reparable')
        self.assertEqual(ingreso.estado_visual_display, 'Entregado al cliente')
        self.assertEqual(ingreso.subestado_visual_display, 'Sin solución')

        response = self.client.get(reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk}))

        self.assertContains(response, 'badge-no_reparable')
        self.assertContains(response, 'estado-no_reparable')
        self.assertContains(response, 'estado-subestado-no_reparable')

    def test_estado_visual_cliente_no_quiso_usa_color_amarillo(self):
        ingreso = self.crear_ingreso_reparacion(
            estado='entregado',
            subestado_entregado='no_quiso_reparar',
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='cliente_no_acepta',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        ingreso.refresh_from_db()

        self.assertEqual(ingreso.estado_visual_key, 'cliente_no_acepta')
        self.assertEqual(ingreso.estado_visual_display, 'Entregado al cliente')
        self.assertEqual(ingreso.subestado_visual_display, 'No quiso repararlo')

        response = self.client.get(reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk}))

        self.assertContains(response, 'badge-cliente_no_acepta')
        self.assertContains(response, 'estado-cliente_no_acepta')
        self.assertContains(response, 'estado-subestado-cliente_no_acepta')

    def test_detalle_ingreso_retirado_bloquea_enlace_de_edicion(self):
        ingreso = self.crear_ingreso_reparacion(estado='entregado')
        SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            fecha_retiro_real=date(2026, 7, 10),
            estado_reparacion='retirado',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        ingreso.refresh_from_db()
        edit_url = reverse('econotec:ingreso_editar', kwargs={'pk': ingreso.pk})

        response = self.client.get(reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk}))

        self.assertTrue(ingreso.retirado_por_cliente)
        self.assertContains(response, 'Ya este equipo fue retirado por el cliente')
        self.assertContains(response, 'Hoja de ingreso cerrada')
        self.assertNotContains(response, f'href="{edit_url}"')

    def test_editar_ingreso_retirado_redirige_y_no_guarda_cambios(self):
        ingreso = self.crear_ingreso_reparacion(estado='entregado')
        SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            fecha_retiro_real=date(2026, 7, 10),
            estado_reparacion='retirado',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        ingreso.refresh_from_db()
        detalle_url = reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk})
        editar_url = reverse('econotec:ingreso_editar', kwargs={'pk': ingreso.pk})

        response_get = self.client.get(editar_url)
        response_post = self.client.post(
            editar_url,
            self.ingreso_edit_post_data(
                ingreso,
                **{'ing-problema_reportado': 'Cambio no permitido'}
            ),
        )

        self.assertRedirects(response_get, detalle_url)
        self.assertRedirects(response_post, detalle_url)
        ingreso.refresh_from_db()
        self.assertNotEqual(ingreso.problema_reportado, 'Cambio no permitido')

    def test_salida_imprimir_muestra_datos_de_factura_si_fue_realizada(self):
        ingreso = self.crear_ingreso_reparacion(estado='entregado')
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='retirado',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('25.00'),
            metodo_pago_final='efectivo',
            factura_realizada='si',
            factura_nombres='Yandri Guevara',
            factura_cedula='1207342716',
            factura_correo='factura@example.com',
            registrado_por=self.usuario,
        )

        response = self.client.get(reverse('econotec:salida_imprimir', kwargs={'pk': salida.pk}))

        self.assertContains(response, 'FACTURA REALIZADA')
        self.assertContains(response, 'ACTA DE EQUIPO FINALIZADO')
        self.assertContains(response, 'RESULTADO FINAL DEL EQUIPO')
        self.assertContains(response, 'SALIÓ DE LA OFICINA')
        self.assertNotContains(response, 'ACTA DE SALIDA DE EQUIPO')
        self.assertNotContains(response, 'ESTADO DE LA SALIDA')
        self.assertContains(response, 'Nombres / Razón Social')
        self.assertContains(response, 'Yandri Guevara')
        self.assertContains(response, '1207342716')
        self.assertContains(response, 'factura@example.com')

        pdf_response = self.client.get(reverse('econotec:salida_pdf', kwargs={'pk': salida.pk}))
        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(pdf_response['Content-Type'], 'application/pdf')
        self.assertEqual(
            pdf_response['Content-Disposition'],
            f'attachment; filename="equipo_finalizado_{ingreso.codigo_equipo}.pdf"',
        )

    def test_salida_factura_imprimir_usa_plantilla_propia(self):
        ingreso = self.crear_ingreso_reparacion(
            estado='entregado',
            valor_acordado=Decimal('100.00'),
            abono_anticipo=Decimal('40.00'),
            marca='Sony',
            modelo_serie='Playstation 5',
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='retirado',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('60.00'),
            metodo_pago_final='efectivo',
            factura_realizada='si',
            factura_nombres='Yandri Guevara',
            factura_cedula='1207342716',
            factura_correo='factura@example.com',
            registrado_por=self.usuario,
        )

        response = self.client.get(
            reverse('econotec:salida_factura_imprimir', kwargs={'pk': salida.pk})
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Comprobante de Compra/Venta')
        self.assertContains(response, 'Fecha de emisión:')
        self.assertContains(response, 'Bienes/Servicios')
        self.assertContains(response, 'Servicio técnico y reparación')
        self.assertContains(response, 'Subtotal 0%:')
        self.assertContains(response, 'Total:')
        self.assertContains(response, '$100,00')
        self.assertContains(response, 'Pagado:')
        self.assertContains(response, 'Teléfonos:')
        self.assertContains(response, 'Firma del técnico')
        self.assertContains(response, 'Firma del cliente')
        self.assertNotContains(response, '098 075 8747')
        self.assertNotContains(response, 'ACTA DE SALIDA DE EQUIPO')
        self.assertNotContains(response, 'ESTADO DE LA SALIDA')

        pdf_response = self.client.get(
            reverse('econotec:salida_factura_pdf', kwargs={'pk': salida.pk})
        )
        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(pdf_response['Content-Type'], 'application/pdf')
        self.assertEqual(
            pdf_response['Content-Disposition'],
            f'attachment; filename="factura_{ingreso.codigo_equipo}.pdf"',
        )

    def test_salida_factura_imprimir_muestra_linea_firma_cliente_sin_imagen(self):
        ingreso = self.crear_ingreso_reparacion(
            estado='entregado',
            firma_cliente=True,
            firma_cliente_imagen=self.FIRMA_PNG_DATA_URI,
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='retirado',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('25.00'),
            metodo_pago_final='efectivo',
            factura_realizada='si',
            factura_nombres='Yandri Guevara',
            factura_cedula='1207342716',
            factura_correo='factura@example.com',
            registrado_por=self.usuario,
        )

        response = self.client.get(
            reverse('econotec:salida_factura_imprimir', kwargs={'pk': salida.pk})
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Firma del técnico')
        self.assertContains(response, 'Firma del cliente')
        self.assertNotContains(response, self.FIRMA_PNG_DATA_URI)

        with patch('econotec.views_print._draw_signature_image') as draw_signature:
            pdf_response = self.client.get(
                reverse('econotec:salida_factura_pdf', kwargs={'pk': salida.pk})
            )
        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(pdf_response['Content-Type'], 'application/pdf')
        draw_signature.assert_not_called()

    def test_salida_factura_imprimir_no_genera_documento_si_no_hay_factura(self):
        ingreso = self.crear_ingreso_reparacion(estado='entregado')
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='retirado',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('25.00'),
            metodo_pago_final='efectivo',
            factura_realizada='no',
            registrado_por=self.usuario,
        )

        response = self.client.get(
            reverse('econotec:salida_factura_imprimir', kwargs={'pk': salida.pk})
        )

        self.assertEqual(response.status_code, 404)

    def test_salida_imprimir_muestra_saldo_pagos_y_solo_firma_tecnico(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('35.00'),
            abono_anticipo=Decimal('10.00'),
            anticipo_metodo='transferencia',
            anticipo_banco='pichincha',
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )

        response = self.client.get(reverse('econotec:salida_imprimir', kwargs={'pk': salida.pk}))

        self.assertContains(response, 'Su equipo ya está listo para su retiro.')
        self.assertContains(response, 'Su valor pendiente es:')
        self.assertContains(response, '$25,00')
        self.assertContains(response, 'Anticipo / abono inicial')
        self.assertContains(response, 'Banco Pichincha')
        self.assertContains(response, 'FIRMA DEL TÉCNICO')
        self.assertNotContains(response, 'FIRMA DEL CLIENTE')
        self.assertNotContains(response, 'FACTURA REALIZADA')
        self.assertNotContains(response, 'No se registró factura para esta salida.')

    def test_salida_impresa_no_atribuye_reparacion_al_tecnico_del_ingreso(self):
        User = get_user_model()
        tecnico_ingreso = User.objects.create_user(
            username='TecnicoSoloIngresoPdf',
            password='test123',
            first_name='Responsable Ingreso',
        )
        ingreso = self.crear_ingreso_reparacion(
            tecnico_encargado=tecnico_ingreso,
            estado='entregado',
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='retirado',
            tecnico_reparo=None,
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )

        response = self.client.get(
            reverse('econotec:salida_imprimir', kwargs={'pk': salida.pk})
        )

        self.assertContains(response, '— Sin técnico registrado —', count=2)
        self.assertNotContains(response, 'Responsable Ingreso')

        with patch('econotec.views_print._draw_label_value') as draw_label:
            pdf_response = self.client.get(
                reverse('econotec:salida_pdf', kwargs={'pk': salida.pk})
            )

        self.assertEqual(pdf_response.status_code, 200)
        etiqueta_tecnico = next(
            call for call in draw_label.call_args_list
            if len(call.args) > 3 and call.args[3] == 'Técnico que reparó:'
        )
        self.assertEqual(etiqueta_tecnico.args[4], '— Sin técnico registrado —')

    def test_salida_imprimir_detalla_valor_acordado_adicional(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('20.00'),
            abono_anticipo=Decimal('20.00'),
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            aplica_valor_acordado_adicional='si',
            valor_acordado_adicional=Decimal('0.10'),
            motivo_valor_acordado_adicional='Repuesto adicional autorizado por el cliente.',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )

        response = self.client.get(
            reverse('econotec:salida_imprimir', kwargs={'pk': salida.pk})
        )

        self.assertContains(response, 'Detalle del valor acordado adicional')
        self.assertContains(response, 'Valor acordado original')
        self.assertContains(response, '$20,00')
        self.assertContains(response, '$0,10')
        self.assertContains(response, '$20,10')
        self.assertContains(response, 'Repuesto adicional autorizado por el cliente.')
        self.assertContains(response, 'Su valor pendiente es:')

        pdf_response = self.client.get(
            reverse('econotec:salida_pdf', kwargs={'pk': salida.pk})
        )
        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(pdf_response['Content-Type'], 'application/pdf')

    def test_salida_imprimir_detalla_pago_mixto(self):
        ingreso = self.crear_ingreso_reparacion(
            estado='entregado',
            valor_acordado=Decimal('30.00'),
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='retirado',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('30.00'),
            metodo_pago_final='mixto',
            monto_1=Decimal('12.50'),
            metodo_1='transferencia',
            banco_1='guayaquil',
            monto_2=Decimal('17.50'),
            metodo_2='tarjeta',
            numero_recibo='RECS-0099',
            registrado_por=self.usuario,
        )

        response = self.client.get(reverse('econotec:salida_imprimir', kwargs={'pk': salida.pk}))

        self.assertContains(response, 'Pago mixto (2 métodos)')
        self.assertContains(response, 'Parte 1: $ 12.50')
        self.assertContains(response, 'Banco Guayaquil')
        self.assertContains(response, 'Parte 2: $ 17.50')
        self.assertContains(response, 'Tarjeta de crédito/débito')
        self.assertContains(response, 'Recibo: RECS-0099')

        pdf_response = self.client.get(reverse('econotec:salida_pdf', kwargs={'pk': salida.pk}))
        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(pdf_response['Content-Type'], 'application/pdf')

    def test_marcar_retirada_conserva_resultado_negativo_y_registra_fecha(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=Decimal('0.00'))
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='no_reparable',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )

        lista_finalizados = self.client.get(reverse('econotec:salida_lista'))
        self.assertContains(lista_finalizados, ingreso.codigo_equipo)
        self.assertContains(lista_finalizados, 'Confirmar salida de la oficina')
        self.assertContains(lista_finalizados, 'El equipo ya salió · Clic aquí')
        self.assertContains(lista_finalizados, '🚚')
        self.assertContains(lista_finalizados, 'class="salidas-table"')
        self.assertContains(lista_finalizados, '⏳ Reparado — pendiente de retiro')
        self.assertContains(lista_finalizados, 'class="salidas-filter-form"')
        self.assertContains(lista_finalizados, 'Pago mixto (2 métodos)')
        self.assertContains(lista_finalizados, 'Guardar pago y confirmar salida')
        self.assertContains(lista_finalizados, 'acta PDF de salida')
        self.assertNotContains(lista_finalizados, 'id="salida-guide-heading"')

        response = self.client.post(reverse('econotec:salida_marcar_retirada', kwargs={'pk': salida.pk}))

        self.assertRedirects(response, reverse('econotec:salida_retiros_lista'))
        salida.refresh_from_db()
        ingreso.refresh_from_db()
        self.assertEqual(salida.estado_reparacion, 'no_reparable')
        self.assertEqual(salida.fecha_retiro_real, date.today())
        self.assertEqual(ingreso.subestado_entregado, 'sin_solucion')

        lista_finalizados = self.client.get(reverse('econotec:salida_lista'))
        lista_salidas_fisicas = self.client.get(reverse('econotec:salida_retiros_lista'))
        inicio = self.client.get(reverse('econotec:bienvenida'))

        self.assertNotContains(lista_finalizados, ingreso.codigo_equipo)
        self.assertContains(lista_salidas_fisicas, ingreso.codigo_equipo)
        self.assertTrue(lista_salidas_fisicas.context['lista_salidas_confirmadas'])
        self.assertContains(lista_salidas_fisicas, 'Salidas físicas')
        self.assertContains(lista_salidas_fisicas, '¿Cómo registrar la salida física de un equipo?')
        self.assertContains(lista_salidas_fisicas, 'Ir a Lista de equipos finalizados →')
        self.assertContains(lista_salidas_fisicas, 'Abre Lista de equipos finalizados')
        self.assertContains(lista_salidas_fisicas, 'Pulsa Confirmar salida de la oficina')
        self.assertContains(lista_salidas_fisicas, reverse('econotec:salida_lista'))
        self.assertContains(lista_salidas_fisicas, '— Estado de salida física —')
        self.assertContains(lista_salidas_fisicas, '✅ Salió de la oficina')
        self.assertNotContains(lista_salidas_fisicas, '⏳ Reparado — pendiente de retiro')
        self.assertEqual(
            lista_salidas_fisicas.context['estados'],
            [('retirado', '✅ Salió de la oficina')],
        )

        filtro_invalido = self.client.get(
            reverse('econotec:salida_retiros_lista'),
            {'estado': 'pendiente_retiro'},
        )
        self.assertEqual(filtro_invalido.context['estado_filtro'], '')
        self.assertEqual(filtro_invalido.context['total'], 1)
        self.assertContains(inicio, 'Salidas físicas confirmadas')
        self.assertContains(inicio, reverse('econotec:salida_retiros_lista'))
        self.assertEqual(inicio.context['stats']['salidas_fisicas_confirmadas'], 1)

        self.client.force_login(self.admin)
        response_deshacer = self.client.post(
            reverse('econotec:salida_deshacer_retiro', kwargs={'pk': salida.pk}),
        )
        self.assertRedirects(response_deshacer, reverse('econotec:salida_lista'))

        salida.refresh_from_db()
        self.assertIsNone(salida.fecha_retiro_real)
        self.assertContains(self.client.get(reverse('econotec:salida_lista')), ingreso.codigo_equipo)
        self.assertNotContains(
            self.client.get(reverse('econotec:salida_retiros_lista')),
            ingreso.codigo_equipo,
        )

    @override_settings(
        EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
        SALIDA_EMAIL_AUTOMATICO=True,
        EQUIPO_EMAIL_ADJUNTAR_PDF=True,
    )
    def test_salida_fisica_sin_bodegaje_envia_acta_actualizada(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=Decimal('0.00'))
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date.today(),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        mail.outbox.clear()

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                reverse('econotec:salida_marcar_retirada', kwargs={'pk': salida.pk}),
                {'aplicar_bodegaje': ''},
            )

        self.assertRedirects(response, reverse('econotec:salida_retiros_lista'))
        salida.refresh_from_db()
        self.assertTrue(salida.cliente_ya_retiro)
        self.assertEqual(len(mail.outbox), 1)
        correo = mail.outbox[0]
        self.assertIn('Salida de la oficina confirmada', correo.subject)
        html = next(
            alternativa[0]
            for alternativa in correo.alternatives
            if alternativa[1] == 'text/html'
        )
        self.assertIn('Tu equipo salió de la oficina', html)
        self.assertIn('Salida física confirmada', html)
        self.assertIn('Regla de bodegaje', html)
        self.assertEqual(
            correo.attachments[0][0],
            f'Acta_salida_oficina_{ingreso.codigo_equipo}.pdf',
        )
        self.assertTrue(correo.attachments[0][1].startswith(b'%PDF-'))

    @override_settings(
        EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
        SALIDA_EMAIL_AUTOMATICO=True,
        EQUIPO_EMAIL_ADJUNTAR_PDF=True,
    )
    def test_salida_fisica_cobra_bodegaje_mixto_y_envia_resultado(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=Decimal('0.00'))
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date.today() - timedelta(days=10),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        self.assertEqual(salida.calcular_bodegaje()['monto'], Decimal('6.00'))
        mail.outbox.clear()

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                reverse('econotec:salida_marcar_retirada', kwargs={'pk': salida.pk}),
                {
                    'aplicar_bodegaje': 'on',
                    'pago_bod_metodo': 'mixto',
                    'pago_bod_monto_1': '2.50',
                    'pago_bod_metodo_1': 'efectivo',
                    'pago_bod_banco_1': '',
                    'pago_bod_banco_otro_1': '',
                    'pago_bod_tarjeta_app_1': '',
                    'pago_bod_comprobante_url_1': '',
                    'pago_bod_monto_2': '3.50',
                    'pago_bod_metodo_2': 'transferencia',
                    'pago_bod_banco_2': 'guayaquil',
                    'pago_bod_banco_otro_2': '',
                    'pago_bod_tarjeta_app_2': '',
                    'pago_bod_comprobante_url_2': 'https://example.com/comprobante',
                    'pago_bod_banco': '',
                    'pago_bod_banco_otro': '',
                    'pago_bod_tarjeta_app': '',
                    'pago_bod_comprobante_url': '',
                },
            )

        self.assertRedirects(response, reverse('econotec:salida_retiros_lista'))
        salida.refresh_from_db()
        self.assertTrue(salida.cliente_ya_retiro)
        self.assertTrue(salida.bodegaje_aplicado_al_pago)
        self.assertEqual(salida.bodegaje_monto_congelado, Decimal('6.00'))
        abonos = list(ingreso.abonos.order_by('pk'))
        self.assertEqual([abono.monto for abono in abonos], [Decimal('2.50'), Decimal('3.50')])
        self.assertEqual([abono.metodo for abono in abonos], ['efectivo', 'transferencia'])
        self.assertEqual(abonos[0].bodegaje_decision, 'si')
        self.assertEqual(abonos[0].bodegaje_monto_aplicado, Decimal('6.00'))
        self.assertEqual(abonos[1].banco, 'guayaquil')
        self.assertEqual(len(mail.outbox), 1)
        html = next(
            alternativa[0]
            for alternativa in mail.outbox[0].alternatives
            if alternativa[1] == 'text/html'
        )
        self.assertIn('Pago registrado ahora', html)
        self.assertIn('$6,00', html)
        self.assertIn('Cobrado y cerrado', html)
        self.assertIn('Pago mixto (parte 1 de 2)', html)
        self.assertEqual(
            mail.outbox[0].attachments[0][0],
            f'Acta_salida_oficina_{ingreso.codigo_equipo}.pdf',
        )

    @override_settings(
        EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
        SALIDA_EMAIL_AUTOMATICO=True,
    )
    def test_salida_fisica_perdona_bodegaje_y_tambien_envia_acta(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=Decimal('0.00'))
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date.today() - timedelta(days=7),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        mail.outbox.clear()

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                reverse('econotec:salida_marcar_retirada', kwargs={'pk': salida.pk}),
                {'aplicar_bodegaje': ''},
            )

        self.assertRedirects(response, reverse('econotec:salida_retiros_lista'))
        salida.refresh_from_db()
        self.assertFalse(salida.bodegaje_aplicado_al_pago)
        self.assertEqual(salida.bodegaje_monto_congelado, Decimal('3.00'))
        self.assertEqual(ingreso.abonos.count(), 0)
        self.assertEqual(len(mail.outbox), 1)
        html = next(
            alternativa[0]
            for alternativa in mail.outbox[0].alternatives
            if alternativa[1] == 'text/html'
        )
        self.assertIn('Perdonado y cerrado', html)
        self.assertIn('$3,00', html)

    def test_pago_mixto_bodegaje_invalido_no_confirma_la_salida(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=Decimal('0.00'))
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date.today() - timedelta(days=10),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )

        response = self.client.post(
            reverse('econotec:salida_marcar_retirada', kwargs={'pk': salida.pk}),
            {
                'aplicar_bodegaje': 'on',
                'pago_bod_metodo': 'mixto',
                'pago_bod_monto_1': '2.00',
                'pago_bod_metodo_1': 'efectivo',
                'pago_bod_monto_2': '2.00',
                'pago_bod_metodo_2': 'efectivo',
            },
        )

        self.assertRedirects(response, reverse('econotec:salida_lista'))
        salida.refresh_from_db()
        self.assertFalse(salida.cliente_ya_retiro)
        self.assertEqual(ingreso.abonos.count(), 0)

    @override_settings(SALIDA_EMAIL_AUTOMATICO=True)
    @patch('econotec.emails.EmailMultiAlternatives.send', side_effect=OSError('SMTP no disponible'))
    def test_fallo_correo_salida_fisica_no_revierte_el_retiro(self, _send):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=Decimal('0.00'))
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date.today(),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )

        with self.assertLogs('econotec.emails', level='ERROR'):
            with self.captureOnCommitCallbacks(execute=True):
                response = self.client.post(
                    reverse('econotec:salida_marcar_retirada', kwargs={'pk': salida.pk}),
                    {'aplicar_bodegaje': ''},
                )

        self.assertRedirects(response, reverse('econotec:salida_retiros_lista'))
        salida.refresh_from_db()
        self.assertTrue(salida.cliente_ya_retiro)
        self.assertEqual(salida.estado_reparacion, 'retirado')

    def test_salida_facturas_lista_muestra_solo_facturas_realizadas(self):
        User = get_user_model()
        admin = User.objects.create_superuser(
            username='AdminFacturas',
            email='admin-facturas@example.com',
            password='testpass123',
        )
        self.client.force_login(admin)

        ingreso_facturado = self.crear_ingreso_reparacion(
            estado='entregado',
            marca='Sony',
            modelo_serie='Playstation 5',
        )
        salida_facturada = SalidaEquipo.objects.create(
            ingreso=ingreso_facturado,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='retirado',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('100.00'),
            metodo_pago_final='efectivo',
            factura_realizada='si',
            factura_nombres='Yandri Guevara',
            factura_cedula='1207342716',
            factura_correo='factura@example.com',
            registrado_por=admin,
        )
        ingreso_sin_factura = self.crear_ingreso_reparacion(
            estado='entregado',
            marca='HP',
            modelo_serie='Elitebook Factura No',
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso_sin_factura,
            fecha_salida=date(2026, 7, 10),
            estado_reparacion='retirado',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            factura_realizada='no',
            registrado_por=admin,
        )

        response = self.client.get(
            reverse('econotec:salida_facturas_lista'),
            {'ano': '2026', 'mes': '7'},
        )

        self.assertEqual(response.context['total_periodo'], 1)
        self.assertEqual(response.context['total'], 1)
        self.assertContains(response, 'Facturas <span class="accent">Realizadas</span>')
        self.assertContains(response, salida_facturada.ingreso.codigo_equipo)
        self.assertContains(response, 'factura@example.com')
        self.assertContains(
            response,
            reverse('econotec:salida_factura_imprimir', kwargs={'pk': salida_facturada.pk}),
        )
        self.assertContains(
            response,
            reverse('econotec:salida_factura_pdf', kwargs={'pk': salida_facturada.pk}),
        )
        self.assertNotContains(
            response,
            reverse('econotec:salida_imprimir', kwargs={'pk': salida_facturada.pk}),
        )
        self.assertNotContains(
            response,
            reverse('econotec:salida_pdf', kwargs={'pk': salida_facturada.pk}),
        )
        self.assertNotContains(response, 'Todas las salidas')
        self.assertNotContains(response, 'Factura realizada: No')
        self.assertNotContains(response, ingreso_sin_factura.codigo_equipo)

    def test_salida_facturas_lista_filtra_por_rango_fecha_salida(self):
        self.client.force_login(self.admin)
        ingreso_julio = self.crear_ingreso_reparacion(
            estado='entregado',
            marca='Epson',
            modelo_serie='L355 facturada',
        )
        ingreso_junio = self.crear_ingreso_reparacion(
            estado='entregado',
            marca='HP',
            modelo_serie='415 facturada fuera',
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso_julio,
            fecha_salida=date(2026, 7, 20),
            estado_reparacion='retirado',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('80.00'),
            metodo_pago_final='efectivo',
            factura_realizada='si',
            factura_nombres='Yandri Guevara',
            factura_cedula='1207342716',
            factura_correo='julio@example.com',
            registrado_por=self.admin,
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso_junio,
            fecha_salida=date(2026, 6, 25),
            estado_reparacion='retirado',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('90.00'),
            metodo_pago_final='efectivo',
            factura_realizada='si',
            factura_nombres='Yandri Guevara',
            factura_cedula='1207342716',
            factura_correo='junio@example.com',
            registrado_por=self.admin,
        )

        response = self.client.get(reverse('econotec:salida_facturas_lista'), {
            'ano': '2026',
            'mes': 'todos',
            'fecha_desde': '2026-07-01',
            'fecha_hasta': '2026-07-31',
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['total'], 1)
        self.assertEqual(response.context['total_periodo'], 1)
        self.assertContains(response, ingreso_julio.codigo_equipo)
        self.assertContains(response, 'julio@example.com')
        self.assertNotContains(response, ingreso_junio.codigo_equipo)
        self.assertNotContains(response, 'junio@example.com')
        self.assertContains(response, 'Fecha factura: 01/07/2026 - 31/07/2026')

    def test_salida_menu_muestra_acceso_a_facturas_realizadas(self):
        ingreso = self.crear_ingreso_reparacion(estado='entregado')
        SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='retirado',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('25.00'),
            metodo_pago_final='efectivo',
            factura_realizada='si',
            factura_nombres='Yandri Guevara',
            factura_cedula='1207342716',
            factura_correo='factura@example.com',
            registrado_por=self.usuario,
        )

        response = self.client.get(reverse('econotec:salida_menu'))

        self.assertContains(response, 'Facturas Realizadas')
        self.assertContains(response, reverse('econotec:salida_facturas_lista'))
        self.assertEqual(response.context['facturas_realizadas'], 1)
        self.assertContains(response, 'equipo finalizado con factura.')
        self.assertContains(response, 'Lista de equipos finalizados')
        self.assertNotContains(response, 'Salidas físicas confirmadas')
        self.assertNotContains(response, 'Ranking de Técnicos')
        self.assertNotContains(response, reverse('econotec:salida_retiros_lista'))
        self.assertNotContains(response, reverse('econotec:salida_totales'))
        self.assertNotContains(response, 'Buscar salidas por fecha')
        self.assertNotContains(response, 'Buscar facturas por fecha')

        inicio_tecnico = self.client.get(reverse('econotec:bienvenida'))
        self.assertContains(inicio_tecnico, 'Salidas físicas confirmadas')
        self.assertContains(inicio_tecnico, reverse('econotec:salida_retiros_lista'))
        self.assertContains(inicio_tecnico, 'Ranking de Técnicos')
        self.assertContains(inicio_tecnico, reverse('econotec:salida_totales'))
        self.assertContains(
            inicio_tecnico,
            'Consulta trabajos asignados, equipos finalizados, resultados y efectividad de cada técnico.',
        )

        self.client.force_login(self.vendedor)
        inicio_asesor = self.client.get(reverse('econotec:bienvenida'))
        self.assertContains(inicio_asesor, 'Ranking de Técnicos')
        self.assertContains(inicio_asesor, reverse('econotec:salida_totales'))

        self.client.force_login(self.admin)
        inicio_admin = self.client.get(reverse('econotec:bienvenida'))
        self.assertContains(inicio_admin, 'Ranking de Técnicos')
        self.assertContains(inicio_admin, reverse('econotec:salida_totales'))
        self.assertContains(
            inicio_admin,
            f'<a href="{reverse("econotec:salida_totales")}" class="card-box card-box-ranking">',
            html=False,
        )
        self.assertContains(inicio_admin, 'Vista completa')

    def test_busqueda_pagos_ignora_tildes_y_mayusculas(self):
        self.cliente_existente.nombres = 'Yandri Guevará'
        self.cliente_existente.save(update_fields=['nombres'])
        ingreso = self.crear_ingreso_reparacion()

        response = self.client.get(reverse('econotec:pagos_lista'), {'q': 'guevara'})

        self.assertEqual(response.context['total_count'], 1)
        self.assertContains(response, ingreso.codigo_equipo)
        self.assertContains(response, 'Yandri Guevará')

    def test_busqueda_ventas_ignora_tildes_y_mayusculas(self):
        self.cliente_existente.nombres = 'Yandri Guevará'
        self.cliente_existente.save(update_fields=['nombres'])
        venta = IngresoEquipo.objects.create(
            sede='ventas',
            asesor_comercial='Kimberly',
            fecha_ingreso=date(2026, 7, 9),
            cliente=self.cliente_existente,
            tipo_equipo='otro',
            marca='N/A',
            modelo_serie='N/A',
            accesorios_entregados='Ninguno',
            problema_reportado='Tinta Epson',
            valor_acordado=Decimal('25.00'),
            abono_anticipo=Decimal('25.00'),
            anticipo_metodo='efectivo',
            tecnico_encargado=self.usuario,
            estado='entregado',
            subestado_entregado='con_solucion',
            registrado_por=self.usuario,
        )

        response = self.client.get(reverse('econotec:venta_lista'), {'q': 'GUEVARA'})

        self.assertEqual(response.context['total'], 1)
        self.assertContains(response, venta.codigo_equipo)
        self.assertContains(response, 'Yandri Guevará')

    def test_lista_ventas_muestra_tecnico_vendio_y_filtra_personal(self):
        User = get_user_model()
        tecnico_alt = User.objects.create_user(username='Carlos')
        tecnico_alt.groups.add(Group.objects.get(name='Tecnicos'))
        venta_yandri = self.crear_venta_producto(
            problema_reportado='Cable HDMI',
            valor_acordado=Decimal('20.00'),
            abono_anticipo=Decimal('20.00'),
            anticipo_metodo='efectivo',
            tecnico_encargado=self.usuario,
            registrado_por=self.usuario,
        )
        venta_carlos = self.crear_venta_producto(
            problema_reportado='Mouse',
            valor_acordado=Decimal('20.00'),
            abono_anticipo=Decimal('20.00'),
            anticipo_metodo='efectivo',
            tecnico_encargado=tecnico_alt,
            registrado_por=self.vendedor,
        )

        response = self.client.get(
            reverse('econotec:venta_lista'),
            {'tecnico_vendio': str(self.usuario.pk)},
        )

        self.assertContains(response, 'Técnico vendió')
        self.assertContains(response, venta_yandri.codigo_equipo)
        self.assertNotContains(response, venta_carlos.codigo_equipo)

        response = self.client.get(
            reverse('econotec:venta_lista'),
            {'registrador': str(self.vendedor.pk)},
        )

        self.assertContains(response, venta_carlos.codigo_equipo)
        self.assertNotContains(response, venta_yandri.codigo_equipo)

    def test_lista_ventas_no_muestra_pago_pendiente(self):
        venta = self.crear_venta_producto(
            problema_reportado='Cable HDMI',
            valor_acordado=Decimal('20.00'),
            abono_anticipo=Decimal('20.00'),
            anticipo_metodo='efectivo',
        )

        response = self.client.get(reverse('econotec:venta_lista'))

        self.assertContains(response, venta.codigo_equipo)
        self.assertContains(response, 'Pagado')
        self.assertNotContains(response, 'Saldo Pendiente')
        self.assertNotContains(response, '💵 Pagar')

    def test_lista_ventas_pago_parcial_muestra_solo_ventas_con_abono(self):
        venta_completa = self.crear_venta_producto(
            problema_reportado='Cable HDMI',
            valor_acordado=Decimal('20.00'),
            abono_anticipo=Decimal('20.00'),
            anticipo_metodo='efectivo',
        )
        venta_parcial = self.crear_venta_producto(
            problema_reportado='Tarjeta grafica',
            valor_acordado=Decimal('200.00'),
            abono_anticipo=Decimal('100.00'),
            anticipo_metodo='transferencia',
            anticipo_banco='pichincha',
        )

        response = self.client.get(reverse('econotec:venta_lista_parciales'))

        self.assertContains(response, 'Ventas con')
        self.assertContains(response, 'Pago Parcial')
        self.assertContains(response, venta_parcial.codigo_equipo)
        self.assertContains(response, 'Tarjeta grafica')
        self.assertNotContains(response, venta_completa.codigo_equipo)
        self.assertContains(response, 'pago=parcial')

        response = self.client.get(reverse('econotec:venta_lista'))

        self.assertContains(response, 'Pago Completo')
        self.assertContains(response, venta_completa.codigo_equipo)
        self.assertNotContains(response, venta_parcial.codigo_equipo)
        self.assertContains(response, 'pago=completo')

    def test_lista_ventas_permite_editar_a_roles_y_eliminar_solo_admin(self):
        venta = self.crear_venta_producto(
            problema_reportado='Tarjeta',
            valor_acordado=Decimal('20.00'),
            abono_anticipo=Decimal('20.00'),
            anticipo_metodo='efectivo',
        )
        editar_url = reverse('econotec:venta_editar', kwargs={'pk': venta.pk})
        eliminar_url = reverse('econotec:venta_eliminar', kwargs={'pk': venta.pk})

        response = self.client.get(reverse('econotec:venta_lista'))

        self.assertContains(response, editar_url)
        self.assertContains(response, 'Editar')
        self.assertNotContains(response, eliminar_url)
        self.assertNotContains(response, 'Eliminar Venta')

        self.client.force_login(self.vendedor)
        response = self.client.get(reverse('econotec:venta_lista'))

        self.assertContains(response, editar_url)
        self.assertContains(response, 'Editar')
        self.assertNotContains(response, eliminar_url)
        self.assertNotContains(response, 'Eliminar Venta')
        self.assertEqual(self.client.get(editar_url).status_code, 200)

        self.client.force_login(self.admin)
        response = self.client.get(reverse('econotec:venta_lista'))

        self.assertContains(response, editar_url)
        self.assertContains(response, eliminar_url)
        self.assertContains(response, 'Eliminar Venta')

    def test_lista_ventas_muestra_envio_whatsapp_y_ayuda(self):
        venta = self.crear_venta_producto(
            problema_reportado='1 x Tarjeta',
            valor_acordado=Decimal('20.00'),
            abono_anticipo=Decimal('20.00'),
            anticipo_metodo='efectivo',
        )

        response = self.client.get(reverse('econotec:venta_lista'))

        self.assertContains(response, venta.codigo_equipo)
        self.assertContains(response, 'Enviar WhatsApp')
        self.assertContains(response, 'Cómo enviar la hoja por WhatsApp')
        self.assertContains(response, 'Primero presiona')
        self.assertContains(response, 'Descargar 📄')
        self.assertContains(response, 'api.whatsapp.com/send')

    def test_whatsapp_venta_producto_usa_saludo_por_hora_y_detalle(self):
        venta = self.crear_venta_producto(
            problema_reportado='1 x Tarjeta',
            valor_acordado=Decimal('20.00'),
            abono_anticipo=Decimal('20.00'),
        )
        ahora = datetime(2026, 7, 25, 15, 30, tzinfo=ZoneInfo('America/Guayaquil'))

        link = whatsapp_link_venta_producto(venta, ahora=ahora)
        texto = parse_qs(urlparse(link).query)['text'][0]

        self.assertIn('Buenas tardes, *Yandri Guevara*.', texto)
        self.assertIn('Le adjunto la hoja correspondiente a su producto comprado en Econotec.', texto)
        self.assertIn(f'Venta: *{venta.codigo_equipo}*', texto)
        self.assertIn('Producto(s): 1 x Tarjeta', texto)
        self.assertIn('Valor: $20.00', texto)

    def test_export_ventas_respeta_filtro_tecnico_vendio(self):
        User = get_user_model()
        tecnico_alt = User.objects.create_user(username='Carlos')
        tecnico_alt.groups.add(Group.objects.get(name='Tecnicos'))
        venta_yandri = self.crear_venta_producto(
            problema_reportado='Cable HDMI',
            valor_acordado=Decimal('20.00'),
            abono_anticipo=Decimal('20.00'),
            anticipo_metodo='efectivo',
            tecnico_encargado=self.usuario,
        )
        venta_carlos = self.crear_venta_producto(
            problema_reportado='Mouse',
            valor_acordado=Decimal('20.00'),
            abono_anticipo=Decimal('20.00'),
            anticipo_metodo='efectivo',
            tecnico_encargado=tecnico_alt,
        )

        response = self.client.get(
            reverse('econotec:venta_export'),
            {'tecnico_vendio': str(self.usuario.pk)},
        )

        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        codigos = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]

        self.assertIn('Técnico vendió', headers)
        self.assertIn(venta_yandri.codigo_equipo, codigos)
        self.assertNotIn(venta_carlos.codigo_equipo, codigos)

    def test_busqueda_control_pago_ventas_ignora_tildes_y_mayusculas(self):
        self.cliente_existente.nombres = 'Yandri Guevará'
        self.cliente_existente.save(update_fields=['nombres'])
        venta = IngresoEquipo.objects.create(
            sede='ventas',
            asesor_comercial='Kimberly',
            fecha_ingreso=date(2026, 7, 9),
            cliente=self.cliente_existente,
            tipo_equipo='otro',
            marca='N/A',
            modelo_serie='N/A',
            accesorios_entregados='Ninguno',
            problema_reportado='Tinta Epson',
            valor_acordado=Decimal('25.00'),
            tecnico_encargado=self.usuario,
            estado='entregado',
            subestado_entregado='con_solucion',
            registrado_por=self.usuario,
        )

        response = self.client.get(
            reverse('econotec:pagos_ventas_lista'),
            {'q': 'GUEVARA'},
        )

        self.assertEqual(response.context['total_count'], 1)
        self.assertContains(response, venta.codigo_equipo)
        self.assertContains(response, 'Yandri Guevará')

    def test_control_pago_ventas_usa_solo_boton_ver_pago(self):
        venta = self.crear_venta_producto(
            problema_reportado='Tarjeta',
            valor_acordado=Decimal('20.00'),
            abono_anticipo=Decimal('20.00'),
            anticipo_metodo='efectivo',
        )

        response = self.client.get(reverse('econotec:pagos_ventas_lista'))

        self.assertContains(response, venta.codigo_equipo)
        self.assertContains(response, 'Ver pago')
        self.assertNotContains(response, 'Ver / Ingresar Abonos')
        self.assertNotContains(response, 'Gestionar Pagos')

    def test_control_pago_ventas_menu_muestra_pago_completo_y_parcial(self):
        self.crear_venta_producto(
            problema_reportado='Tarjeta',
            valor_acordado=Decimal('20.00'),
            abono_anticipo=Decimal('20.00'),
            anticipo_metodo='efectivo',
        )
        self.crear_venta_producto(
            problema_reportado='Neutron',
            valor_acordado=Decimal('200.00'),
            abono_anticipo=Decimal('100.00'),
            anticipo_metodo='efectivo',
        )

        response = self.client.get(reverse('econotec:pagos_ventas_menu'))

        self.assertContains(response, 'Control de Pago de')
        self.assertContains(response, 'Pago completo')
        self.assertContains(response, 'Pagos parciales')
        self.assertContains(response, reverse('econotec:pagos_ventas_completos'))
        self.assertContains(response, reverse('econotec:pagos_ventas_parciales'))
        self.assertContains(response, 'Ventas: 1')

    def test_control_pago_ventas_filtra_completos_y_parciales(self):
        venta_completa = self.crear_venta_producto(
            problema_reportado='Tarjeta',
            valor_acordado=Decimal('20.00'),
            abono_anticipo=Decimal('20.00'),
            anticipo_metodo='efectivo',
        )
        venta_parcial = self.crear_venta_producto(
            problema_reportado='Neutron',
            valor_acordado=Decimal('200.00'),
            abono_anticipo=Decimal('100.00'),
            anticipo_metodo='efectivo',
        )

        response = self.client.get(reverse('econotec:pagos_ventas_completos'))

        self.assertContains(response, 'Pago Completo')
        self.assertContains(response, venta_completa.codigo_equipo)
        self.assertNotContains(response, venta_parcial.codigo_equipo)
        self.assertContains(response, 'Ver pago')

        response = self.client.get(reverse('econotec:pagos_ventas_parciales'))

        self.assertContains(response, 'Pagos Parciales')
        self.assertContains(response, venta_parcial.codigo_equipo)
        self.assertNotContains(response, venta_completa.codigo_equipo)
        self.assertContains(response, 'Ver abono / Historial')

    def test_registrar_venta_con_abono_activa_historial_de_pagos(self):
        response = self.client.post(
            reverse('econotec:venta_registrar'),
            self.venta_post_data(
                **{
                    'ing-valor_acordado': '25.00',
                    'venta_pago_modalidad': 'abono',
                    'ing-abono_anticipo': '10.00',
                    'ing-anticipo_metodo': 'efectivo',
                }
            ),
        )

        self.assertRedirects(response, reverse('econotec:venta_lista_parciales'))
        venta = IngresoEquipo.objects.get(sede='ventas')
        self.assertEqual(venta.abono_anticipo, Decimal('10.00'))
        self.assertEqual(venta.diferencia, Decimal('15.00'))
        self.assertEqual(venta.estado_pago, 'Parcial')

        response = self.client.get(reverse('econotec:pagos_ventas_lista'))
        self.assertContains(response, venta.codigo_equipo)
        self.assertContains(response, 'Ver abono / Historial')
        self.assertNotContains(response, 'Gestionar Pagos')

        response = self.client.get(reverse('econotec:ingreso_abonos', kwargs={'pk': venta.pk}))
        self.assertContains(response, 'Abonos / Historial de la Venta')
        self.assertContains(response, '+ Nuevo Abono')
        self.assertContains(response, 'Abono inicial de venta')

    def test_venta_con_abono_permite_registrar_pago_posterior(self):
        venta = self.crear_venta_producto(
            problema_reportado='Tarjeta',
            valor_acordado=Decimal('25.00'),
            abono_anticipo=Decimal('10.00'),
            anticipo_metodo='efectivo',
        )

        response = self.client.post(
            reverse('econotec:abono_crear', kwargs={'ingreso_pk': venta.pk}),
            {
                'fecha': '2026-07-25',
                'monto': '15.00',
                'metodo': 'transferencia',
                'banco': 'pichincha',
                'banco_otro': '',
                'tarjeta_app': '',
                'comprobante_url': '',
                'numero_recibo': '',
                'observaciones': 'Pago de saldo de venta.',
                'factura_realizada': 'no',
                'factura_nombres': '',
                'factura_cedula': '',
                'factura_correo': '',
                'bodegaje_decision': 'na',
                'bodegaje_monto_aplicado': '0.00',
            },
        )

        self.assertRedirects(response, reverse('econotec:ingreso_abonos', kwargs={'pk': venta.pk}))
        venta.refresh_from_db()
        self.assertEqual(venta.abonos.count(), 1)
        self.assertEqual(venta.total_abonado, Decimal('25.00'))
        self.assertEqual(venta.diferencia, Decimal('0.00'))

        response = self.client.get(reverse('econotec:pagos_ventas_lista'))
        self.assertContains(response, 'Ver abono / Historial')

    def test_ingreso_abonos_muestra_editor_de_valor_acordado(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('20.00'),
            abono_anticipo=Decimal('0.00'),
        )

        response = self.client.get(reverse('econotec:ingreso_abonos', kwargs={'pk': ingreso.pk}))

        self.assertContains(response, 'Valor acordado de este equipo')
        self.assertContains(response, 'id="btn-editar-valor-acordado"')
        self.assertContains(response, 'Editar valor acordado')
        self.assertContains(response, 'id="valor-acordado-confirm-modal"')
        self.assertContains(response, '¿Está seguro que quieres guardar los cambios?')
        self.assertContains(
            response,
            reverse('econotec:ingreso_valor_acordado_editar', kwargs={'pk': ingreso.pk}),
        )

    def test_editar_valor_acordado_desde_abonos_actualiza_saldo(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('20.00'),
            abono_anticipo=Decimal('5.00'),
        )

        response = self.client.post(
            reverse('econotec:ingreso_valor_acordado_editar', kwargs={'pk': ingreso.pk}),
            {'valor_acordado': '35.50'},
        )

        self.assertRedirects(response, reverse('econotec:ingreso_abonos', kwargs={'pk': ingreso.pk}))
        ingreso.refresh_from_db()
        self.assertEqual(ingreso.valor_acordado, Decimal('35.50'))
        self.assertEqual(ingreso.diferencia, Decimal('30.50'))
        evento = BitacoraTecnico.objects.get(ingreso=ingreso, tipo='ingreso_editado')
        self.assertIn('Valor acordado actualizado', evento.texto)
        self.assertIn('$20.00 a $35.50', evento.texto)

    def test_editar_valor_acordado_rechaza_monto_invalido(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=Decimal('20.00'))

        response = self.client.post(
            reverse('econotec:ingreso_valor_acordado_editar', kwargs={'pk': ingreso.pk}),
            {'valor_acordado': '20.999'},
        )

        self.assertRedirects(response, reverse('econotec:ingreso_abonos', kwargs={'pk': ingreso.pk}))
        ingreso.refresh_from_db()
        self.assertEqual(ingreso.valor_acordado, Decimal('20.00'))
        self.assertFalse(
            BitacoraTecnico.objects.filter(ingreso=ingreso, tipo='ingreso_editado').exists()
        )

    def test_formulario_abono_mixto_conserva_el_monto_total_a_dividir(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('5.00'),
            abono_anticipo=Decimal('0.00'),
        )

        response = self.client.get(
            reverse('econotec:abono_crear', kwargs={'ingreso_pk': ingreso.pk})
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Monto total a dividir (USD)')
        self.assertContains(response, 'id="id_monto"')
        self.assertContains(response, 'value="5.00"')
        self.assertContains(response, 'id="abono_monto_1"')
        self.assertContains(response, 'id="abono_monto_2"')
        self.assertContains(response, 'resumen-pago-mixto')

    def test_formulario_abono_muestra_accion_de_registrar_y_dar_salida(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('28.00'),
            abono_anticipo=Decimal('0.00'),
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date.today(),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )

        response = self.client.get(
            reverse('econotec:abono_crear', kwargs={'ingreso_pk': ingreso.pk})
        )

        self.assertContains(response, 'Registrar Abono y Dar Salida del Equipo')
        self.assertContains(response, 'value="registrar_y_salida"')
        self.assertContains(response, 'value="registrar"')

    def test_abono_completo_puede_confirmar_salida_fisica(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('28.00'),
            abono_anticipo=Decimal('0.00'),
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date.today(),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )

        response = self.client.post(
            reverse('econotec:abono_crear', kwargs={'ingreso_pk': ingreso.pk}),
            {
                'fecha': '2026-08-17',
                'monto': '28.00',
                'metodo': 'efectivo',
                'banco': '',
                'banco_otro': '',
                'tarjeta_app': '',
                'comprobante_url': '',
                'numero_recibo': '',
                'observaciones': '',
                'factura_realizada': 'no',
                'factura_nombres': '',
                'factura_cedula': '',
                'factura_correo': '',
                'bodegaje_decision': 'na',
                'bodegaje_monto_aplicado': '0.00',
                'accion_abono': 'registrar_y_salida',
            },
        )

        self.assertRedirects(response, reverse('econotec:salida_retiros_lista'))
        ingreso.refresh_from_db()
        salida.refresh_from_db()
        self.assertEqual(ingreso.total_abonado, Decimal('28.00'))
        self.assertEqual(ingreso.diferencia, Decimal('0.00'))
        self.assertEqual(salida.fecha_retiro_real, date.today())
        self.assertEqual(salida.estado_reparacion, 'retirado')

    def test_editar_abono_despues_de_salida_no_altera_estado_cerrado(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('28.00'),
            abono_anticipo=Decimal('0.00'),
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 8, 17),
            fecha_retiro_real=date(2026, 8, 18),
            estado_reparacion='retirado',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        abono = Abono.objects.create(
            ingreso=ingreso,
            fecha=date(2026, 8, 17),
            monto=Decimal('28.00'),
            metodo='efectivo',
            registrado_por=self.usuario,
        )

        response = self.client.post(
            reverse('econotec:abono_editar', kwargs={
                'ingreso_pk': ingreso.pk,
                'abono_pk': abono.pk,
            }),
            {
                'fecha': '2026-08-19',
                'monto': '28.00',
                'metodo': 'transferencia',
                'banco': 'pichincha',
                'banco_otro': '',
                'tarjeta_app': '',
                'comprobante_url': '',
                'numero_recibo': abono.numero_recibo,
                'observaciones': 'Pago actualizado después de la entrega.',
                'factura_realizada': 'no',
                'factura_nombres': '',
                'factura_cedula': '',
                'factura_correo': '',
                'bodegaje_decision': 'na',
                'bodegaje_monto_aplicado': '0.00',
            },
        )

        self.assertRedirects(
            response,
            reverse('econotec:ingreso_abonos', kwargs={'pk': ingreso.pk}),
        )
        salida.refresh_from_db()
        self.assertEqual(salida.estado_reparacion, 'retirado')
        self.assertEqual(salida.fecha_retiro_real, date(2026, 8, 18))

    def test_abono_parcial_no_confirma_salida_fisica(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('28.00'),
            abono_anticipo=Decimal('0.00'),
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date.today(),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )

        response = self.client.post(
            reverse('econotec:abono_crear', kwargs={'ingreso_pk': ingreso.pk}),
            {
                'fecha': '2026-08-17',
                'monto': '10.00',
                'metodo': 'efectivo',
                'banco': '',
                'banco_otro': '',
                'tarjeta_app': '',
                'comprobante_url': '',
                'numero_recibo': '',
                'observaciones': '',
                'factura_realizada': 'no',
                'factura_nombres': '',
                'factura_cedula': '',
                'factura_correo': '',
                'bodegaje_decision': 'na',
                'bodegaje_monto_aplicado': '0.00',
                'accion_abono': 'registrar_y_salida',
            },
        )

        self.assertRedirects(response, reverse('econotec:ingreso_abonos', kwargs={'pk': ingreso.pk}))
        ingreso.refresh_from_db()
        salida.refresh_from_db()
        self.assertEqual(ingreso.total_abonado, Decimal('10.00'))
        self.assertEqual(ingreso.diferencia, Decimal('18.00'))
        self.assertIsNone(salida.fecha_retiro_real)
        self.assertEqual(salida.estado_reparacion, 'pendiente_retiro')

    def test_abono_mixto_guarda_dos_partes_que_suman_el_total(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('5.00'),
            abono_anticipo=Decimal('0.00'),
        )

        response = self.client.post(
            reverse('econotec:abono_crear', kwargs={'ingreso_pk': ingreso.pk}),
            {
                'fecha': '2026-08-20',
                'monto': '5.00',
                'metodo': 'mixto',
                'abono_monto_1': '2.50',
                'abono_metodo_1': 'efectivo',
                'abono_banco_1': '',
                'abono_monto_2': '2.50',
                'abono_metodo_2': 'transferencia',
                'abono_banco_2': 'pichincha',
                'banco': '',
                'banco_otro': '',
                'tarjeta_app': '',
                'comprobante_url': '',
                'numero_recibo': '',
                'observaciones': 'Pago dividido.',
                'factura_realizada': 'no',
                'factura_nombres': '',
                'factura_cedula': '',
                'factura_correo': '',
                'bodegaje_decision': 'na',
                'bodegaje_monto_aplicado': '0.00',
            },
        )

        self.assertRedirects(
            response,
            reverse('econotec:ingreso_abonos', kwargs={'pk': ingreso.pk}),
        )
        partes = list(ingreso.abonos.order_by('pk'))
        self.assertEqual(len(partes), 2)
        self.assertEqual([parte.monto for parte in partes], [Decimal('2.50'), Decimal('2.50')])
        self.assertEqual([parte.metodo for parte in partes], ['efectivo', 'transferencia'])
        self.assertEqual(partes[1].banco, 'pichincha')
        self.assertIn('Parte 1 de 2', partes[0].observaciones)
        self.assertIn('Parte 2 de 2', partes[1].observaciones)
        self.assertEqual(ingreso.total_abonado, Decimal('5.00'))
        self.assertEqual(ingreso.diferencia, Decimal('0.00'))

    def test_abono_mixto_rechaza_partes_que_no_suman_el_total(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('5.00'),
            abono_anticipo=Decimal('0.00'),
        )

        response = self.client.post(
            reverse('econotec:abono_crear', kwargs={'ingreso_pk': ingreso.pk}),
            {
                'fecha': '2026-08-20',
                'monto': '5.00',
                'metodo': 'mixto',
                'abono_monto_1': '2.00',
                'abono_metodo_1': 'efectivo',
                'abono_banco_1': '',
                'abono_monto_2': '2.00',
                'abono_metodo_2': 'efectivo',
                'abono_banco_2': '',
                'banco': '',
                'banco_otro': '',
                'tarjeta_app': '',
                'comprobante_url': '',
                'numero_recibo': '',
                'observaciones': '',
                'factura_realizada': 'no',
                'factura_nombres': '',
                'factura_cedula': '',
                'factura_correo': '',
                'bodegaje_decision': 'na',
                'bodegaje_monto_aplicado': '0.00',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            'La suma de las dos partes debe ser exactamente $5.00.',
        )
        self.assertFalse(ingreso.abonos.exists())

    def test_ingreso_permite_detalle_simple_en_reparacion(self):
        form = IngresoEquipoForm(data={
            'numero_factura': '',
            'asesor_comercial': 'Kimberly',
            'tecnico_encargado': str(self.usuario.pk),
            'fecha_ingreso': '2026-07-09',
            'tipo_equipo': 'laptop',
            'tipo_equipo_otro': '',
            'marca': 'HP',
            'modelo_serie': 'Elitebook',
            'serie': '',
            'accesorios_entregados': '',
            'problema_reportado': 'No enciende',
            'firma_cliente_opcion': 'no',
            'firma_cliente_imagen': '',
            'diagnostico_inmediato': 'no',
            'valor_diagnostico': '0.00',
            'valor_acordado': '25',
            'abono_anticipo': '0.00',
            'diagnostico_metodo': 'efectivo',
            'diagnostico_banco': '',
            'diagnostico_banco_otro': '',
            'diagnostico_tarjeta_app': '',
            'diagnostico_comprobante_url': '',
            'diagnostico_monto_1': '',
            'diagnostico_metodo_1': '',
            'diagnostico_banco_1': '',
            'diagnostico_monto_2': '',
            'diagnostico_metodo_2': '',
            'diagnostico_banco_2': '',
            'anticipo_metodo': 'efectivo',
            'anticipo_banco': '',
            'anticipo_banco_otro': '',
            'anticipo_tarjeta_app': '',
            'anticipo_comprobante_url': '',
            'anticipo_monto_1': '',
            'anticipo_metodo_1': '',
            'anticipo_banco_1': '',
            'anticipo_monto_2': '',
            'anticipo_metodo_2': '',
            'anticipo_banco_2': '',
            'estado': 'en_reparacion',
            'subestado_reparacion': 'en_reparacion',
            'subestado_entregado': '',
            'equipo_garantia': '',
            'equipo_garantia_manual': '',
            'motivo_garantia': '',
        })

        self.assertTrue(form.is_valid(), form.errors.as_json())
        self.assertEqual(form.cleaned_data['subestado_reparacion'], 'en_reparacion')

    def test_valor_acordado_no_guarda_como_pendiente(self):
        form = IngresoEquipoForm(data=self.ingreso_form_data(
            valor_acordado_estado='no',
            valor_acordado='99.00',
        ))

        self.assertTrue(form.is_valid(), form.errors.as_json())
        self.assertIsNone(form.cleaned_data['valor_acordado'])

    def test_valor_acordado_no_con_punto_guarda_como_pendiente(self):
        form = IngresoEquipoForm(data=self.ingreso_form_data(
            valor_acordado_estado='no',
            valor_acordado='.',
        ))

        self.assertTrue(form.is_valid(), form.errors.as_json())
        self.assertIsNone(form.cleaned_data['valor_acordado'])

    def test_valor_acordado_si_exige_monto(self):
        form = IngresoEquipoForm(data=self.ingreso_form_data(
            valor_acordado_estado='si',
            valor_acordado='',
        ))

        self.assertFalse(form.is_valid())
        self.assertIn('valor_acordado', form.errors)

    def test_valor_acordado_si_exige_minimo_un_dolar(self):
        for valor in ('0.00', '0.99', '01.00'):
            with self.subTest(valor=valor):
                form = IngresoEquipoForm(data=self.ingreso_form_data(
                    valor_acordado_estado='si',
                    valor_acordado=valor,
                ))

                self.assertFalse(form.is_valid())
                self.assertIn('valor_acordado', form.errors)
                self.assertIn(
                    'El valor debe ser igual o mayor a $1.00.',
                    form.errors['valor_acordado']
                )

    def test_valor_acordado_si_acepta_un_dolar_exacto(self):
        form = IngresoEquipoForm(data=self.ingreso_form_data(
            valor_acordado_estado='si',
            valor_acordado='1.00',
        ))

        self.assertTrue(form.is_valid(), form.errors.as_json())
        self.assertEqual(form.cleaned_data['valor_acordado'], Decimal('1.00'))

    def test_ingreso_garantia_fuerza_valor_acordado_cero(self):
        form = IngresoEquipoForm(data=self.ingreso_form_data(
            estado='garantia',
            subestado_reparacion='',
            valor_acordado_estado='si',
            valor_acordado='99.00',
            motivo_garantia='Falla cubierta por garantía',
        ))

        self.assertTrue(form.is_valid(), form.errors.as_json())
        self.assertEqual(form.cleaned_data['valor_acordado'], Decimal('0.00'))

    def test_ingreso_garantia_no_valida_monto_manual(self):
        form = IngresoEquipoForm(data=self.ingreso_form_data(
            estado='garantia',
            subestado_reparacion='',
            valor_acordado_estado='si',
            valor_acordado='valor indebido',
            motivo_garantia='Falla cubierta por garantía',
        ))

        self.assertTrue(form.is_valid(), form.errors.as_json())
        self.assertEqual(form.cleaned_data['valor_acordado'], Decimal('0.00'))

    def test_ingreso_garantia_fuerza_diagnostico_sin_cobro(self):
        form = IngresoEquipoForm(data=self.ingreso_form_data(
            estado='garantia',
            subestado_reparacion='',
            valor_acordado_estado='si',
            valor_acordado='99.00',
            diagnostico_inmediato='si',
            valor_diagnostico='25.00',
            diagnostico_metodo='mixto',
            diagnostico_monto_1='5.00',
            diagnostico_metodo_1='transferencia',
            diagnostico_banco_1='pichincha',
            diagnostico_monto_2='20.00',
            diagnostico_metodo_2='efectivo',
            motivo_garantia='Falla cubierta por garantía',
        ))

        self.assertTrue(form.is_valid(), form.errors.as_json())
        self.assertEqual(form.cleaned_data['diagnostico_inmediato'], 'no')
        self.assertEqual(form.cleaned_data['valor_diagnostico'], Decimal('0.00'))
        self.assertEqual(form.cleaned_data['diagnostico_metodo'], 'efectivo')
        self.assertIsNone(form.cleaned_data['diagnostico_monto_1'])
        self.assertIsNone(form.cleaned_data['diagnostico_monto_2'])

    def test_ingreso_cortesia_fuerza_todos_los_valores_a_cero(self):
        form = IngresoEquipoForm(data=self.ingreso_form_data(
            estado='cortesia',
            subestado_reparacion='',
            valor_acordado_estado='si',
            valor_acordado='125.00',
            diagnostico_inmediato='si',
            valor_diagnostico='25.00',
            diagnostico_metodo='transferencia',
            diagnostico_banco='pichincha',
            abono_anticipo='40.00',
            anticipo_metodo='tarjeta',
            anticipo_tarjeta_app='payphone',
        ))

        self.assertTrue(form.is_valid(), form.errors.as_json())
        self.assertEqual(form.cleaned_data['valor_acordado'], Decimal('0.00'))
        self.assertEqual(form.cleaned_data['diagnostico_inmediato'], 'no')
        self.assertEqual(form.cleaned_data['valor_diagnostico'], Decimal('0.00'))
        self.assertEqual(form.cleaned_data['abono_anticipo'], Decimal('0.00'))

    def test_ingreso_cortesia_ignora_valor_manual_invalido(self):
        form = IngresoEquipoForm(data=self.ingreso_form_data(
            estado='cortesia',
            subestado_reparacion='',
            valor_acordado_estado='si',
            valor_acordado='valor manipulado',
        ))

        self.assertTrue(form.is_valid(), form.errors.as_json())
        self.assertEqual(form.cleaned_data['valor_acordado'], Decimal('0.00'))

    def test_ingreso_donado_limpia_firma_diagnostico_anticipo_y_valor(self):
        data = self.ingreso_form_data(
            estado='donado',
            subestado_reparacion='',
            valor_acordado='99.00',
            diagnostico_inmediato='si',
            valor_diagnostico='25.00',
            abono_anticipo='15.00',
        )
        data.pop('firma_cliente_opcion')

        form = IngresoEquipoForm(data=data)

        self.assertTrue(form.is_valid(), form.errors.as_json())
        self.assertEqual(form.cleaned_data['valor_acordado'], Decimal('0.00'))
        self.assertEqual(form.cleaned_data['diagnostico_inmediato'], 'no')
        self.assertEqual(form.cleaned_data['valor_diagnostico'], Decimal('0.00'))
        self.assertEqual(form.cleaned_data['abono_anticipo'], Decimal('0.00'))
        self.assertFalse(form.cleaned_data['firma_cliente'])

    def test_compra_mixta_acepta_transferencia_y_tarjeta_y_limpia_anticipo(self):
        data = self.ingreso_form_data(
            estado='equipo_a_comprar',
            subestado_reparacion='',
            valor_acordado='100.00',
            abono_anticipo='20.00',
            compra_metodo_pago='mixto',
            compra_monto_1='60.00',
            compra_metodo_1='transferencia',
            compra_banco_1='pichincha',
            compra_monto_2='40.00',
            compra_metodo_2='tarjeta',
            compra_tarjeta_app_2='deuna',
        )
        data.pop('firma_cliente_opcion')

        form = IngresoEquipoForm(data=data)

        self.assertTrue(form.is_valid(), form.errors.as_json())
        self.assertEqual(form.cleaned_data['abono_anticipo'], Decimal('0.00'))
        self.assertEqual(form.cleaned_data['compra_monto_1'], Decimal('60.00'))
        self.assertEqual(form.cleaned_data['compra_monto_2'], Decimal('40.00'))

    def test_compra_mixta_no_permite_metodo_mixto_anidado(self):
        form = IngresoEquipoForm(data=self.ingreso_form_data(
            estado='equipo_a_comprar',
            subestado_reparacion='',
            valor_acordado='100.00',
            compra_metodo_pago='mixto',
            compra_monto_1='60.00',
            compra_metodo_1='mixto',
            compra_monto_2='40.00',
            compra_metodo_2='efectivo',
        ))

        self.assertFalse(form.is_valid())
        self.assertIn('compra_metodo_1', form.errors)

    def test_registrar_compra_mixta_crea_un_solo_egreso_vinculado(self):
        self.activar_sede_guayaquil()
        data = self.ingreso_registro_post_data(
            **{
                'ing-estado': 'equipo_a_comprar',
                'ing-subestado_reparacion': '',
                'ing-valor_acordado': '100.00',
                'ing-diagnostico_inmediato': 'si',
                'ing-valor_diagnostico': '25.00',
                'ing-abono_anticipo': '20.00',
                'ing-compra_metodo_pago': 'mixto',
                'ing-compra_monto_1': '60.00',
                'ing-compra_metodo_1': 'transferencia',
                'ing-compra_banco_1': 'pichincha',
                'ing-compra_monto_2': '40.00',
                'ing-compra_metodo_2': 'tarjeta',
                'ing-compra_tarjeta_app_2': 'deuna',
            }
        )
        data.pop('ing-firma_cliente_opcion')

        response = self.client.post(reverse('econotec:ingreso_registrar'), data)

        ingreso = IngresoEquipo.objects.get(cliente=self.cliente_existente)
        self.assertRedirects(
            response,
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk}),
        )
        self.assertEqual(ingreso.estado, 'equipo_a_comprar')
        self.assertEqual(ingreso.abono_anticipo, Decimal('0.00'))
        self.assertEqual(ingreso.diagnostico_inmediato, 'no')
        self.assertEqual(Egreso.objects.filter(ingreso_compra=ingreso).count(), 1)
        egreso = ingreso.egreso_compra
        self.assertEqual(egreso.monto, Decimal('100.00'))
        self.assertEqual(egreso.metodo, 'mixto')
        self.assertEqual(egreso.monto_1, Decimal('60.00'))
        self.assertEqual(egreso.banco_1, 'pichincha')
        self.assertEqual(egreso.monto_2, Decimal('40.00'))
        self.assertEqual(egreso.tarjeta_app_2, 'deuna')

    def test_detalle_bloquea_boton_salida_si_valor_acordado_pendiente(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=None)

        response = self.client.get(
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk})
        )

        self.assertContains(response, 'btn-salida-bloqueada')
        self.assertContains(
            response,
            'Por favor registra un valor acordado para finalizar el equipo.'
        )

    def test_no_permite_registrar_salida_sin_valor_acordado(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=None)

        response = self.client.get(
            reverse('econotec:salida_registrar', kwargs={'ingreso_pk': ingreso.pk})
        )

        self.assertRedirects(
            response,
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk})
        )
        self.assertFalse(SalidaEquipo.objects.filter(ingreso=ingreso).exists())

    def test_finalizacion_nueva_no_ofrece_resultados_gestionados_en_ingreso(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=Decimal('40.00'))

        response = self.client.get(
            reverse('econotec:salida_registrar', kwargs={'ingreso_pk': ingreso.pk})
        )

        self.assertEqual(response.status_code, 200)
        opciones = dict(response.context['form'].fields['estado_reparacion'].choices)
        self.assertNotIn('cliente_no_acepta', opciones)
        self.assertNotIn('no_reparable', opciones)
        self.assertIn('pendiente_retiro', opciones)
        self.assertIn('revision', opciones)

    def test_finalizacion_nueva_rechaza_resultado_negativo_enviado_manualmente(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=Decimal('40.00'))

        response = self.client.post(
            reverse('econotec:salida_registrar', kwargs={'ingreso_pk': ingreso.pk}),
            self.salida_post_data(
                estado_reparacion='no_reparable',
                metodo_pago_final='sin_pago',
            ),
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(SalidaEquipo.objects.filter(ingreso=ingreso).exists())
        self.assertIn('estado_reparacion', response.context['form'].errors)

    def test_edicion_conserva_resultado_negativo_existente(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=Decimal('0.00'))
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 17),
            estado_reparacion='no_reparable',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )

        response = self.client.get(
            reverse('econotec:salida_editar', kwargs={'pk': salida.pk})
        )

        self.assertEqual(response.status_code, 200)
        opciones = dict(response.context['form'].fields['estado_reparacion'].choices)
        self.assertIn('no_reparable', opciones)
        self.assertNotIn('cliente_no_acepta', opciones)

    def test_aviso_post_finalizacion_muestra_confirmacion_o_tutorial_segun_saldo(self):
        ingreso_pagado = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('0.00'),
        )
        salida_pagada = SalidaEquipo.objects.create(
            ingreso=ingreso_pagado,
            fecha_salida=date(2026, 7, 17),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        session = self.client.session
        session['confirmar_ubicacion_salida_id'] = salida_pagada.pk
        session.save()

        aviso_pagado = self.client.get(
            reverse('econotec:salida_listo_aviso', kwargs={'pk': salida_pagada.pk}),
        )
        self.assertTrue(aviso_pagado.context['mostrar_confirmacion_guardado'])
        self.assertContains(aviso_pagado, '¿El equipo se encuentra aquí en la oficina?')
        self.assertContains(aviso_pagado, 'Sí, está en la oficina')
        self.assertContains(aviso_pagado, 'No, ya salió')

        salida_pagada.fecha_retiro_real = date(2026, 7, 18)
        salida_pagada.save(update_fields=['fecha_retiro_real'])
        session = self.client.session
        session['confirmar_ubicacion_salida_id'] = salida_pagada.pk
        session.save()
        aviso_ya_retirado = self.client.get(
            reverse('econotec:salida_listo_aviso', kwargs={'pk': salida_pagada.pk}),
        )
        self.assertTrue(aviso_ya_retirado.context['mostrar_confirmacion_guardado'])
        self.assertTrue(aviso_ya_retirado.context['salida_ya_confirmada'])
        self.assertContains(aviso_ya_retirado, 'Cambios guardados')
        self.assertContains(aviso_ya_retirado, 'Su salida de la oficina ya está confirmada.')

        ingreso_pendiente = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('25.00'),
        )
        salida_pendiente = SalidaEquipo.objects.create(
            ingreso=ingreso_pendiente,
            fecha_salida=date(2026, 7, 18),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        session = self.client.session
        session['confirmar_ubicacion_salida_id'] = salida_pendiente.pk
        session.save()

        aviso_pendiente = self.client.get(
            reverse('econotec:salida_listo_aviso', kwargs={'pk': salida_pendiente.pk}),
        )
        self.assertContains(aviso_pendiente, 'Este equipo tiene un saldo pendiente')
        self.assertContains(aviso_pendiente, '📖 Guía rápida')
        self.assertContains(aviso_pendiente, 'Siguiente →')
        self.assertContains(aviso_pendiente, 'Ingresar saldo pendiente')
        self.assertContains(aviso_pendiente, 'id="btn-ingresar-saldo-pendiente"')
        self.assertContains(aviso_pendiente, 'saldo-pendiente-directo')
        self.assertContains(aviso_pendiente, 'No quiero que se muestre de nuevo')
        self.assertContains(aviso_pendiente, 'type="checkbox"')
        self.assertContains(aviso_pendiente, '¿Está seguro?')
        self.assertContains(aviso_pendiente, "confirmButtonText: 'Sí'")
        self.assertContains(aviso_pendiente, "cancelButtonText: 'No'")
        self.assertNotContains(aviso_pendiente, 'Omitir')
        self.assertContains(
            aviso_pendiente,
            reverse('econotec:abono_crear', kwargs={'ingreso_pk': ingreso_pendiente.pk}),
        )
        lista_con_saldo = self.client.get(reverse('econotec:salida_lista'))
        self.assertContains(lista_con_saldo, '💳')
        self.assertContains(lista_con_saldo, 'Primero debe pagarse el saldo pendiente')

        rechazo = self.client.post(
            reverse('econotec:salida_marcar_retirada', kwargs={'pk': salida_pendiente.pk}),
        )
        self.assertRedirects(
            rechazo,
            reverse('econotec:salida_listo_aviso', kwargs={'pk': salida_pendiente.pk}),
        )
        salida_pendiente.refresh_from_db()
        self.assertIsNone(salida_pendiente.fecha_retiro_real)

    def test_ocultar_guia_saldo_pendiente_se_guarda_por_usuario(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=Decimal('25.00'))
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 18),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )

        guardar = self.client.post(
            reverse('econotec:salida_ocultar_guia_saldo_pendiente'),
        )

        self.assertEqual(guardar.status_code, 200)
        self.assertJSONEqual(guardar.content, {'ok': True})
        actividad = UsuarioActividad.objects.get(user=self.usuario)
        self.assertTrue(actividad.ocultar_guia_saldo_pendiente)

        session = self.client.session
        session['confirmar_ubicacion_salida_id'] = salida.pk
        session.save()
        aviso_mismo_usuario = self.client.get(
            reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}),
        )
        self.assertContains(aviso_mismo_usuario, 'Este equipo tiene un saldo pendiente')
        self.assertNotContains(aviso_mismo_usuario, '📖 Guía rápida')
        self.assertNotContains(aviso_mismo_usuario, 'No quiero que se muestre de nuevo')

        otro_usuario = get_user_model().objects.create_user(username='OtroTecnico')
        otro_usuario.groups.add(Group.objects.get(name='Tecnicos'))
        self.client.force_login(otro_usuario)
        session = self.client.session
        session['confirmar_ubicacion_salida_id'] = salida.pk
        session.save()
        aviso_otro_usuario = self.client.get(
            reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}),
        )
        self.assertContains(aviso_otro_usuario, '📖 Guía rápida')
        self.assertContains(aviso_otro_usuario, 'No quiero que se muestre de nuevo')

    def test_salida_garantia_fallos_adicionales_deja_valor_pendiente_y_notifica_asesora(self):
        ingreso = self.crear_ingreso_reparacion(
            estado='garantia',
            valor_acordado=Decimal('0.00'),
            motivo_garantia='Garantía por retorno',
            equipo_garantia_manual='G1000',
        )

        response = self.client.post(
            reverse('econotec:salida_registrar', kwargs={'ingreso_pk': ingreso.pk}),
            {
                'fecha_salida': '2026-07-17',
                'estado_reparacion': 'garantia_fallos_adicionales',
                'tecnico_reparo': str(self.usuario.pk),
                'reporte_tecnico': 'Se detectaron fallos adicionales.',
                'observaciones': '',
                'valor_final_cobrado': '75.00',
                'metodo_pago_final': 'sin_pago',
                'numero_recibo': '',
                'banco': '',
                'banco_otro': '',
                'tarjeta_app': '',
                'comprobante_url': '',
                'monto_1': '',
                'metodo_1': '',
                'banco_1': '',
                'monto_2': '',
                'metodo_2': '',
                'banco_2': '',
                'factura_realizada': 'no',
                'factura_nombres': '',
                'factura_cedula': '',
                'factura_correo': '',
                'asesora_notificacion': str(self.vendedor.pk),
                'mensaje_notificacion': 'Cobrar fallos adicionales antes del retiro.',
            },
        )

        salida = SalidaEquipo.objects.get(ingreso=ingreso)
        self.assertRedirects(response, reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}))
        ingreso.refresh_from_db()
        self.assertEqual(salida.estado_reparacion, 'garantia_fallos_adicionales')
        self.assertEqual(salida.valor_final_cobrado, Decimal('0.00'))
        self.assertEqual(salida.metodo_pago_final, 'sin_pago')
        self.assertEqual(ingreso.valor_acordado, Decimal('75.00'))
        self.assertEqual(ingreso.diferencia, Decimal('75.00'))
        self.assertEqual(ingreso.estado_pago, 'Pendiente')

        notificacion = NotificacionAsesora.objects.get(salida=salida)
        self.assertEqual(notificacion.asesora, self.vendedor)
        self.assertEqual(notificacion.valor_acordado, Decimal('75.00'))
        self.assertFalse(notificacion.leida)

    def test_salida_cortesia_es_unica_y_limpia_todo_cobro(self):
        ingreso = self.crear_ingreso_reparacion(
            estado='cortesia',
            subestado_reparacion='',
            valor_acordado=Decimal('99.00'),
            diagnostico_inmediato='si',
            valor_diagnostico=Decimal('15.00'),
            abono_anticipo=Decimal('20.00'),
        )
        ingreso.refresh_from_db()
        self.assertEqual(ingreso.valor_acordado, Decimal('0.00'))
        self.assertEqual(ingreso.valor_diagnostico, Decimal('0.00'))
        self.assertEqual(ingreso.abono_anticipo, Decimal('0.00'))

        response_get = self.client.get(
            reverse('econotec:salida_registrar', kwargs={'ingreso_pk': ingreso.pk})
        )
        self.assertEqual(response_get.status_code, 200)
        self.assertContains(response_get, 'Finalización de cortesía')
        self.assertNotContains(response_get, 'Cierre Económico')
        self.assertNotContains(response_get, '¿Factura realizada?')

        response = self.client.post(
            reverse('econotec:salida_registrar', kwargs={'ingreso_pk': ingreso.pk}),
            self.salida_post_data(
                estado_reparacion='cortesia',
                valor_final_cobrado='80.00',
                metodo_pago_final='transferencia',
                banco='pichincha',
                numero_recibo='REC-MANIPULADO',
                factura_realizada='si',
                factura_nombres='No debe guardarse',
                factura_cedula='0999999999',
                factura_correo='factura@example.com',
            ),
        )

        salida = SalidaEquipo.objects.get(ingreso=ingreso)
        self.assertRedirects(
            response,
            reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}),
        )
        self.assertEqual(salida.estado_reparacion, 'cortesia')
        self.assertEqual(salida.valor_final_cobrado, Decimal('0.00'))
        self.assertEqual(salida.metodo_pago_final, 'cortesia')
        self.assertEqual(salida.numero_recibo, '')
        self.assertEqual(salida.factura_realizada, 'no')
        self.assertEqual(salida.factura_nombres, '')
        self.assertEqual(ingreso.diferencia, Decimal('0.00'))

        response_print = self.client.get(
            reverse('econotec:salida_imprimir', kwargs={'pk': salida.pk})
        )
        self.assertContains(response_print, 'EQUIPO DE CORTESÍA FINALIZADO')
        self.assertNotContains(response_print, 'CIERRE ECONÓMICO')

    def test_ingreso_cortesia_no_admite_abonos(self):
        ingreso = self.crear_ingreso_reparacion(
            estado='cortesia',
            subestado_reparacion='',
            valor_acordado=Decimal('0.00'),
        )

        response = self.client.get(
            reverse('econotec:abono_crear', kwargs={'ingreso_pk': ingreso.pk})
        )

        self.assertRedirects(
            response,
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk}),
        )
        self.assertFalse(Abono.objects.filter(ingreso=ingreso).exists())

    def test_admin_cortesia_muestra_ingreso_y_salida_en_apartado_propio(self):
        self.client.force_login(self.admin)
        ingreso_con_salida = self.crear_ingreso_reparacion(
            fecha_ingreso=date(2026, 7, 10),
            estado='cortesia',
            subestado_reparacion='',
            valor_acordado=Decimal('0.00'),
            modelo_serie='Cortesía con salida',
        )
        ingreso_pendiente = self.crear_ingreso_reparacion(
            fecha_ingreso=date(2026, 7, 11),
            estado='cortesia',
            subestado_reparacion='',
            valor_acordado=Decimal('0.00'),
            modelo_serie='Cortesía pendiente',
        )
        SalidaEquipo.objects.create(
            ingreso=ingreso_con_salida,
            fecha_salida=date(2026, 7, 20),
            estado_reparacion='cortesia',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('90.00'),
            metodo_pago_final='efectivo',
            registrado_por=self.usuario,
        )

        response = self.client.get(
            reverse('econotec:admin_equipos_cortesia'),
            {'ano': '2026', 'mes': '7'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['ingresos_mes'], 2)
        self.assertEqual(response.context['salidas_mes'], 1)
        self.assertEqual(response.context['pendientes_salida'], 1)
        self.assertContains(response, ingreso_con_salida.codigo_equipo)
        self.assertContains(response, ingreso_pendiente.codigo_equipo)
        self.assertContains(response, 'Equipo de cortesía finalizado')
        self.assertContains(response, 'Pendiente de registrar')

    def test_salida_cliente_no_acepta_cobro_adicional_pendiente_notifica_asesora(self):
        response, ingreso, salida = self.registrar_ingreso_negativo(
            IngresoEquipoForm.ESTADO_NO_QUISO_REPARAR,
            estado_reparacion='cliente_no_acepta',
            aplica_valor_acordado_adicional='si',
            valor_acordado_adicional='12.00',
            motivo_valor_acordado_adicional='Revisión profunda autorizada.',
            valor_final_cobrado='0.00',
            metodo_pago_final='sin_pago',
            asesora_notificacion=str(self.vendedor.pk),
            mensaje_notificacion='Cobrar valor adicional antes del retiro.',
        )

        self.assertRedirects(response, reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}))
        ingreso.refresh_from_db()
        self.assertEqual(salida.aplica_valor_acordado_adicional, 'si')
        self.assertEqual(salida.valor_acordado_adicional, Decimal('12.00'))
        self.assertEqual(salida.valor_final_cobrado, Decimal('0.00'))
        self.assertEqual(salida.metodo_pago_final, 'sin_pago')
        self.assertEqual(ingreso.valor_efectivo_a_cobrar, Decimal('12.00'))
        self.assertEqual(ingreso.diferencia, Decimal('12.00'))
        self.assertEqual(ingreso.estado_pago, 'Pendiente')

        notificacion = NotificacionAsesora.objects.get(salida=salida)
        self.assertEqual(notificacion.tipo, NotificacionAsesora.TIPO_COBRO_ADICIONAL)
        self.assertEqual(notificacion.asesora, self.vendedor)
        self.assertEqual(notificacion.valor_acordado, Decimal('12.00'))

    def test_salida_no_reparable_pago_parcial_transferencia_bloquea_retiro(self):
        response, ingreso, salida = self.registrar_ingreso_negativo(
            IngresoEquipoForm.ESTADO_NO_SE_PUDO_REPARAR,
            estado_reparacion='no_reparable',
            aplica_valor_acordado_adicional='si',
            valor_acordado_adicional='7.00',
            motivo_valor_acordado_adicional='Trabajo adicional autorizado.',
            valor_final_cobrado='3.00',
            metodo_pago_final='transferencia',
            banco='pichincha',
            asesora_notificacion=str(self.vendedor.pk),
            mensaje_notificacion='Quedan cuatro dólares pendientes.',
        )

        self.assertRedirects(response, reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}))
        ingreso.refresh_from_db()
        self.assertEqual(salida.valor_final_cobrado, Decimal('3.00'))
        self.assertEqual(salida.metodo_pago_final, 'transferencia')
        self.assertEqual(salida.banco, 'pichincha')
        self.assertEqual(ingreso.valor_efectivo_a_cobrar, Decimal('7.00'))
        self.assertEqual(ingreso.diferencia, Decimal('4.00'))

        notificacion = NotificacionAsesora.objects.get(salida=salida)
        self.assertEqual(notificacion.tipo, NotificacionAsesora.TIPO_COBRO_ADICIONAL)
        self.assertEqual(notificacion.valor_acordado, Decimal('4.00'))

        retiro = self.client.post(
            reverse('econotec:salida_marcar_retirada', kwargs={'pk': salida.pk}),
        )
        self.assertRedirects(
            retiro,
            reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}),
        )
        salida.refresh_from_db()
        self.assertIsNone(salida.fecha_retiro_real)
        self.assertEqual(salida.estado_reparacion, 'no_reparable')

    def test_salida_sin_reparacion_acepta_cobro_adicional_desde_un_centavo(self):
        response, ingreso, salida = self.registrar_ingreso_negativo(
            IngresoEquipoForm.ESTADO_NO_QUISO_REPARAR,
            estado_reparacion='cliente_no_acepta',
            aplica_valor_acordado_adicional='si',
            valor_acordado_adicional='0.01',
            motivo_valor_acordado_adicional='Material consumido autorizado.',
            valor_final_cobrado='0.00',
            metodo_pago_final='sin_pago',
            asesora_notificacion=str(self.vendedor.pk),
            mensaje_notificacion='Cobrar valor adicional antes del retiro.',
        )

        self.assertRedirects(response, reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}))
        ingreso.refresh_from_db()
        self.assertEqual(ingreso.diferencia, Decimal('0.01'))

        notificacion = NotificacionAsesora.objects.get(salida=salida)
        self.assertEqual(notificacion.tipo, NotificacionAsesora.TIPO_COBRO_ADICIONAL)
        self.assertEqual(notificacion.valor_acordado, Decimal('0.01'))

    def test_salida_negativa_sin_cobro_adicional_permanece_en_cero(self):
        response, ingreso, salida = self.registrar_ingreso_negativo(
            IngresoEquipoForm.ESTADO_NO_SE_PUDO_REPARAR,
            estado_reparacion='no_reparable',
            aplica_valor_acordado_adicional='no',
            valor_final_cobrado='20.00',
            metodo_pago_final='transferencia',
            banco='guayaquil',
        )

        self.assertRedirects(
            response,
            reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}),
        )
        ingreso.refresh_from_db()
        self.assertEqual(salida.aplica_valor_acordado_adicional, 'no')
        self.assertEqual(salida.valor_acordado_adicional, Decimal('0.00'))
        self.assertEqual(salida.valor_final_cobrado, Decimal('0.00'))
        self.assertEqual(salida.metodo_pago_final, 'sin_pago')
        self.assertEqual(ingreso.valor_efectivo_a_cobrar, Decimal('0.00'))
        self.assertEqual(ingreso.diferencia, Decimal('0.00'))
        self.assertFalse(NotificacionAsesora.objects.filter(salida=salida).exists())

    def test_salida_negativa_pago_mixto_completo_permite_retiro_y_conserva_resultado(self):
        response, ingreso, salida = self.registrar_ingreso_negativo(
            IngresoEquipoForm.ESTADO_NO_QUISO_REPARAR,
            estado_reparacion='cliente_no_acepta',
            aplica_valor_acordado_adicional='si',
            valor_acordado_adicional='12.00',
            motivo_valor_acordado_adicional='Diagnóstico avanzado autorizado.',
            valor_final_cobrado='12.00',
            metodo_pago_final='mixto',
            monto_1='5.00',
            metodo_1='efectivo',
            monto_2='7.00',
            metodo_2='transferencia',
            banco_2='guayaquil',
        )

        self.assertRedirects(
            response,
            reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}),
        )
        ingreso.refresh_from_db()
        self.assertEqual(salida.valor_final_cobrado, Decimal('12.00'))
        self.assertEqual(salida.metodo_pago_final, 'mixto')
        self.assertEqual(len(salida.pago_mixto_partes), 2)
        self.assertEqual(ingreso.diferencia, Decimal('0.00'))
        self.assertFalse(NotificacionAsesora.objects.filter(salida=salida).exists())

        retiro = self.client.post(
            reverse('econotec:salida_marcar_retirada', kwargs={'pk': salida.pk}),
        )
        self.assertRedirects(retiro, reverse('econotec:salida_retiros_lista'))
        salida.refresh_from_db()
        ingreso.refresh_from_db()
        self.assertEqual(salida.estado_reparacion, 'cliente_no_acepta')
        self.assertEqual(salida.fecha_retiro_real, date.today())
        self.assertEqual(ingreso.subestado_entregado, 'no_quiso_reparar')
        self.assertEqual(ingreso.diferencia, Decimal('0.00'))

    def test_salida_revision_usa_valor_acordado_propio_para_saldo(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('40.00'),
            abono_anticipo=Decimal('5.00'),
        )

        response = self.client.post(
            reverse('econotec:salida_registrar', kwargs={'ingreso_pk': ingreso.pk}),
            self.salida_post_data(
                estado_reparacion='revision',
                valor_acordado_revision='12.00',
                valor_final_cobrado='99.00',
                metodo_pago_final='efectivo',
                asesora_notificacion=str(self.vendedor.pk),
                mensaje_notificacion='Cobrar revisión antes del retiro.',
            ),
        )

        salida = SalidaEquipo.objects.get(ingreso=ingreso)
        self.assertRedirects(response, reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}))
        ingreso.refresh_from_db()
        salida.refresh_from_db()
        self.assertEqual(salida.estado_reparacion, 'revision')
        self.assertEqual(salida.valor_acordado_revision, Decimal('12.00'))
        self.assertEqual(salida.valor_final_cobrado, Decimal('0.00'))
        self.assertEqual(salida.metodo_pago_final, 'sin_pago')
        self.assertEqual(ingreso.valor_acordado, Decimal('40.00'))
        self.assertEqual(ingreso.valor_efectivo_a_cobrar, Decimal('12.00'))
        self.assertEqual(ingreso.diferencia, Decimal('7.00'))

        notificacion = NotificacionAsesora.objects.get(salida=salida)
        self.assertEqual(notificacion.tipo, NotificacionAsesora.TIPO_REVISION_PENDIENTE)
        self.assertEqual(notificacion.valor_acordado, Decimal('12.00'))

    def test_salida_revision_exige_valor_acordado_propio(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=Decimal('40.00'))

        response = self.client.post(
            reverse('econotec:salida_registrar', kwargs={'ingreso_pk': ingreso.pk}),
            self.salida_post_data(
                estado_reparacion='revision',
                valor_acordado_revision='',
                metodo_pago_final='sin_pago',
                asesora_notificacion=str(self.vendedor.pk),
            ),
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(SalidaEquipo.objects.filter(ingreso=ingreso).exists())
        self.assertContains(response, 'Ingresa el valor acordado por revisión')

    def test_salida_pendiente_retiro_con_saldo_notifica_asesora(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=Decimal('100.00'))

        response = self.client.post(
            reverse('econotec:salida_registrar', kwargs={'ingreso_pk': ingreso.pk}),
            self.salida_post_data(
                estado_reparacion='pendiente_retiro',
                valor_final_cobrado='20.00',
                metodo_pago_final='efectivo',
                asesora_notificacion=str(self.vendedor.pk),
                mensaje_notificacion='Equipo listo, falta saldo.',
            ),
        )

        salida = SalidaEquipo.objects.get(ingreso=ingreso)
        self.assertRedirects(response, reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}))
        ingreso.refresh_from_db()
        self.assertEqual(salida.valor_final_cobrado, Decimal('0.00'))
        self.assertEqual(salida.metodo_pago_final, 'sin_pago')
        self.assertEqual(ingreso.diferencia, Decimal('100.00'))

        notificacion = NotificacionAsesora.objects.get(salida=salida)
        self.assertEqual(notificacion.tipo, NotificacionAsesora.TIPO_SALDO_RETIRO)
        self.assertEqual(notificacion.valor_acordado, Decimal('100.00'))

    def test_salida_pendiente_retiro_pagada_no_crea_notificacion(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('25.00'),
            abono_anticipo=Decimal('25.00'),
        )

        response = self.client.post(
            reverse('econotec:salida_registrar', kwargs={'ingreso_pk': ingreso.pk}),
            self.salida_post_data(
                estado_reparacion='pendiente_retiro',
                valor_final_cobrado='25.00',
                metodo_pago_final='efectivo',
            ),
        )

        salida = SalidaEquipo.objects.get(ingreso=ingreso)
        self.assertRedirects(response, reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}))
        ingreso.refresh_from_db()
        self.assertEqual(salida.valor_final_cobrado, Decimal('0.00'))
        self.assertEqual(salida.metodo_pago_final, 'sin_pago')
        self.assertEqual(ingreso.diferencia, Decimal('0.00'))
        self.assertFalse(NotificacionAsesora.objects.filter(salida=salida).exists())

    def test_salida_form_valor_adicional_oculto_no_nace_con_min_invalidante(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=Decimal('20.00'))

        response = self.client.get(
            reverse('econotec:salida_registrar', kwargs={'ingreso_pk': ingreso.pk})
        )

        self.assertEqual(response.status_code, 200)
        html = response.content.decode()
        match = re.search(
            r'<input[^>]+name="valor_acordado_adicional"[^>]*>',
            html,
        )
        self.assertIsNotNone(match)
        self.assertNotIn('min="0.01"', match.group(0))

    def test_salida_pendiente_retiro_suma_valor_adicional_al_saldo(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=Decimal('20.00'))

        response = self.client.post(
            reverse('econotec:salida_registrar', kwargs={'ingreso_pk': ingreso.pk}),
            self.salida_post_data(
                aplica_valor_acordado_adicional='si',
                valor_acordado_adicional='0.10',
                motivo_valor_acordado_adicional='Repuesto adicional autorizado.',
                asesora_notificacion=str(self.vendedor.pk),
            ),
        )

        salida = SalidaEquipo.objects.get(ingreso=ingreso)
        self.assertRedirects(
            response,
            reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}),
        )
        ingreso.refresh_from_db()
        self.assertEqual(salida.aplica_valor_acordado_adicional, 'si')
        self.assertEqual(salida.valor_acordado_adicional, Decimal('0.10'))
        self.assertEqual(
            salida.motivo_valor_acordado_adicional,
            'Repuesto adicional autorizado.',
        )
        self.assertEqual(ingreso.valor_efectivo_a_cobrar, Decimal('20.10'))
        self.assertEqual(ingreso.diferencia, Decimal('20.10'))
        notificacion = NotificacionAsesora.objects.get(salida=salida)
        self.assertEqual(notificacion.valor_acordado, Decimal('20.10'))

    def test_valor_adicional_crea_saldo_si_valor_original_ya_esta_pagado(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('20.00'),
            abono_anticipo=Decimal('20.00'),
        )

        response = self.client.post(
            reverse('econotec:salida_registrar', kwargs={'ingreso_pk': ingreso.pk}),
            self.salida_post_data(
                aplica_valor_acordado_adicional='si',
                valor_acordado_adicional='0.01',
                motivo_valor_acordado_adicional='Material adicional.',
                asesora_notificacion=str(self.vendedor.pk),
            ),
        )

        salida = SalidaEquipo.objects.get(ingreso=ingreso)
        self.assertRedirects(
            response,
            reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}),
        )
        ingreso.refresh_from_db()
        self.assertEqual(ingreso.diferencia, Decimal('0.01'))

    def test_valor_adicional_rechaza_cero_y_exige_motivo(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=Decimal('20.00'))

        response = self.client.post(
            reverse('econotec:salida_registrar', kwargs={'ingreso_pk': ingreso.pk}),
            self.salida_post_data(
                aplica_valor_acordado_adicional='si',
                valor_acordado_adicional='0.00',
                motivo_valor_acordado_adicional='',
                asesora_notificacion=str(self.vendedor.pk),
            ),
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(SalidaEquipo.objects.filter(ingreso=ingreso).exists())
        self.assertContains(response, 'Ingresa un valor adicional de al menos $0.01.')
        self.assertContains(
            response,
            'Explica por qué se aplica el valor acordado adicional.',
        )

    def test_valor_adicional_no_aplica_en_estado_revision(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=Decimal('20.00'))

        response = self.client.post(
            reverse('econotec:salida_registrar', kwargs={'ingreso_pk': ingreso.pk}),
            self.salida_post_data(
                estado_reparacion='revision',
                valor_acordado_revision='10.00',
                aplica_valor_acordado_adicional='si',
                valor_acordado_adicional='10.00',
                motivo_valor_acordado_adicional='No debe conservarse.',
                metodo_pago_final='sin_pago',
                asesora_notificacion=str(self.vendedor.pk),
            ),
        )

        salida = SalidaEquipo.objects.get(ingreso=ingreso)
        self.assertRedirects(
            response,
            reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}),
        )
        self.assertEqual(salida.aplica_valor_acordado_adicional, 'no')
        self.assertEqual(salida.valor_acordado_adicional, Decimal('0.00'))
        self.assertEqual(salida.motivo_valor_acordado_adicional, '')

    def test_salida_retirado_oculta_y_limpia_cierre_economico(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('25.00'),
            abono_anticipo=Decimal('25.00'),
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 17),
            fecha_retiro_real=date(2026, 7, 18),
            estado_reparacion='retirado',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )

        response = self.client.get(reverse('econotec:salida_editar', kwargs={'pk': salida.pk}))
        self.assertContains(response, 'id="section-cierre-economico" style="display:none;"')
        self.assertContains(response, "v === 'retirado'")

        response = self.client.post(
            reverse('econotec:salida_editar', kwargs={'pk': salida.pk}),
            self.salida_post_data(
                fecha_salida='2026-07-22',
                estado_reparacion='retirado',
                tecnico_reparo=str(self.usuario.pk),
                valor_final_cobrado='99.00',
                metodo_pago_final='transferencia',
                numero_recibo='REC-RETIRADO',
                banco='pichincha',
                comprobante_url='https://example.com/comprobante',
            ),
        )

        self.assertRedirects(
            response,
            reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}),
        )
        salida.refresh_from_db()
        self.assertEqual(salida.valor_final_cobrado, Decimal('0.00'))
        self.assertEqual(salida.metodo_pago_final, 'sin_pago')
        self.assertEqual(salida.numero_recibo, '')
        self.assertEqual(salida.banco, '')
        self.assertEqual(salida.comprobante_url, '')

    def test_editar_finalizacion_cerrada_ignora_regreso_a_pendiente(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('25.00'),
            abono_anticipo=Decimal('25.00'),
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 8, 20),
            fecha_retiro_real=date(2026, 8, 21),
            estado_reparacion='retirado',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )

        response = self.client.post(
            reverse('econotec:salida_editar', kwargs={'pk': salida.pk}),
            self.salida_post_data(
                fecha_salida='2026-08-22',
                estado_reparacion='pendiente_retiro',
                tecnico_reparo=str(self.usuario.pk),
            ),
        )

        self.assertRedirects(
            response,
            reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}),
        )
        salida.refresh_from_db()
        ingreso.refresh_from_db()
        self.assertEqual(salida.estado_reparacion, 'retirado')
        self.assertEqual(salida.fecha_retiro_real, date(2026, 8, 21))
        self.assertEqual(ingreso.subestado_entregado, 'con_solucion')

    def test_salida_confirmada_heredada_no_aparece_como_pendiente(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('25.00'),
            abono_anticipo=Decimal('25.00'),
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 8, 20),
            estado_reparacion='pendiente_retiro',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        # Simula un registro anterior a la corrección: tenía fecha de retiro,
        # pero conservaba el código "pendiente_retiro".
        SalidaEquipo.objects.filter(pk=salida.pk).update(
            fecha_retiro_real=date(2026, 8, 21),
            estado_reparacion='pendiente_retiro',
        )
        ingreso.refresh_from_db()

        inicio = self.client.get(reverse('econotec:bienvenida'))
        pendientes = self.client.get(
            reverse('econotec:dashboard_details', kwargs={'tipo': 'pendientes'}),
        )
        detalle = self.client.get(
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk}),
        )

        self.assertEqual(inicio.context['stats']['pendientes_retiro'], 0)
        self.assertNotContains(pendientes, ingreso.codigo_equipo)
        self.assertEqual(ingreso.estado_visual_key, 'retirado')
        self.assertEqual(ingreso.estado_visual_display, 'Salió de la oficina')
        self.assertContains(detalle, 'Salió de la oficina')
        self.assertNotContains(detalle, 'Reparado — pendiente de retiro')

    def test_tecnico_no_puede_reabrir_equipo_que_ya_salio(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('25.00'),
            abono_anticipo=Decimal('25.00'),
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 8, 20),
            fecha_retiro_real=date(2026, 8, 21),
            estado_reparacion='retirado',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        token = token_para_ingreso(ingreso.pk)

        response = self.client.post(
            reverse('econotec:tecnico_hoja', kwargs={'token': token}),
            {
                'reporte_tecnico': 'Se agregó una observación posterior.',
                'estado_movil': 'en_reparacion',
                'subestado_reparacion': 'en_reparacion',
                'accion': 'guardar',
            },
        )

        self.assertRedirects(
            response,
            reverse('econotec:tecnico_hoja', kwargs={'token': token}),
        )
        ingreso.refresh_from_db()
        salida.refresh_from_db()
        self.assertEqual(ingreso.reporte_tecnico, 'Se agregó una observación posterior.')
        self.assertEqual(ingreso.estado, 'entregado')
        self.assertEqual(ingreso.subestado_entregado, 'con_solucion')
        self.assertEqual(salida.estado_reparacion, 'retirado')
        self.assertEqual(salida.fecha_retiro_real, date(2026, 8, 21))

    def test_aviso_salida_negativa_no_dice_equipo_reparado(self):
        for estado in ('no_reparable', 'cliente_no_acepta'):
            with self.subTest(estado=estado):
                ingreso = self.crear_ingreso_reparacion()
                salida = SalidaEquipo.objects.create(
                    ingreso=ingreso,
                    fecha_salida=date(2026, 7, 17),
                    estado_reparacion=estado,
                    tecnico_reparo=self.usuario,
                    valor_final_cobrado=Decimal('0.00'),
                    metodo_pago_final='sin_pago',
                    registrado_por=self.usuario,
                )

                response = self.client.get(reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}))

                self.assertContains(response, '¡Equipo finalizado y listo para retiro!')
                self.assertNotContains(response, '¡Equipo reparado y listo para retiro!')

    def test_whatsapp_retirado_usa_mensaje_de_cierre_sin_bodegaje(self):
        User = get_user_model()
        tecnico_ingreso = User.objects.create_user(
            username='TecnicoIngresoWhatsApp',
            first_name='Tecnico',
            last_name='Ingreso',
        )
        tecnico_reparo = User.objects.create_user(
            username='TecnicoReparoWhatsApp',
            first_name='Tecnico',
            last_name='Reparador',
        )
        ingreso = self.crear_ingreso_reparacion(
            tecnico_encargado=tecnico_ingreso,
            valor_acordado=Decimal('25.00'),
            abono_anticipo=Decimal('25.00'),
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 17),
            fecha_retiro_real=date(2026, 7, 18),
            estado_reparacion='retirado',
            tecnico_reparo=tecnico_reparo,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )

        link = whatsapp_link_equipo_listo(salida)
        texto = parse_qs(urlparse(link).query)['text'][0]

        self.assertIn('entregado y retirado satisfactoriamente', texto)
        self.assertIn('Gracias por confiar su equipo a Econotec', texto)
        self.assertIn('reparación de sus próximos equipos', texto)
        self.assertIn('Fecha de retiro: 18/07/2026', texto)
        self.assertIn('Técnico que reparó: Tecnico Reparador', texto)
        self.assertNotIn('Técnico encargado', texto)
        self.assertNotIn('Tecnico Ingreso', texto)
        self.assertNotIn('listo para retiro', texto)
        self.assertNotIn('coordine con nosotros', texto)
        self.assertNotIn('Política de bodegaje', texto)

    def test_notificacion_asesora_se_puede_marcar_como_vista(self):
        ingreso = self.crear_ingreso_reparacion(
            estado='garantia',
            valor_acordado=Decimal('60.00'),
            motivo_garantia='Garantía por retorno',
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 17),
            estado_reparacion='garantia_fallos_adicionales',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        notificacion = NotificacionAsesora.objects.create(
            salida=salida,
            ingreso=ingreso,
            asesora=self.vendedor,
            creado_por=self.usuario,
            valor_acordado=Decimal('60.00'),
            mensaje='Pendiente por cobrar.',
        )

        self.client.force_login(self.vendedor)
        response = self.client.get(reverse('econotec:notificaciones_asesora'))
        self.assertContains(response, ingreso.codigo_equipo)
        self.assertContains(response, 'Pendiente por cobrar.')

        response = self.client.post(
            reverse('econotec:notificacion_asesora_marcar_vista', kwargs={'pk': notificacion.pk})
        )
        self.assertRedirects(response, reverse('econotec:notificaciones_asesora'))
        notificacion.refresh_from_db()
        self.assertTrue(notificacion.leida)
        self.assertIsNotNone(notificacion.leida_en)

    def test_notificacion_asesora_muestra_hecho_si_saldo_esta_pagado(self):
        ingreso = self.crear_ingreso_reparacion(
            estado='garantia',
            valor_acordado=Decimal('60.00'),
            motivo_garantia='Garantía por retorno',
        )
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 17),
            estado_reparacion='garantia_fallos_adicionales',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        NotificacionAsesora.objects.create(
            salida=salida,
            ingreso=ingreso,
            asesora=self.vendedor,
            creado_por=self.usuario,
            valor_acordado=Decimal('60.00'),
            mensaje='Pendiente por cobrar.',
            leida=True,
        )
        ingreso.abonos.create(
            fecha=date(2026, 7, 17),
            monto=Decimal('60.00'),
            metodo='efectivo',
            registrado_por=self.vendedor,
        )

        self.client.force_login(self.vendedor)
        response = self.client.get(reverse('econotec:notificaciones_asesora'), {'estado': 'todas'})

        self.assertContains(response, 'Hecho')
        self.assertContains(response, 'noti-status-done')
        self.assertContains(response, 'noti-card hecha')
        self.assertContains(response, 'noti-valor pagado')

    def test_notificacion_asesora_limpiar_bandeja_borra_solo_sus_notificaciones(self):
        grupo_asesores = Group.objects.get(name='Asesores')
        otra_asesora = get_user_model().objects.create_user(
            username='OtraAsesora',
            email='otra@example.com',
        )
        otra_asesora.groups.add(grupo_asesores)

        ingreso_1 = self.crear_ingreso_reparacion(
            estado='garantia',
            valor_acordado=Decimal('60.00'),
            motivo_garantia='Garantía por retorno',
        )
        salida_1 = SalidaEquipo.objects.create(
            ingreso=ingreso_1,
            fecha_salida=date(2026, 7, 17),
            estado_reparacion='garantia_fallos_adicionales',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        NotificacionAsesora.objects.create(
            salida=salida_1,
            ingreso=ingreso_1,
            asesora=self.vendedor,
            creado_por=self.usuario,
            valor_acordado=Decimal('60.00'),
            mensaje='Pendiente por cobrar.',
        )

        ingreso_2 = self.crear_ingreso_reparacion(
            estado='garantia',
            valor_acordado=Decimal('40.00'),
            marca='Lenovo',
            motivo_garantia='Garantía por retorno',
        )
        salida_2 = SalidaEquipo.objects.create(
            ingreso=ingreso_2,
            fecha_salida=date(2026, 7, 17),
            estado_reparacion='garantia_fallos_adicionales',
            tecnico_reparo=self.usuario,
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )
        notificacion_otra = NotificacionAsesora.objects.create(
            salida=salida_2,
            ingreso=ingreso_2,
            asesora=otra_asesora,
            creado_por=self.usuario,
            valor_acordado=Decimal('40.00'),
            mensaje='Pendiente de otra asesora.',
        )

        self.client.force_login(self.vendedor)
        response = self.client.post(reverse('econotec:notificacion_asesora_limpiar_bandeja'))

        self.assertRedirects(response, reverse('econotec:notificaciones_asesora'))
        self.assertFalse(NotificacionAsesora.objects.filter(asesora=self.vendedor).exists())
        self.assertTrue(NotificacionAsesora.objects.filter(pk=notificacion_otra.pk).exists())

    def test_admin_ve_todas_las_notificaciones_de_asesoras_y_filtra_por_asesora(self):
        grupo_asesores = Group.objects.get(name='Asesores')
        otra_asesora = get_user_model().objects.create_user(
            username='OtraAsesora',
            email='otra@example.com',
        )
        otra_asesora.groups.add(grupo_asesores)
        self.crear_notificacion_asesora(self.vendedor, 'Pendiente de Kimberly.')
        self.crear_notificacion_asesora(
            otra_asesora,
            'Pendiente de otra asesora.',
            marca='Lenovo',
            valor_acordado=Decimal('40.00'),
        )

        self.client.force_login(self.admin)
        response = self.client.get(reverse('econotec:notificaciones_asesora'))

        self.assertContains(response, 'Control de notificaciones de asesoras')
        self.assertContains(response, 'Pendiente de Kimberly.')
        self.assertContains(response, 'Pendiente de otra asesora.')
        self.assertContains(response, 'Responder / gestionar pago')

        response = self.client.get(
            reverse('econotec:notificaciones_asesora'),
            {'asesora': str(otra_asesora.pk), 'estado': 'todas'},
        )

        self.assertNotContains(response, 'Pendiente de Kimberly.')
        self.assertContains(response, 'Pendiente de otra asesora.')

    def test_admin_puede_marcar_notificacion_de_asesora_como_gestionada(self):
        notificacion = self.crear_notificacion_asesora(
            self.vendedor,
            'Gestionar desde admin.',
        )

        self.client.force_login(self.admin)
        next_url = reverse('econotec:notificaciones_asesora') + '?estado=todas'
        response = self.client.post(
            reverse('econotec:notificacion_asesora_marcar_vista', kwargs={'pk': notificacion.pk}),
            {'next': next_url},
        )

        self.assertRedirects(response, next_url)
        notificacion.refresh_from_db()
        self.assertTrue(notificacion.leida)
        self.assertIsNotNone(notificacion.leida_en)

    def test_admin_ve_acceso_a_notificaciones_asesoras_en_inicio(self):
        self.crear_notificacion_asesora(self.vendedor, 'Pendiente visible para admin.')
        self.client.force_login(self.admin)

        response = self.client.get(reverse('econotec:bienvenida'))

        self.assertContains(response, 'Notificaciones asesoras')
        self.assertContains(response, 'Asesoras')
        self.assertContains(response, 'Pendientes: 1')

    def test_detalle_muestra_alerta_si_valor_acordado_pendiente(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=None)

        response = self.client.get(
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk})
        )

        self.assertContains(response, 'pending-value-alert')
        self.assertContains(response, 'Pendiente de valor acordado')

    def test_menu_muestra_apartado_pendientes_de_valor_acordado(self):
        self.crear_ingreso_reparacion(valor_acordado=None)
        self.crear_ingreso_reparacion(
            valor_acordado=Decimal('45.00'),
            marca='Dell',
            modelo_serie='Inspiron',
        )

        response = self.client.get(reverse('econotec:ingreso_menu'))

        self.assertContains(response, 'Pendiente de Valores Acordados')
        self.assertContains(response, 'Lista de equipos sin valor acordado.')
        self.assertContains(response, '?sede=todas&valor=pendiente')
        self.assertContains(response, '1 pendiente')
        self.assertNotContains(response, 'Buscar equipos por fecha')
        self.assertNotContains(response, 'Filtrar equipos')

    def test_menu_muestra_donados_compras_a_tecnico_en_ruta_general(self):
        response = self.client.get(reverse('econotec:ingreso_menu'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Donados / Equipos a comprar')
        self.assertContains(
            response,
            reverse('econotec:equipos_administrativos_general'),
        )
        self.assertNotContains(
            response,
            f'href="{reverse("econotec:admin_equipos_administrativos")}"',
        )

    def test_menu_admin_conserva_enlace_a_bandeja_administrativa(self):
        self.client.force_login(self.admin)

        response = self.client.get(reverse('econotec:ingreso_menu'))

        self.assertContains(
            response,
            reverse('econotec:admin_equipos_administrativos'),
        )
        self.assertNotContains(
            response,
            f'href="{reverse("econotec:equipos_administrativos_general")}"',
        )

    def test_asesora_puede_abrir_vista_general_pero_no_bandeja_admin(self):
        self.client.force_login(self.vendedor)

        general = self.client.get(
            reverse('econotec:equipos_administrativos_general')
        )
        admin = self.client.get(
            reverse('econotec:admin_equipos_administrativos')
        )

        self.assertEqual(general.status_code, 200)
        self.assertRedirects(admin, reverse('econotec:bienvenida'))

    def test_vista_general_donados_compras_oculta_resumen_y_egreso_interno(self):
        compra = self.crear_ingreso_reparacion(
            estado='equipo_a_comprar',
            subestado_reparacion='',
            marca='Lenovo Compra Privada',
            modelo_serie='ThinkPad Compra',
            valor_acordado=Decimal('350.00'),
            compra_metodo_pago='transferencia',
            compra_banco='pichincha',
        )
        donado = self.crear_ingreso_reparacion(
            estado='donado',
            subestado_reparacion='',
            marca='Sony Donado Privado',
            modelo_serie='PlayStation Donada',
            valor_acordado=Decimal('0.00'),
        )

        response = self.client.get(
            reverse('econotec:equipos_administrativos_general')
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, compra.codigo_equipo)
        self.assertContains(response, donado.codigo_equipo)
        self.assertContains(response, 'Ver información completa del equipo')
        self.assertNotContains(response, 'Total en bandeja')
        self.assertNotContains(response, 'Equipos comprados')
        self.assertNotContains(response, 'Total pagado')
        self.assertNotContains(response, 'Egreso automático vinculado')
        self.assertNotContains(response, 'Ver / editar egreso')
        self.assertNotContains(response, 'Editar ingreso')

    def test_detalle_compra_oculta_egreso_a_tecnico_y_lo_conserva_para_admin(self):
        compra = self.crear_ingreso_reparacion(
            estado='equipo_a_comprar',
            subestado_reparacion='',
            valor_acordado=Decimal('90.00'),
            compra_metodo_pago='efectivo',
        )

        tecnico = self.client.get(
            reverse('econotec:ingreso_detalle', kwargs={'pk': compra.pk})
        )
        self.assertNotContains(tecnico, 'Egreso automático')
        self.assertNotContains(tecnico, 'Ver egreso')
        self.assertContains(
            tecnico,
            reverse('econotec:equipos_administrativos_general'),
        )

        self.client.force_login(self.admin)
        admin = self.client.get(
            reverse('econotec:ingreso_detalle', kwargs={'pk': compra.pk})
        )
        self.assertContains(admin, 'Egreso automático')
        self.assertContains(
            admin,
            reverse('econotec:admin_equipos_administrativos'),
        )

    def test_admin_conserva_resumen_completo_y_redirige_desde_vista_general(self):
        self.crear_ingreso_reparacion(
            estado='equipo_a_comprar',
            subestado_reparacion='',
            valor_acordado=Decimal('125.00'),
            compra_metodo_pago='efectivo',
        )
        self.client.force_login(self.admin)

        redireccion = self.client.get(
            reverse('econotec:equipos_administrativos_general')
        )
        self.assertRedirects(
            redireccion,
            reverse('econotec:admin_equipos_administrativos'),
        )

        response = self.client.get(
            reverse('econotec:admin_equipos_administrativos')
        )
        self.assertContains(response, 'Total en bandeja')
        self.assertContains(response, 'Equipos comprados')
        self.assertContains(response, 'Total pagado')
        self.assertContains(response, 'Ver toda la información administrativa')

    def test_lista_general_excluye_donados_y_equipos_a_comprar(self):
        normal = self.crear_ingreso_reparacion(
            marca='HP Operativo',
            modelo_serie='Equipo normal',
        )
        compra = self.crear_ingreso_reparacion(
            estado='equipo_a_comprar',
            subestado_reparacion='',
            marca='Lenovo Administrativo',
            modelo_serie='Equipo comprado',
        )
        donado = self.crear_ingreso_reparacion(
            estado='donado',
            subestado_reparacion='',
            marca='Sony Administrativo',
            modelo_serie='Equipo donado',
            valor_acordado=Decimal('0.00'),
        )

        response = self.client.get(
            reverse('econotec:ingreso_lista'),
            {'sede': 'todas'},
        )

        self.assertEqual(response.context['total'], 1)
        self.assertEqual(list(response.context['ingresos']), [normal])
        self.assertContains(response, normal.codigo_equipo)
        self.assertNotContains(response, compra.codigo_equipo)
        self.assertNotContains(response, donado.codigo_equipo)
        self.assertNotContains(response, 'Donado — pasa a Administrativo')
        self.assertNotContains(response, 'Equipo a comprar — pasa a Administrativo')

    def test_lista_filtra_valor_acordado_pendiente(self):
        pendiente = self.crear_ingreso_reparacion(valor_acordado=None)
        con_valor = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('80.00'),
            marca='Lenovo',
            modelo_serie='ThinkPad',
        )

        response = self.client.get(
            reverse('econotec:ingreso_lista'),
            {'sede': 'todas', 'valor': 'pendiente'},
        )

        ingresos = list(response.context['ingresos'])
        self.assertEqual(ingresos, [pendiente])
        self.assertNotIn(con_valor, ingresos)
        self.assertContains(response, 'Equipos <span class="accent">Pendientes de Valor</span>')
        self.assertContains(response, 'value="pendiente" selected')
        self.assertContains(response, 'Pendiente valor acordado')

    def test_hoja_tecnico_muestra_valor_acordado_y_registro_salida(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=None)
        token = token_para_ingreso(ingreso.pk)

        response = self.client.get(reverse('econotec:tecnico_hoja', kwargs={'token': token}))

        self.assertContains(response, 'name="valor_acordado"')
        self.assertContains(response, 'Valor acordado')
        self.assertContains(response, 'id="valor-toggle"')
        self.assertContains(response, 'aria-expanded="false"')
        self.assertContains(response, 'id="valor-content" hidden')
        self.assertContains(response, 'readonly')
        self.assertContains(response, 'solo lectura')
        self.assertNotContains(response, 'Actualizar valor')
        self.assertContains(response, 'name="valor_pendiente_reporte"')
        self.assertContains(response, 'Reportar por qué está pendiente el valor acordado')
        self.assertContains(response, 'Registrar equipo listo / finalizado')
        self.assertContains(response, 'id="btn-perfil-movil"')
        self.assertContains(response, 'Ver perfil')
        self.assertContains(response, 'id="perfil-mobile-modal"')
        self.assertContains(response, 'id="btn-bitacora-mobile"')
        self.assertContains(response, 'id="bitacora-mobile-modal"')

    def test_hoja_tecnico_reporta_motivo_valor_acordado_pendiente(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=None)
        token = token_para_ingreso(ingreso.pk)

        response = self.client.post(
            reverse('econotec:tecnico_hoja', kwargs={'token': token}),
            {
                'reporte_tecnico': 'Equipo sigue en diagnostico',
                'valor_pendiente_reporte': 'Pendiente confirmar repuesto con proveedor.',
                'estado_movil': 'en_reparacion',
                'subestado_reparacion': '',
                'accion': 'reportar_valor_pendiente',
            },
        )

        self.assertRedirects(response, reverse('econotec:tecnico_hoja', kwargs={'token': token}))
        ingreso.refresh_from_db()
        self.assertEqual(ingreso.valor_pendiente_reporte, 'Pendiente confirmar repuesto con proveedor.')
        self.assertEqual(ingreso.valor_pendiente_reporte_por, self.usuario)
        self.assertIsNotNone(ingreso.valor_pendiente_reporte_actualizado)
        self.assertEqual(ingreso.reporte_tecnico, 'Equipo sigue en diagnostico')

        response = self.client.get(
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk})
        )
        self.assertContains(response, 'Pendiente de valor acordado')
        self.assertContains(response, 'Reporte del técnico')
        self.assertContains(response, 'Ver reporte del técnico')
        self.assertContains(response, 'Pendiente confirmar repuesto con proveedor.')

    def test_hoja_tecnico_no_actualiza_valor_acordado_desde_movil(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=None,
            abono_anticipo=Decimal('5.00'),
        )
        token = token_para_ingreso(ingreso.pk)

        response = self.client.post(
            reverse('econotec:tecnico_hoja', kwargs={'token': token}),
            {
                'reporte_tecnico': 'Valor confirmado con el cliente',
                'valor_acordado': '100',
                'estado_movil': 'en_reparacion',
                'subestado_reparacion': '',
                'accion': 'actualizar_valor',
            },
        )

        self.assertRedirects(response, reverse('econotec:tecnico_hoja', kwargs={'token': token}))
        ingreso.refresh_from_db()
        self.assertIsNone(ingreso.valor_acordado)
        self.assertEqual(ingreso.estado_pago, 'Pendiente')

        response = self.client.get(reverse('econotec:tecnico_hoja', kwargs={'token': token}))
        self.assertContains(response, 'Sin valor acordado registrado')
        self.assertNotContains(response, 'Valor acordado total')

        response = self.client.get(
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk})
        )
        self.assertContains(response, 'btn-salida-bloqueada')

    def test_hoja_tecnico_bloquea_valor_si_esta_pagado_completo(self):
        ingreso = self.crear_ingreso_reparacion(
            valor_acordado=Decimal('5.00'),
            abono_anticipo=Decimal('5.00'),
        )
        token = token_para_ingreso(ingreso.pk)

        response = self.client.get(reverse('econotec:tecnico_hoja', kwargs={'token': token}))

        self.assertContains(response, 'Ya está pagado todo')
        self.assertContains(response, 'readonly')
        self.assertContains(response, 'Pagado')
        self.assertNotContains(response, 'Actualizar valor')

        response = self.client.post(
            reverse('econotec:tecnico_hoja', kwargs={'token': token}),
            {
                'reporte_tecnico': 'Se intentó cambiar el valor pagado',
                'valor_acordado': '200',
                'estado_movil': 'en_reparacion',
                'subestado_reparacion': '',
                'accion': 'actualizar_valor',
            },
        )

        self.assertRedirects(response, reverse('econotec:tecnico_hoja', kwargs={'token': token}))
        ingreso.refresh_from_db()
        self.assertEqual(ingreso.valor_acordado, Decimal('5.00'))

        response = self.client.post(
            reverse('econotec:tecnico_hoja', kwargs={'token': token}),
            {
                'reporte_tecnico': 'Listo para salida',
                'valor_acordado': '200',
                'estado_movil': 'en_reparacion',
                'subestado_reparacion': '',
                'accion': 'registrar_salida',
            },
        )

        self.assertRedirects(
            response,
            reverse('econotec:salida_registrar', kwargs={'ingreso_pk': ingreso.pk})
        )
        ingreso.refresh_from_db()
        self.assertEqual(ingreso.valor_acordado, Decimal('5.00'))
        self.assertFalse(SalidaEquipo.objects.filter(ingreso=ingreso).exists())

    def test_hoja_tecnico_no_registra_salida_con_valor_pendiente(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=None)
        token = token_para_ingreso(ingreso.pk)

        response = self.client.post(
            reverse('econotec:tecnico_hoja', kwargs={'token': token}),
            {
                'reporte_tecnico': 'Listo para salir',
                'valor_acordado': '100',
                'estado_movil': 'en_reparacion',
                'subestado_reparacion': '',
                'accion': 'registrar_salida',
            },
        )

        self.assertRedirects(response, reverse('econotec:tecnico_hoja', kwargs={'token': token}))
        ingreso.refresh_from_db()
        self.assertIsNone(ingreso.valor_acordado)
        self.assertEqual(ingreso.reporte_tecnico, 'Listo para salir')
        self.assertFalse(SalidaEquipo.objects.filter(ingreso=ingreso).exists())

    def test_hoja_tecnico_redirige_a_registrar_salida_con_valor_real_existente(self):
        ingreso = self.crear_ingreso_reparacion(valor_acordado=Decimal('100.00'))
        token = token_para_ingreso(ingreso.pk)

        response = self.client.post(
            reverse('econotec:tecnico_hoja', kwargs={'token': token}),
            {
                'reporte_tecnico': 'Reparado y probado',
                'valor_acordado': '100',
                'estado_movil': 'en_reparacion',
                'subestado_reparacion': '',
                'accion': 'registrar_salida',
            },
        )

        self.assertRedirects(
            response,
            reverse('econotec:salida_registrar', kwargs={'ingreso_pk': ingreso.pk})
        )
        ingreso.refresh_from_db()
        self.assertEqual(ingreso.valor_acordado, Decimal('100.00'))
        self.assertFalse(SalidaEquipo.objects.filter(ingreso=ingreso).exists())

    def test_lista_ingresos_oculta_boton_detalle_salida_a_no_admin(self):
        ingreso = self.crear_ingreso_reparacion(estado='entregado')
        SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='pendiente_retiro',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=self.usuario,
        )

        response = self.client.get(
            reverse('econotec:ingreso_lista'),
            {'sede': 'todas'},
        )

        self.assertNotContains(response, 'Ver finalización')

    def test_lista_ingresos_muestra_boton_detalle_salida_solo_admin(self):
        User = get_user_model()
        admin = User.objects.create_superuser(
            username='Admin',
            email='admin@example.com',
            password='testpass123',
        )
        self.client.force_login(admin)

        ingreso = self.crear_ingreso_reparacion(estado='entregado')
        salida = SalidaEquipo.objects.create(
            ingreso=ingreso,
            fecha_salida=date(2026, 7, 9),
            estado_reparacion='pendiente_retiro',
            cliente_recibe_conforme='si',
            valor_final_cobrado=Decimal('0.00'),
            metodo_pago_final='sin_pago',
            registrado_por=admin,
        )

        response = self.client.get(
            reverse('econotec:ingreso_lista'),
            {'sede': 'todas'},
        )

        self.assertContains(response, 'Ver finalización')
        self.assertContains(
            response,
            reverse('econotec:salida_editar', kwargs={'pk': salida.pk})
        )

    def test_lista_ingresos_muestra_y_filtra_garantia_de_ingreso(self):
        self.activar_sede_guayaquil()
        ingreso_garantia = self.crear_ingreso_reparacion(
            estado='garantia',
            marca='Epson',
            modelo_serie='L3250 Garantia',
            motivo_garantia='Garantia de ingreso',
        )
        self.crear_ingreso_reparacion(
            marca='HP',
            modelo_serie='Elitebook',
            estado='en_reparacion',
        )

        response = self.client.get(
            reverse('econotec:ingreso_lista'),
            {'estado': 'garantia'},
        )

        self.assertContains(response, 'Garantía (Ingreso)')
        self.assertContains(response, ingreso_garantia.codigo_equipo)
        self.assertContains(response, 'value="garantia" selected')
        self.assertEqual(response.context['total'], 1)

    def test_detalle_garantia_muestra_equipo_manual(self):
        ingreso = self.crear_ingreso_reparacion(
            estado='garantia',
            equipo_garantia_manual='G980',
            motivo_garantia='Garantia por revision',
        )

        response = self.client.get(
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk})
        )

        self.assertContains(response, 'Garantía de G980')

    def test_detalle_garantia_muestra_equipo_anterior_seleccionado(self):
        equipo_anterior = self.crear_ingreso_reparacion(
            marca='Epson',
            modelo_serie='L3250',
        )
        ingreso = self.crear_ingreso_reparacion(
            marca='Epson',
            modelo_serie='L3250 Garantia',
            estado='garantia',
            equipo_garantia=equipo_anterior,
            motivo_garantia='Garantia por equipo anterior',
        )

        response = self.client.get(
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk})
        )

        self.assertContains(response, f'Garantía de {equipo_anterior.codigo_equipo}')

    def test_editar_mismo_equipo_no_se_detecta_como_duplicado(self):
        ingreso = self.crear_ingreso_reparacion(
            marca='MacBook M4 S',
            modelo_serie='MacBook M4 S',
        )

        response = self.client.post(
            reverse('econotec:ingreso_editar', kwargs={'pk': ingreso.pk}),
            self.ingreso_edit_post_data(ingreso),
        )

        self.assertRedirects(
            response,
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk})
        )

    def test_editar_reingreso_igual_sin_cambiar_identidad_no_bloquea(self):
        ingreso = self.crear_ingreso_reparacion(
            marca='Sony',
            modelo_serie='Playstation 5',
        )
        self.crear_ingreso_reparacion(
            marca='Sony',
            modelo_serie='Playstation 5',
            problema_reportado='Reingreso anterior',
        )

        response = self.client.post(
            reverse('econotec:ingreso_editar', kwargs={'pk': ingreso.pk}),
            self.ingreso_edit_post_data(
                ingreso,
                **{'ing-problema_reportado': 'Solo actualizo el reporte'}
            ),
        )

        self.assertRedirects(
            response,
            reverse('econotec:ingreso_detalle', kwargs={'pk': ingreso.pk})
        )
        ingreso.refresh_from_db()
        self.assertEqual(ingreso.problema_reportado, 'Solo actualizo el reporte')

    def test_editar_a_otro_equipo_igual_si_se_detecta_como_duplicado(self):
        duplicado = self.crear_ingreso_reparacion(
            marca='MacBook M4 S',
            modelo_serie='MacBook M4 S',
        )
        ingreso = self.crear_ingreso_reparacion(
            marca='HP',
            modelo_serie='Elitebook',
        )

        response = self.client.post(
            reverse('econotec:ingreso_editar', kwargs={'pk': ingreso.pk}),
            self.ingreso_edit_post_data(
                ingreso,
                **{
                    'ing-marca': duplicado.marca,
                    'ing-modelo_serie': duplicado.modelo_serie,
                }
            ),
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('modelo_serie', response.context['ing_form'].errors)

    def test_registrar_mismo_equipo_y_cliente_sin_confirmacion_bloquea(self):
        self.activar_sede_guayaquil()
        self.crear_ingreso_reparacion(
            marca='MacBook M4 S',
            modelo_serie='MacBook M4 S',
        )

        response = self.client.post(
            reverse('econotec:ingreso_registrar'),
            self.ingreso_registro_post_data(),
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(IngresoEquipo.objects.filter(cliente=self.cliente_existente).count(), 1)
        self.assertIn('modelo_serie', response.context['ing_form'].errors)

    def test_registrar_mismo_modelo_ignora_mayusculas_y_tildes(self):
        self.activar_sede_guayaquil()
        self.crear_ingreso_reparacion(
            marca='Canon',
            modelo_serie='Cámara Pró',
            serie='ABC-001',
        )

        response = self.client.post(
            reverse('econotec:ingreso_registrar'),
            self.ingreso_registro_post_data(
                **{
                    'ing-marca': 'Canon',
                    'ing-modelo_serie': 'CAMARA PRO',
                    'ing-serie': 'XYZ-999',
                }
            ),
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(IngresoEquipo.objects.filter(cliente=self.cliente_existente).count(), 1)
        self.assertIn('modelo_serie', response.context['ing_form'].errors)

    def test_registrar_mismo_equipo_y_cliente_confirmado_crea_reingreso(self):
        self.activar_sede_guayaquil()
        ingreso_anterior = self.crear_ingreso_reparacion(
            marca='MacBook M4 S',
            modelo_serie='MacBook M4 S',
        )

        response = self.client.post(
            reverse('econotec:ingreso_registrar'),
            self.ingreso_registro_post_data(
                **{'confirmar_mismo_equipo_cliente': '1'}
            ),
        )

        nuevo_ingreso = IngresoEquipo.objects.exclude(pk=ingreso_anterior.pk).get()
        self.assertRedirects(
            response,
            reverse('econotec:ingreso_detalle', kwargs={'pk': nuevo_ingreso.pk})
        )
        self.assertEqual(nuevo_ingreso.cliente, self.cliente_existente)
        self.assertEqual(nuevo_ingreso.marca, 'MacBook M4 S')

    def test_nueva_solicitud_no_restaura_borrador_localstorage(self):
        self.activar_sede_guayaquil()

        response = self.client.get(reverse('econotec:ingreso_registrar'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="confirmar_mismo_equipo_cliente"')
        self.assertContains(response, 'name="ing-valor_acordado_estado"')
        self.assertContains(response, '¿El técnico ya tiene el valor acordado?')
        self.assertContains(response, 'No / pendiente de valor')
        self.assertContains(response, "localStorage.removeItem('econotec_ingreso_form_nuevo')")
        self.assertNotContains(response, "localStorage.getItem('econotec_ingreso_form_nuevo')")

    def test_nueva_solicitud_muestra_resultados_negativos_y_cobro_opcional(self):
        self.activar_sede_guayaquil()

        response = self.client.get(reverse('econotec:ingreso_registrar'))

        opciones = dict(response.context['ing_form'].fields['estado'].choices)
        self.assertNotIn('entregado', opciones)
        self.assertEqual(
            opciones[IngresoEquipoForm.ESTADO_NO_QUISO_REPARAR],
            'Cliente no quiso reparar',
        )
        self.assertEqual(
            opciones[IngresoEquipoForm.ESTADO_NO_SE_PUDO_REPARAR],
            'No se pudo reparar',
        )
        self.assertContains(response, 'id="finalizacion-rapida"')
        self.assertContains(response, 'Finalización y cobro del equipo')
        self.assertContains(response, 'Cobro adicional opcional')
        self.assertNotContains(response, 'Resultado final del equipo')
        self.assertNotContains(response, '>Equipo finalizado<')

    def test_nueva_solicitud_no_permite_equipo_finalizado(self):
        self.activar_sede_guayaquil()
        cantidad_antes = IngresoEquipo.objects.count()
        data = self.ingreso_registro_post_data(**{
            'ing-estado': 'entregado',
            'ing-subestado_reparacion': '',
            'ing-subestado_entregado': '',
            'ing-valor_acordado_estado': 'si',
            'ing-valor_acordado': '25.00',
        })
        data.update(self.salida_rapida_post_data())

        response = self.client.post(reverse('econotec:ingreso_registrar'), data)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(IngresoEquipo.objects.count(), cantidad_antes)
        self.assertIn('estado', response.context['ing_form'].errors)

    def test_nueva_solicitud_finaliza_sin_solucion_sin_cobro_por_defecto(self):
        self.activar_sede_guayaquil()
        data = self.ingreso_registro_post_data(**{
            'ing-estado': IngresoEquipoForm.ESTADO_NO_SE_PUDO_REPARAR,
            'ing-subestado_reparacion': '',
            'ing-subestado_entregado': '',
            'ing-valor_acordado_estado': 'no',
            'ing-valor_acordado': '',
        })
        data.update(self.salida_rapida_post_data(
            estado_reparacion='no_reparable',
            aplica_valor_acordado_adicional='no',
            valor_acordado_adicional='0.00',
            valor_final_cobrado='0.00',
            metodo_pago_final='sin_pago',
        ))

        response = self.client.post(reverse('econotec:ingreso_registrar'), data)

        ingreso = IngresoEquipo.objects.get(cliente=self.cliente_existente)
        salida = SalidaEquipo.objects.get(ingreso=ingreso)
        self.assertRedirects(
            response,
            reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}),
        )
        self.assertEqual(ingreso.estado, 'entregado')
        self.assertEqual(ingreso.subestado_entregado, 'sin_solucion')
        self.assertEqual(salida.estado_reparacion, 'no_reparable')
        self.assertEqual(ingreso.valor_efectivo_a_cobrar, Decimal('0.00'))
        self.assertEqual(ingreso.diferencia, Decimal('0.00'))
        self.assertFalse(NotificacionAsesora.objects.filter(salida=salida).exists())

    def test_nueva_solicitud_cobro_mixto_parcial_bloquea_retiro(self):
        self.activar_sede_guayaquil()
        data = self.ingreso_registro_post_data(**{
            'ing-estado': IngresoEquipoForm.ESTADO_NO_QUISO_REPARAR,
            'ing-subestado_reparacion': '',
            'ing-subestado_entregado': '',
            'ing-valor_acordado_estado': 'no',
            'ing-valor_acordado': '',
        })
        data.update(self.salida_rapida_post_data(
            estado_reparacion='cliente_no_acepta',
            aplica_valor_acordado_adicional='si',
            valor_acordado_adicional='12.00',
            motivo_valor_acordado_adicional='Revisión avanzada autorizada.',
            valor_final_cobrado='8.00',
            metodo_pago_final='mixto',
            monto_1='5.00',
            metodo_1='efectivo',
            monto_2='3.00',
            metodo_2='transferencia',
            banco_2='guayaquil',
            asesora_notificacion=str(self.vendedor.pk),
        ))

        response = self.client.post(reverse('econotec:ingreso_registrar'), data)

        ingreso = IngresoEquipo.objects.get(cliente=self.cliente_existente)
        salida = SalidaEquipo.objects.get(ingreso=ingreso)
        self.assertRedirects(
            response,
            reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}),
        )
        self.assertEqual(salida.estado_reparacion, 'cliente_no_acepta')
        self.assertEqual(salida.metodo_pago_final, 'mixto')
        self.assertEqual(salida.valor_final_cobrado, Decimal('8.00'))
        self.assertEqual(ingreso.valor_efectivo_a_cobrar, Decimal('12.00'))
        self.assertEqual(ingreso.diferencia, Decimal('4.00'))
        notificacion = NotificacionAsesora.objects.get(salida=salida)
        self.assertEqual(notificacion.tipo, NotificacionAsesora.TIPO_COBRO_ADICIONAL)
        self.assertEqual(notificacion.valor_acordado, Decimal('4.00'))

        retiro = self.client.post(
            reverse('econotec:salida_marcar_retirada', kwargs={'pk': salida.pk}),
        )
        self.assertRedirects(
            retiro,
            reverse('econotec:salida_listo_aviso', kwargs={'pk': salida.pk}),
        )
        salida.refresh_from_db()
        self.assertIsNone(salida.fecha_retiro_real)

    def test_lista_filtra_subestado_en_reparacion_simple(self):
        ingreso_reparacion = self.crear_ingreso_reparacion(
            marca='MacBook M4 S',
            subestado_reparacion='en_reparacion',
        )
        self.crear_ingreso_reparacion(
            marca='HP',
            subestado_reparacion='espera_repuesto',
        )

        response = self.client.get(
            reverse('econotec:ingreso_lista'),
            {'estado': 'reparacion_en_reparacion', 'sede': 'todas'},
        )

        ingresos = list(response.context['ingresos'])
        self.assertEqual(ingresos, [ingreso_reparacion])
        self.assertContains(response, '↳ En reparación')
