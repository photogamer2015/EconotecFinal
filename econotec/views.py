"""
Vistas principales de Econotec.
Maneja: bienvenida, ayuda, ingresos de equipos, salidas y clientes.
"""
from datetime import date, timedelta
from decimal import Decimal as D
from io import BytesIO
import json
import unicodedata

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db import transaction
from django.db.models import Count, Prefetch, Q, Sum
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.html import format_html
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_GET, require_POST

from .forms import (
    ClienteForm, CobroBodegajeForm, IngresoEquipoForm, SalidaEquipoForm,
    InventarioItemForm,
)
from .busqueda import (
    filtrar_objetos_normalizado,
    texto_cliente_busqueda,
    texto_ingreso_busqueda,
    texto_salida_busqueda,
    total_resultados,
)
from .models import (
    Cliente, IngresoEquipo, SalidaEquipo, Abono, SEDES_EQUIPOS,
    UsuarioActividad, NotificacionAsesora, BitacoraTecnico, InventarioItem,
    NotificacionInventarioAdmin, VentaInventarioItem, Egreso, CategoriaEgreso,
)
from .bitacora import registrar_bitacora, nombre_corto_usuario, construir_bitacora_usuario
from .date_filters import aplicar_rango_fecha, contexto_rango_fecha, obtener_rango_fecha
from .emails import (
    enviar_correo_finalizacion_seguro,
    enviar_correo_ingreso_seguro,
    enviar_correo_salida_fisica_seguro,
)
from .pagination import paginar_resultados
from .permisos import admin_requerido, tecnico_requerido, es_admin, es_asesor, es_tecnico
from .gamificacion import (
    SALIDA_BUENA_ESTADOS,
    SALIDA_GARANTIA_ESTADOS,
    SALIDA_MALA_ESTADOS,
    calcular_puntaje_gamificacion,
)
from .alertas import (
    equipos_demorados_qs,
    salidas_bodegaje_qs,
    dias_en_taller,
    dias_desde_salida,
    whatsapp_link_demora,
    whatsapp_link_equipo_listo,
    whatsapp_link_bodegaje,
    whatsapp_link_hoja_ingreso,
    whatsapp_link_venta_producto,
    UMBRAL_DIAS_DIAGNOSTICO,
    UMBRAL_DIAS_BODEGAJE,
    COSTO_BODEGAJE_DIA,
)


def _sincronizar_notificacion_asesora(form, salida, user):
    """Crea/actualiza la notificación interna para cobros pendientes de asesora."""
    tipos_controlados = [
        NotificacionAsesora.TIPO_FALLOS_ADICIONALES,
        NotificacionAsesora.TIPO_REVISION_PENDIENTE,
        NotificacionAsesora.TIPO_SALDO_RETIRO,
    ]
    tipo = getattr(form, 'notificacion_asesora_tipo', None)
    valor = getattr(form, 'notificacion_asesora_valor', D('0.00')) or D('0.00')

    NotificacionAsesora.objects.filter(
        salida=salida,
        tipo__in=[tipo_actual for tipo_actual in tipos_controlados if tipo_actual != tipo],
    ).delete()

    if not tipo or valor <= 0:
        NotificacionAsesora.objects.filter(
            salida=salida,
            tipo__in=tipos_controlados,
        ).delete()
        return

    asesora = form.cleaned_data.get('asesora_notificacion')
    if not asesora:
        return

    mensaje = (form.cleaned_data.get('mensaje_notificacion') or '').strip()
    if not mensaje:
        mensaje_default = getattr(form, 'notificacion_asesora_mensaje_default', '')
        mensaje = mensaje_default.format(
            codigo=salida.ingreso.codigo_equipo,
            valor=valor,
            cliente=salida.ingreso.cliente.nombres,
        )

    NotificacionAsesora.objects.update_or_create(
        salida=salida,
        tipo=tipo,
        defaults={
            'ingreso': salida.ingreso,
            'asesora': asesora,
            'creado_por': user,
            'valor_acordado': valor,
            'mensaje': mensaje,
            'leida': False,
            'leida_en': None,
        }
    )


def _normalizar_comparacion(valor):
    texto = ' '.join((valor or '').strip().casefold().split())
    texto = unicodedata.normalize('NFD', texto)
    return ''.join(c for c in texto if unicodedata.category(c) != 'Mn')


def _identidad_equipo_normalizada(datos_ingreso):
    tipo = _normalizar_comparacion(datos_ingreso.get('tipo_equipo'))
    tipo_otro = _normalizar_comparacion(datos_ingreso.get('tipo_equipo_otro')) if tipo == 'otro' else ''
    return (
        tipo,
        tipo_otro,
        _normalizar_comparacion(datos_ingreso.get('marca')),
        _normalizar_comparacion(datos_ingreso.get('modelo_serie')),
    )


def _identidad_equipo_de_ingreso(ingreso):
    return _identidad_equipo_normalizada({
        'tipo_equipo': ingreso.tipo_equipo,
        'tipo_equipo_otro': ingreso.tipo_equipo_otro,
        'marca': ingreso.marca,
        'modelo_serie': ingreso.modelo_serie,
    })


def _equipos_duplicados_para_cliente(cliente, datos_ingreso, excluir_pk=None):
    modelo = _normalizar_comparacion(datos_ingreso.get('modelo_serie'))

    if not cliente or not modelo:
        return []

    qs = cliente.ingresos.order_by('-creado')
    if excluir_pk:
        qs = qs.exclude(pk=excluir_pk)

    duplicados = []
    for equipo in qs:
        modelo_existente = _normalizar_comparacion(equipo.modelo_serie)
        if modelo_existente == modelo:
            duplicados.append(equipo)

    return duplicados


def _equipo_duplicado_para_cliente(cliente, datos_ingreso, excluir_pk=None):
    duplicados = _equipos_duplicados_para_cliente(cliente, datos_ingreso, excluir_pk)
    return duplicados[0] if duplicados else None


def _confirmo_mismo_equipo_cliente(request):
    return request.POST.get('confirmar_mismo_equipo_cliente') == '1'


# ═════════════════════════════════════════════════════════════════
# Páginas base
# ═════════════════════════════════════════════════════════════════


def ingresos_de_equipo_qs():
    return IngresoEquipo.objects.filter(sede__in=SEDES_EQUIPOS)


def ingresos_operativos_qs():
    """Equipos del flujo normal de reparación, sin donaciones ni compras."""
    return ingresos_de_equipo_qs().exclude(
        estado__in=('donado', 'equipo_a_comprar')
    )


def home(request):
    if request.user.is_authenticated:
        return redirect('econotec:bienvenida')
    return redirect('login')


@login_required
def bienvenida(request):
    """Dashboard de inicio."""
    hoy = date.today()
    mes_actual = hoy.month
    ano_actual = hoy.year
    ingresos_equipos = ingresos_de_equipo_qs()

    stats = {
        'total_ingresos': ingresos_equipos.count(),
        'ingresos_mes': ingresos_equipos.filter(
            fecha_ingreso__year=ano_actual, fecha_ingreso__month=mes_actual,
        ).count(),
        'total_salidas': SalidaEquipo.objects.count(),
        'salidas_mes': SalidaEquipo.objects.filter(
            fecha_salida__year=ano_actual, fecha_salida__month=mes_actual,
        ).count(),
        'salidas_fisicas_confirmadas': SalidaEquipo.objects.filter(
            ingreso__sede__in=SEDES_EQUIPOS,
            fecha_retiro_real__isnull=False,
        ).count(),
        'total_clientes': Cliente.objects.count(),
        'pendientes_retiro': ingresos_equipos.filter(
            estado__in=['ingresado', 'en_reparacion'],
            salida__isnull=True,
        ).count() + SalidaEquipo.objects.filter(
            ingreso__sede__in=SEDES_EQUIPOS,
            estado_reparacion='pendiente_retiro',
            fecha_retiro_real__isnull=True,
        ).count(),
    }
    salidas_equipos = SalidaEquipo.objects.filter(ingreso__sede__in=SEDES_EQUIPOS)
    ingresos_por_sede = dict(
        ingresos_equipos
        .values('sede')
        .annotate(total=Count('id'))
        .values_list('sede', 'total')
    )
    salidas_por_sede = dict(
        salidas_equipos
        .values('ingreso__sede')
        .annotate(total=Count('id'))
        .values_list('ingreso__sede', 'total')
    )
    resumen_movimientos = {
        'ingresos': {
            'guayaquil': ingresos_por_sede.get('guayaquil', 0),
            'quito': ingresos_por_sede.get('quito', 0),
            'total': stats['total_ingresos'],
        },
        'salidas': {
            'guayaquil': salidas_por_sede.get('guayaquil', 0),
            'quito': salidas_por_sede.get('quito', 0),
            'total': salidas_equipos.count(),
        },
    }
    resumen_movimientos['total_general'] = (
        resumen_movimientos['ingresos']['total'] + resumen_movimientos['salidas']['total']
    )

    # ── Equipos más ingresados ──────────────────────────────
    from .models import TIPOS_EQUIPO

    qs_equipos = (
        ingresos_equipos.values('tipo_equipo', 'tipo_equipo_otro')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    
    dict_tipos = dict(TIPOS_EQUIPO)
    equipos_stats = []
    
    # We might have multiple "otro" with different text, they are already grouped by (tipo_equipo, tipo_equipo_otro)
    for row in qs_equipos:
        if row['tipo_equipo'] == 'otro' and row['tipo_equipo_otro']:
            nombre = row['tipo_equipo_otro'].title()
        else:
            nombre = dict_tipos.get(row['tipo_equipo'], row['tipo_equipo']).title()
            
        # To avoid duplicates if casing differs in 'otro'
        found = False
        for stat in equipos_stats:
            if stat['nombre'].lower() == nombre.lower():
                stat['total'] += row['total']
                found = True
                break
        if not found:
            equipos_stats.append({'nombre': nombre, 'total': row['total']})
            
    # Re-sort and take top 5
    equipos_stats.sort(key=lambda x: x['total'], reverse=True)
    equipos_top = equipos_stats[:5]

    # ── Alertas: dos tipos independientes ────────────────────────
    es_admin_user = es_admin(request.user)

    # 1. Equipos demorados en diagnóstico (4+ días sin diagnosticar)
    demorados_qs = equipos_demorados_qs(usuario=None)

    demorados = []
    for ing in demorados_qs[:10]:
        demorados.append({
            'ingreso': ing,
            'dias': dias_en_taller(ing, hoy=hoy),
            'wa_link': whatsapp_link_demora(ing),
        })

    # 2. Salidas con bodegaje pendiente (5+ días sin que el cliente venga)
    bodegaje_qs = salidas_bodegaje_qs(usuario=None)

    bodegajes = []
    for sal in bodegaje_qs[:10]:
        bod = sal.calcular_bodegaje(hoy=hoy)
        bodegajes.append({
            'salida': sal,
            'ingreso': sal.ingreso,
            'dias_desde_salida': dias_desde_salida(sal, hoy=hoy),
            'bodegaje_dias': bod['dias'],
            'bodegaje_monto': bod['monto'],
            'wa_link': whatsapp_link_bodegaje(sal),
        })

    # 1b. Diagnósticos silenciados
    from datetime import timedelta as _td
    fecha_limite_diag = date.today() - _td(days=UMBRAL_DIAS_DIAGNOSTICO)
    qs_diag_silenciados = (
        IngresoEquipo.objects
        .select_related('cliente', 'tecnico_encargado')
        .filter(fecha_ingreso__lte=fecha_limite_diag)
        .filter(estado='ingresado')
        .filter(salida__isnull=True)
        .filter(diagnostico_silenciado=True)
    )
    # No filtramos por usuario para los técnicos, ven todo
    
    demorados_silenciados = []
    for ing in qs_diag_silenciados:
        demorados_silenciados.append({
            'ingreso': ing,
            'dias': dias_en_taller(ing, hoy=hoy),
        })

    # 2b. Bodegajes silenciados
    fecha_limite_bod = date.today() - _td(days=UMBRAL_DIAS_BODEGAJE)
    qs_bod_silenciados = (
        SalidaEquipo.objects
        .select_related('ingreso', 'ingreso__cliente', 'ingreso__tecnico_encargado', 'tecnico_reparo')
        .filter(fecha_salida__lte=fecha_limite_bod)
        .filter(fecha_retiro_real__isnull=True)
        .filter(bodegaje_silenciado=True)
    )
    # No filtramos por usuario para los técnicos, ven todo
    
    bodegajes_silenciados = []
    for sal in qs_bod_silenciados:
        bod = sal.calcular_bodegaje(hoy=hoy)
        bodegajes_silenciados.append({
            'salida': sal,
            'ingreso': sal.ingreso,
            'dias_desde_salida': dias_desde_salida(sal, hoy=hoy),
            'bodegaje_dias': bod['dias'],
            'bodegaje_monto': bod['monto'],
        })
    # ── Top Clientes ──────────────────────────────
    clientes_top = (
        Cliente.objects
        .annotate(total_ingresos=Count('ingresos', filter=Q(ingresos__sede__in=SEDES_EQUIPOS)))
        .filter(total_ingresos__gt=0)
        .order_by('-total_ingresos')[:5]
    )

    ctx = {
        'usuario': request.user,
        'es_admin': es_admin_user,
        'stats': stats,
        'resumen_movimientos': resumen_movimientos,
        'equipos_top': equipos_top,
        'clientes_top': clientes_top,
        'demorados': demorados,
        'demorados_total': demorados_qs.count(),
        'demorados_silenciados': demorados_silenciados,
        'bodegajes': bodegajes,
        'bodegajes_total': bodegaje_qs.count(),
        'bodegajes_silenciados': bodegajes_silenciados,
        'total_silenciados': len(demorados_silenciados) + len(bodegajes_silenciados),
        'umbral_diagnostico': UMBRAL_DIAS_DIAGNOSTICO,
        'umbral_bodegaje': UMBRAL_DIAS_BODEGAJE,
        'costo_bodegaje_dia': COSTO_BODEGAJE_DIA,
    }
    return render(request, 'bienvenida.html', ctx)

@login_required
def dashboard_details(request, tipo):
    """Devuelve el HTML parcial para el modal del dashboard."""
    hoy = date.today()
    mes_actual = hoy.month
    ano_actual = hoy.year

    titulo = ""
    columnas = []
    filas = []
    tipos_disponibles = {}
    es_modal_clientes = tipo == 'clientes'

    link_ver_todos = ""

    def estado_ingreso_para_modal(ingreso):
        return ingreso.estado_visual_display

    sedes_dashboard = {
        'guayaquil': ('Guayaquil', 'G'),
        'quito': ('Quito', 'U'),
    }

    def dinero_modal(valor):
        if valor is None:
            return '—'
        return f'${valor:.2f}'

    def boton_modal(url, texto, clase):
        return format_html(
            '<a href="{}" class="badge {} dashboard-modal-action">{}</a>',
            url,
            clase,
            texto,
        )

    def agregar_fila(
        valores,
        *,
        codigo='',
        sedes=(),
        tipos=(),
        texto_busqueda='',
        fecha_orden=None,
        clave_orden='',
        destacar_codigo=True,
    ):
        etiquetas_tipo = []
        for etiqueta in tipos:
            etiqueta = str(etiqueta or '').strip()
            if not etiqueta:
                continue
            tipos_disponibles[etiqueta.casefold()] = etiqueta
            etiquetas_tipo.append(etiqueta)

        filas.append({
            'celdas': [
                {
                    'label': columnas[indice],
                    'valor': valor,
                    'es_codigo': destacar_codigo and indice == 0,
                }
                for indice, valor in enumerate(valores)
            ],
            'codigo': codigo,
            'sedes': '|||'.join(sede for sede in sedes if sede),
            'tipos': '|||'.join(etiquetas_tipo),
            'busqueda': texto_busqueda,
            'orden_fecha': fecha_orden.isoformat() if fecha_orden else '',
            'orden_codigo': clave_orden or codigo,
        })

    if tipo == 'equipos_total':
        titulo = "Total de Equipos Ingresados"
        link_ver_todos = reverse('econotec:ingreso_lista')
        qs = ingresos_de_equipo_qs().select_related('cliente', 'salida').order_by(
            '-fecha_ingreso', '-creado',
        )
        columnas = ['Código', 'Cliente', 'Equipo', 'Fecha Ingreso', 'Estado', 'Acción']
        for eq in qs:
            btn = boton_modal(
                reverse('econotec:ingreso_detalle', kwargs={'pk': eq.pk}),
                'Ver detalles',
                'badge-ingresado',
            )
            agregar_fila(
                [
                    eq.codigo_equipo,
                    eq.cliente.nombres,
                    eq.tipo_equipo_display,
                    eq.fecha_ingreso.strftime('%d/%m/%Y'),
                    estado_ingreso_para_modal(eq),
                    btn,
                ],
                codigo=eq.codigo_equipo,
                sedes=(eq.sede,),
                tipos=(eq.tipo_equipo_display,),
                texto_busqueda=(
                    f'{eq.codigo_equipo} {eq.cliente.nombres} '
                    f'{eq.tipo_equipo_display} {eq.marca} {eq.modelo_serie_detalle} '
                    f'{estado_ingreso_para_modal(eq)}'
                ),
                fecha_orden=eq.fecha_ingreso,
            )

    elif tipo.startswith('ingresos_sede_'):
        sede = tipo.replace('ingresos_sede_', '', 1)
        sede_nombre, sede_codigo = sedes_dashboard.get(sede, ('Sede', ''))
        titulo = f"Ingresos de Equipo {sede_nombre} ({sede_codigo})"
        qs = (
            ingresos_de_equipo_qs()
            .select_related('cliente', 'tecnico_encargado', 'salida')
            .filter(sede=sede)
            .order_by('-fecha_ingreso', '-creado')
        )
        columnas = ['Código', 'Cliente', 'Equipo', 'Fecha', 'Técnico', 'Estado', 'Valor', 'Acción']
        for eq in qs:
            btn = boton_modal(
                reverse('econotec:ingreso_detalle', kwargs={'pk': eq.pk}),
                'Ver detalle',
                'badge-ingresado',
            )
            estado = estado_ingreso_para_modal(eq)
            agregar_fila(
                [
                    eq.codigo_equipo,
                    eq.cliente.nombres,
                    f'{eq.tipo_equipo_display} — {eq.marca} {eq.modelo_serie_detalle}',
                    eq.fecha_ingreso.strftime('%d/%m/%Y'),
                    eq.tecnico_encargado_nombre or '—',
                    estado,
                    dinero_modal(eq.valor_acordado),
                    btn,
                ],
                codigo=eq.codigo_equipo,
                sedes=(eq.sede,),
                tipos=(eq.tipo_equipo_display,),
                texto_busqueda=(
                    f'{eq.codigo_equipo} {eq.cliente.nombres} {eq.tipo_equipo_display} '
                    f'{eq.marca} {eq.modelo_serie_detalle} '
                    f'{eq.tecnico_encargado_nombre} {estado}'
                ),
                fecha_orden=eq.fecha_ingreso,
            )

    elif tipo == 'ingresos_mes':
        titulo = f"Ingresos del Mes ({hoy.strftime('%B %Y').capitalize()})"
        link_ver_todos = reverse('econotec:ingreso_lista')
        qs = ingresos_de_equipo_qs().select_related('cliente', 'salida').filter(
            fecha_ingreso__year=ano_actual, fecha_ingreso__month=mes_actual
        ).order_by('-fecha_ingreso', '-creado')
        columnas = ['Código', 'Cliente', 'Equipo', 'Fecha Ingreso', 'Estado', 'Acción']
        for eq in qs:
            btn = boton_modal(
                reverse('econotec:ingreso_detalle', kwargs={'pk': eq.pk}),
                'Ver detalles',
                'badge-ingresado',
            )
            estado = estado_ingreso_para_modal(eq)
            agregar_fila(
                [
                    eq.codigo_equipo,
                    eq.cliente.nombres,
                    eq.tipo_equipo_display,
                    eq.fecha_ingreso.strftime('%d/%m/%Y'),
                    estado,
                    btn,
                ],
                codigo=eq.codigo_equipo,
                sedes=(eq.sede,),
                tipos=(eq.tipo_equipo_display,),
                texto_busqueda=(
                    f'{eq.codigo_equipo} {eq.cliente.nombres} '
                    f'{eq.tipo_equipo_display} {eq.marca} {eq.modelo_serie_detalle} {estado}'
                ),
                fecha_orden=eq.fecha_ingreso,
            )

    elif tipo == 'pendientes':
        titulo = "Equipos Pendientes en Taller"
        link_ver_todos = reverse('econotec:ingreso_lista')
        ingresos = list(ingresos_de_equipo_qs().select_related('cliente').filter(
            estado__in=['ingresado', 'en_reparacion'],
            salida__isnull=True,
        ))
        salidas = list(SalidaEquipo.objects.select_related('ingreso__cliente').filter(
            ingreso__sede__in=SEDES_EQUIPOS,
            estado_reparacion='pendiente_retiro',
            fecha_retiro_real__isnull=True,
        ))
        columnas = ['Código', 'Cliente', 'Equipo', 'Fase', 'Estado', 'Acción']

        for eq in ingresos:
            btn = boton_modal(
                reverse('econotec:ingreso_detalle', kwargs={'pk': eq.pk}),
                'Ver detalles',
                'badge-ingresado',
            )
            estado = estado_ingreso_para_modal(eq)
            agregar_fila(
                [
                    eq.codigo_equipo,
                    eq.cliente.nombres,
                    eq.tipo_equipo_display,
                    'En proceso',
                    estado,
                    btn,
                ],
                codigo=eq.codigo_equipo,
                sedes=(eq.sede,),
                tipos=(eq.tipo_equipo_display,),
                texto_busqueda=(
                    f'{eq.codigo_equipo} {eq.cliente.nombres} '
                    f'{eq.tipo_equipo_display} {eq.marca} {eq.modelo_serie_detalle} {estado}'
                ),
                fecha_orden=eq.fecha_ingreso,
            )
        for sal in salidas:
            eq = sal.ingreso
            btn = boton_modal(
                reverse('econotec:ingreso_detalle', kwargs={'pk': eq.pk}),
                'Ver detalles',
                'badge-ingresado',
            )
            agregar_fila(
                [
                    eq.codigo_equipo,
                    eq.cliente.nombres,
                    eq.tipo_equipo_display,
                    'Terminado',
                    'Listo (pendiente de retiro)',
                    btn,
                ],
                codigo=eq.codigo_equipo,
                sedes=(eq.sede,),
                tipos=(eq.tipo_equipo_display,),
                texto_busqueda=(
                    f'{eq.codigo_equipo} {eq.cliente.nombres} '
                    f'{eq.tipo_equipo_display} {eq.marca} {eq.modelo_serie_detalle} '
                    'Terminado listo pendiente de retiro'
                ),
                fecha_orden=sal.fecha_salida,
            )

    elif tipo == 'salidas_mes':
        titulo = f"Equipos Finalizados del Mes ({hoy.strftime('%B %Y').capitalize()})"
        link_ver_todos = reverse('econotec:salida_lista')
        qs = SalidaEquipo.objects.select_related('ingreso__cliente').filter(
            ingreso__sede__in=SEDES_EQUIPOS,
            fecha_salida__year=ano_actual, fecha_salida__month=mes_actual
        ).order_by('-fecha_salida', '-creado')
        columnas = ['Código', 'Cliente', 'Equipo', 'Fecha de finalización', 'Resultado', 'Acción']
        for sal in qs:
            eq = sal.ingreso
            btn = boton_modal(
                reverse('econotec:salida_imprimir', kwargs={'pk': sal.pk}),
                'Ver PDF',
                'badge-entregado',
            )
            resultado = sal.estado_operativo_display
            agregar_fila(
                [
                    eq.codigo_equipo,
                    eq.cliente.nombres,
                    eq.tipo_equipo_display,
                    sal.fecha_salida.strftime('%d/%m/%Y'),
                    resultado,
                    btn,
                ],
                codigo=eq.codigo_equipo,
                sedes=(eq.sede,),
                tipos=(eq.tipo_equipo_display,),
                texto_busqueda=(
                    f'{eq.codigo_equipo} {eq.cliente.nombres} '
                    f'{eq.tipo_equipo_display} {eq.marca} {eq.modelo_serie_detalle} {resultado}'
                ),
                fecha_orden=sal.fecha_salida,
            )

    elif tipo.startswith('salidas_sede_'):
        sede = tipo.replace('salidas_sede_', '', 1)
        sede_nombre, sede_codigo = sedes_dashboard.get(sede, ('Sede', ''))
        titulo = f"Equipos Finalizados {sede_nombre} ({sede_codigo})"
        qs = (
            SalidaEquipo.objects
            .select_related('ingreso', 'ingreso__cliente', 'tecnico_reparo')
            .filter(ingreso__sede=sede)
            .order_by('-fecha_salida', '-creado')
        )
        columnas = ['Código', 'Cliente', 'Equipo', 'Fecha', 'Técnico', 'Resultado', 'Cobrado', 'Acción']
        for sal in qs:
            eq = sal.ingreso
            btn = boton_modal(
                reverse('econotec:salida_imprimir', kwargs={'pk': sal.pk}),
                'Ver PDF',
                'badge-entregado',
            )
            tecnico = sal.tecnico_reparo.get_username() if sal.tecnico_reparo else '—'
            resultado = sal.estado_operativo_display
            agregar_fila(
                [
                    eq.codigo_equipo,
                    eq.cliente.nombres,
                    f'{eq.tipo_equipo_display} — {eq.marca} {eq.modelo_serie_detalle}',
                    sal.fecha_salida.strftime('%d/%m/%Y'),
                    tecnico,
                    resultado,
                    dinero_modal(sal.valor_final_cobrado),
                    btn,
                ],
                codigo=eq.codigo_equipo,
                sedes=(eq.sede,),
                tipos=(eq.tipo_equipo_display,),
                texto_busqueda=(
                    f'{eq.codigo_equipo} {eq.cliente.nombres} {eq.tipo_equipo_display} '
                    f'{eq.marca} {eq.modelo_serie_detalle} {tecnico} {resultado}'
                ),
                fecha_orden=sal.fecha_salida,
            )

    elif tipo == 'clientes':
        titulo = "Directorio de Clientes"
        link_ver_todos = ""
        ingresos_cliente = ingresos_de_equipo_qs().only(
            'cliente_id',
            'sede',
            'numero_equipo',
            'tipo_equipo',
            'tipo_equipo_otro',
            'fecha_ingreso',
        ).order_by('-fecha_ingreso', '-creado')
        qs = Cliente.objects.prefetch_related(
            Prefetch('ingresos', queryset=ingresos_cliente, to_attr='equipos_dashboard')
        ).order_by('nombres', 'id')
        columnas = ['Nombre / Razón Social', 'Cédula / RUC', 'WhatsApp', 'Email', 'Equipos']
        for cli in qs:
            equipos = cli.equipos_dashboard
            codigos = [equipo.codigo_equipo for equipo in equipos]
            sedes = sorted({equipo.sede for equipo in equipos})
            tipos = sorted({equipo.tipo_equipo_display for equipo in equipos}, key=str.casefold)
            codigos_resumen = ', '.join(codigos[:6]) or 'Sin equipos'
            if len(codigos) > 6:
                codigos_resumen += f' · +{len(codigos) - 6}'
            agregar_fila(
                [
                    cli.nombres,
                    cli.cedula,
                    cli.whatsapp or '—',
                    cli.correo or '—',
                    codigos_resumen,
                ],
                codigo=' '.join(codigos),
                sedes=sedes,
                tipos=tipos,
                texto_busqueda=' '.join([
                    cli.nombres,
                    cli.cedula,
                    cli.whatsapp or '',
                    cli.correo or '',
                    *codigos,
                    *tipos,
                ]),
                fecha_orden=equipos[0].fecha_ingreso if equipos else None,
                clave_orden=cli.nombres,
                destacar_codigo=False,
            )

    if es_modal_clientes:
        filas.sort(key=lambda fila: fila['orden_codigo'].casefold())
    else:
        filas.sort(
            key=lambda fila: (fila['orden_fecha'], fila['orden_codigo']),
            reverse=True,
        )

    ctx = {
        'titulo': titulo,
        'columnas': columnas,
        'filas': filas,
        'total_filas': len(filas),
        'tipos_filtro': sorted(tipos_disponibles.values(), key=str.casefold),
        'es_modal_clientes': es_modal_clientes,
        'link_ver_todos': link_ver_todos,
    }
    return render(request, 'includes/dashboard_modal_content.html', ctx)

@login_required
def ayuda(request):
    return render(request, 'ayuda.html')


@login_required
def reproductor_musica(request):
    """Mini-reproductor de música de YouTube para técnicos."""
    return render(request, 'musica.html')


INVENTARIO_SEDES = {
    'guayaquil': 'Guayaquil',
    'quito': 'Quito',
}

INVENTARIO_UBICACIONES_POR_SEDE = {
    'guayaquil': (
        ('guayaquil_norte', 'Guayaquil - Norte'),
        ('guayaquil_centro', 'Guayaquil - Centro'),
    ),
    'quito': (
        ('quito', 'Quito'),
    ),
}

INVENTARIO_UBICACIONES_LABELS = {
    valor: etiqueta
    for opciones in INVENTARIO_UBICACIONES_POR_SEDE.values()
    for valor, etiqueta in opciones
}

INVENTARIO_CATEGORIAS = {
    'impresora': {
        'nombre': 'Impresora',
        'tipos': (
            ('impresora-laser', 'Impresora Laser'),
            ('impresora-inyeccion', 'Impresora Inyección'),
        ),
    },
    'computadora': {
        'nombre': 'Computadora',
        'tipos': (
            ('pc', 'PC'),
            ('laptops', 'Laptops'),
        ),
    },
    'consola': {
        'nombre': 'Consola',
        'tipos': (
            ('consola-de-mesa', 'consola de mesa'),
            ('portatil', 'Portatil'),
        ),
    },
    'celular': {
        'nombre': 'Celular',
        'tipos': (
            ('celular', 'Celular'),
        ),
    },
    'tablet': {
        'nombre': 'Tablet',
        'tipos': (
            ('tablet', 'Tablet'),
        ),
    },
    'mando': {
        'nombre': 'Mando',
        'tipos': (
            ('mando', 'Mando'),
        ),
    },
    'otros-equipos-materiales': {
        'nombre': 'Otros equipos/materiales',
        'tipos': (
            ('otros-equipos-materiales', 'Otros equipos/materiales'),
        ),
    },
}


def _inventario_tipo_por_slug(categoria_info, tipo_slug):
    for slug, nombre in categoria_info['tipos']:
        if slug == tipo_slug:
            return {'slug': slug, 'nombre': nombre}
    return None


def _inventario_ubicaciones_para_sede(sede_slug):
    return INVENTARIO_UBICACIONES_POR_SEDE.get(sede_slug, ())


def _texto_inventario_item_busqueda(item):
    ubicacion_label = INVENTARIO_UBICACIONES_LABELS.get(
        item.ubicacion,
        item.get_ubicacion_display(),
    )
    return ' '.join(str(valor or '') for valor in (
        item.producto,
        item.codigo,
        item.marca,
        item.modelo,
        item.serie,
        item.estado,
        item.get_estado_display(),
        item.causa_no_disponible,
        item.get_causa_no_disponible_display(),
        item.ubicacion,
        item.get_ubicacion_display(),
        ubicacion_label,
        item.observacion,
    ))


def _puede_gestionar_inventario(user):
    return es_admin(user) or es_tecnico(user)


def _puede_notificar_admin_inventario(user):
    return es_tecnico(user) or es_asesor(user)


def _inventario_listado_filtrado(request, sede_slug, categoria_slug, tipo_slug):
    filtro_q = (request.GET.get('q') or '').strip()
    estado_filtro = (request.GET.get('estado') or '').strip()
    ubicacion_filtro = (request.GET.get('ubicacion') or '').strip()
    estado_valores = {valor for valor, _ in InventarioItem.ESTADOS}
    ubicaciones_opciones = _inventario_ubicaciones_para_sede(sede_slug)
    ubicacion_valores = {valor for valor, _ in ubicaciones_opciones}

    if estado_filtro not in estado_valores:
        estado_filtro = ''
    if ubicacion_filtro not in ubicacion_valores:
        ubicacion_filtro = ''

    items_qs = (
        InventarioItem.objects
        .filter(sede=sede_slug, categoria=categoria_slug, tipo=tipo_slug)
        .select_related('registrado_por')
        .order_by('-creado')
    )
    total_general = items_qs.count()

    if estado_filtro:
        items_qs = items_qs.filter(estado=estado_filtro)
    if ubicacion_filtro:
        items_qs = items_qs.filter(ubicacion=ubicacion_filtro)

    items = list(items_qs)
    if filtro_q:
        items = filtrar_objetos_normalizado(
            items,
            filtro_q,
            _texto_inventario_item_busqueda,
        )

    return {
        'items': items,
        'filtro_q': filtro_q,
        'estado_filtro': estado_filtro,
        'ubicacion_filtro': ubicacion_filtro,
        'ubicacion_opciones': ubicaciones_opciones,
        'filtros_activos': bool(filtro_q or estado_filtro or ubicacion_filtro),
        'total_general': total_general,
        'total_filtrado': len(items),
    }


def _inventario_redirect_tabla(item):
    return redirect(
        'econotec:inventario_tabla',
        sede=item.sede,
        categoria=item.categoria,
        tipo=item.tipo,
    )


def _inventario_contexto_base(sede_slug, categoria_slug, tipo_slug):
    sede_label = INVENTARIO_SEDES.get(sede_slug)
    categoria_info = INVENTARIO_CATEGORIAS.get(categoria_slug)
    if not sede_label or not categoria_info:
        return None, None, None
    return sede_label, categoria_info, _inventario_tipo_por_slug(categoria_info, tipo_slug)


def _inventario_item_contexto(request, item, box_size=8):
    from .qr_utils import qr_data_uri
    from qrcode.constants import ERROR_CORRECT_H

    detalle_url = reverse('econotec:inventario_detalle_item', kwargs={'codigo': item.codigo})
    qr_url = request.build_absolute_uri(detalle_url)
    categoria_info = INVENTARIO_CATEGORIAS.get(item.categoria, {'nombre': item.categoria})
    tipo_info = {'slug': item.tipo, 'nombre': item.tipo}
    if 'tipos' in categoria_info:
        tipo_info = _inventario_tipo_por_slug(categoria_info, item.tipo) or tipo_info
    return {
        'item': item,
        'categoria': categoria_info,
        'tipo': tipo_info,
        'detalle_url': detalle_url,
        'detalle_url_absoluta': qr_url,
        # El logo central cubre parte de la matriz. La corrección alta permite
        # que cámaras móviles recuperen el enlace, y el margen de 4 módulos es
        # la zona silenciosa estándar necesaria al mostrar o imprimir el QR.
        'qr_data_uri': qr_data_uri(
            qr_url,
            box_size=box_size,
            border=4,
            error_correction=ERROR_CORRECT_H,
        ),
        'puede_gestionar_inventario': _puede_gestionar_inventario(request.user),
    }


@login_required
def inventario_menu(request):
    """Pantalla inicial del inventario por sede."""
    return render(request, 'inventario/menu.html')


@login_required
def inventario_categoria(request, sede, categoria):
    sede_slug = (sede or '').strip().lower()
    categoria_slug = (categoria or '').strip().lower()
    sede_label = INVENTARIO_SEDES.get(sede_slug)
    categoria_info = INVENTARIO_CATEGORIAS.get(categoria_slug)
    if not sede_label or not categoria_info:
        messages.error(request, 'Selecciona una sede y una categoría válida para inventario.')
        return redirect('econotec:inventario_menu')

    tipos = [
        {'slug': slug, 'nombre': nombre}
        for slug, nombre in categoria_info['tipos']
    ]
    if len(tipos) == 1:
        return redirect(
            'econotec:inventario_tabla',
            sede=sede_slug,
            categoria=categoria_slug,
            tipo=tipos[0]['slug'],
        )

    return render(request, 'inventario/categoria.html', {
        'sede_slug': sede_slug,
        'sede_label': sede_label,
        'categoria_slug': categoria_slug,
        'categoria': categoria_info,
        'tipos': tipos,
    })


@login_required
def inventario_tabla(request, sede, categoria, tipo):
    sede_slug = (sede or '').strip().lower()
    categoria_slug = (categoria or '').strip().lower()
    tipo_slug = (tipo or '').strip().lower()
    sede_label = INVENTARIO_SEDES.get(sede_slug)
    categoria_info = INVENTARIO_CATEGORIAS.get(categoria_slug)
    if not sede_label or not categoria_info:
        messages.error(request, 'Selecciona una sede y una categoría válida para inventario.')
        return redirect('econotec:inventario_menu')

    tipo_info = _inventario_tipo_por_slug(categoria_info, tipo_slug)
    if not tipo_info:
        messages.error(request, 'Selecciona un tipo válido para inventario.')
        return redirect('econotec:inventario_categoria', sede=sede_slug, categoria=categoria_slug)

    listado = _inventario_listado_filtrado(request, sede_slug, categoria_slug, tipo_slug)
    items_filtrados = listado['items']
    page_obj, querystring = paginar_resultados(request, items_filtrados)
    items = list(page_obj.object_list)

    for item in items:
        detalle_url = reverse('econotec:inventario_detalle_item', kwargs={'codigo': item.codigo})
        item.detalle_url = detalle_url
        item.imprimir_qr_url = reverse('econotec:inventario_qr_imprimir', kwargs={'codigo': item.codigo})
        item.qr_data_uri = _inventario_item_contexto(request, item, box_size=3)['qr_data_uri']

    items_bajo_stock = [item for item in items_filtrados if item.cantidad == 1]
    for item in items_bajo_stock:
        item.detalle_url = reverse('econotec:inventario_detalle_item', kwargs={'codigo': item.codigo})
    notificados_ids = set()
    if items_bajo_stock:
        notificados_ids = set(
            NotificacionInventarioAdmin.objects
            .filter(inventario_item__in=items_bajo_stock, leida=False)
            .values_list('inventario_item_id', flat=True)
        )
    for item in items_bajo_stock:
        item.notificacion_admin_pendiente = item.pk in notificados_ids

    return render(request, 'inventario/tabla.html', {
        'sede_slug': sede_slug,
        'sede_label': sede_label,
        'categoria_slug': categoria_slug,
        'categoria': categoria_info,
        'tipo': tipo_info,
        'items': items,
        'page_obj': page_obj,
        'querystring': querystring,
        'tiene_subtipos': len(categoria_info['tipos']) > 1,
        'puede_gestionar_inventario': _puede_gestionar_inventario(request.user),
        'puede_notificar_admin_inventario': _puede_notificar_admin_inventario(request.user),
        'items_bajo_stock': items_bajo_stock,
        'filtro_q': listado['filtro_q'],
        'estado_filtro': listado['estado_filtro'],
        'ubicacion_filtro': listado['ubicacion_filtro'],
        'estado_opciones': InventarioItem.ESTADOS,
        'ubicacion_opciones': listado['ubicacion_opciones'],
        'filtros_activos': listado['filtros_activos'],
        'total_general': listado['total_general'],
        'total_filtrado': listado['total_filtrado'],
    })


@login_required
def inventario_export(request, sede, categoria, tipo):
    sede_slug = (sede or '').strip().lower()
    categoria_slug = (categoria or '').strip().lower()
    tipo_slug = (tipo or '').strip().lower()
    sede_label = INVENTARIO_SEDES.get(sede_slug)
    categoria_info = INVENTARIO_CATEGORIAS.get(categoria_slug)
    if not sede_label or not categoria_info:
        messages.error(request, 'Selecciona una sede y una categoría válida para exportar inventario.')
        return redirect('econotec:inventario_menu')

    tipo_info = _inventario_tipo_por_slug(categoria_info, tipo_slug)
    if not tipo_info:
        messages.error(request, 'Selecciona un tipo válido para exportar inventario.')
        return redirect('econotec:inventario_categoria', sede=sede_slug, categoria=categoria_slug)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    listado = _inventario_listado_filtrado(request, sede_slug, categoria_slug, tipo_slug)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario'
    headers = [
        'Código',
        'Producto',
        'Marca',
        'Modelo',
        'Serie',
        'Estado',
        'Causa no disponible',
        'Cantidad',
        'Costo (USD)',
        'Ubicación',
        'Observación',
        'Sede',
        'Categoría',
        'Tipo',
        'Registrado por',
        'Creado',
        'Actualizado',
    ]
    ws.append(headers)

    header_fill = PatternFill('solid', fgColor='F97316')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')

    for item in listado['items']:
        registrado_por = item.registrado_por.get_full_name() or item.registrado_por.username if item.registrado_por else '—'
        ws.append([
            item.codigo,
            item.producto,
            item.marca,
            item.modelo,
            item.serie or '—',
            item.get_estado_display(),
            item.get_causa_no_disponible_display() if item.causa_no_disponible else '—',
            item.cantidad,
            float(item.costo or 0),
            item.get_ubicacion_display(),
            item.observacion or '—',
            sede_label,
            categoria_info['nombre'],
            tipo_info['nombre'],
            registrado_por,
            timezone.localtime(item.creado).strftime('%d/%m/%Y %H:%M') if item.creado else '—',
            timezone.localtime(item.actualizado).strftime('%d/%m/%Y %H:%M') if item.actualizado else '—',
        ])

    widths = [24, 28, 18, 18, 18, 18, 24, 12, 14, 22, 32, 16, 18, 20, 22, 18, 18]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'inventario_{sede_slug}_{categoria_slug}_{tipo_slug}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@require_POST
def inventario_notificar_admin(request, codigo):
    if not _puede_notificar_admin_inventario(request.user):
        messages.error(request, 'Solo técnicos y asesoras pueden notificar stock crítico al admin.')
        return redirect('econotec:inventario_detalle_item', codigo=codigo)

    item = get_object_or_404(InventarioItem, codigo=codigo)
    next_url = (request.POST.get('next') or '').strip()
    if not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        next_url = reverse('econotec:inventario_tabla', kwargs={
            'sede': item.sede,
            'categoria': item.categoria,
            'tipo': item.tipo,
        })

    if item.cantidad != 1:
        messages.warning(request, 'La notificación al admin solo se activa cuando queda exactamente 1 unidad.')
        return redirect(next_url)

    mensaje = (
        f'Solo queda 1 unidad de {item.producto} ({item.codigo}) '
        f'en {item.get_ubicacion_display()}.'
    )
    notificacion, creada = NotificacionInventarioAdmin.objects.get_or_create(
        inventario_item=item,
        leida=False,
        defaults={
            'creado_por': request.user,
            'mensaje': mensaje,
        },
    )
    if creada:
        messages.success(request, 'Admin notificado por stock crítico.')
    else:
        messages.info(request, 'Este producto ya tiene una notificación pendiente para el admin.')
    return redirect(next_url)


@admin_requerido
def notificacion_inventario_admin_ver(request, pk):
    notificacion = get_object_or_404(
        NotificacionInventarioAdmin.objects.select_related('inventario_item'),
        pk=pk,
    )
    item = notificacion.inventario_item
    notificacion.marcar_vista()
    return redirect('econotec:inventario_detalle_item', codigo=item.codigo)


@login_required
def inventario_registrar(request, sede, categoria, tipo):
    sede_slug = (sede or '').strip().lower()
    categoria_slug = (categoria or '').strip().lower()
    tipo_slug = (tipo or '').strip().lower()
    sede_label = INVENTARIO_SEDES.get(sede_slug)
    categoria_info = INVENTARIO_CATEGORIAS.get(categoria_slug)
    if not sede_label or not categoria_info:
        messages.error(request, 'Selecciona una sede y una categoría válida para inventario.')
        return redirect('econotec:inventario_menu')

    tipo_info = _inventario_tipo_por_slug(categoria_info, tipo_slug)
    if not tipo_info:
        messages.error(request, 'Selecciona un tipo válido para inventario.')
        return redirect('econotec:inventario_categoria', sede=sede_slug, categoria=categoria_slug)

    if request.method == 'POST':
        form = InventarioItemForm(
            request.POST,
            sede_slug=sede_slug,
            categoria_slug=categoria_slug,
            tipo_slug=tipo_slug,
        )
        if form.is_valid():
            item = form.save(commit=False)
            item.sede = sede_slug
            item.categoria = categoria_slug
            item.tipo = tipo_slug
            item.registrado_por = request.user
            item.save()
            messages.success(request, f'Inventario registrado con código {item.codigo}.')
            return redirect(
                'econotec:inventario_tabla',
                sede=sede_slug,
                categoria=categoria_slug,
                tipo=tipo_slug,
            )
    else:
        form = InventarioItemForm(
            sede_slug=sede_slug,
            categoria_slug=categoria_slug,
            tipo_slug=tipo_slug,
            initial={
                'cantidad': 1,
                'estado': 'disponible',
            },
        )

    return render(request, 'inventario/form.html', {
        'form': form,
        'sede_slug': sede_slug,
        'sede_label': sede_label,
        'categoria_slug': categoria_slug,
        'categoria': categoria_info,
        'tipo': tipo_info,
        'tiene_subtipos': len(categoria_info['tipos']) > 1,
        'modo_formulario': 'crear',
    })


@login_required
def inventario_editar(request, codigo):
    if not _puede_gestionar_inventario(request.user):
        messages.error(request, 'Solo admin y técnicos pueden editar inventario.')
        return redirect('econotec:inventario_detalle_item', codigo=codigo)

    item = get_object_or_404(InventarioItem, codigo=codigo)
    sede_label, categoria_info, tipo_info = _inventario_contexto_base(
        item.sede,
        item.categoria,
        item.tipo,
    )
    if not sede_label or not categoria_info or not tipo_info:
        messages.error(request, 'Este producto tiene una categoría de inventario inválida.')
        return redirect('econotec:inventario_menu')

    if request.method == 'POST':
        form = InventarioItemForm(
            request.POST,
            instance=item,
            sede_slug=item.sede,
            categoria_slug=item.categoria,
            tipo_slug=item.tipo,
        )
        if form.is_valid():
            form.save()
            messages.success(request, 'Producto de inventario actualizado.')
            return redirect('econotec:inventario_detalle_item', codigo=item.codigo)
    else:
        form = InventarioItemForm(
            instance=item,
            sede_slug=item.sede,
            categoria_slug=item.categoria,
            tipo_slug=item.tipo,
        )

    return render(request, 'inventario/form.html', {
        'form': form,
        'item': item,
        'sede_slug': item.sede,
        'sede_label': sede_label,
        'categoria_slug': item.categoria,
        'categoria': categoria_info,
        'tipo': tipo_info,
        'tiene_subtipos': len(categoria_info['tipos']) > 1,
        'modo_formulario': 'editar',
    })


@login_required
def inventario_detalle_item(request, codigo):
    item = get_object_or_404(InventarioItem, codigo=codigo)
    return render(request, 'inventario/detalle.html', _inventario_item_contexto(request, item, box_size=6))


@login_required
@require_POST
def inventario_actualizar_cantidad(request, codigo):
    if not _puede_gestionar_inventario(request.user):
        messages.error(request, 'Solo admin y técnicos pueden actualizar inventario.')
        return redirect('econotec:inventario_detalle_item', codigo=codigo)

    accion = (request.POST.get('accion') or '').strip().lower()
    cantidad_final = (request.POST.get('cantidad') or '').strip()
    if accion not in {'sumar', 'restar', ''}:
        messages.error(request, 'No se pudo actualizar la cantidad.')
        return redirect('econotec:inventario_detalle_item', codigo=codigo)

    with transaction.atomic():
        item = get_object_or_404(
            InventarioItem.objects.select_for_update(),
            codigo=codigo,
        )
        if cantidad_final:
            try:
                nueva_cantidad = int(cantidad_final)
            except ValueError:
                messages.error(request, 'La cantidad debe ser un número válido.')
                return redirect('econotec:inventario_detalle_item', codigo=item.codigo)
            item.cantidad = max(0, nueva_cantidad)
        else:
            if accion == 'restar':
                item.cantidad = max(0, item.cantidad - 1)
            elif accion == 'sumar':
                item.cantidad += 1
        item.save(update_fields=['cantidad', 'actualizado'])

    messages.success(request, f'Cantidad actualizada a {item.cantidad}.')
    return redirect('econotec:inventario_detalle_item', codigo=item.codigo)


@login_required
@require_POST
def inventario_eliminar(request, codigo):
    if not _puede_gestionar_inventario(request.user):
        messages.error(request, 'Solo admin y técnicos pueden eliminar inventario.')
        return redirect('econotec:inventario_detalle_item', codigo=codigo)

    item = get_object_or_404(InventarioItem, codigo=codigo)
    tabla_kwargs = {
        'sede': item.sede,
        'categoria': item.categoria,
        'tipo': item.tipo,
    }
    producto = item.producto
    codigo_item = item.codigo
    try:
        item.delete()
    except ProtectedError:
        messages.error(
            request,
            'No se puede eliminar este producto porque está asociado a una venta. '
            'Primero revierte o elimina el producto desde la venta.',
        )
        return redirect('econotec:inventario_detalle_item', codigo=codigo_item)
    messages.success(request, f'Producto eliminado: {producto} ({codigo_item}).')
    return redirect('econotec:inventario_tabla', **tabla_kwargs)


@login_required
def inventario_qr_imprimir(request, codigo):
    item = get_object_or_404(InventarioItem, codigo=codigo)
    return render(request, 'inventario/imprimir_qr.html', _inventario_item_contexto(request, item, box_size=8))


# ═════════════════════════════════════════════════════════════════
# Ingreso de Equipos
# ═════════════════════════════════════════════════════════════════

@tecnico_requerido
def ingreso_menu(request):
    """Menú de ingresos: registrar nuevo / ver lista."""
    ingresos_equipos = ingresos_operativos_qs()
    total = ingresos_equipos.count()
    pendientes = ingresos_equipos.filter(
        estado__in=['ingresado', 'en_reparacion'],
        salida__isnull=True,
    ).count() + SalidaEquipo.objects.filter(
        ingreso__sede__in=SEDES_EQUIPOS,
        estado_reparacion='pendiente_retiro',
        fecha_retiro_real__isnull=True,
    ).count()
    pendientes_valor = _ingresos_pendientes_valor_qs().count()
    return render(request, 'ingresos/menu.html', {
        'total': total,
        'pendientes': pendientes,
        'pendientes_valor': pendientes_valor,
    })


def _ingresos_pendientes_valor_qs():
    return (
        ingresos_operativos_qs()
        .filter(valor_acordado__isnull=True)
        .exclude(estado='entregado')
    )


def _sincronizar_egreso_compra(ingreso, usuario):
    """Crea o actualiza el único egreso asociado a una compra de equipo."""
    if ingreso.estado != 'equipo_a_comprar' or not ingreso.valor_acordado:
        return None

    categoria, _ = CategoriaEgreso.objects.get_or_create(
        nombre='Compra de equipo',
        defaults={
            'descripcion': 'Compras de equipos registradas desde Ingreso de Equipo.',
            'color': '#00838f',
            'icono': '🖥️',
        },
    )
    datos = {
        'fecha': ingreso.fecha_ingreso,
        'categoria': categoria,
        'concepto': (
            f'Compra de equipo {ingreso.codigo_equipo} — '
            f'{ingreso.tipo_equipo_display}: {ingreso.marca} {ingreso.modelo_serie}'
        )[:200],
        'monto': ingreso.valor_acordado,
        'metodo': ingreso.compra_metodo_pago or 'efectivo',
        'banco': ingreso.compra_banco or '',
        'banco_otro': ingreso.compra_banco_otro or '',
        'tarjeta_app': ingreso.compra_tarjeta_app or '',
        'comprobante_url': ingreso.compra_comprobante_url or '',
        'monto_1': ingreso.compra_monto_1,
        'metodo_1': ingreso.compra_metodo_1 or '',
        'banco_1': ingreso.compra_banco_1 or '',
        'banco_otro_1': ingreso.compra_banco_otro_1 or '',
        'tarjeta_app_1': ingreso.compra_tarjeta_app_1 or '',
        'monto_2': ingreso.compra_monto_2,
        'metodo_2': ingreso.compra_metodo_2 or '',
        'banco_2': ingreso.compra_banco_2 or '',
        'banco_otro_2': ingreso.compra_banco_otro_2 or '',
        'tarjeta_app_2': ingreso.compra_tarjeta_app_2 or '',
        'registrado_por': usuario,
        'notas': 'Egreso generado automáticamente desde Equipo a comprar.',
    }
    egreso, creado = Egreso.objects.update_or_create(
        ingreso_compra=ingreso,
        defaults=datos,
    )
    return egreso, creado


@tecnico_requerido
@require_GET
def cliente_buscar_por_cedula(request):
    """
    Endpoint AJAX: busca un cliente por cédula y devuelve sus datos en JSON.
    Lo usa el formulario de ingreso para autocompletar nombre, WhatsApp, correo
    y sector cuando el técnico escribe una cédula que ya existe en el sistema.

    Respuesta cuando existe el cliente:
        {"existe": true, "cliente": {...campos...}, "num_equipos_anteriores": N}
    Respuesta cuando NO existe:
        {"existe": false}
    """
    cedula = (request.GET.get('cedula') or '').strip()
    if not cedula:
        return JsonResponse({'existe': False})

    cliente = Cliente.objects.filter(cedula=cedula).first()
    if not cliente:
        return JsonResponse({'existe': False})

    equipos_qs = cliente.ingresos.order_by('-creado')
    equipos_data = [
        {
            'id': eq.id, 
            'codigo': eq.codigo_equipo,
            'label': f"{eq.codigo_equipo} — {eq.marca} {eq.modelo_serie_detalle} ({eq.creado.strftime('%d/%m/%Y')})",
            'tipo_equipo': eq.tipo_equipo,
            'tipo_equipo_otro': eq.tipo_equipo_otro,
            'marca': eq.marca,
            'modelo_serie': eq.modelo_serie,
            'serie': eq.serie,
            'detalle_url': reverse('econotec:ingreso_detalle', args=[eq.pk]),
        }
        for eq in equipos_qs
    ]

    return JsonResponse({
        'existe': True,
        'cliente': {
            'nombres': cliente.nombres,
            'whatsapp': cliente.whatsapp,
            'correo': cliente.correo,
            'sector': cliente.sector,
            'sector_otro': cliente.sector_otro,
        },
        'equipos': equipos_data,
        'num_equipos_anteriores': equipos_qs.count(),
    })


def _programar_correo_ingreso_automatico(request, ingreso):
    """Programa el envío solo después de guardar correctamente la transacción."""
    if not getattr(settings, 'INGRESO_EMAIL_AUTOMATICO', True):
        return
    if not (ingreso.cliente.correo or '').strip():
        messages.warning(
            request,
            'El equipo quedó registrado, pero el cliente no tiene un correo para enviarle el comprobante.',
        )
        return

    transaction.on_commit(
        lambda ingreso_pk=ingreso.pk: enviar_correo_ingreso_seguro(ingreso_pk)
    )


@tecnico_requerido
@transaction.atomic
def ingreso_registrar(request):
    """Registra un nuevo ingreso.
    - La SEDE se toma de la sesión (la elegida en el login).
    - Si la cédula del cliente ya existe, se reutiliza.
    - El número de equipo es correlativo dentro de la sede (G1/U1...).
    """
    # La sede se toma de la sesión. Si no hay sede, mandamos al login.
    sede_actual = (request.session.get('sede_actual') or '').strip().lower()
    if sede_actual not in ('guayaquil', 'quito'):
        messages.error(request, 'Tu sesión no tiene una sede asignada. Vuelve a iniciar sesión.')
        return redirect('login')

    confirmar_mismo_equipo_cliente = (
        _confirmo_mismo_equipo_cliente(request) if request.method == 'POST' else False
    )

    if request.method == 'POST':
        cli_form = ClienteForm(request.POST, prefix='cli')
        ing_form = IngresoEquipoForm(request.POST, prefix='ing')

        cedula = (request.POST.get('cli-cedula') or '').strip()
        cliente_existente = Cliente.objects.filter(cedula=cedula).first() if cedula else None

        if cliente_existente:
            # Actualizar datos del cliente con la nueva info, si hay cambios
            cli_form_existente = ClienteForm(request.POST, prefix='cli', instance=cliente_existente)
            if ing_form.is_valid() and cli_form_existente.is_valid():
                duplicado = _equipo_duplicado_para_cliente(cliente_existente, ing_form.cleaned_data)
                if duplicado and not confirmar_mismo_equipo_cliente:
                    mensaje_duplicado = (
                        'ESTE EQUIPO YA SE ENCUENTRA REGISTRADO, POR FAVOR VERIFICA EN LA LISTA DE EQUIPOS.'
                    )
                    ing_form.add_error('modelo_serie', f'{mensaje_duplicado} Coincide con el equipo {duplicado.codigo_equipo}.')
                    messages.error(
                        request,
                        f'{mensaje_duplicado} Coincide con {duplicado.codigo_equipo}.'
                    )
                    cli_form = cli_form_existente
                else:
                    cliente = cli_form_existente.save()
                    ingreso = ing_form.save(commit=False)
                    ingreso.cliente = cliente
                    ingreso.sede = sede_actual           # ← sede de la sesión
                    ingreso.registrado_por = request.user
                    ingreso.save()
                    _sincronizar_egreso_compra(ingreso, request.user)
                    registrar_bitacora(
                        request.user,
                        'ingreso',
                        _texto_ingreso_bitacora(ingreso),
                        ingreso=ingreso,
                        dedupe_key=f'ingreso:{ingreso.pk}:creado',
                    )
                    messages.success(
                        request,
                        f'Equipo {ingreso.codigo_equipo} ingresado para {cliente.nombres}.'
                    )
                    _programar_correo_ingreso_automatico(request, ingreso)
                    if duplicado and confirmar_mismo_equipo_cliente:
                        messages.info(
                            request,
                            f'Reingreso confirmado: mismo cliente y mismo equipo que {duplicado.codigo_equipo}.'
                        )
                    return redirect('econotec:ingreso_detalle', pk=ingreso.pk)
            else:
                cli_form = cli_form_existente
        else:
            if cli_form.is_valid() and ing_form.is_valid():
                cliente = cli_form.save()
                ingreso = ing_form.save(commit=False)
                ingreso.cliente = cliente
                ingreso.sede = sede_actual           # ← sede de la sesión
                ingreso.registrado_por = request.user
                ingreso.save()
                _sincronizar_egreso_compra(ingreso, request.user)
                registrar_bitacora(
                    request.user,
                    'ingreso',
                    _texto_ingreso_bitacora(ingreso),
                    ingreso=ingreso,
                    dedupe_key=f'ingreso:{ingreso.pk}:creado',
                )
                messages.success(
                    request,
                    f'Equipo {ingreso.codigo_equipo} ingresado para {cliente.nombres}.'
                )
                _programar_correo_ingreso_automatico(request, ingreso)
                return redirect('econotec:ingreso_detalle', pk=ingreso.pk)

    else:
        cli_initial = {
            'cedula': request.GET.get('cedula', ''),
            'nombres': request.GET.get('nombres', ''),
            'whatsapp': request.GET.get('whatsapp', ''),
            'correo': request.GET.get('correo', ''),
            'sector': request.GET.get('sector', ''),
            'sector_otro': request.GET.get('sector_otro', ''),
        }
        
        # Validar tipo_equipo contra las opciones disponibles
        from .models import TIPOS_EQUIPO
        tipo_get = request.GET.get('tipo_equipo', '').lower()
        tipos_validos = [t[0] for t in TIPOS_EQUIPO]
        if tipo_get and tipo_get not in tipos_validos:
            tipo_get = 'otro'
            
        ing_initial = {
            'fecha_ingreso': timezone.now().date(),
            'tecnico_encargado': request.user if request.user.groups.filter(name='Tecnicos').exists() else None,
            'tipo_equipo': tipo_get,
            'tipo_equipo_otro': request.GET.get('tipo_equipo_otro', ''),
            'marca': request.GET.get('marca', ''),
            'modelo_serie': request.GET.get('modelo_serie', ''),
            'serie': request.GET.get('serie', ''),
            'problema_reportado': request.GET.get('problema_reportado', ''),
            'accesorios_entregados': request.GET.get('accesorios_entregados', ''),
            'numero_factura': request.GET.get('numero_factura', ''),
            'asesor_comercial': request.GET.get('asesor_comercial', ''),
            'reporte_tecnico': request.GET.get('reporte_tecnico', ''),
            'diagnostico_inmediato': request.GET.get('diagnostico_inmediato', 'no'),
            'valor_diagnostico': request.GET.get('valor_diagnostico', '0.00'),
            'valor_acordado': request.GET.get('valor_acordado', ''),
            'abono_anticipo': request.GET.get('abono_anticipo', '0.00'),
        }
        cli_form = ClienteForm(prefix='cli', initial=cli_initial)
        ing_form = IngresoEquipoForm(prefix='ing', initial=ing_initial)

    # El siguiente código se calcula dentro de la sede actual
    from .models import SEDE_PREFIJOS
    siguiente_numero = IngresoEquipo.siguiente_numero_equipo(sede_actual)
    siguiente_codigo = f"{SEDE_PREFIJOS.get(sede_actual, '?')}{siguiente_numero}"

    return render(request, 'ingresos/form.html', {
        'cli_form': cli_form,
        'ing_form': ing_form,
        'modo': 'registrar',
        'titulo': 'Nueva Solicitud de Ingreso',
        'siguiente_numero': siguiente_numero,
        'siguiente_codigo': siguiente_codigo,
        'confirmar_mismo_equipo_cliente': confirmar_mismo_equipo_cliente,
    })


@tecnico_requerido
@transaction.atomic
def ingreso_editar(request, pk):
    """Edita un ingreso existente."""
    ingreso = get_object_or_404(IngresoEquipo, pk=pk)
    if ingreso.retirado_por_cliente:
        messages.warning(
            request,
            f'Ya este equipo fue retirado por el cliente. '
            f'La hoja de ingreso {ingreso.codigo_equipo} queda cerrada y no se puede editar.'
        )
        return redirect('econotec:ingreso_detalle', pk=ingreso.pk)

    identidad_original = _identidad_equipo_de_ingreso(ingreso)
    confirmar_mismo_equipo_cliente = (
        _confirmo_mismo_equipo_cliente(request) if request.method == 'POST' else False
    )

    # Mapeo subestado_entregado → estado_reparacion de SalidaEquipo
    _MAPA_SALIDA = {
        'con_solucion': {
            'estado_reparacion': 'pendiente_retiro',
            'cliente_recibe_conforme': 'si',
        },
        'sin_solucion': {
            'estado_reparacion': 'no_reparable',
            'cliente_recibe_conforme': 'no',
        },
        'no_quiso_reparar': {
            'estado_reparacion': 'cliente_no_acepta',
            'cliente_recibe_conforme': 'no',
        },
        'pendiente_retiro': {
            'estado_reparacion': 'pendiente_retiro',
            'cliente_recibe_conforme': 'si',
        },
    }

    if request.method == 'POST':
        cli_form = ClienteForm(request.POST, prefix='cli', instance=ingreso.cliente)
        ing_form = IngresoEquipoForm(request.POST, prefix='ing', instance=ingreso)
        valores_cliente_antes = _snapshot_form_original(cli_form)
        valores_ingreso_antes = _snapshot_form_original(ing_form)
        campos_cliente_cambiados = set(cli_form.changed_data or [])
        campos_ingreso_cambiados = set(ing_form.changed_data or [])
        estado_original = ingreso.estado
        subestado_reparacion_original = ingreso.subestado_reparacion
        subestado_entregado_original = ingreso.subestado_entregado
        reporte_original = (ingreso.reporte_tecnico or '').strip()
        if cli_form.is_valid() and ing_form.is_valid():
            cliente_editado = cli_form.save(commit=False)
            identidad_sin_cambios = (
                _identidad_equipo_normalizada(ing_form.cleaned_data)
                == identidad_original
            )
            duplicado = None
            if not identidad_sin_cambios:
                duplicado = _equipo_duplicado_para_cliente(
                    cliente_editado,
                    ing_form.cleaned_data,
                    excluir_pk=ingreso.pk,
                )
            if duplicado and not confirmar_mismo_equipo_cliente:
                mensaje_duplicado = (
                    'ESTE EQUIPO YA SE ENCUENTRA REGISTRADO, POR FAVOR VERIFICA EN LA LISTA DE EQUIPOS.'
                )
                ing_form.add_error('modelo_serie', f'{mensaje_duplicado} Coincide con el equipo {duplicado.codigo_equipo}.')
                messages.error(
                    request,
                    f'{mensaje_duplicado} Coincide con {duplicado.codigo_equipo}.'
                )
                return render(request, 'ingresos/form.html', {
                    'cli_form': cli_form,
                    'ing_form': ing_form,
                    'modo': 'editar',
                    'titulo': f'Editar equipo {ingreso.codigo_equipo}',
                    'siguiente_numero': ingreso.numero_equipo,
                    'siguiente_codigo': ingreso.codigo_equipo,
                    'ingreso': ingreso,
                    'confirmar_mismo_equipo_cliente': confirmar_mismo_equipo_cliente,
                })

            estado_nuevo = ing_form.cleaned_data.get('estado')
            valor_acordado_nuevo = ing_form.cleaned_data.get('valor_acordado')
            if estado_nuevo == 'entregado' and valor_acordado_nuevo is None:
                ing_form.add_error(
                    'valor_acordado',
                    'Por favor registra un valor acordado para finalizar el equipo.'
                )
                return render(request, 'ingresos/form.html', {
                    'cli_form': cli_form,
                    'ing_form': ing_form,
                    'modo': 'editar',
                    'titulo': f'Editar equipo {ingreso.codigo_equipo}',
                    'siguiente_numero': ingreso.numero_equipo,
                    'siguiente_codigo': ingreso.codigo_equipo,
                    'ingreso': ingreso,
                    'confirmar_mismo_equipo_cliente': confirmar_mismo_equipo_cliente,
                })

            cliente_editado.save()
            ingreso = ing_form.save()
            _sincronizar_egreso_compra(ingreso, request.user)

            # ── Auto-crear Salida si estado=entregado + subestado definido ──
            subestado = ingreso.subestado_entregado
            if ingreso.estado == 'entregado' and subestado in _MAPA_SALIDA:
                datos_salida = _MAPA_SALIDA[subestado]
                salida_existente = getattr(ingreso, 'salida', None)
                if salida_existente is None:
                    # Calcular saldo pendiente para sugerir el valor final
                    saldo = ingreso.diferencia
                    salida_auto = SalidaEquipo.objects.create(
                        ingreso=ingreso,
                        fecha_salida=date.today(),
                        estado_reparacion=datos_salida['estado_reparacion'],
                        cliente_recibe_conforme=datos_salida['cliente_recibe_conforme'],
                        valor_final_cobrado=saldo if saldo > 0 else 0,
                        metodo_pago_final='efectivo' if saldo > 0 else 'sin_pago',
                        registrado_por=request.user,
                    )
                    registrar_bitacora(
                        request.user,
                        'salida',
                        _texto_salida_bitacora(salida_auto),
                        ingreso=ingreso,
                        salida=salida_auto,
                        dedupe_key=f'salida:{salida_auto.pk}:creada',
                    )
                    if subestado == 'con_solucion':
                        etiqueta = 'Con solución — cliente conforme'
                    elif subestado == 'sin_solucion':
                        etiqueta = 'Sin solución — no se pudo reparar'
                    elif subestado == 'pendiente_retiro':
                        etiqueta = 'Pendiente de retiro'
                    else:
                        etiqueta = 'Sin reparación — cliente no quiso repararlo'
                    
                    messages.success(
                        request,
                        f'✅ Equipo {ingreso.codigo_equipo} actualizado. '
                        f'Equipo finalizado automáticamente: {etiqueta}.'
                    )
                else:
                    messages.success(request, f'Equipo {ingreso.codigo_equipo} actualizado.')
            else:
                messages.success(request, f'Equipo {ingreso.codigo_equipo} actualizado.')

            campos_ingreso_bitacora = campos_ingreso_cambiados - {'reporte_tecnico'}
            cambios_datos = campos_cliente_cambiados or campos_ingreso_bitacora
            if cambios_datos:
                registrar_bitacora(
                    request.user,
                    'ingreso_editado',
                    _texto_actualizacion_ingreso_bitacora(
                        ingreso,
                        cli_form,
                        ing_form,
                        campos_cliente_cambiados,
                        campos_ingreso_bitacora,
                        valores_cliente_antes,
                        valores_ingreso_antes,
                    ),
                    ingreso=ingreso,
                )

            if (ingreso.reporte_tecnico or '').strip() != reporte_original:
                reporte = _texto_limpio_bitacora(ingreso.reporte_tecnico)
                if reporte:
                    registrar_bitacora(
                        request.user,
                        'reporte',
                        f'Actualización de reporte técnico en {_equipo_bitacora(ingreso)} #{ingreso.codigo_equipo}: {reporte}.',
                        ingreso=ingreso,
                    )

            if (
                not cambios_datos
                and (
                    ingreso.estado != estado_original
                    or ingreso.subestado_reparacion != subestado_reparacion_original
                    or ingreso.subestado_entregado != subestado_entregado_original
                )
            ):
                registrar_bitacora(
                    request.user,
                    'estado',
                    _texto_actualizacion_ingreso_bitacora(
                        ingreso,
                        cli_form,
                        ing_form,
                        set(),
                        {'estado', 'subestado_reparacion', 'subestado_entregado'},
                        valores_cliente_antes,
                        valores_ingreso_antes,
                    ),
                    ingreso=ingreso,
                )

            return redirect('econotec:ingreso_detalle', pk=ingreso.pk)
    else:
        cli_form = ClienteForm(prefix='cli', instance=ingreso.cliente)
        ing_form = IngresoEquipoForm(prefix='ing', instance=ingreso)

    return render(request, 'ingresos/form.html', {
        'cli_form': cli_form,
        'ing_form': ing_form,
        'ingreso': ingreso,
        'modo': 'editar',
        'titulo': f'Editar Equipo {ingreso.codigo_equipo}',
        'siguiente_numero': ingreso.numero_equipo,
        'siguiente_codigo': ingreso.codigo_equipo,
        'confirmar_mismo_equipo_cliente': confirmar_mismo_equipo_cliente,
    })


@tecnico_requerido
def ingreso_lista(request):
    """Listado de ingresos con filtros.
    Por defecto se filtra por la sede actual de la sesión, pero el usuario
    puede ver "Todas" o cambiar entre Guayaquil/Quito con el filtro.
    """
    q = (request.GET.get('q') or '').strip()
    estado = (request.GET.get('estado') or '').strip()
    tipo = (request.GET.get('tipo') or '').strip()
    valor = (request.GET.get('valor') or '').strip()
    firma = (request.GET.get('firma') or '').strip()
    fecha_desde, fecha_hasta, fecha_preset = obtener_rango_fecha(request)

    # Filtro por sede:
    # - Si el querystring trae explícitamente sede=todas → no filtrar
    # - Si trae sede=guayaquil/quito → filtrar por esa
    # - Si NO viene en el querystring → usar la sede de la sesión
    sede_sesion = (request.session.get('sede_actual') or '').strip().lower()
    if 'sede' in request.GET:
        sede_filtro = (request.GET.get('sede') or '').strip().lower()
    else:
        sede_filtro = sede_sesion

    qs = (ingresos_operativos_qs()
          .select_related('cliente', 'registrado_por', 'salida')
          .prefetch_related('abonos'))

    if sede_filtro in ('guayaquil', 'quito'):
        qs = qs.filter(sede=sede_filtro)
    # si sede_filtro es 'todas' o vacío → no se filtra

    estados_salida_filtro = {
        'salida_pendiente_retiro': 'pendiente_retiro',
        'salida_entregado_cliente': 'retirado',
        'salida_cliente_no_acepta': 'cliente_no_acepta',
        'salida_no_reparable': 'no_reparable',
        'salida_garantia': 'garantia',
    }

    subestados_reparacion_filtro = {
        'reparacion_en_reparacion': 'en_reparacion',
        'espera_cliente': 'espera_cliente',
        'espera_repuesto': 'espera_repuesto',
    }

    if estado:
        if estado in subestados_reparacion_filtro:
            qs = qs.filter(
                estado='en_reparacion',
                subestado_reparacion=subestados_reparacion_filtro[estado],
            )
        elif estado == 'con_salida':
            qs = qs.filter(salida__isnull=False)
        elif estado == 'salida_entregado_cliente':
            qs = qs.filter(salida__fecha_retiro_real__isnull=False)
        elif estado in estados_salida_filtro:
            qs = qs.filter(salida__estado_reparacion=estados_salida_filtro[estado])
            if estado == 'salida_pendiente_retiro':
                qs = qs.filter(salida__fecha_retiro_real__isnull=True)
        else:
            qs = qs.filter(estado=estado)
    if tipo:
        qs = qs.filter(tipo_equipo=tipo)

    if valor == 'pendiente':
        qs = qs.filter(
            sede__in=['guayaquil', 'quito'],
            valor_acordado__isnull=True,
        ).exclude(estado='entregado')
    elif valor == 'con_valor':
        qs = qs.filter(valor_acordado__isnull=False)

    if firma == 'con_firma':
        qs = qs.filter(firma_cliente=True).exclude(firma_cliente_imagen='')
    elif firma == 'sin_firma':
        qs = qs.filter(Q(firma_cliente=False) | Q(firma_cliente_imagen=''))

    qs = aplicar_rango_fecha(qs, 'fecha_ingreso', fecha_desde, fecha_hasta)

    tecnico_filtro = (request.GET.get('tecnico') or '').strip()
    registrador_filtro = (request.GET.get('registrador') or '').strip()
    asesor_filtro = (request.GET.get('asesor') or '').strip()

    if tecnico_filtro.isdigit():
        qs = qs.filter(tecnico_encargado_id=tecnico_filtro)
    if registrador_filtro.isdigit():
        qs = qs.filter(registrado_por_id=registrador_filtro)
    if asesor_filtro:
        qs = qs.filter(asesor_comercial=asesor_filtro)

    qs = filtrar_objetos_normalizado(qs, q, texto_ingreso_busqueda)

    from django.contrib.auth import get_user_model
    User = get_user_model()
    usuarios_all = User.objects.filter(is_active=True).order_by('first_name', 'username')
    from .forms import _queryset_asesores
    asesores_qs = _queryset_asesores()
    asesores_choices = [f'{u.first_name} {u.last_name}'.strip() or u.username for u in asesores_qs]

    estados_filtro = [
        ('', '— Estado —'),
        ('ingresado', 'Ingresado / En diagnóstico'),
        ('en_reparacion', 'En reparación (Todos)'),
        ('reparacion_en_reparacion', '   ↳ En reparación'),
        ('espera_cliente', '   ↳ En reparación - Cliente'),
        ('espera_repuesto', '   ↳ En reparación - Repuestos'),
        ('entregado', 'Entregado al cliente (Ingreso)'),
        ('garantia', 'Garantía (Ingreso)'),
        ('con_salida', 'Equipo finalizado (Todos)'),
        ('salida_pendiente_retiro', '   ↳ Reparado - pendiente de retiro'),
        ('salida_entregado_cliente', '   ↳ Entregado / retirado por cliente'),
        ('salida_cliente_no_acepta', '   ↳ Cliente no quiso reparar'),
        ('salida_no_reparable', '   ↳ No se pudo reparar'),
        ('salida_garantia', '   ↳ Garantía finalizada'),
    ]

    total = total_resultados(qs)
    page_obj, querystring = paginar_resultados(request, qs)

    context = {
        'ingresos': page_obj.object_list,
        'page_obj': page_obj,
        'querystring': querystring,
        'q': q,
        'estado_filtro': estado,
        'tipo_filtro': tipo,
        'valor_filtro': valor,
        'firma_filtro': firma,
        'sede_filtro': sede_filtro,
        'sede_sesion': sede_sesion,
        'tecnico_filtro': tecnico_filtro,
        'registrador_filtro': registrador_filtro,
        'asesor_filtro': asesor_filtro,
        'usuarios_all': usuarios_all,
        'asesores_choices': asesores_choices,
        'estados': estados_filtro,
        'tipos': IngresoEquipo._meta.get_field('tipo_equipo').choices,
        'total': total,
    }
    context.update(contexto_rango_fecha(
        fecha_desde,
        fecha_hasta,
        fecha_preset,
        etiqueta='Fecha ingreso',
    ))
    return render(request, 'ingresos/lista.html', context)


@tecnico_requerido
def ingreso_detalle(request, pk):
    """Vista de detalle de un ingreso con todas sus relaciones."""
    ingreso = get_object_or_404(
        IngresoEquipo.objects.select_related(
            'cliente',
            'registrado_por',
            'tecnico_encargado',
            'equipo_garantia',
            'valor_pendiente_reporte_por',
        ),
        pk=pk,
    )
    abonos = ingreso.abonos.all().order_by('-fecha', '-creado')
    salida = getattr(ingreso, 'salida', None)
    reparacion_check_fecha = timezone.localdate()
    usuario_bitacora_check = _usuario_reparacion_check_bitacora(request.user, ingreso)
    mostrar_check_reparacion = _ingreso_permite_reparacion_check(ingreso) and usuario_bitacora_check is not None
    reparacion_check_hecho = False
    if mostrar_check_reparacion:
        reparacion_check_hecho = _reparacion_check_ya_registrado(
            usuario_bitacora_check,
            ingreso,
            reparacion_check_fecha,
        )
    from .qr_utils import qr_data_uri_para_ingreso, url_hoja_movil
    return render(request, 'ingresos/detalle.html', {
        'ingreso': ingreso,
        'abonos': abonos,
        'salida': salida,
        'qr_data_uri': qr_data_uri_para_ingreso(request, ingreso, box_size=6),
        'qr_url': url_hoja_movil(request, ingreso),
        'wa_link': whatsapp_link_hoja_ingreso(request, ingreso),
        'mostrar_check_reparacion': mostrar_check_reparacion,
        'reparacion_check_hecho': reparacion_check_hecho,
        'reparacion_check_fecha': reparacion_check_fecha.isoformat(),
        'reparacion_check_tecnico_nombre': nombre_corto_usuario(usuario_bitacora_check),
    })


@tecnico_requerido
@require_POST
def ingreso_reparacion_check(request, pk):
    """Marca una vez al día que el técnico sigue reparando el equipo."""
    ingreso = get_object_or_404(
        IngresoEquipo.objects.select_related('cliente', 'tecnico_encargado'),
        pk=pk,
    )
    if not _ingreso_permite_reparacion_check(ingreso):
        return JsonResponse({
            'ok': False,
            'error': 'Este check solo aplica cuando el equipo está en En reparación -> En reparación o Garantía.',
        }, status=400)

    dia = timezone.localdate()
    usuario_bitacora = _usuario_reparacion_check_bitacora(request.user, ingreso)
    if usuario_bitacora is None:
        return JsonResponse({
            'ok': False,
            'error': 'No hay un técnico asignado para registrar este check.',
        }, status=400)

    dedupe_key = _reparacion_check_dedupe_key(usuario_bitacora, ingreso, dia)
    if _reparacion_check_ya_registrado(usuario_bitacora, ingreso, dia):
        return JsonResponse({
            'ok': True,
            'already': True,
            'message': 'Ya diste el check de hoy para este equipo.',
            'fecha': dia.isoformat(),
            'bitacora_total': construir_bitacora_usuario(usuario_bitacora)['total'],
        })

    evento = registrar_bitacora(
        usuario_bitacora,
        'estado',
        _texto_reparacion_check_bitacora(ingreso, usuario_bitacora),
        ingreso=ingreso,
        dedupe_key=dedupe_key,
        metadata={'accion': 'reparacion_check', 'dia': dia.isoformat()},
    )
    return JsonResponse({
        'ok': True,
        'already': False,
        'message': 'Check registrado en la bitácora de hoy.',
        'texto': evento.texto if evento else '',
        'fecha': dia.isoformat(),
        'bitacora_total': construir_bitacora_usuario(usuario_bitacora)['total'],
    })


@admin_requerido
@require_POST
def ingreso_eliminar(request, pk):
    """Solo admin: eliminar ingreso (con confirmación)."""
    ingreso = get_object_or_404(IngresoEquipo, pk=pk)
    numero = ingreso.numero_equipo
    if hasattr(ingreso, 'salida'):
        messages.error(
            request,
            f'No se puede eliminar el equipo #{numero}: ya está finalizado. '
            'Elimina primero la finalización.'
        )
        return redirect('econotec:ingreso_detalle', pk=ingreso.pk)
    if hasattr(ingreso, 'egreso_compra'):
        messages.error(
            request,
            f'No se puede eliminar el equipo #{numero}: tiene un egreso automático de compra. '
            'Conserva el registro para no perder el historial administrativo.'
        )
        return redirect('econotec:ingreso_detalle', pk=ingreso.pk)
    codigo = ingreso.codigo_equipo
    cliente = ingreso.cliente.nombres
    registrar_bitacora(
        request.user,
        'eliminacion',
        f'Equipo eliminado del sistema: #{codigo} de {cliente}.',
        ingreso=ingreso,
        codigo=codigo,
    )
    ingreso.delete()
    messages.success(request, f'Equipo #{numero} eliminado.')
    return redirect('econotec:ingreso_lista')


# ═════════════════════════════════════════════════════════════════
# Ventas de Productos
# ═════════════════════════════════════════════════════════════════

VENTA_INVENTARIO_UBICACIONES = [
    {'slug': 'guayaquil_norte', 'nombre': 'Guayaquil - Norte', 'sede': 'guayaquil'},
    {'slug': 'guayaquil_centro', 'nombre': 'Guayaquil - Centro', 'sede': 'guayaquil'},
    {'slug': 'quito', 'nombre': 'Quito', 'sede': 'quito'},
]


def _venta_inventario_ubicacion_nombre(slug, default):
    for opcion in VENTA_INVENTARIO_UBICACIONES:
        if opcion['slug'] == slug:
            return opcion['nombre']
    return default


def _venta_observacion_producto_limpia(valor):
    texto = str(valor or '').replace('\r\n', '\n').replace('\r', '\n').strip()
    if len(texto) > 500:
        raise ValueError('La observación del producto no puede superar 500 caracteres.')
    return texto


def _inventario_item_venta_json(item, cantidad=None, relacion_id=None, observacion=''):
    categoria_info = INVENTARIO_CATEGORIAS.get(item.categoria, {})
    categoria_nombre = categoria_info.get('nombre', item.categoria)
    tipo_info = _inventario_tipo_por_slug(categoria_info, item.tipo) if categoria_info else None
    seleccionable = item.estado == 'disponible' and item.cantidad > 0
    try:
        observacion_venta = _venta_observacion_producto_limpia(observacion)
    except ValueError:
        observacion_venta = ''
    return {
        'relacion_id': relacion_id,
        'item_id': item.pk,
        'codigo': item.codigo,
        'producto': item.producto,
        'categoria': categoria_nombre,
        'tipo': (tipo_info or {}).get('nombre', item.tipo),
        'marca': item.marca,
        'modelo': item.modelo,
        'serie': item.serie,
        'sede': item.get_sede_display(),
        'ubicacion': _venta_inventario_ubicacion_nombre(item.ubicacion, item.get_ubicacion_display()),
        'cantidad': item.cantidad if cantidad is None else cantidad,
        'disponible': item.cantidad,
        'costo': f'{(item.costo or D("0.00")):.2f}',
        'observacion': observacion_venta,
        'observacion_inventario': item.observacion or '',
        'estado': item.estado,
        'estado_label': item.get_estado_display(),
        'causa_no_disponible': item.causa_no_disponible,
        'causa_no_disponible_label': item.get_causa_no_disponible_display() if item.causa_no_disponible else '',
        'seleccionable': seleccionable,
    }


def _venta_inventario_item_json(relacion):
    return _inventario_item_venta_json(
        relacion.inventario_item,
        cantidad=relacion.cantidad,
        relacion_id=relacion.pk,
        observacion=relacion.observacion,
    )


def _venta_inventario_categorias_json():
    return [
        {
            'slug': slug,
            'nombre': info['nombre'],
            'tipos': [
                {'slug': tipo_slug, 'nombre': tipo_nombre}
                for tipo_slug, tipo_nombre in info['tipos']
            ],
        }
        for slug, info in INVENTARIO_CATEGORIAS.items()
    ]


def _venta_inventario_selecciones(post_data):
    """Lee la selección temporal del formulario de nueva venta."""
    raw = (post_data.get('inventario_seleccionado') or '').strip()
    if not raw:
        return []
    try:
        datos = json.loads(raw)
    except (TypeError, ValueError):
        raise ValueError('La selección de inventario no es válida. Vuelve a intentarlo.')
    if not isinstance(datos, list):
        raise ValueError('La selección de inventario no es válida. Vuelve a intentarlo.')

    resultado = []
    vistos = set()
    for dato in datos:
        if not isinstance(dato, dict):
            raise ValueError('La selección de inventario no es válida. Vuelve a intentarlo.')
        try:
            item_id = int(dato.get('item_id'))
            cantidad = int(dato.get('cantidad'))
        except (TypeError, ValueError):
            raise ValueError('La cantidad seleccionada no es válida.')
        observacion = _venta_observacion_producto_limpia(dato.get('observacion'))
        if item_id in vistos or cantidad < 1:
            raise ValueError('No se puede repetir un producto ni seleccionar una cantidad inválida.')
        vistos.add(item_id)
        resultado.append((item_id, cantidad, observacion))
    return resultado


def _venta_inventario_contexto_desde_post(post_data):
    """Reconstruye la seleccion visible si el formulario vuelve con errores."""
    try:
        selecciones = _venta_inventario_selecciones(post_data)
    except ValueError:
        return []
    if not selecciones:
        return []
    items = InventarioItem.objects.in_bulk([item_id for item_id, _cantidad, _observacion in selecciones])
    contexto = []
    for item_id, cantidad, observacion in selecciones:
        item = items.get(item_id)
        if item:
            contexto.append(_inventario_item_venta_json(item, cantidad=cantidad, observacion=observacion))
    return contexto


def _venta_usa_valor_desde_inventario(post_data):
    return (post_data.get('venta_valor_desde_inventario') or '').strip() == 'si'


def _venta_total_desde_inventario(selecciones):
    items = InventarioItem.objects.in_bulk([item_id for item_id, _cantidad, _observacion in selecciones])
    total = D('0.00')
    for item_id, cantidad, _observacion in selecciones:
        item = items.get(item_id)
        if not item:
            continue
        total += (item.costo or D('0.00')) * D(cantidad)
    return total.quantize(D('0.01'))


def _resumen_venta_inventario(venta):
    relaciones = venta.productos_inventario.select_related('inventario_item').all()
    return ', '.join(
        f'{relacion.cantidad} x {relacion.inventario_item.producto}'
        for relacion in relaciones
    )


def _aplicar_inventario_a_venta(venta, selecciones):
    """Descuenta stock y crea las relaciones dentro de la transacción de venta."""
    for item_id, cantidad, observacion in selecciones:
        item = InventarioItem.objects.select_for_update().filter(pk=item_id).first()
        if not item:
            raise ValueError('Uno de los productos seleccionados ya no existe en inventario.')
        if item.estado != 'disponible' or item.cantidad < cantidad:
            raise ValueError(
                f'No hay suficiente disponibilidad para «{item.producto}». '
                f'Solo quedan {item.cantidad} unidad(es).'
            )
        item.cantidad -= cantidad
        item.save(update_fields=['cantidad', 'actualizado'])
        VentaInventarioItem.objects.create(
            venta=venta,
            inventario_item=item,
            cantidad=cantidad,
            observacion=observacion,
        )


def _actualizar_observaciones_inventario_venta(venta, selecciones):
    """Actualiza solo las notas de productos ya vinculados a una venta."""
    relaciones = {
        relacion.inventario_item_id: relacion
        for relacion in VentaInventarioItem.objects.filter(venta=venta)
    }
    for item_id, _cantidad, observacion in selecciones:
        relacion = relaciones.get(item_id)
        if relacion and relacion.observacion != observacion:
            relacion.observacion = observacion
            relacion.save(update_fields=['observacion', 'actualizado'])


def _devolver_inventario_de_venta(venta, relacion_ids=None):
    """Devuelve al inventario las unidades de una relación o de toda la venta."""
    relaciones = VentaInventarioItem.objects.select_for_update().select_related('inventario_item').filter(
        venta=venta,
    )
    if relacion_ids is not None:
        relaciones = relaciones.filter(pk__in=relacion_ids)
    for relacion in relaciones:
        item = InventarioItem.objects.select_for_update().get(pk=relacion.inventario_item_id)
        item.cantidad += relacion.cantidad
        item.save(update_fields=['cantidad', 'actualizado'])
        relacion.delete()


@tecnico_requerido
@require_GET
def venta_inventario_catalogo(request):
    """Catálogo compacto para el selector de inventario de ventas."""
    q = ' '.join((request.GET.get('q') or '').strip().split())
    ubicacion = (request.GET.get('ubicacion') or '').strip()
    sede = (request.GET.get('sede') or '').strip()
    categoria = (request.GET.get('categoria') or '').strip()
    tipo = (request.GET.get('tipo') or '').strip()
    items = InventarioItem.objects.filter(
        sede__in=['guayaquil', 'quito'],
    ).order_by('producto', 'marca', 'modelo')
    ubicaciones_validas = {opcion['slug'] for opcion in VENTA_INVENTARIO_UBICACIONES}
    if ubicacion in ubicaciones_validas:
        items = items.filter(ubicacion=ubicacion)
    elif sede in INVENTARIO_SEDES:
        items = items.filter(sede=sede)
    if categoria in INVENTARIO_CATEGORIAS:
        items = items.filter(categoria=categoria)
    if tipo:
        items = items.filter(tipo=tipo)
    if q:
        items = items.filter(
            Q(producto__icontains=q)
            | Q(codigo__icontains=q)
            | Q(categoria__icontains=q)
            | Q(tipo__icontains=q)
            | Q(marca__icontains=q)
            | Q(modelo__icontains=q)
            | Q(observacion__icontains=q)
            | Q(causa_no_disponible__icontains=q)
        )
    datos = [_inventario_item_venta_json(item) for item in items[:200]]
    return JsonResponse({'items': datos})


@tecnico_requerido
@require_POST
@transaction.atomic
def venta_inventario_agregar(request, pk):
    """Añade un producto a una venta ya creada y descuenta sus unidades."""
    venta = get_object_or_404(IngresoEquipo, pk=pk, sede='ventas')
    try:
        item_id = int(request.POST.get('item_id'))
        cantidad = int(request.POST.get('cantidad'))
    except (TypeError, ValueError):
        return JsonResponse({'ok': False, 'error': 'Selecciona una cantidad válida.'}, status=400)
    if cantidad < 1:
        return JsonResponse({'ok': False, 'error': 'La cantidad debe ser mayor que cero.'}, status=400)
    try:
        observacion = _venta_observacion_producto_limpia(request.POST.get('observacion'))
    except ValueError as exc:
        return JsonResponse({'ok': False, 'error': str(exc)}, status=400)

    item = get_object_or_404(InventarioItem.objects.select_for_update(), pk=item_id)
    if item.estado != 'disponible' or item.cantidad < cantidad:
        return JsonResponse({
            'ok': False,
            'error': f'Solo quedan {item.cantidad} unidad(es) disponibles de «{item.producto}».',
        }, status=409)

    relacion, creada = VentaInventarioItem.objects.select_for_update().get_or_create(
        venta=venta,
        inventario_item=item,
        defaults={'cantidad': cantidad, 'observacion': observacion},
    )
    if not creada:
        relacion.cantidad += cantidad
        update_fields = ['cantidad', 'actualizado']
        if observacion:
            relacion.observacion = observacion
            update_fields.append('observacion')
        relacion.save(update_fields=update_fields)
    item.cantidad -= cantidad
    item.save(update_fields=['cantidad', 'actualizado'])
    return JsonResponse({
        'ok': True,
        'producto': _venta_inventario_item_json(relacion),
        'disponible': item.cantidad,
    })


@tecnico_requerido
@require_POST
@transaction.atomic
def venta_inventario_actualizar_cantidad(request, pk, relacion_pk):
    """Actualiza la cantidad vendida y compensa la diferencia en inventario."""
    venta = get_object_or_404(IngresoEquipo, pk=pk, sede='ventas')
    relacion = get_object_or_404(
        VentaInventarioItem.objects.select_for_update(),
        pk=relacion_pk,
        venta=venta,
    )
    try:
        nueva_cantidad = int(request.POST.get('cantidad'))
    except (TypeError, ValueError):
        return JsonResponse({'ok': False, 'error': 'Selecciona una cantidad válida.'}, status=400)
    if nueva_cantidad < 1:
        return JsonResponse({'ok': False, 'error': 'La cantidad debe ser mayor que cero.'}, status=400)

    item = InventarioItem.objects.select_for_update().get(pk=relacion.inventario_item_id)
    cantidad_actual = relacion.cantidad
    diferencia = nueva_cantidad - cantidad_actual

    if diferencia > 0:
        if item.estado != 'disponible' or item.cantidad < diferencia:
            return JsonResponse({
                'ok': False,
                'error': f'Solo puedes aumentar {item.cantidad} unidad(es) adicionales de «{item.producto}».',
            }, status=409)
        item.cantidad -= diferencia
    elif diferencia < 0:
        item.cantidad += abs(diferencia)

    if diferencia:
        item.save(update_fields=['cantidad', 'actualizado'])
        relacion.cantidad = nueva_cantidad
        relacion.save(update_fields=['cantidad', 'actualizado'])

    return JsonResponse({
        'ok': True,
        'producto': _venta_inventario_item_json(relacion),
        'disponible': item.cantidad,
    })


@tecnico_requerido
@require_POST
@transaction.atomic
def venta_inventario_actualizar_observacion(request, pk, relacion_pk):
    """Actualiza la observación de un producto vendido sin mover stock."""
    venta = get_object_or_404(IngresoEquipo, pk=pk, sede='ventas')
    relacion = get_object_or_404(
        VentaInventarioItem.objects.select_for_update().select_related('inventario_item'),
        pk=relacion_pk,
        venta=venta,
    )
    try:
        observacion = _venta_observacion_producto_limpia(request.POST.get('observacion'))
    except ValueError as exc:
        return JsonResponse({'ok': False, 'error': str(exc)}, status=400)

    relacion.observacion = observacion
    relacion.save(update_fields=['observacion', 'actualizado'])
    return JsonResponse({
        'ok': True,
        'producto': _venta_inventario_item_json(relacion),
    })


@tecnico_requerido
@require_POST
@transaction.atomic
def venta_inventario_quitar(request, pk, relacion_pk):
    """Quita un producto de una venta y devuelve sus unidades al inventario."""
    venta = get_object_or_404(IngresoEquipo, pk=pk, sede='ventas')
    relacion = get_object_or_404(
        VentaInventarioItem.objects.select_for_update(),
        pk=relacion_pk,
        venta=venta,
    )
    cantidad_devuelta = relacion.cantidad
    item = InventarioItem.objects.select_for_update().get(pk=relacion.inventario_item_id)
    item.cantidad += cantidad_devuelta
    item.save(update_fields=['cantidad', 'actualizado'])
    relacion.delete()
    return JsonResponse({'ok': True, 'cantidad_devuelta': cantidad_devuelta})

@tecnico_requerido
def venta_menu(request):
    """Menú de ventas: registrar nueva / ver lista."""
    ventas = list(IngresoEquipo.objects.filter(sede='ventas').prefetch_related('abonos'))
    total = len(ventas)
    total_parciales = sum(1 for venta in ventas if _venta_con_pago_parcial(venta))
    total_completas = total - total_parciales
    return render(request, 'ventas/menu.html', {
        'total': total,
        'total_parciales': total_parciales,
        'total_completas': total_completas,
    })


def _venta_con_pago_parcial(venta):
    """
    Una venta pertenece al flujo parcial si nació con abono inicial menor al
    valor total o si ya tiene abonos posteriores. Se mantiene en esa lista
    aunque luego el saldo quede pagado.
    """
    if venta.sede != 'ventas':
        return False
    valor = venta.valor_efectivo_a_cobrar or D('0.00')
    anticipo = venta.abono_anticipo or D('0.00')
    return venta.abonos.exists() or (valor > D('0.00') and anticipo < valor)


def _preparar_post_venta(post_data):
    """
    Completa los campos del formulario de ingreso que no se muestran en ventas.
    Ventas reutiliza IngresoEquipoForm, pero omite diagnóstico/técnico/estado
    en la pantalla; sin estos defaults el formulario puede fallar en silencio.
    """
    defaults_si_falta = {
        'ing-marca': 'N/A',
        'ing-modelo_serie': 'N/A',
        'ing-serie': '',
        'ing-tipo_equipo': 'otro',
        'ing-tipo_equipo_otro': '',
        'ing-accesorios_entregados': 'Ninguno',
        'ing-abono_anticipo': '0.00',
        'ing-compra_metodo_pago': 'efectivo',
        'ing-compra_banco': '',
        'ing-compra_banco_otro': '',
        'ing-compra_tarjeta_app': '',
        'ing-compra_comprobante_url': '',
        'ing-compra_monto_1': '',
        'ing-compra_metodo_1': '',
        'ing-compra_banco_1': '',
        'ing-compra_banco_otro_1': '',
        'ing-compra_tarjeta_app_1': '',
        'ing-compra_monto_2': '',
        'ing-compra_metodo_2': '',
        'ing-compra_banco_2': '',
        'ing-compra_banco_otro_2': '',
        'ing-compra_tarjeta_app_2': '',
        'ing-anticipo_metodo': 'efectivo',
        'ing-anticipo_banco': '',
        'ing-anticipo_banco_otro': '',
        'ing-anticipo_tarjeta_app': '',
        'ing-anticipo_comprobante_url': '',
        'ing-anticipo_monto_1': '',
        'ing-anticipo_metodo_1': '',
        'ing-anticipo_banco_1': '',
        'ing-anticipo_monto_2': '',
        'ing-anticipo_metodo_2': '',
        'ing-anticipo_banco_2': '',
        'ing-equipo_garantia': '',
        'ing-equipo_garantia_manual': '',
        'ing-motivo_garantia': '',
        'ing-firma_cliente_opcion': 'no',
        'ing-firma_cliente_imagen': '',
    }
    for campo, valor in defaults_si_falta.items():
        if not post_data.get(campo):
            post_data[campo] = valor

    if _venta_usa_valor_desde_inventario(post_data):
        try:
            total_inventario = _venta_total_desde_inventario(_venta_inventario_selecciones(post_data))
        except ValueError:
            total_inventario = None
        if total_inventario is not None:
            post_data['ing-valor_acordado'] = f'{total_inventario:.2f}'

    # En ventas de producto el pago puede ser directo o por abono.
    # Pago directo: el abono inicial refleja el valor total a cobrar.
    # Pago por abono: se respeta el monto parcial ingresado por el usuario.
    valor_venta = (post_data.get('ing-valor_acordado') or '').strip()
    modalidad_pago = (post_data.get('venta_pago_modalidad') or 'directo').strip()
    if valor_venta and modalidad_pago != 'abono':
        post_data['ing-abono_anticipo'] = valor_venta.replace(',', '.')

    # El diagnóstico no aplica a ventas, pero IngresoEquipoForm exige el método.
    post_data['ing-diagnostico_inmediato'] = 'no'
    post_data['ing-valor_diagnostico'] = '0.00'
    post_data['ing-diagnostico_metodo'] = 'efectivo'
    post_data['ing-diagnostico_banco'] = ''
    post_data['ing-diagnostico_banco_otro'] = ''
    post_data['ing-diagnostico_tarjeta_app'] = ''
    post_data['ing-diagnostico_comprobante_url'] = ''
    post_data['ing-diagnostico_monto_1'] = ''
    post_data['ing-diagnostico_metodo_1'] = ''
    post_data['ing-diagnostico_banco_1'] = ''
    post_data['ing-diagnostico_monto_2'] = ''
    post_data['ing-diagnostico_metodo_2'] = ''
    post_data['ing-diagnostico_banco_2'] = ''

    # IngresoEquipoForm oculta la opción "entregado"; validamos como ingreso
    # normal y luego forzamos el estado final de venta antes de guardar.
    post_data['ing-estado'] = 'ingresado'
    post_data['ing-subestado_reparacion'] = ''
    post_data['ing-subestado_entregado'] = ''


def _configurar_form_venta(ing_form):
    ing_form.fields['tecnico_encargado'].required = True
    ing_form.fields['tecnico_encargado'].label = 'Técnico vendió'
    ing_form.fields['tecnico_encargado'].empty_label = '— Selecciona el técnico que vendió —'
    if 'equipo_garantia' in ing_form.fields:
        ing_form.fields['equipo_garantia'].required = False
    # En ventas el detalle se genera desde los productos elegidos en inventario.
    # El campo queda oculto para conservar listados, exportes y bitacora.
    ing_form.fields['problema_reportado'].required = False
    ing_form.fields['problema_reportado'].widget.attrs['placeholder'] = 'Ej.: 1 Tinta Epson Negra, 2 Cables USB'


def _venta_pago_contexto(post_data=None, venta=None):
    if post_data is not None:
        factura_realizada = (post_data.get('venta_factura_realizada') or 'no').strip()
        modalidad = (post_data.get('venta_pago_modalidad') or 'directo').strip()
        return {
            'modalidad': modalidad if modalidad in ('directo', 'abono') else 'directo',
            'valor_desde_inventario': _venta_usa_valor_desde_inventario(post_data),
            'factura_realizada': factura_realizada if factura_realizada in ('si', 'no') else 'no',
            'factura_nombres': (post_data.get('venta_factura_nombres') or '').strip(),
            'factura_cedula': (post_data.get('venta_factura_cedula') or '').strip(),
            'factura_correo': (post_data.get('venta_factura_correo') or '').strip(),
        }
    if venta is not None:
        valor = venta.valor_efectivo_a_cobrar or D('0.00')
        abono = venta.abono_anticipo or D('0.00')
        modalidad = 'abono' if (valor > D('0.00') and abono < valor) or venta.abonos.exists() else 'directo'
        return {
            'modalidad': modalidad,
            'valor_desde_inventario': False,
            'factura_realizada': venta.factura_realizada or 'no',
            'factura_nombres': venta.factura_nombres or '',
            'factura_cedula': venta.factura_cedula or '',
            'factura_correo': venta.factura_correo or '',
        }
    return {
        'modalidad': 'directo',
        'valor_desde_inventario': True,
        'factura_realizada': 'no',
        'factura_nombres': '',
        'factura_cedula': '',
        'factura_correo': '',
    }


def _agregar_error_pago_venta(ing_form, mensaje):
    ing_form.add_error(None, mensaje)


def _validar_pago_venta(post_data, ing_form):
    valor = ing_form.cleaned_data.get('valor_acordado')
    if valor is None or valor <= D('0.00'):
        ing_form.add_error('valor_acordado', 'Ingresa el valor total de la venta.')
        return

    modalidad_pago = (post_data.get('venta_pago_modalidad') or 'directo').strip()
    if modalidad_pago not in ('directo', 'abono'):
        modalidad_pago = 'directo'

    abono_inicial = ing_form.cleaned_data.get('abono_anticipo') or D('0.00')
    if modalidad_pago == 'abono':
        if abono_inicial <= D('0.00'):
            ing_form.add_error('abono_anticipo', 'Ingresa el monto del abono inicial.')
        if abono_inicial >= valor:
            ing_form.add_error(
                'abono_anticipo',
                'Para pago por abono, el monto inicial debe ser menor al valor total de la venta.',
            )
    else:
        abono_inicial = valor

    monto_a_validar = abono_inicial

    metodo = ing_form.cleaned_data.get('anticipo_metodo')
    metodos_validos = {codigo for codigo, _ in IngresoEquipo.METODOS_PAGO}
    if metodo not in metodos_validos:
        _agregar_error_pago_venta(ing_form, 'Selecciona el método de pago de la venta.')
        return

    banco = ing_form.cleaned_data.get('anticipo_banco')
    banco_otro = (ing_form.cleaned_data.get('anticipo_banco_otro') or '').strip()
    tarjeta_app = ing_form.cleaned_data.get('anticipo_tarjeta_app')

    if metodo == 'transferencia':
        if not banco:
            _agregar_error_pago_venta(ing_form, 'Indica el banco usado para la transferencia.')
        if banco == 'otro' and not banco_otro:
            _agregar_error_pago_venta(ing_form, 'Escribe el nombre del banco usado en la transferencia.')
    elif metodo == 'tarjeta':
        if not tarjeta_app:
            _agregar_error_pago_venta(ing_form, 'Selecciona la aplicación o tarjeta usada para el pago.')
    elif metodo == 'mixto':
        monto_1 = ing_form.cleaned_data.get('anticipo_monto_1') or D('0.00')
        monto_2 = ing_form.cleaned_data.get('anticipo_monto_2') or D('0.00')
        metodo_1 = ing_form.cleaned_data.get('anticipo_metodo_1')
        metodo_2 = ing_form.cleaned_data.get('anticipo_metodo_2')
        banco_1 = ing_form.cleaned_data.get('anticipo_banco_1')
        banco_2 = ing_form.cleaned_data.get('anticipo_banco_2')

        if monto_1 <= 0:
            _agregar_error_pago_venta(ing_form, 'Ingresa el primer monto del pago mixto.')
        if monto_2 <= 0:
            _agregar_error_pago_venta(ing_form, 'Ingresa el segundo monto del pago mixto.')
        if not metodo_1 or metodo_1 == 'mixto':
            _agregar_error_pago_venta(ing_form, 'Selecciona un método válido para la primera parte del pago mixto.')
        elif metodo_1 == 'transferencia' and not banco_1:
            _agregar_error_pago_venta(ing_form, 'Selecciona el banco de la primera transferencia.')
        if not metodo_2 or metodo_2 == 'mixto':
            _agregar_error_pago_venta(ing_form, 'Selecciona un método válido para la segunda parte del pago mixto.')
        elif metodo_2 == 'transferencia' and not banco_2:
            _agregar_error_pago_venta(ing_form, 'Selecciona el banco de la segunda transferencia.')
        if (monto_1 + monto_2) != monto_a_validar:
            _agregar_error_pago_venta(
                ing_form,
                f'La suma del pago mixto debe ser igual al monto cobrado ahora: ${monto_a_validar:.2f}.',
            )

    factura = _venta_pago_contexto(post_data)
    if factura['factura_realizada'] == 'si':
        if not factura['factura_nombres']:
            _agregar_error_pago_venta(ing_form, 'Completa los nombres o razón social para la factura.')
        if not factura['factura_cedula']:
            _agregar_error_pago_venta(ing_form, 'Completa la cédula o RUC para la factura.')
        if not factura['factura_correo']:
            _agregar_error_pago_venta(ing_form, 'Completa el correo para la factura.')
        else:
            try:
                validate_email(factura['factura_correo'])
            except ValidationError:
                _agregar_error_pago_venta(ing_form, 'Ingresa un correo válido para la factura.')


def _limpiar_pago_venta(venta):
    metodo = venta.anticipo_metodo
    if metodo == 'efectivo':
        venta.anticipo_banco = ''
        venta.anticipo_banco_otro = ''
        venta.anticipo_tarjeta_app = ''
        venta.anticipo_comprobante_url = ''
    elif metodo == 'transferencia':
        venta.anticipo_tarjeta_app = ''
    elif metodo == 'tarjeta':
        venta.anticipo_banco = ''
        venta.anticipo_banco_otro = ''
        venta.anticipo_comprobante_url = ''

    if metodo != 'mixto':
        venta.anticipo_monto_1 = None
        venta.anticipo_metodo_1 = ''
        venta.anticipo_banco_1 = ''
        venta.anticipo_monto_2 = None
        venta.anticipo_metodo_2 = ''
        venta.anticipo_banco_2 = ''
    else:
        venta.anticipo_banco = ''
        venta.anticipo_banco_otro = ''
        venta.anticipo_tarjeta_app = ''
        venta.anticipo_comprobante_url = ''


def _aplicar_pago_venta(venta, post_data):
    modalidad_pago = (post_data.get('venta_pago_modalidad') or 'directo').strip()
    if modalidad_pago != 'abono':
        venta.abono_anticipo = venta.valor_acordado or D('0.00')
    _limpiar_pago_venta(venta)

    factura = _venta_pago_contexto(post_data)
    venta.factura_realizada = factura['factura_realizada']
    if factura['factura_realizada'] == 'si':
        venta.factura_nombres = factura['factura_nombres']
        venta.factura_cedula = factura['factura_cedula']
        venta.factura_correo = factura['factura_correo']
    else:
        venta.factura_nombres = ''
        venta.factura_cedula = ''
        venta.factura_correo = ''


@tecnico_requerido
@transaction.atomic
def venta_registrar(request):
    """Registra una nueva venta de producto."""
    if request.method == 'POST':
        post_data = request.POST.copy()
        _preparar_post_venta(post_data)

        cli_form = ClienteForm(post_data, prefix='cli')
        ing_form = IngresoEquipoForm(post_data, prefix='ing')
        _configurar_form_venta(ing_form)

        cedula = (post_data.get('cli-cedula') or '').strip()
        cliente_existente = Cliente.objects.filter(cedula=cedula).first() if cedula else None

        cliente = None
        venta = None

        if cliente_existente:
            cli_form_existente = ClienteForm(post_data, prefix='cli', instance=cliente_existente)
            cli_form = cli_form_existente
        else:
            cli_form_existente = cli_form

        if cli_form_existente.is_valid() and ing_form.is_valid():
            _validar_pago_venta(post_data, ing_form)

        if cli_form_existente.is_valid() and ing_form.is_valid():
            try:
                selecciones = _venta_inventario_selecciones(post_data)
                if not selecciones:
                    raise ValueError('Selecciona al menos un producto del inventario para registrar la venta.')
                with transaction.atomic():
                    cliente = cli_form_existente.save()
                    venta = ing_form.save(commit=False)
                    venta.cliente = cliente
                    venta.sede = 'ventas'
                    venta.registrado_por = request.user
                    venta.estado = 'entregado'
                    venta.subestado_entregado = 'con_solucion'
                    _aplicar_pago_venta(venta, post_data)
                    venta.save()
                    _aplicar_inventario_a_venta(venta, selecciones)
                    if selecciones and not (venta.problema_reportado or '').strip():
                        venta.problema_reportado = _resumen_venta_inventario(venta)
                        venta.save(update_fields=['problema_reportado', 'actualizado'])
            except ValueError as exc:
                ing_form.add_error(None, str(exc))
                cliente = None
                venta = None
            else:
                registrar_bitacora(
                    request.user,
                    'venta_producto',
                    _texto_venta_bitacora(venta),
                    ingreso=venta,
                    dedupe_key=f'venta:{venta.pk}:creada',
                )

                messages.success(request, f'Venta {venta.codigo_equipo} registrada para {cliente.nombres}.')
                if _venta_con_pago_parcial(venta):
                    return redirect('econotec:venta_lista_parciales')
                return redirect('econotec:venta_lista')
            
    else:
        cli_form = ClienteForm(prefix='cli')
        initial = {
            'fecha_ingreso': date.today(),
            'estado': 'entregado',
            'subestado_entregado': 'con_solucion', # Asumimos la venta está finalizada y entregada
            'diagnostico_inmediato': 'no',
            'accesorios_entregados': 'Ninguno',
            'marca': 'N/A',
            'modelo_serie': 'N/A',
            'serie': '',
            'tipo_equipo': 'otro',
            'firma_cliente_opcion': 'no',
        }
        if es_tecnico(request.user):
            initial['tecnico_encargado'] = request.user
        ing_form = IngresoEquipoForm(prefix='ing', initial=initial)
        _configurar_form_venta(ing_form)

    from .models import SEDE_PREFIJOS
    siguiente_numero = IngresoEquipo.siguiente_numero_equipo('ventas')
    siguiente_codigo = f"{siguiente_numero:03d}"
    inventario_seleccionados_json = (
        _venta_inventario_contexto_desde_post(request.POST)
        if request.method == 'POST' else []
    )

    return render(request, 'ventas/form.html', {
        'cli_form': cli_form,
        'ing_form': ing_form,
        'venta': None,
        'inventario_seleccionados_json': inventario_seleccionados_json,
        'venta_inventario_sedes_json': VENTA_INVENTARIO_UBICACIONES,
        'venta_inventario_categorias_json': _venta_inventario_categorias_json(),
        'venta_pago_form': _venta_pago_contexto(request.POST if request.method == 'POST' else None),
        'modo': 'registrar',
        'titulo': 'Nueva Venta de Producto',
        'siguiente_numero': siguiente_numero,
        'siguiente_codigo': siguiente_codigo,
    })

@tecnico_requerido
@transaction.atomic
def venta_editar(request, pk):
    """Edita una venta existente."""
    venta = get_object_or_404(IngresoEquipo, pk=pk, sede='ventas')

    if request.method == 'POST':
        post_data = request.POST.copy()
        _preparar_post_venta(post_data)
            
        cli_form = ClienteForm(post_data, prefix='cli', instance=venta.cliente)
        ing_form = IngresoEquipoForm(post_data, prefix='ing', instance=venta)
        _configurar_form_venta(ing_form)
        selecciones = []
        
        if cli_form.is_valid() and ing_form.is_valid():
            _validar_pago_venta(post_data, ing_form)
            try:
                selecciones = _venta_inventario_selecciones(post_data)
            except ValueError as exc:
                ing_form.add_error(None, str(exc))

        if cli_form.is_valid() and ing_form.is_valid():
            campos_cambiados = set(cli_form.changed_data or []) | set(ing_form.changed_data or [])
            cli_form.save()
            venta = ing_form.save(commit=False)
            venta.sede = 'ventas'
            venta.estado = 'entregado'
            venta.subestado_entregado = 'con_solucion'
            _aplicar_pago_venta(venta, post_data)
            venta.save()
            _actualizar_observaciones_inventario_venta(venta, selecciones)
            if campos_cambiados:
                registrar_bitacora(
                    request.user,
                    'venta_editada',
                    f'Venta de producto actualizada: #{venta.codigo_equipo} para {venta.cliente.nombres}.',
                    ingreso=venta,
                )
            messages.success(request, f'Venta {venta.codigo_equipo} actualizada.')
            if _venta_con_pago_parcial(venta):
                return redirect('econotec:venta_lista_parciales')
            return redirect('econotec:venta_lista')
    else:
        cli_form = ClienteForm(prefix='cli', instance=venta.cliente)
        ing_form = IngresoEquipoForm(prefix='ing', instance=venta)
        
    _configurar_form_venta(ing_form)

    relaciones_inventario = venta.productos_inventario.select_related('inventario_item').all()

    return render(request, 'ventas/form.html', {
        'cli_form': cli_form,
        'ing_form': ing_form,
        'venta': venta,
        'inventario_seleccionados_json': [
            _venta_inventario_item_json(relacion) for relacion in relaciones_inventario
        ],
        'venta_inventario_sedes_json': VENTA_INVENTARIO_UBICACIONES,
        'venta_inventario_categorias_json': _venta_inventario_categorias_json(),
        'venta_pago_form': _venta_pago_contexto(request.POST if request.method == 'POST' else None, venta),
        'modo': 'editar',
        'titulo': f'Editar Venta {venta.codigo_equipo}',
        'siguiente_numero': venta.numero_equipo,
        'siguiente_codigo': venta.codigo_equipo,
    })

@admin_requerido
@require_POST
def venta_eliminar(request, pk):
    """Elimina una venta."""
    venta = get_object_or_404(IngresoEquipo, pk=pk, sede='ventas')
    codigo = venta.codigo_equipo
    cliente = venta.cliente.nombres
    descripcion = _texto_limpio_bitacora(venta.problema_reportado, max_len=120)
    registrar_bitacora(
        request.user,
        'eliminacion',
        f'Venta de producto eliminada: {descripcion} #{codigo} de {cliente}.',
        ingreso=venta,
        codigo=codigo,
    )
    with transaction.atomic():
        _devolver_inventario_de_venta(venta)
        venta.delete()
    messages.success(request, 'Venta eliminada correctamente.')
    return redirect('econotec:venta_lista')

@tecnico_requerido
def venta_export(request):
    """Exportar ventas a Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from io import BytesIO
    q = (request.GET.get('q') or '').strip()
    tecnico_vendio_filtro = (request.GET.get('tecnico_vendio') or '').strip()
    registrador_filtro = (request.GET.get('registrador') or '').strip()
    pago_filtro = (request.GET.get('pago') or '').strip()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Ventas Econotec'

    headers = ['Código', 'Fecha', 'Cliente', 'Cédula', 'Descripción', 'Técnico vendió', 'Registrado por', 'Valor']
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='F97618')
        c.alignment = Alignment(horizontal='center')

    ventas = (
        IngresoEquipo.objects
        .select_related('cliente', 'tecnico_encargado', 'registrado_por')
        .prefetch_related('abonos')
        .filter(sede='ventas')
        .order_by('-fecha_ingreso', '-pk')
    )
    if tecnico_vendio_filtro.isdigit():
        ventas = ventas.filter(tecnico_encargado_id=tecnico_vendio_filtro)
    if registrador_filtro.isdigit():
        ventas = ventas.filter(registrado_por_id=registrador_filtro)
    ventas = filtrar_objetos_normalizado(ventas, q, texto_ingreso_busqueda)
    ventas = list(ventas)
    if pago_filtro == 'parcial':
        ventas = [v for v in ventas if _venta_con_pago_parcial(v)]
    elif pago_filtro != 'todos':
        ventas = [v for v in ventas if not _venta_con_pago_parcial(v)]

    for row, v in enumerate(ventas, start=2):
        ws.cell(row=row, column=1, value=v.codigo_equipo)
        ws.cell(row=row, column=2, value=v.fecha_ingreso.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=3, value=v.cliente.nombres)
        ws.cell(row=row, column=4, value=v.cliente.cedula)
        ws.cell(row=row, column=5, value=v.problema_reportado)
        tecnico_vendio = f"{v.tecnico_encargado.first_name} {v.tecnico_encargado.last_name}".strip() if v.tecnico_encargado else 'N/A'
        tecnico_vendio = tecnico_vendio or (v.tecnico_encargado.username if v.tecnico_encargado else 'N/A')
        ws.cell(row=row, column=6, value=tecnico_vendio)
        registrador = f"{v.registrado_por.first_name} {v.registrado_por.last_name}".strip() if v.registrado_por else 'N/A'
        registrador = registrador or (v.registrado_por.username if v.registrado_por else 'N/A')
        ws.cell(row=row, column=7, value=registrador)
        ws.cell(row=row, column=8, value=v.valor_acordado)

    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 22

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="ventas_econotec.xlsx"'
    return response


@tecnico_requerido
def venta_lista(request):
    """Listado de ventas pagadas completas."""
    return _render_venta_lista(request, filtro_pago='completo')


@tecnico_requerido
def venta_lista_parciales(request):
    """Listado exclusivo de ventas que se registraron con pago parcial."""
    return _render_venta_lista(request, filtro_pago='parcial')


def _render_venta_lista(request, filtro_pago='completo'):
    """Renderiza lista de ventas completa o filtrada por pago parcial."""
    q = (request.GET.get('q') or '').strip()
    tecnico_vendio_filtro = (request.GET.get('tecnico_vendio') or '').strip()
    registrador_filtro = (request.GET.get('registrador') or '').strip()

    qs = (IngresoEquipo.objects
          .select_related('cliente', 'tecnico_encargado', 'registrado_por')
          .prefetch_related('abonos')
          .filter(sede='ventas')
          .order_by('-fecha_ingreso', '-pk'))

    if tecnico_vendio_filtro.isdigit():
        qs = qs.filter(tecnico_encargado_id=tecnico_vendio_filtro)
    if registrador_filtro.isdigit():
        qs = qs.filter(registrado_por_id=registrador_filtro)

    qs = filtrar_objetos_normalizado(qs, q, texto_ingreso_busqueda)

    from django.contrib.auth import get_user_model
    User = get_user_model()
    usuarios_all = User.objects.filter(is_active=True).order_by('first_name', 'username')
    from .forms import _queryset_tecnicos
    tecnicos_solo = _queryset_tecnicos()
    ventas = list(qs)
    if filtro_pago == 'parcial':
        ventas = [venta for venta in ventas if _venta_con_pago_parcial(venta)]
    else:
        ventas = [venta for venta in ventas if not _venta_con_pago_parcial(venta)]
    total = len(ventas)
    page_obj, querystring = paginar_resultados(request, ventas)
    for venta in page_obj.object_list:
        venta.wa_venta_link = whatsapp_link_venta_producto(venta)

    export_params = request.GET.copy()
    export_params.pop('pagina', None)
    if filtro_pago == 'parcial':
        export_params['pago'] = 'parcial'
    else:
        export_params['pago'] = 'completo'

    return render(request, 'ventas/lista.html', {
        'ingresos': page_obj.object_list,
        'page_obj': page_obj,
        'querystring': querystring,
        'q': q,
        'tecnico_vendio_filtro': tecnico_vendio_filtro,
        'registrador_filtro': registrador_filtro,
        'usuarios_all': usuarios_all,
        'tecnicos_solo': tecnicos_solo,
        'total': total,
        'lista_pago_parcial': filtro_pago == 'parcial',
        'lista_pago_completo': filtro_pago != 'parcial',
        'export_querystring': export_params.urlencode(),
    })


# ═════════════════════════════════════════════════════════════════
# Salida de Equipos
# ═════════════════════════════════════════════════════════════════

@tecnico_requerido
def salida_menu(request):
    """Menú de salidas."""
    total = SalidaEquipo.objects.filter(fecha_retiro_real__isnull=True).count()
    listos_para_entregar = SalidaEquipo.objects.filter(
        estado_reparacion='pendiente_retiro',
        fecha_retiro_real__isnull=True,
    ).count()
    facturas_realizadas = SalidaEquipo.objects.filter(
        factura_realizada='si',
    ).count()
    return render(request, 'salidas/menu.html', {
        'total': total,
        'listos_para_entregar': listos_para_entregar,
        'facturas_realizadas': facturas_realizadas,
    })


@tecnico_requerido
def salida_lista(request, solo_fuera_oficina=False):
    """Lista separada de equipos en oficina o con salida física confirmada."""
    q = (request.GET.get('q') or '').strip()
    estado = (request.GET.get('estado') or '').strip()
    sede_filtro = (request.GET.get('sede') or '').strip().lower()
    tecnico_registro_filtro = (request.GET.get('tecnico_registro') or '').strip()
    tecnico_salida_filtro = (request.GET.get('tecnico_salida') or '').strip()
    fecha_desde, fecha_hasta, fecha_preset = obtener_rango_fecha(request)

    # Cada lista ofrece únicamente estados coherentes con su etapa física.
    # Dentro de la oficina se conservan los resultados de finalización;
    # fuera de la oficina solo corresponde el estado de salida confirmada.
    if solo_fuera_oficina:
        estados_filtro = [
            e for e in SalidaEquipo.ESTADO_REPARACION
            if e[0] == 'retirado'
        ]
    else:
        estados_filtro = [
            e for e in SalidaEquipo.ESTADO_REPARACION
            if e[0] not in ('retirado', 'chatarrerizacion')
        ]
    estados_validos = {valor for valor, _etiqueta in estados_filtro}
    if estado not in estados_validos:
        estado = ''

    qs = SalidaEquipo.objects.select_related(
        'ingreso', 'ingreso__cliente', 'registrado_por', 'tecnico_reparo',
    )
    if solo_fuera_oficina:
        qs = qs.filter(fecha_retiro_real__isnull=False).order_by(
            '-fecha_retiro_real', '-fecha_salida', '-creado',
        )
        campo_fecha_filtro = 'fecha_retiro_real'
        etiqueta_fecha_filtro = 'Fecha de salida física'
    else:
        qs = qs.filter(fecha_retiro_real__isnull=True).order_by(
            '-fecha_salida', '-creado',
        )
        campo_fecha_filtro = 'fecha_salida'
        etiqueta_fecha_filtro = 'Fecha de finalización'

    if estado and not (solo_fuera_oficina and estado == 'retirado'):
        qs = qs.filter(estado_reparacion=estado)
    if sede_filtro in ('guayaquil', 'quito'):
        qs = qs.filter(ingreso__sede=sede_filtro)
    if tecnico_registro_filtro:
        qs = qs.filter(registrado_por_id=tecnico_registro_filtro)
    if tecnico_salida_filtro:
        qs = qs.filter(tecnico_reparo_id=tecnico_salida_filtro)

    qs = aplicar_rango_fecha(qs, campo_fecha_filtro, fecha_desde, fecha_hasta)

    qs = filtrar_objetos_normalizado(qs, q, texto_salida_busqueda)

    from django.contrib.auth import get_user_model
    User = get_user_model()
    tecnicos_all = User.objects.filter(is_active=True).order_by('first_name', 'username')
    from .forms import _queryset_tecnicos
    tecnicos_solo = _queryset_tecnicos()

    total = total_resultados(qs)
    page_obj, querystring = paginar_resultados(request, qs)

    context = {
        'salidas': page_obj.object_list,
        'page_obj': page_obj,
        'querystring': querystring,
        'q': q,
        'estado_filtro': estado,
        'sede_filtro': sede_filtro,
        'tecnico_registro_filtro': tecnico_registro_filtro,
        'tecnico_salida_filtro': tecnico_salida_filtro,
        'tecnicos_all': tecnicos_all,
        'tecnicos_solo': tecnicos_solo,
        'estados': estados_filtro,
        'total': total,
        'lista_salidas_confirmadas': solo_fuera_oficina,
    }
    context.update(contexto_rango_fecha(
        fecha_desde,
        fecha_hasta,
        fecha_preset,
        etiqueta=etiqueta_fecha_filtro,
    ))
    return render(request, 'salidas/lista.html', context)


@tecnico_requerido
@transaction.atomic
def salida_registrar(request, ingreso_pk):
    """Registrar la salida de un equipo (cierre del ciclo de reparación)."""
    ingreso = get_object_or_404(IngresoEquipo, pk=ingreso_pk)

    if ingreso.estado in ('donado', 'equipo_a_comprar'):
        messages.warning(
            request,
            f'El equipo {ingreso.codigo_equipo} está en "{ingreso.get_estado_display()}" '
            'y se gestiona directamente desde el Registro Administrativo. No usa la finalización de reparación.'
        )
        return redirect('econotec:ingreso_detalle', pk=ingreso.pk)

    if hasattr(ingreso, 'salida'):
        messages.info(
            request,
            f'El equipo {ingreso.codigo_equipo} ya está finalizado. '
            'Puedes editarla aquí.'
        )
        return redirect('econotec:salida_editar', pk=ingreso.salida.pk)

    if ingreso.valor_acordado is None:
        messages.warning(
            request,
            'Por favor registra un valor acordado para finalizar el equipo.'
        )
        return redirect('econotec:ingreso_detalle', pk=ingreso.pk)

    if request.method == 'POST':
        salida_inst = SalidaEquipo(ingreso=ingreso)
        form = SalidaEquipoForm(request.POST, instance=salida_inst)
        if form.is_valid():
            salida = form.save(commit=False)
            salida.registrado_por = request.user
            salida.save()
            _sincronizar_notificacion_asesora(form, salida, request.user)
            registrar_bitacora(
                request.user,
                'salida',
                _texto_salida_bitacora(salida),
                ingreso=ingreso,
                salida=salida,
                dedupe_key=f'salida:{salida.pk}:creada',
            )
            messages.success(
                request,
                f'Equipo {ingreso.codigo_equipo} finalizado como '
                f'"{salida.get_estado_reparacion_display()}".'
            )
            # La pantalla siguiente confirma, una sola vez, si el equipo sigue
            # dentro de la oficina o si su salida física ya ocurrió.
            request.session['confirmar_ubicacion_salida_id'] = salida.pk
            return redirect('econotec:salida_listo_aviso', pk=salida.pk)
    else:
        # Saldo pendiente sugerido como valor a cobrar
        saldo = ingreso.diferencia
        estado_salida_inicial = 'cortesia' if ingreso.estado == 'cortesia' else 'pendiente_retiro'
        metodo_salida_inicial = 'cortesia' if ingreso.estado == 'cortesia' else 'efectivo'
        salida_inst = SalidaEquipo(ingreso=ingreso)
        form = SalidaEquipoForm(instance=salida_inst, initial={
            'fecha_salida': date.today(),
            'estado_reparacion': estado_salida_inicial,
            'metodo_pago_final': metodo_salida_inicial,
            'valor_final_cobrado': 0,
            'tecnico_reparo': request.user if es_tecnico(request.user) else None,
        })

    return render(request, 'salidas/form.html', {
        'form': form,
        'ingreso': ingreso,
        'modo': 'registrar',
        'titulo': f'Finalizar Equipo — {ingreso.codigo_equipo}',
    })


@tecnico_requerido
@transaction.atomic
def salida_editar(request, pk):
    salida = get_object_or_404(
        SalidaEquipo.objects.select_related('ingreso', 'ingreso__cliente'),
        pk=pk,
    )
    if request.method == 'POST':
        estado_anterior = salida.estado_reparacion
        tecnico_anterior = salida.tecnico_reparo_id
        valor_anterior = salida.valor_final_cobrado
        form = SalidaEquipoForm(request.POST, instance=salida)
        if form.is_valid():
            salida = form.save()
            _sincronizar_notificacion_asesora(form, salida, request.user)
            if (
                salida.estado_reparacion != estado_anterior
                or salida.tecnico_reparo_id != tecnico_anterior
                or salida.valor_final_cobrado != valor_anterior
                or form.changed_data
            ):
                registrar_bitacora(
                    request.user,
                    'salida_editada',
                    f'Finalización actualizada en #{salida.ingreso.codigo_equipo}: {salida.get_estado_reparacion_display()}.',
                    ingreso=salida.ingreso,
                    salida=salida,
                )
            messages.success(request, 'Finalización actualizada correctamente.')
            request.session['confirmar_ubicacion_salida_id'] = salida.pk
            return redirect('econotec:salida_listo_aviso', pk=salida.pk)
    else:
        form = SalidaEquipoForm(instance=salida)
    return render(request, 'salidas/form.html', {
        'form': form,
        'ingreso': salida.ingreso,
        'salida': salida,
        'modo': 'editar',
        'titulo': f'Editar Finalización — Equipo {salida.ingreso.codigo_equipo}',
        # Si la salida ya está marcada como positiva (retirado/garantía/parcial)
        # y el cliente tiene WhatsApp, generamos el link para reenviar el aviso
        # junto con el PDF de la hoja de salida.
        'wa_link': whatsapp_link_equipo_listo(salida),
    })


@login_required
def notificaciones_asesora(request):
    admin_mode = es_admin(request.user)
    if not (admin_mode or es_asesor(request.user)):
        messages.warning(request, 'No tienes acceso a las notificaciones de asesoras.')
        return redirect('econotec:bienvenida')

    qs = (
        NotificacionAsesora.objects
        .select_related('ingreso', 'ingreso__cliente', 'salida', 'asesora', 'creado_por')
    )
    asesora_filtro_id = None
    asesoras_filtro = []

    if admin_mode:
        from django.contrib.auth import get_user_model

        ids_asesoras = (
            NotificacionAsesora.objects
            .exclude(asesora_id__isnull=True)
            .values_list('asesora_id', flat=True)
            .distinct()
        )
        asesoras_filtro = (
            get_user_model().objects
            .filter(pk__in=ids_asesoras)
            .order_by('first_name', 'last_name', 'username')
        )

        asesora_param = (request.GET.get('asesora') or '').strip()
        if asesora_param and asesora_param != 'todas':
            try:
                asesora_filtro_id = int(asesora_param)
            except (TypeError, ValueError):
                asesora_filtro_id = None
            if asesora_filtro_id:
                qs = qs.filter(asesora_id=asesora_filtro_id)
    else:
        qs = qs.filter(asesora=request.user)

    total_bandeja = qs.count()
    total_pendientes = qs.filter(leida=False).count()
    total_vistas = qs.filter(leida=True).count()

    estado = (request.GET.get('estado') or 'pendientes').strip()
    if estado == 'vistas':
        qs = qs.filter(leida=True)
    elif estado != 'todas':
        estado = 'pendientes'
        qs = qs.filter(leida=False)

    return render(request, 'notificaciones/asesoras.html', {
        'notificaciones': qs,
        'estado_filtro': estado,
        'total_notificaciones': total_resultados(qs),
        'total_bandeja': total_bandeja,
        'total_pendientes': total_pendientes,
        'total_vistas': total_vistas,
        'admin_notificaciones': admin_mode,
        'asesoras_filtro': asesoras_filtro,
        'asesora_filtro_id': asesora_filtro_id,
    })


@login_required
@require_POST
def notificacion_asesora_marcar_vista(request, pk):
    admin_mode = es_admin(request.user)
    if not (admin_mode or es_asesor(request.user)):
        messages.warning(request, 'No tienes acceso a las notificaciones de asesoras.')
        return redirect('econotec:bienvenida')

    qs = NotificacionAsesora.objects.all() if admin_mode else NotificacionAsesora.objects.filter(asesora=request.user)
    notificacion = get_object_or_404(qs, pk=pk)
    notificacion.leida = True
    notificacion.leida_en = timezone.now()
    notificacion.save(update_fields=['leida', 'leida_en', 'actualizado'])
    if admin_mode:
        messages.success(request, 'Notificación marcada como gestionada.')
    else:
        messages.success(request, 'Notificación marcada como vista.')

    next_url = request.POST.get('next') or ''
    if next_url and url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return redirect(next_url)
    return redirect('econotec:notificaciones_asesora')


@login_required
@require_POST
def notificacion_asesora_limpiar_bandeja(request):
    if not es_asesor(request.user):
        messages.warning(request, 'No tienes acceso a las notificaciones de asesoras.')
        return redirect('econotec:bienvenida')

    total, _ = NotificacionAsesora.objects.filter(asesora=request.user).delete()
    if total:
        messages.success(request, f'Bandeja limpiada. Se eliminaron {total} notificación(es).')
    else:
        messages.info(request, 'La bandeja ya estaba vacía.')
    return redirect('econotec:notificaciones_asesora')


@admin_requerido
@require_POST
def salida_eliminar(request, pk):
    salida = get_object_or_404(SalidaEquipo, pk=pk)
    ingreso = salida.ingreso
    salida.delete()
    # Volver el equipo al estado anterior
    ingreso.estado = 'en_reparacion'
    ingreso.save(update_fields=['estado'])
    messages.success(
        request,
        f'Finalización del equipo {ingreso.codigo_equipo} eliminada. '
        'El equipo vuelve a estado "Pendiente de retiro".'
    )
    return redirect('econotec:salida_lista')


# ═════════════════════════════════════════════════════════════════
# Clientes
# ═════════════════════════════════════════════════════════════════

@tecnico_requerido
def cliente_lista(request):
    q = (request.GET.get('q') or '').strip()
    sede_filtro = (request.GET.get('sede') or '').strip().lower()

    qs = Cliente.objects.prefetch_related('ingresos').annotate(
        equipos_total=Count('ingresos'),
    ).order_by('nombres')

    if sede_filtro in ('guayaquil', 'quito'):
        qs = qs.filter(ingresos__sede=sede_filtro).distinct()

    qs = filtrar_objetos_normalizado(qs, q, texto_cliente_busqueda)
    total = total_resultados(qs)
    page_obj, querystring = paginar_resultados(request, qs)
    return render(request, 'clientes/lista.html', {
        'clientes': page_obj.object_list,
        'page_obj': page_obj,
        'querystring': querystring,
        'q': q,
        'sede_filtro': sede_filtro,
        'total': total,
    })


@tecnico_requerido
def cliente_top_recurrentes(request):
    """Ranking de clientes recurrentes separados por sede (Top 10 cada una)."""
    clientes_guayaquil = (
        Cliente.objects
        .annotate(total_ingresos=Count(
            'ingresos',
            filter=Q(ingresos__sede='guayaquil'),
            distinct=True,
        ))
        .filter(total_ingresos__gt=0)
        .order_by('-total_ingresos', 'nombres')[:10]
    )
    clientes_quito = (
        Cliente.objects
        .annotate(total_ingresos=Count(
            'ingresos',
            filter=Q(ingresos__sede='quito'),
            distinct=True,
        ))
        .filter(total_ingresos__gt=0)
        .order_by('-total_ingresos', 'nombres')[:10]
    )

    return render(request, 'clientes/top.html', {
        'clientes_guayaquil': clientes_guayaquil,
        'clientes_quito': clientes_quito,
    })


@tecnico_requerido
def cliente_detalle(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    ingresos = (cliente.ingresos
                .select_related('cliente')
                .prefetch_related('abonos', 'salida')
                .order_by('-fecha_ingreso'))

    total_pagado = sum((ing.total_abonado for ing in ingresos), 0)
    total_acordado = sum((ing.valor_acordado or 0 for ing in ingresos), 0)

    return render(request, 'clientes/detalle.html', {
        'cliente': cliente,
        'ingresos': ingresos,
        'total_equipos': ingresos.count(),
        'total_pagado': total_pagado,
        'total_acordado': total_acordado,
    })


@tecnico_requerido
def cliente_export(request):
    """Exportar clientes a Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Clientes Econotec'

    headers = ['Cédula/RUC', 'Nombres', 'WhatsApp', 'Correo', 'Sector', 'Equipos', 'Registrado']
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='F97618')
        c.alignment = Alignment(horizontal='center')

    clientes = Cliente.objects.annotate(
        equipos_total=Count('ingresos'),
    ).order_by('nombres')

    for row, cli in enumerate(clientes, start=2):
        ws.cell(row=row, column=1, value=cli.cedula)
        ws.cell(row=row, column=2, value=cli.nombres)
        ws.cell(row=row, column=3, value=cli.whatsapp)
        ws.cell(row=row, column=4, value=cli.correo)
        ws.cell(row=row, column=5, value=cli.sector_display)
        ws.cell(row=row, column=6, value=cli.equipos_total)
        ws.cell(row=row, column=7, value=cli.creado.strftime('%d/%m/%Y'))

    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 22

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="clientes_econotec.xlsx"'
    return response


# ═════════════════════════════════════════════════════════════════
# Ranking de Técnicos (totales por técnico)
# ═════════════════════════════════════════════════════════════════

from .permisos import (
    puede_ver_valores_ranking as _puede_ver_valores_ranking,
    ranking_requerido as _ranking_requerido,
)

@_ranking_requerido
def salida_totales(request):
    """
    Ranking de técnicos por salidas: cuenta al técnico que reparó el equipo,
    registrado en `SalidaEquipo.tecnico_reparo`.

    Filtros opcionales por rango de fechas.
    """
    from decimal import Decimal as D

    desde = (request.GET.get('desde') or '').strip()
    hasta = (request.GET.get('hasta') or '').strip()
    mostrar_valores_monetarios = _puede_ver_valores_ranking(request.user)

    estados_positivos = SALIDA_BUENA_ESTADOS
    estados_negativos = SALIDA_MALA_ESTADOS

    # Base de ingresos de equipos. Las ventas de productos no pertenecen al
    # ranking de técnicos responsables de un ingreso de reparación.
    qs_ing = IngresoEquipo.objects.filter(
        sede__in=SEDES_EQUIPOS,
    ).select_related('tecnico_encargado')
    if desde:
        qs_ing = qs_ing.filter(fecha_ingreso__gte=desde)
    if hasta:
        qs_ing = qs_ing.filter(fecha_ingreso__lte=hasta)

    total_ingresos_ranking = qs_ing.count()
    ranking_ingresos_qs = (
        qs_ing
        .order_by()
        .values(
            'tecnico_encargado_id',
            'tecnico_encargado__first_name',
            'tecnico_encargado__last_name',
            'tecnico_encargado__username',
        )
        .annotate(
            num_ingresos=Count('id'),
            sin_salida=Count('id', filter=Q(salida__isnull=True)),
            con_salida=Count('id', filter=Q(salida__isnull=False)),
        )
    )
    if mostrar_valores_monetarios:
        ranking_ingresos_qs = ranking_ingresos_qs.annotate(
            total_acordado=Sum('valor_acordado'),
            total_anticipo=Sum('abono_anticipo'),
        ).order_by('-num_ingresos', '-total_acordado')
    else:
        ranking_ingresos_qs = ranking_ingresos_qs.order_by(
            '-num_ingresos',
            'tecnico_encargado__username',
        )
    ranking_ingresos = []
    for posicion, row in enumerate(ranking_ingresos_qs, start=1):
        tid = row['tecnico_encargado_id']
        nombre = (
            f"{row['tecnico_encargado__first_name'] or ''} "
            f"{row['tecnico_encargado__last_name'] or ''}"
        ).strip() or row['tecnico_encargado__username'] or '— Sin asignar —'
        item_ranking_ingreso = {
            'posicion': posicion,
            'tecnico_id': tid,
            'nombre': nombre,
            'sin_asignar': tid is None,
            'num_ingresos': row['num_ingresos'],
            'sin_salida': row['sin_salida'],
            'con_salida': row['con_salida'],
            'participacion': round(
                (row['num_ingresos'] / total_ingresos_ranking) * 100,
                1,
            ) if total_ingresos_ranking else 0,
        }
        if mostrar_valores_monetarios:
            item_ranking_ingreso.update({
                'total_acordado': row['total_acordado'] or D('0.00'),
                'total_anticipo': row['total_anticipo'] or D('0.00'),
            })
        ranking_ingresos.append(item_ranking_ingreso)

    # Base del ranking: salidas en el rango. La productividad se atribuye al
    # técnico que reparó en la salida, no al técnico encargado del ingreso.
    salidas_ranking_qs = SalidaEquipo.objects.select_related('ingreso', 'tecnico_reparo')
    if desde:
        salidas_ranking_qs = salidas_ranking_qs.filter(fecha_salida__gte=desde)
    if hasta:
        salidas_ranking_qs = salidas_ranking_qs.filter(fecha_salida__lte=hasta)

    ranking = (
        salidas_ranking_qs
        .values('tecnico_reparo_id', 'tecnico_reparo__first_name',
                'tecnico_reparo__last_name', 'tecnico_reparo__username')
        .annotate(
            num_equipos=Count('id'),
            # La ubicación física se determina por la fecha real de retiro.
            # El resultado de reparación puede tener otros valores válidos
            # (revisión, no reparable, cortesía, etc.) y no debe confundirse
            # con que el equipo continúe o no dentro de la oficina.
            entregados=Count('id', filter=Q(fecha_retiro_real__isnull=False)),
            pendientes=Count('id', filter=Q(fecha_retiro_real__isnull=True)),
        )
    )
    if mostrar_valores_monetarios:
        ranking = ranking.annotate(
            total_acordado=Sum('ingreso__valor_acordado'),
            total_anticipo=Sum('ingreso__abono_anticipo'),
        ).order_by('-num_equipos', '-total_acordado')
    else:
        ranking = ranking.order_by('-num_equipos', 'tecnico_reparo__username')

    ranking_list = []
    for row in ranking:
        tid = row['tecnico_reparo_id']
        nombre = (
            f"{row['tecnico_reparo__first_name'] or ''} {row['tecnico_reparo__last_name'] or ''}".strip()
            or row['tecnico_reparo__username']
            or '— Sin asignar —'
        )

        sal_qs = salidas_ranking_qs.filter(tecnico_reparo_id=tid)

        total_salidas = row['num_equipos']
        
        salidas_positivas = 0
        salidas_negativas = 0
        cobrado_final = D('0.00')

        if mostrar_valores_monetarios:
            total_acordado = row['total_acordado'] or D('0.00')
            total_anticipo = row['total_anticipo'] or D('0.00')

        for salida in sal_qs:
            estado = salida.estado_reparacion
            
            # Conteo de salidas
            if estado in estados_positivos:
                salidas_positivas += 1
            elif estado in estados_negativos:
                salidas_negativas += 1
                
            # Cobrado
            if mostrar_valores_monetarios:
                cobrado_final += (salida.valor_final_cobrado or D('0.00'))

                # Ajuste de Venta (Acordado)
                if estado == 'cliente_no_acepta':
                    total_acordado -= (salida.ingreso.valor_acordado or D('0.00'))
                    total_acordado += D('5.00')
                elif estado == 'no_reparable':
                    total_acordado -= (salida.ingreso.valor_acordado or D('0.00'))
                elif estado == 'revision':
                    total_acordado -= (salida.ingreso.valor_acordado or D('0.00'))
                    total_acordado += (salida.valor_acordado_revision or D('0.00'))

        item_ranking = {
            'tecnico_id': tid,
            'nombre': nombre,
            'sin_asignar': tid is None,
            'num_equipos': row['num_equipos'],
            'entregados': row['entregados'],
            'pendientes': row['pendientes'],
            'total_salidas': total_salidas,
            'salidas_positivas': salidas_positivas,
            'salidas_negativas': salidas_negativas,
            'efectividad': round((salidas_positivas / total_salidas * 100) if total_salidas else 0, 1),
        }
        if mostrar_valores_monetarios:
            item_ranking.update({
                'total_acordado': total_acordado,
                'total_anticipo': total_anticipo,
                'cobrado_final': cobrado_final,
                # Recaudado para el técnico: excluye anticipos.
                'total_recaudado': cobrado_final,
            })
        ranking_list.append(item_ranking)

    # Totales globales
    total_equipos = qs_ing.count()

    # Salidas globales
    sal_global = SalidaEquipo.objects.all()
    if desde:
        sal_global = sal_global.filter(fecha_salida__gte=desde)
    if hasta:
        sal_global = sal_global.filter(fecha_salida__lte=hasta)
    total_salidas_global = sal_global.count()
    total_positivas_global = sal_global.filter(
        estado_reparacion__in=estados_positivos
    ).count()
    # Top tipos de equipo trabajados
    por_tipo = qs_ing.values('tipo_equipo').annotate(num=Count('id'))
    if mostrar_valores_monetarios:
        por_tipo = por_tipo.annotate(suma=Sum('valor_acordado'))
    por_tipo = por_tipo.order_by('-num')

    map_tipos = dict(IngresoEquipo._meta.get_field('tipo_equipo').choices)
    por_tipo_list = []
    for tipo in por_tipo:
        item_tipo = {
            'tipo': map_tipos.get(tipo['tipo_equipo'], tipo['tipo_equipo']),
            'num': tipo['num'],
        }
        if mostrar_valores_monetarios:
            item_tipo['suma'] = tipo['suma'] or D('0.00')
        por_tipo_list.append(item_tipo)

    contexto = {
        'ranking_ingresos': ranking_ingresos,
        'ranking': ranking_list,
        'por_tipo': por_tipo_list,
        'total_equipos': total_equipos,
        'total_salidas_global': total_salidas_global,
        'total_positivas_global': total_positivas_global,
        'mostrar_valores_monetarios': mostrar_valores_monetarios,
        'filtros': {'desde': desde, 'hasta': hasta},
    }
    if mostrar_valores_monetarios:
        total_acordado_global = qs_ing.aggregate(s=Sum('valor_acordado'))['s'] or D('0.00')
        total_anticipos_global = qs_ing.aggregate(s=Sum('abono_anticipo'))['s'] or D('0.00')
        cobrado_final_global = sal_global.aggregate(s=Sum('valor_final_cobrado'))['s'] or D('0.00')
        total_diag_no_reparado = sal_global.filter(
            estado_reparacion__in=['no_reparable', 'cliente_no_acepta']
        ).aggregate(s=Sum('valor_final_cobrado'))['s'] or D('0.00')
        contexto.update({
            'total_acordado_global': total_acordado_global,
            'total_anticipos_global': total_anticipos_global,
            'cobrado_final_global': cobrado_final_global,
            'total_recaudado_global': total_anticipos_global + cobrado_final_global,
            'total_diag_no_reparado': total_diag_no_reparado,
        })

    return render(request, 'salidas/totales.html', contexto)


# ═════════════════════════════════════════════════════════════════
# Vistas de Alertas (demora en taller + bodegaje post-salida)
# ═════════════════════════════════════════════════════════════════

@login_required
def alertas_demora(request):
    """
    Lista completa de equipos demorados en diagnóstico (4+ días sin diagnosticar).
    Muestra dos secciones: activos y silenciados.
    """
    es_admin_user = request.user.is_superuser or request.user.groups.filter(
        name__in=['Administradores', 'Admin']
    ).exists()

    qs_activos = equipos_demorados_qs(usuario=None)

    # Silenciados: mismo filtro de estado pero con diagnostico_silenciado=True
    from datetime import timedelta as _td
    fecha_limite = date.today() - _td(days=UMBRAL_DIAS_DIAGNOSTICO)
    qs_silenciados = (
        IngresoEquipo.objects
        .select_related('cliente', 'tecnico_encargado')
        .filter(fecha_ingreso__lte=fecha_limite)
        .filter(estado='ingresado')
        .filter(salida__isnull=True)
        .filter(diagnostico_silenciado=True)
        .order_by('fecha_ingreso', 'numero_equipo')
    )
    # Todos ven todo

    hoy = date.today()

    def _build(qs):
        return [{
            'ingreso': ing,
            'dias': dias_en_taller(ing, hoy=hoy),
            'wa_link': whatsapp_link_demora(ing),
        } for ing in qs]

    items = _build(qs_activos)
    items_silenciados = _build(qs_silenciados)

    return render(request, 'alertas_demora.html', {
        'items': items,
        'total': len(items),
        'items_silenciados': items_silenciados,
        'total_silenciados': len(items_silenciados),
        'umbral_dias': UMBRAL_DIAS_DIAGNOSTICO,
        'es_admin_view': es_admin_user,
    })


@login_required
def alertas_bodegaje(request):
    """
    Lista completa de salidas con bodegaje pendiente
    (5+ días sin que el cliente venga a retirar).

    Muestra dos secciones:
      - Activos: alertas visibles
      - Silenciados: alertas que el usuario marcó como "no molestar"
    """
    es_admin_user = request.user.is_superuser or request.user.groups.filter(
        name__in=['Administradores', 'Admin']
    ).exists()

    qs_activos = salidas_bodegaje_qs(usuario=None)

    # Para los silenciados: incluimos todos los del usuario/admin que estén silenciados
    from django.db.models import Q as _Q
    from datetime import timedelta as _td
    fecha_limite = date.today() - _td(days=UMBRAL_DIAS_BODEGAJE)
    qs_silenciados = (
        SalidaEquipo.objects
        .select_related('ingreso', 'ingreso__cliente', 'ingreso__tecnico_encargado', 'tecnico_reparo')
        .filter(fecha_salida__lte=fecha_limite)
        .filter(fecha_retiro_real__isnull=True)
        .filter(bodegaje_silenciado=True)
        .order_by('fecha_salida')
    )
    # Todos ven todo

    hoy = date.today()

    def _build_items(qs):
        out = []
        total = D('0.00')
        for sal in qs:
            bod = sal.calcular_bodegaje(hoy=hoy)
            out.append({
                'salida': sal,
                'ingreso': sal.ingreso,
                'dias_desde_salida': dias_desde_salida(sal, hoy=hoy),
                'bodegaje_dias': bod['dias'],
                'bodegaje_monto': bod['monto'],
                'wa_link': whatsapp_link_bodegaje(sal),
            })
            total += bod['monto']
        return out, total

    items, total_acumulado = _build_items(qs_activos)
    items_silenciados, total_silenciados = _build_items(qs_silenciados)

    return render(request, 'alertas_bodegaje.html', {
        'items': items,
        'total': len(items),
        'total_acumulado': total_acumulado,
        'items_silenciados': items_silenciados,
        'total_silenciados': len(items_silenciados),
        'total_acumulado_silenciado': total_silenciados,
        'umbral_dias': UMBRAL_DIAS_BODEGAJE,
        'costo_dia': COSTO_BODEGAJE_DIA,
        'es_admin_view': es_admin_user,
    })


@tecnico_requerido
def salida_listo_aviso(request, pk):
    """
    Pantalla post-salida positiva: muestra el botón "Avisar al cliente
    por WhatsApp que su equipo está listo".
    """
    salida = get_object_or_404(
        SalidaEquipo.objects.select_related('ingreso', 'ingreso__cliente'),
        pk=pk,
    )
    confirmacion_pendiente = request.session.pop(
        'confirmar_ubicacion_salida_id',
        None,
    )
    actividad, _ = UsuarioActividad.objects.get_or_create(user=request.user)
    return render(request, 'salidas/listo_aviso.html', {
        'salida': salida,
        'ingreso': salida.ingreso,
        'wa_link': whatsapp_link_equipo_listo(salida),
        'mostrar_guia_saldo_pendiente': not actividad.ocultar_guia_saldo_pendiente,
        # Después de cada Guardar se muestra una confirmación una sola vez.
        # Si el equipo ya salió, la plantilla informa ese estado en lugar de
        # volver a preguntar algo que podría contradecir la fecha de retiro.
        'mostrar_confirmacion_guardado': confirmacion_pendiente == salida.pk,
        'salida_ya_confirmada': salida.cliente_ya_retiro,
    })


@tecnico_requerido
@require_POST
def salida_ocultar_guia_saldo_pendiente(request):
    """Guarda, para el usuario actual, que la guía no debe mostrarse otra vez."""
    actividad, _ = UsuarioActividad.objects.get_or_create(user=request.user)
    if not actividad.ocultar_guia_saldo_pendiente:
        actividad.ocultar_guia_saldo_pendiente = True
        actividad.save(update_fields=['ocultar_guia_saldo_pendiente'])
    return JsonResponse({'ok': True})


@tecnico_requerido
@require_POST
def salida_enviar_correo_finalizacion(request, pk):
    """Envía el acta cuando el usuario confirma el aviso de finalización."""
    salida = get_object_or_404(
        SalidaEquipo.objects.select_related('ingreso', 'ingreso__cliente'),
        pk=pk,
    )
    destino = (salida.ingreso.cliente.correo or '').strip()
    if not destino:
        return JsonResponse({
            'ok': False,
            'codigo': 'sin_correo',
            'mensaje': 'El cliente no tiene un correo registrado.',
        })
    if not getattr(settings, 'SALIDA_EMAIL_AUTOMATICO', True):
        return JsonResponse({
            'ok': False,
            'codigo': 'desactivado',
            'mensaje': 'El envío automático de finalización está desactivado.',
        })

    enviado = enviar_correo_finalizacion_seguro(salida.pk)
    if not enviado:
        return JsonResponse({
            'ok': False,
            'codigo': 'fallo_envio',
            'mensaje': 'No se pudo enviar el correo. La finalización quedó guardada correctamente.',
        })
    return JsonResponse({
        'ok': True,
        'mensaje': f'Acta enviada correctamente a {destino}.',
    })


def _url_retorno_salida(request, nombre_default):
    """Respeta `next` solo cuando apunta al propio sistema."""
    next_url = request.POST.get('next') or request.GET.get('next') or ''
    if next_url and url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return next_url
    return reverse(f'econotec:{nombre_default}')


def _programar_correo_salida_fisica(request, salida, abono_pks=()):
    """Programa el acta de salida tras el commit sin arriesgar la operación."""
    if not getattr(settings, 'SALIDA_EMAIL_AUTOMATICO', True):
        return

    destino = (salida.ingreso.cliente.correo or '').strip()
    if not destino:
        messages.warning(
            request,
            'La salida quedó confirmada, pero no se envió correo porque el cliente '
            'no tiene uno registrado.',
        )
        return

    pks = tuple(abono_pks)
    transaction.on_commit(
        lambda salida_pk=salida.pk, abonos=pks: enviar_correo_salida_fisica_seguro(
            salida_pk,
            abonos,
        )
    )
    messages.info(
        request,
        f'El acta de salida actualizada será enviada a {destino}.',
    )


def _crear_abonos_bodegaje(salida, pago, monto_bodegaje, dias_bodegaje, usuario):
    """Registra uno o dos abonos según el método elegido para el bodegaje."""
    metodo = pago['pago_bod_metodo']
    if metodo == 'mixto':
        partes = [
            {
                'monto': pago['pago_bod_monto_1'],
                'metodo': pago['pago_bod_metodo_1'],
                'banco': pago.get('pago_bod_banco_1') or '',
                'banco_otro': pago.get('pago_bod_banco_otro_1') or '',
                'tarjeta_app': pago.get('pago_bod_tarjeta_app_1') or '',
                'comprobante_url': pago.get('pago_bod_comprobante_url_1') or '',
            },
            {
                'monto': pago['pago_bod_monto_2'],
                'metodo': pago['pago_bod_metodo_2'],
                'banco': pago.get('pago_bod_banco_2') or '',
                'banco_otro': pago.get('pago_bod_banco_otro_2') or '',
                'tarjeta_app': pago.get('pago_bod_tarjeta_app_2') or '',
                'comprobante_url': pago.get('pago_bod_comprobante_url_2') or '',
            },
        ]
    else:
        partes = [{
            'monto': monto_bodegaje,
            'metodo': metodo,
            'banco': pago.get('pago_bod_banco') or '',
            'banco_otro': pago.get('pago_bod_banco_otro') or '',
            'tarjeta_app': pago.get('pago_bod_tarjeta_app') or '',
            'comprobante_url': pago.get('pago_bod_comprobante_url') or '',
        }]

    abonos = []
    total_partes = len(partes)
    for indice, parte in enumerate(partes, start=1):
        metodo_parte = parte['metodo']
        detalle_mixto = (
            f' Pago mixto (parte {indice} de {total_partes}).'
            if total_partes > 1
            else ''
        )
        abono = Abono.objects.create(
            ingreso=salida.ingreso,
            monto=parte['monto'],
            fecha=date.today(),
            metodo=metodo_parte,
            banco=parte['banco'] if metodo_parte == 'transferencia' else '',
            banco_otro=(
                parte['banco_otro']
                if metodo_parte == 'transferencia' and parte['banco'] == 'otro'
                else ''
            ),
            tarjeta_app=parte['tarjeta_app'] if metodo_parte == 'tarjeta' else '',
            comprobante_url=(
                parte['comprobante_url'] if metodo_parte == 'transferencia' else ''
            ),
            observaciones=(
                f'Cobro por {dias_bodegaje} día(s) de bodegaje al retirar el equipo.'
                f'{detalle_mixto}'
            ),
            bodegaje_decision='si' if indice == 1 else 'na',
            bodegaje_monto_aplicado=(
                monto_bodegaje if indice == 1 else D('0.00')
            ),
            registrado_por=usuario,
        )
        registrar_bitacora(
            usuario,
            'abono',
            (
                f'Cobro de bodegaje {abono.numero_recibo} por ${abono.monto:.2f} '
                f'en #{salida.ingreso.codigo_equipo}.'
            ),
            ingreso=salida.ingreso,
            salida=salida,
            abono=abono,
            dedupe_key=f'abono:{abono.pk}:bodegaje-salida',
        )
        abonos.append(abono)
    return abonos


@tecnico_requerido
@require_POST
@transaction.atomic
def salida_marcar_retirada(request, pk):
    """
    Marca la salida como "Cliente ya retiró", congelando el bodegaje
    acumulado hasta hoy. Esto cierra el caso.

    Si en el POST viene `aplicar_bodegaje=on`, valida y registra el pago
    del bodegaje antes de cerrar el expediente.
    """
    salida = get_object_or_404(
        SalidaEquipo.objects.select_related('ingreso', 'ingreso__cliente'),
        pk=pk,
    )

    if salida.cliente_ya_retiro:
        messages.info(
            request,
            f'El equipo {salida.ingreso.codigo_equipo} ya consta fuera de la oficina.'
        )
        return redirect('econotec:salida_retiros_lista')

    saldo_pendiente = salida.ingreso.diferencia
    if saldo_pendiente > 0:
        messages.error(
            request,
            f'El equipo {salida.ingreso.codigo_equipo} tiene un saldo pendiente de '
            f'${saldo_pendiente:.2f}. Debe pagarse antes de confirmar la salida de la oficina.'
        )
        return redirect('econotec:salida_listo_aviso', pk=salida.pk)

    bod = salida.calcular_bodegaje()
    aplicar = request.POST.get('aplicar_bodegaje') == 'on'

    pago_bodegaje = None
    if bod['monto'] > 0 and aplicar:
        pago_bodegaje = CobroBodegajeForm(
            request.POST,
            monto_esperado=bod['monto'],
        )
        if not pago_bodegaje.is_valid():
            errores = ' '.join(
                str(error)
                for lista_errores in pago_bodegaje.errors.values()
                for error in lista_errores
            )
            messages.error(
                request,
                f'No se confirmó la salida: revisa el pago de bodegaje. {errores}',
            )
            return redirect(_url_retorno_salida(request, 'salida_lista'))

    salida.fecha_retiro_real = date.today()
    salida.estado_reparacion = 'retirado'
    salida.bodegaje_dias_congelado = bod['dias']
    salida.bodegaje_monto_congelado = bod['monto']
    salida.bodegaje_aplicado_al_pago = aplicar
    salida.save(update_fields=[
        'fecha_retiro_real',
        'estado_reparacion',
        'bodegaje_dias_congelado',
        'bodegaje_monto_congelado',
        'bodegaje_aplicado_al_pago',
    ])

    abonos_bodegaje = []
    if pago_bodegaje is not None:
        abonos_bodegaje = _crear_abonos_bodegaje(
            salida,
            pago_bodegaje.cleaned_data,
            bod['monto'],
            bod['dias'],
            request.user,
        )

    _programar_correo_salida_fisica(
        request,
        salida,
        [abono.pk for abono in abonos_bodegaje],
    )

    if bod['monto'] > 0:
        if aplicar:
            messages.success(
                request,
                f'Salida de la oficina confirmada para el equipo {salida.ingreso.codigo_equipo}. '
                f'Se cobraron ${bod["monto"]} de bodegaje ({bod["dias"]} días).'
            )
        else:
            messages.success(
                request,
                f'Salida de la oficina confirmada para el equipo {salida.ingreso.codigo_equipo}. '
                f'Bodegaje de ${bod["monto"]} ({bod["dias"]} días) NO cobrado al cliente.'
            )
    else:
        messages.success(
            request,
            f'Salida de la oficina confirmada para el equipo {salida.ingreso.codigo_equipo}.'
        )

    return redirect(_url_retorno_salida(request, 'salida_retiros_lista'))


@admin_requerido
@require_POST
def salida_deshacer_retiro(request, pk):
    """
    Deshace el retiro físico de un equipo. Solo para administradores.
    """
    salida = get_object_or_404(SalidaEquipo, pk=pk)

    if not salida.cliente_ya_retiro:
        messages.info(
            request,
            f'El equipo {salida.ingreso.codigo_equipo} no tenía confirmada la salida de la oficina.'
        )
        return redirect('econotec:salida_lista')

    salida.fecha_retiro_real = None
    if salida.estado_reparacion == 'retirado':
        salida.estado_reparacion = 'pendiente_retiro'
    salida.bodegaje_dias_congelado = None
    salida.bodegaje_monto_congelado = None
    salida.bodegaje_aplicado_al_pago = False
    salida.save(update_fields=[
        'fecha_retiro_real',
        'estado_reparacion',
        'bodegaje_dias_congelado',
        'bodegaje_monto_congelado',
        'bodegaje_aplicado_al_pago',
    ])
    messages.success(
        request,
        f'Deshecho: el equipo {salida.ingreso.codigo_equipo} vuelve a constar dentro de la oficina.'
    )
    return redirect('econotec:salida_lista')


@login_required
@require_POST
def salida_bodegaje_silenciar(request, pk):
    """
    Activa/desactiva el modo 'no molestar' para la alerta de bodegaje
    de un equipo específico. El bodegaje sigue acumulándose; solo se
    oculta del banner del dashboard.

    El parámetro POST `accion` puede ser 'silenciar' o 'reactivar'.
    Si no viene, hace toggle.
    """
    salida = get_object_or_404(SalidaEquipo, pk=pk)
    accion = (request.POST.get('accion') or '').strip().lower()

    if accion == 'silenciar':
        salida.bodegaje_silenciado = True
    elif accion == 'reactivar':
        salida.bodegaje_silenciado = False
    else:
        # Toggle
        salida.bodegaje_silenciado = not salida.bodegaje_silenciado

    salida.save(update_fields=['bodegaje_silenciado', 'actualizado'])

    codigo = salida.ingreso.codigo_equipo
    if salida.bodegaje_silenciado:
        messages.success(
            request,
            f'🔕 Alerta silenciada para el equipo {codigo}. '
            f'El bodegaje sigue acumulándose, pero no aparecerá en el dashboard.'
        )
    else:
        messages.success(
            request,
            f'🔔 Alerta reactivada para el equipo {codigo}.'
        )

    # Volver a donde venía: alerta detallada o dashboard
    next_url = request.POST.get('next', '')
    if next_url:
        return redirect(next_url)
    return redirect('econotec:bienvenida')


@login_required
@require_POST
def ingreso_diagnostico_silenciar(request, pk):
    """
    Activa/desactiva el modo 'no molestar' para la alerta de diagnóstico
    pendiente de un equipo específico. El equipo sigue pendiente; solo
    se oculta del banner del dashboard.

    Se reactiva automáticamente cuando el estado del equipo cambia
    (lógica implementada en IngresoEquipo.save()).

    El parámetro POST `accion` puede ser 'silenciar' o 'reactivar'.
    Si no viene, hace toggle.
    """
    ingreso = get_object_or_404(IngresoEquipo, pk=pk)
    accion = (request.POST.get('accion') or '').strip().lower()

    if accion == 'silenciar':
        ingreso.diagnostico_silenciado = True
    elif accion == 'reactivar':
        ingreso.diagnostico_silenciado = False
    else:
        # Toggle
        ingreso.diagnostico_silenciado = not ingreso.diagnostico_silenciado

    ingreso.save(update_fields=['diagnostico_silenciado', 'actualizado'])

    codigo = ingreso.codigo_equipo
    if ingreso.diagnostico_silenciado:
        messages.success(
            request,
            f'🔕 Alerta silenciada para el equipo {codigo}. '
            f'Se reactivará automáticamente cuando cambie el estado del equipo.'
        )
    else:
        messages.success(
            request,
            f'🔔 Alerta de diagnóstico reactivada para el equipo {codigo}.'
        )

    next_url = request.POST.get('next', '')
    if next_url:
        return redirect(next_url)
    return redirect('econotec:bienvenida')


# ═════════════════════════════════════════════════════════════════
# API Perfil (Gamificación)
# ═════════════════════════════════════════════════════════════════

COLORES_PERFIL_ASESOR = {
    '#0d47a1': 'Azul',
    '#ec4899': 'Rosa',
    '#c62828': 'Rojo',
    '#f97618': 'Naranja',
    '#2e7d32': 'Verde',
    '#f9c74f': 'Amarillo',
}


def _hora_bitacora(dt):
    local = timezone.localtime(dt)
    hora = local.hour % 12 or 12
    return f'{hora}:{local.minute:02d}'


def _periodo_bitacora(dt):
    return 'AM' if timezone.localtime(dt).hour < 12 else 'PM'


def _texto_evento_bitacora_para_copiar(texto):
    texto = (texto or '').strip()
    marcador_detalles = '. Detalles: '
    if marcador_detalles not in texto:
        return texto

    base, detalles_txt = texto.split(marcador_detalles, 1)
    detalles = [
        detalle.strip().rstrip('.')
        for detalle in detalles_txt.rstrip('.').split(';')
        if detalle.strip()
    ]
    if not detalles:
        return texto

    lineas_detalles = '\n'.join(f'  - {detalle}.' for detalle in detalles)
    return f'{base.strip()}.\nDetalles:\n{lineas_detalles}'


def _linea_bitacora_para_copiar(evento):
    hora = f'{_hora_bitacora(evento["momento"])} {_periodo_bitacora(evento["momento"])}'
    return f'*{hora}* - {_texto_evento_bitacora_para_copiar(evento["texto"])}'


def _texto_limpio_bitacora(texto, max_len=170):
    texto = ' '.join((texto or '').split())
    if len(texto) <= max_len:
        return texto
    return texto[:max_len - 1].rstrip() + '…'


def _equipo_bitacora(ingreso):
    partes = [
        ingreso.tipo_equipo_display,
        ingreso.marca,
        ingreso.modelo_serie_detalle,
    ]
    return ' '.join(p for p in partes if p).strip()


def _texto_ingreso_bitacora(ingreso):
    equipo = _equipo_bitacora(ingreso)
    return f'Ingreso de equipo registrado: {equipo} #{ingreso.codigo_equipo} para {ingreso.cliente.nombres}.'


def _texto_venta_bitacora(venta):
    descripcion = _texto_limpio_bitacora(venta.problema_reportado, max_len=120)
    if not descripcion:
        descripcion = _equipo_bitacora(venta) or 'producto'
    valor = venta.valor_acordado or D('0.00')
    return f'Venta de producto registrada: {descripcion} #{venta.codigo_equipo} para {venta.cliente.nombres} por ${valor:.2f}.'


def _texto_estado_ingreso_bitacora(ingreso):
    estado = ingreso.estado_visual_display
    subestado = ingreso.subestado_visual_display
    detalle = f'{estado} - {subestado}' if subestado and subestado != estado else estado
    return f'Cambio de estado en {_equipo_bitacora(ingreso)} #{ingreso.codigo_equipo}: {detalle}.'


_LABELS_CAMPOS_BITACORA = {
    'cedula': 'Cédula / RUC',
    'nombres': 'Cliente',
    'whatsapp': 'WhatsApp',
    'correo': 'Correo',
    'sector': 'Sector',
    'sector_otro': 'Sector otro',
    'numero_factura': 'Factura N°',
    'asesor_comercial': 'Asesora comercial',
    'tecnico_encargado': 'Técnico recibido',
    'fecha_ingreso': 'Fecha de ingreso',
    'tipo_equipo': 'Tipo de equipo',
    'tipo_equipo_otro': 'Tipo especificado',
    'marca': 'Marca',
    'modelo_serie': 'Modelo',
    'serie': 'Serie',
    'accesorios_entregados': 'Accesorios',
    'problema_reportado': 'Problema reportado',
    'firma_cliente': 'Firma del cliente',
    'firma_cliente_opcion': 'Firma del cliente',
    'firma_cliente_imagen': 'Imagen de firma',
    'diagnostico_inmediato': 'Diagnóstico inmediato',
    'valor_diagnostico': 'Valor diagnóstico',
    'valor_acordado': 'Valor acordado',
    'valor_acordado_estado': 'Estado del valor acordado',
    'abono_anticipo': 'Anticipo',
    'diagnostico_metodo': 'Método diagnóstico',
    'diagnostico_banco': 'Banco diagnóstico',
    'diagnostico_banco_otro': 'Banco diagnóstico otro',
    'diagnostico_tarjeta_app': 'Tarjeta/App diagnóstico',
    'diagnostico_comprobante_url': 'Comprobante diagnóstico',
    'diagnostico_monto_1': 'Monto diagnóstico 1',
    'diagnostico_metodo_1': 'Método diagnóstico 1',
    'diagnostico_banco_1': 'Banco diagnóstico 1',
    'diagnostico_monto_2': 'Monto diagnóstico 2',
    'diagnostico_metodo_2': 'Método diagnóstico 2',
    'diagnostico_banco_2': 'Banco diagnóstico 2',
    'anticipo_metodo': 'Método anticipo',
    'anticipo_banco': 'Banco anticipo',
    'anticipo_banco_otro': 'Banco anticipo otro',
    'anticipo_tarjeta_app': 'Tarjeta/App anticipo',
    'anticipo_comprobante_url': 'Comprobante anticipo',
    'anticipo_monto_1': 'Monto anticipo 1',
    'anticipo_metodo_1': 'Método anticipo 1',
    'anticipo_banco_1': 'Banco anticipo 1',
    'anticipo_monto_2': 'Monto anticipo 2',
    'anticipo_metodo_2': 'Método anticipo 2',
    'anticipo_banco_2': 'Banco anticipo 2',
    'estado': 'Estado del equipo',
    'subestado_reparacion': 'Detalle de reparación',
    'subestado_entregado': 'Detalle de entrega',
    'equipo_garantia': 'Equipo de garantía',
    'equipo_garantia_manual': 'Equipo garantía manual',
    'motivo_garantia': 'Motivo de garantía',
}

_ORDEN_CAMPOS_BITACORA = [
    'estado', 'subestado_reparacion', 'subestado_entregado',
    'tipo_equipo', 'tipo_equipo_otro', 'marca', 'modelo_serie', 'serie',
    'tecnico_encargado', 'asesor_comercial', 'fecha_ingreso', 'numero_factura',
    'problema_reportado', 'accesorios_entregados',
    'valor_acordado_estado', 'valor_acordado', 'diagnostico_inmediato',
    'valor_diagnostico', 'abono_anticipo',
    'diagnostico_metodo', 'diagnostico_banco', 'diagnostico_banco_otro',
    'diagnostico_tarjeta_app', 'diagnostico_comprobante_url',
    'diagnostico_monto_1', 'diagnostico_metodo_1', 'diagnostico_banco_1',
    'diagnostico_monto_2', 'diagnostico_metodo_2', 'diagnostico_banco_2',
    'anticipo_metodo', 'anticipo_banco', 'anticipo_banco_otro',
    'anticipo_tarjeta_app', 'anticipo_comprobante_url',
    'anticipo_monto_1', 'anticipo_metodo_1', 'anticipo_banco_1',
    'anticipo_monto_2', 'anticipo_metodo_2', 'anticipo_banco_2',
    'equipo_garantia', 'equipo_garantia_manual', 'motivo_garantia',
    'firma_cliente', 'firma_cliente_opcion', 'firma_cliente_imagen',
    'cedula', 'nombres', 'whatsapp', 'correo', 'sector', 'sector_otro',
]

_CAMPOS_DINERO_BITACORA = {
    'valor_diagnostico', 'valor_acordado', 'abono_anticipo',
    'diagnostico_monto_1', 'diagnostico_monto_2',
    'anticipo_monto_1', 'anticipo_monto_2',
}

_CAMPOS_AUXILIARES_NO_BITACORA = {
    'valor_acordado_estado',
    'firma_cliente_opcion',
}


def _snapshot_form_original(form):
    instancia = getattr(form, 'instance', None)
    valores = {}
    for campo in form.fields:
        if instancia is not None and hasattr(instancia, campo):
            valores[campo] = getattr(instancia, campo)
        else:
            valores[campo] = (getattr(form, 'initial', {}) or {}).get(campo)
    return valores


def _label_cambio_bitacora(form, campo):
    if campo in _LABELS_CAMPOS_BITACORA:
        return _LABELS_CAMPOS_BITACORA[campo]
    field = form.fields.get(campo)
    if field and field.label:
        return str(field.label)
    modelo = getattr(getattr(form, '_meta', None), 'model', None)
    if modelo:
        try:
            return str(modelo._meta.get_field(campo).verbose_name).capitalize()
        except Exception:
            pass
    return campo.replace('_', ' ').capitalize()


def _display_cambio_bitacora(form, campo, valor):
    if campo in ('tecnico_encargado',):
        return nombre_corto_usuario(valor) if valor else '—'
    if campo == 'equipo_garantia':
        return valor.codigo_equipo if valor else '—'
    if campo in ('firma_cliente',):
        return 'Sí' if bool(valor) else 'No'
    if campo in ('firma_cliente_imagen',):
        return 'Guardada' if valor else 'Sin imagen'
    if valor in (None, ''):
        return '—'
    if campo in _CAMPOS_DINERO_BITACORA:
        try:
            return f'${D(str(valor)):.2f}'
        except Exception:
            return str(valor)
    if isinstance(valor, date):
        return valor.strftime('%d/%m/%Y')

    modelo = getattr(getattr(form, '_meta', None), 'model', None)
    if modelo:
        try:
            model_field = modelo._meta.get_field(campo)
            if model_field.choices:
                choices = dict(model_field.flatchoices)
                return str(choices.get(valor, choices.get(str(valor), valor)))
        except Exception:
            pass

    field = form.fields.get(campo)
    choices = getattr(field, 'choices', None)
    if choices and campo != 'tecnico_encargado':
        try:
            choices_dict = dict(choices)
            return str(choices_dict.get(valor, choices_dict.get(str(valor), valor)))
        except Exception:
            pass

    return _texto_limpio_bitacora(str(valor), max_len=90)


def _campos_ordenados_bitacora(campos):
    orden = {campo: idx for idx, campo in enumerate(_ORDEN_CAMPOS_BITACORA)}
    return sorted(campos, key=lambda campo: (orden.get(campo, 999), campo))


def _detalles_cambios_form_bitacora(form, campos, valores_antes):
    detalles = []
    for campo in _campos_ordenados_bitacora(campos):
        if campo in _CAMPOS_AUXILIARES_NO_BITACORA:
            continue
        if campo not in form.fields:
            continue
        antes = _display_cambio_bitacora(form, campo, valores_antes.get(campo))
        despues = _display_cambio_bitacora(form, campo, form.cleaned_data.get(campo))
        if antes == despues:
            continue
        detalles.append(f'{_label_cambio_bitacora(form, campo)}: {antes} -> {despues}')
    return detalles


def _texto_actualizacion_ingreso_bitacora(
    ingreso,
    cli_form,
    ing_form,
    campos_cliente,
    campos_ingreso,
    valores_cliente_antes,
    valores_ingreso_antes,
):
    detalles = []
    detalles.extend(_detalles_cambios_form_bitacora(
        ing_form,
        campos_ingreso,
        valores_ingreso_antes,
    ))
    detalles.extend(_detalles_cambios_form_bitacora(
        cli_form,
        campos_cliente,
        valores_cliente_antes,
    ))

    texto = f'Datos actualizados en {_equipo_bitacora(ingreso)} #{ingreso.codigo_equipo} para {ingreso.cliente.nombres}.'
    if detalles:
        texto += f' Detalles: {"; ".join(detalles)}.'
    return texto


def _ingreso_permite_reparacion_check(ingreso):
    return (
        (
            ingreso.estado == 'en_reparacion'
            and ingreso.subestado_reparacion == 'en_reparacion'
        )
        or ingreso.estado == 'garantia'
    )


def _usuario_reparacion_check_bitacora(user, ingreso):
    if es_tecnico(user):
        return user
    if ingreso.tecnico_encargado_id:
        return ingreso.tecnico_encargado
    return None


def _reparacion_check_dedupe_key(user, ingreso, dia):
    return f'reparacion-check:{dia.isoformat()}:{user.pk}:{ingreso.pk}'


def _reparacion_check_ya_registrado(user, ingreso, dia):
    return BitacoraTecnico.objects.filter(
        dedupe_key=_reparacion_check_dedupe_key(user, ingreso, dia)
    ).exists()


def _texto_reparacion_check_bitacora(ingreso, user):
    tecnico = nombre_corto_usuario(user)
    equipo = _equipo_bitacora(ingreso) or 'equipo'
    problema = _texto_limpio_bitacora(ingreso.problema_reportado, max_len=150)
    accesorios = _texto_limpio_bitacora(ingreso.accesorios_entregados, max_len=90)
    cliente = _texto_limpio_bitacora(ingreso.cliente.nombres, max_len=80)
    garantia_ref = _texto_limpio_bitacora(ingreso.equipo_garantia_referencia, max_len=90)
    motivo_garantia = _texto_limpio_bitacora(ingreso.motivo_garantia, max_len=150)

    partes = [
        f'El técnico {tecnico} aún sigue reparando este equipo: {equipo} #{ingreso.codigo_equipo}',
    ]
    if cliente:
        partes.append(f'cliente {cliente}')
    if problema and problema.strip(' .').lower() not in ('no', 'n/a', 'na', 'ninguno', 'ninguna', '-'):
        partes.append(f'problema reportado: {problema}')
    if accesorios:
        partes.append(f'accesorios: {accesorios}')
    if ingreso.estado == 'garantia':
        if garantia_ref:
            partes.append(f'garantía de {garantia_ref}')
        if motivo_garantia:
            partes.append(f'motivo de garantía: {motivo_garantia}')
        partes.append('estado confirmado: Garantía')
    else:
        partes.append('estado confirmado: En reparación -> En reparación')
    return '. '.join(partes) + '.'


def _texto_salida_bitacora(salida):
    ingreso = salida.ingreso
    equipo = _equipo_bitacora(ingreso)
    reporte = _texto_limpio_bitacora(ingreso.reporte_tecnico)

    if reporte:
        base = reporte.rstrip('.')
    else:
        base = f'Trabajo registrado en {equipo}'.strip()

    if salida.estado_reparacion in ('pendiente_retiro', 'retirado'):
        return f'{base} #{ingreso.codigo_equipo} lista, cliente notificado.'
    if salida.estado_reparacion in ('garantia', 'garantia_fallos_adicionales'):
        return f'{base} #{ingreso.codigo_equipo} salida por garantía.'
    if salida.estado_reparacion == 'cortesia':
        return f'{base} #{ingreso.codigo_equipo} salida de cortesía, sin cobro.'
    if salida.estado_reparacion == 'cliente_no_acepta':
        return f'{base} #{ingreso.codigo_equipo} cliente no quiso reparar.'
    if salida.estado_reparacion == 'no_reparable':
        return f'{base} #{ingreso.codigo_equipo} no se pudo reparar.'
    if salida.estado_reparacion == 'revision':
        return f'{base} #{ingreso.codigo_equipo} salió en revisión.'
    return f'{base} #{ingreso.codigo_equipo} {salida.get_estado_reparacion_display()}.'


def _eventos_bitacora_usuario(user, dia=None):
    dia = dia or timezone.localdate()
    eventos = []
    eventos_guardados = (
        BitacoraTecnico.objects
        .select_related('ingreso', 'salida', 'abono')
        .filter(user=user, momento__date=dia)
        .order_by('momento', 'pk')
    )
    ingresos_con_evento_por_tipo = {}
    salidas_con_evento = set()
    abonos_con_evento = set()

    for evento in eventos_guardados:
        if evento.ingreso_id:
            ingresos_con_evento_por_tipo.setdefault(evento.tipo, set()).add(evento.ingreso_id)
        if evento.salida_id:
            salidas_con_evento.add(evento.salida_id)
        if evento.abono_id:
            abonos_con_evento.add(evento.abono_id)

        eventos.append({
            'momento': evento.momento,
            'texto': evento.texto,
            'tipo': evento.tipo,
            'codigo': evento.codigo,
        })

    ingresos_con_evento = set()
    for tipos in ingresos_con_evento_por_tipo.values():
        ingresos_con_evento.update(tipos)

    salidas = (
        SalidaEquipo.objects
        .select_related('ingreso', 'ingreso__cliente', 'tecnico_reparo', 'registrado_por')
        .filter(creado__date=dia)
        .filter(Q(tecnico_reparo=user) | Q(registrado_por=user))
        .exclude(pk__in=salidas_con_evento)
        .order_by('creado', 'pk')
    )
    ingresos_con_salida = set()
    for salida in salidas:
        ingresos_con_salida.add(salida.ingreso_id)
        eventos.append({
            'momento': salida.creado,
            'texto': _texto_salida_bitacora(salida),
            'tipo': 'salida',
            'codigo': salida.ingreso.codigo_equipo,
        })

    ingresos = (
        IngresoEquipo.objects
        .select_related('cliente', 'tecnico_encargado', 'registrado_por')
        .filter(creado__date=dia)
        .filter(Q(registrado_por=user) | Q(tecnico_encargado=user))
        .exclude(pk__in=ingresos_con_salida)
        .exclude(pk__in=ingresos_con_evento)
        .order_by('creado', 'pk')
    )
    for ingreso in ingresos:
        equipo = _equipo_bitacora(ingreso)
        if ingreso.sede == 'ventas':
            texto = f'Venta de producto registrada: {equipo} #{ingreso.codigo_equipo}.'
        elif ingreso.registrado_por_id == user.id:
            texto = f'Recepción y registro de {equipo} #{ingreso.codigo_equipo} para {ingreso.cliente.nombres}.'
        else:
            texto = f'Equipo asignado para revisión: {equipo} #{ingreso.codigo_equipo}.'
        eventos.append({
            'momento': ingreso.creado,
            'texto': texto,
            'tipo': 'ingreso',
            'codigo': ingreso.codigo_equipo,
        })

    reportes = (
        IngresoEquipo.objects
        .select_related('cliente')
        .filter(reporte_por=user, reporte_actualizado__date=dia)
        .exclude(pk__in=ingresos_con_salida)
        .exclude(pk__in=ingresos_con_evento_por_tipo.get('reporte', set()))
        .order_by('reporte_actualizado', 'pk')
    )
    for ingreso in reportes:
        reporte = _texto_limpio_bitacora(ingreso.reporte_tecnico)
        if not reporte:
            continue
        eventos.append({
            'momento': ingreso.reporte_actualizado,
            'texto': f'Actualización de reporte técnico en {_equipo_bitacora(ingreso)} #{ingreso.codigo_equipo}: {reporte}.',
            'tipo': 'reporte',
            'codigo': ingreso.codigo_equipo,
        })

    reportes_valor = (
        IngresoEquipo.objects
        .select_related('cliente')
        .filter(valor_pendiente_reporte_por=user, valor_pendiente_reporte_actualizado__date=dia)
        .exclude(pk__in=ingresos_con_evento_por_tipo.get('valor_pendiente', set()))
        .order_by('valor_pendiente_reporte_actualizado', 'pk')
    )
    for ingreso in reportes_valor:
        motivo = _texto_limpio_bitacora(ingreso.valor_pendiente_reporte)
        if not motivo:
            continue
        eventos.append({
            'momento': ingreso.valor_pendiente_reporte_actualizado,
            'texto': f'Reporte de valor acordado pendiente en #{ingreso.codigo_equipo}: {motivo}.',
            'tipo': 'valor_pendiente',
            'codigo': ingreso.codigo_equipo,
        })

    abonos = (
        Abono.objects
        .select_related('ingreso', 'ingreso__cliente')
        .filter(registrado_por=user, creado__date=dia)
        .exclude(pk__in=abonos_con_evento)
        .order_by('creado', 'pk')
    )
    for abono in abonos:
        eventos.append({
            'momento': abono.creado,
            'texto': f'Registro de abono {abono.numero_recibo} por ${abono.monto:.2f} en #{abono.ingreso.codigo_equipo}.',
            'tipo': 'abono',
            'codigo': abono.ingreso.codigo_equipo,
        })

    eventos.sort(key=lambda e: (e['momento'], e['texto']))
    return eventos


def _construir_bitacora_usuario(user, dia=None):
    dia = dia or timezone.localdate()
    eventos = _eventos_bitacora_usuario(user, dia=dia)
    fecha_txt = dia.strftime('%d/%m/%Y')
    nombre = nombre_corto_usuario(user)
    encabezado = '\n'.join(['Reporte del día', f'Técnico: {nombre}', f'Fecha: {fecha_txt}'])

    if not eventos:
        return {
            'fecha': fecha_txt,
            'total': 0,
            'tiene_datos': False,
            'encabezado': encabezado,
            'detalle': '',
            'texto': encabezado,
            'eventos': [],
        }

    lineas = []
    eventos_json = []
    for evento in eventos:
        hora_inicio = _hora_bitacora(evento['momento'])
        periodo_inicio = _periodo_bitacora(evento['momento'])
        lineas.append(_linea_bitacora_para_copiar(evento))
        eventos_json.append({
            'hora_inicio': hora_inicio,
            'periodo_inicio': periodo_inicio,
            'texto': evento['texto'],
            'tipo': evento['tipo'],
            'codigo': evento['codigo'],
        })

    detalle = '\n\n'.join(lineas)
    return {
        'fecha': fecha_txt,
        'total': len(eventos),
        'tiene_datos': True,
        'encabezado': encabezado,
        'detalle': detalle,
        'texto': f'{encabezado}\n\n{detalle}',
        'eventos': eventos_json,
    }


@login_required
def api_perfil(request):
    user = request.user

    if es_asesor(user) and not es_tecnico(user) and not user.is_superuser:
        actividad, _ = UsuarioActividad.objects.get_or_create(user=user)
        color = actividad.perfil_color_asesor
        if color not in COLORES_PERFIL_ASESOR:
            color = '#0d47a1'

        return JsonResponse({
            'username': user.username,
            'nombre': user.first_name or user.username,
            'email': user.email or '',
            'tipo_perfil': 'asesor',
            'rol': 'Asesor registrado',
            'nivel': 'Asesor registrado',
            'color': color,
            'colores_disponibles': COLORES_PERFIL_ASESOR,
            'ingresos': 0,
            'salidas_buenas': 0,
            'salidas_producto': 0,
            'salidas_malas': 0,
            'total': 0,
            'proximo': None,
            'bitacora_total': construir_bitacora_usuario(user)['total'],
        })
    
    # Verificar si el usuario tiene una fecha de reinicio
    fecha_reinicio = None
    if hasattr(user, 'actividad') and user.actividad.fecha_reinicio_perfil:
        fecha_reinicio = user.actividad.fecha_reinicio_perfil

    # Base querysets
    #
    # IMPORTANTE (regla del negocio): el NIVEL del técnico se calcula SOLO por
    # las SALIDAS que él reparó (campo `tecnico_reparo`), NO por los ingresos.
    # El técnico seleccionado en la salida asume la responsabilidad del resultado:
    #   • salida buena  → suma
    #   • salida mala    → resta
    #   • garantía       → cuenta como salida positiva
    # Los ingresos se siguen mostrando como dato informativo, pero ya NO cuentan
    # para subir de nivel.
    ingresos_qs = IngresoEquipo.objects.filter(registrado_por=user)
    salidas_qs = SalidaEquipo.objects.filter(tecnico_reparo=user)
    ventas_producto_qs = IngresoEquipo.objects.filter(
        sede='ventas',
        tecnico_encargado=user,
    )

    if fecha_reinicio:
        ingresos_qs = ingresos_qs.filter(creado__gte=fecha_reinicio)
        salidas_qs = salidas_qs.filter(creado__gte=fecha_reinicio)
        ventas_producto_qs = ventas_producto_qs.filter(creado__gte=fecha_reinicio)

    # Ingresos registrados por el usuario (solo informativo, no suma nivel)
    ingresos_count = ingresos_qs.count()
    salidas_producto = ventas_producto_qs.count()
    
    # Salidas buenas positivas (reparadas por el técnico)
    salidas_buenas = salidas_qs.filter(
        estado_reparacion__in=SALIDA_BUENA_ESTADOS
    ).count()
    
    # Salidas negativas (restan 1 punto)
    salidas_malas = salidas_qs.filter(
        estado_reparacion__in=SALIDA_MALA_ESTADOS
    ).count()
    
    # Compatibilidad histórica: las garantías ya están incluidas como buenas.
    salidas_garantia = salidas_qs.filter(
        estado_reparacion__in=SALIDA_GARANTIA_ESTADOS
    ).count()
    
    # Calcular total (no puede ser menor a 0).
    # Las ventas de producto cuentan como salida positiva de producto: +1 cada una.
    # Las salidas buenas positivas valen más para equilibrar el perfil.
    total_operaciones = calcular_puntaje_gamificacion(
        salidas_buenas,
        salidas_producto,
        salidas_malas,
        salidas_garantia,
    )
    
    # Gamificación
    if total_operaciones <= 49:
        nivel = 'Novato'
        color = '#8e8e8e' # Gris
        proximo = 50
    elif total_operaciones <= 99:
        nivel = 'Intermedio'
        color = '#cd7f32' # Bronce
        proximo = 100
    elif total_operaciones <= 499:
        nivel = 'Avanzado'
        color = '#c0c0c0' # Plata
        proximo = 500
    elif total_operaciones <= 999:
        nivel = 'Experto'
        color = '#ffd700' # Oro
        proximo = 1000
    elif total_operaciones <= 3999:
        nivel = 'Maestro'
        color = '#b9f2ff' # Diamante brillante
        proximo = 4000
    else:
        nivel = 'God Tec Econotec'
        color = 'linear-gradient(45deg, #FFD700, #ff8c00)' # Oro
        proximo = None
        
    return JsonResponse({
        'username': user.username,
        'nombre': user.first_name or user.username,
        'email': user.email or '',
        'tipo_perfil': 'tecnico',
        'ingresos': ingresos_count,
        'salidas_buenas': salidas_buenas,
        'salidas_producto': salidas_producto,
        'salidas_malas': salidas_malas,
        'total': total_operaciones,
        'nivel': nivel,
        'color': color,
        'proximo': proximo,
        'bitacora_total': construir_bitacora_usuario(user)['total'],
    })


@login_required
@require_GET
def api_bitacora_hoy(request):
    return JsonResponse(construir_bitacora_usuario(request.user))


@login_required
@require_POST
def api_perfil_color(request):
    user = request.user
    if not (es_asesor(user) and not es_tecnico(user) and not user.is_superuser):
        return JsonResponse({'ok': False, 'error': 'No autorizado.'}, status=403)

    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except json.JSONDecodeError:
        payload = {}

    color = (payload.get('color') or '').strip()
    if color not in COLORES_PERFIL_ASESOR:
        return JsonResponse({'ok': False, 'error': 'Color no permitido.'}, status=400)

    actividad, _ = UsuarioActividad.objects.get_or_create(user=user)
    actividad.perfil_color_asesor = color
    actividad.save(update_fields=['perfil_color_asesor'])
    return JsonResponse({'ok': True, 'color': color})
